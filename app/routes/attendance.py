from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from app.utils import permission_required, log_activity
from flask_login import login_required, current_user
from app import db
from app.models import Staff, Attendance
from datetime import datetime, timedelta
from sqlalchemy import func
from app.routes.filters import apply_saved_filter_to_query

bp = Blueprint('attendance', __name__, url_prefix='/attendance')

# --- Attendance Dashboard ---

@bp.route('/')
@login_required
def index():
    """Attendance dashboard with date filtering"""
    # Get date filter from request
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    staff_id_filter = request.args.get('staff_id')
    
    # Set default date range (current month)
    if not date_from_str:
        today = datetime.now()
        date_from = today.replace(day=1)
    else:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
    
    if not date_to_str:
        date_to = datetime.now()
    else:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
    
    # Get attendance records
    query = Attendance.query.filter(
        Attendance.date >= date_from.date(),
        Attendance.date <= date_to.date()
    )
    
    if staff_id_filter:
        query = query.filter(Attendance.staff_id == staff_id_filter)
        
    query = apply_saved_filter_to_query(query, 'attendance', request.args)
    
    attendance_records = query.order_by(Attendance.date.desc(), Attendance.staff_id).all()
    
    # Calculate summary
    total_hours = 0
    total_minutes = 0
    total_earned = 0
    
    for record in attendance_records:
        # Refresh earnings in-memory (no commit) so records saved before overtime
        # existed — or with a stale stored value — display the overtime-inclusive
        # amount. No-op for records without overtime.
        if record.clock_out and record.overtime_total_hours > 0:
            record.calculate_earned_amount()

        if record.clock_in and not record.clock_out:
            # Active shift (Live estimate)
            diff = datetime.now() - record.clock_in
            cur_total_mins = int(diff.total_seconds() / 60)
            
            if record.used_break and cur_total_mins >= 240:
                cur_total_mins -= 60
            
            d_hours = getattr(record, 'deduct_hours', 0) or 0
            d_mins = getattr(record, 'deduct_minutes', 0) or 0
            cur_total_mins -= int(d_hours * 60 + d_mins)
            
            if cur_total_mins < 0: cur_total_mins = 0
            
            total_hours += cur_total_mins // 60
            total_minutes += cur_total_mins % 60
            
            if not record.hourly_rate:
                record.calculate_hourly_rate()
            total_earned += (cur_total_mins / 60.0) * record.hourly_rate
            # Live shift: earned_amount isn't used above, so add overtime pay here.
            total_earned += record.overtime_earned
            total_minutes += int(round(record.overtime_total_hours * 60))
        else:
            # Completed shift — earned_amount already includes overtime pay.
            total_hours += record.hours_worked
            total_minutes += record.minutes_worked
            total_minutes += int(round(record.overtime_total_hours * 60))
            total_earned += record.earned_amount
    
    # Convert minutes to hours
    total_hours += total_minutes // 60
    total_minutes = total_minutes % 60

    # Computed overtime: for each staff shown, how far their actual worked
    # hours in this filtered range exceed the required hours for that range
    # (working days - Sundays - their holidays, times 8h/day). Zero if not over.
    from app.utils import get_required_hours_in_range
    from collections import defaultdict

    records_by_staff = defaultdict(list)
    for record in attendance_records:
        records_by_staff[record.staff_id].append(record)

    total_overtime_hours = 0.0
    total_overtime_amount = 0.0
    total_required_hours = 0.0
    for staff_id, records in records_by_staff.items():
        staff_member = records[0].staff
        actual_hours = sum((r.hours_worked or 0) + (r.minutes_worked or 0) / 60.0 for r in records)
        required_hours = get_required_hours_in_range(staff_member, date_from.date(), date_to.date())
        total_required_hours += required_hours
        overtime_hours = max(0.0, actual_hours - required_hours)
        total_overtime_hours += overtime_hours
        if overtime_hours > 0:
            hourly_rate = (staff_member.daily_salary or 0) / 8.0
            total_overtime_amount += overtime_hours * hourly_rate

    # Get all staff for filter dropdown
    all_staff = Staff.query.filter_by(is_active=True).all()
    
    # Get today's attendance for quick status/prefill
    today = datetime.now().date()
    today_records = Attendance.query.filter_by(date=today).all()
    today_attendance = {a.staff_id: a for a in today_records}
    
    return render_template('salary/attendance_list.html',
                         attendance_records=attendance_records,
                         all_staff=all_staff,
                         today_attendance=today_attendance,
                         date_from=date_from_str or date_from.strftime('%Y-%m-%d'),
                         date_to=date_to_str or date_to.strftime('%Y-%m-%d'),
                         total_hours=total_hours,
                         total_minutes=total_minutes,
                         total_earned=total_earned,
                         total_overtime_hours=total_overtime_hours,
                         total_overtime_amount=total_overtime_amount,
                         total_required_hours=total_required_hours,
                         now=datetime.now(),
                         timedelta=timedelta,
                         active_module='attendance')

# --- Clock In/Out ---

@bp.route('/clock-in/<int:staff_id>', methods=['POST'])
@login_required
@permission_required('attendance', action='add')
def clock_in(staff_id):
    """Staff clock in or update status"""
    staff = Staff.query.get_or_404(staff_id)
    today = datetime.now().date()
    
    used_break = request.form.get('used_break') == 'true'
    deduct_hours = float(request.form.get('deduct_hours') or 0)
    deduct_minutes = int(request.form.get('deduct_minutes') or 0)
    deduct_reason = request.form.get('deduct_reason')
    notes = request.form.get('notes')
    
    # Check if already clocked in today
    existing = Attendance.query.filter_by(
        staff_id=staff_id,
        date=today
    ).first()
    
    if existing and existing.clock_in and not existing.clock_out:
        # Update existing record (Update deductions/notes during shift)
        existing.used_break = used_break
        existing.deduct_hours = deduct_hours
        existing.deduct_minutes = deduct_minutes
        existing.deduct_reason = deduct_reason
        existing.notes = notes
        db.session.commit()
        log_activity('Attendance', f'Updated shift details for {staff.name}', f'Date: {today}')
        flash(f'{staff.name}\'s deduction/notes updated.', 'success')
        return redirect(url_for('attendance.index'))
    
    if not existing:
        # Create new attendance record
        attendance = Attendance(
            staff_id=staff_id,
            date=today,
            clock_in=datetime.now(),
            notes=notes
        )
        attendance.used_break = used_break
        attendance.deduct_hours = deduct_hours
        attendance.deduct_minutes = deduct_minutes
        attendance.deduct_reason = deduct_reason
        attendance.calculate_hourly_rate()
        db.session.add(attendance)
    else:
        # Update existing record (if they clocked out earlier and re-clocking in)
        existing.clock_in = datetime.now()
        existing.used_break = used_break
        existing.deduct_hours = deduct_hours
        existing.deduct_minutes = deduct_minutes
        existing.deduct_reason = deduct_reason
        existing.notes = notes
    
    db.session.commit()
    log_activity('Attendance', f'Clocked In: {staff.name}', f'Date: {today}')
    flash(f'{staff.name} clocked in at {datetime.now().strftime("%H:%M:%S")}', 'success')
    return redirect(url_for('attendance.index'))

@bp.route('/clock-out/<int:staff_id>', methods=['POST'])
@login_required
@permission_required('attendance', action='edit')
def clock_out(staff_id):
    """Staff clock out"""
    staff = Staff.query.get_or_404(staff_id)
    today = datetime.now().date()
    
    # Find today's attendance record
    attendance = Attendance.query.filter_by(
        staff_id=staff_id,
        date=today
    ).first()
    
    if not attendance:
        flash(f'No clock in record found for {staff.name} today!', 'danger')
        return redirect(url_for('attendance.index'))
    
    if not attendance.clock_in:
        flash(f'{staff.name} has not clocked in yet!', 'warning')
        return redirect(url_for('attendance.index'))
    
    if attendance.clock_out:
        flash(f'{staff.name} has already clocked out!', 'info')
        return redirect(url_for('attendance.index'))
    
    # Update deductions/notes at clock out
    attendance.used_break = request.form.get('used_break') == 'true'
    attendance.deduct_hours = float(request.form.get('deduct_hours') or 0)
    attendance.deduct_minutes = int(request.form.get('deduct_minutes') or 0)
    attendance.deduct_reason = request.form.get('deduct_reason')
    attendance.notes = request.form.get('notes') or attendance.notes
    
    # Set clock out time and calculate
    attendance.clock_out = datetime.now()
    attendance.calculate_hours_worked()
    attendance.calculate_earned_amount()
    
    db.session.commit()
    log_activity('Attendance', f'Clocked Out: {staff.name}', f'Date: {today}, Hours: {attendance.get_time_summary()}')
    flash(f'{staff.name} clocked out at {attendance.clock_out.strftime("%H:%M:%S")}. Worked: {attendance.get_time_summary()}', 'success')
    return redirect(url_for('attendance.index'))

@bp.route('/mark-holiday/<int:staff_id>', methods=['POST'])
@login_required
@permission_required('attendance', action='add')
def mark_holiday(staff_id):
    """Mark staff attendance as holiday for a specific date"""
    staff = Staff.query.get_or_404(staff_id)
    
    # 1. Get date from form or default to today
    date_str = request.form.get('holiday_date')
    if date_str:
        try:
            holiday_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            holiday_date = datetime.now().date()
    else:
        holiday_date = datetime.now().date()
    
    # 2. Find or create record for that date
    attendance = Attendance.query.filter_by(
        staff_id=staff_id,
        date=holiday_date
    ).first()
    
    if not attendance:
        attendance = Attendance(
            staff_id=staff_id,
            date=holiday_date,
            notes="Sudden holiday declared"
        )
        db.session.add(attendance)
    
    # 3. Mark as holiday and clear clocks
    attendance.staff = staff
    attendance.is_holiday = True
    attendance.clock_in = None
    attendance.clock_out = None
    attendance.calculate_hours_worked() # Now sets to 0h
    attendance.calculate_earned_amount() # Now sets to 0 PKR
    
    db.session.flush() # Ensure this record is counted in subsequent calculations
    
    # 4. CRITICAL: Recalculate ALL attendance records for this staff in this month
    # because the new holiday reduces working days and thus changes the daily/hourly rate.
    from sqlalchemy import extract
    all_month_records = Attendance.query.filter(
        Attendance.staff_id == staff_id,
        extract('year', Attendance.date) == holiday_date.year,
        extract('month', Attendance.date) == holiday_date.month
    ).all()
    
    for record in all_month_records:
        record.calculate_hourly_rate() # This will call staff.calculate_daily_salary(record.date)
        record.calculate_earned_amount()
    
    db.session.commit()
    log_activity('Attendance', f'Marked Holiday: {staff.name}', f'Date: {holiday_date}')
    flash(f'Marked {holiday_date} as holiday for {staff.name}. Monthly working days and rates recalculated.', 'success')
    return redirect(url_for('attendance.index'))

# --- Absent / No-Show Logic ---

@bp.route('/mark-absent/<int:staff_id>', methods=['POST'])
@login_required
@permission_required('attendance', action='add')
def mark_absent(staff_id):
    """Manually mark a staff member as absent for a given date"""
    staff = Staff.query.get_or_404(staff_id)

    date_str = request.form.get('absent_date')
    if date_str:
        try:
            absent_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            absent_date = datetime.now().date()
    else:
        absent_date = datetime.now().date()

    # Prevent marking today before 5pm (shift end)
    now = datetime.now()
    if absent_date == now.date() and now.hour < 17:
        flash(f'Cannot mark absent before shift ends (5:00 PM). Current time: {now.strftime("%H:%M")}.', 'warning')
        return redirect(url_for('attendance.index'))

    # Don't mark absent on Sundays
    if absent_date.weekday() == 6:
        flash('Cannot mark absent on Sunday (official day off).', 'warning')
        return redirect(url_for('attendance.index'))

    existing = Attendance.query.filter_by(staff_id=staff_id, date=absent_date).first()

    if existing and existing.clock_in:
        flash(f'{staff.name} has a clock-in record for {absent_date}. Cannot mark absent.', 'danger')
        return redirect(url_for('attendance.index'))

    if not existing:
        existing = Attendance(staff_id=staff_id, date=absent_date, notes='Absent - no clock-in')
        db.session.add(existing)

    existing.staff = staff
    existing.is_absent = True
    existing.is_holiday = False
    existing.clock_in = None
    existing.clock_out = None
    existing.hours_worked = 0
    existing.minutes_worked = 0
    existing.earned_amount = 0
    existing.calculate_hourly_rate()

    db.session.commit()
    log_activity('Attendance', f'Marked Absent: {staff.name}', f'Date: {absent_date}')
    flash(f'{staff.name} marked as absent on {absent_date}.', 'warning')
    return redirect(url_for('attendance.index'))


@bp.route('/process-absences', methods=['POST'])
@login_required
@permission_required('attendance', action='add')
def process_absences():
    """
    Auto-detect and mark absent any active staff member who has no attendance
    record for the previous working day (Mon-Sat, after 5:00 PM shift end).
    Can also be triggered manually for a specific date.
    """
    date_from_str = request.form.get('process_date_from')
    date_to_str = request.form.get('process_date_to')
    
    # Fallback to single process_date if provided
    process_date_single = request.form.get('process_date')
    if process_date_single and not date_from_str:
        date_from_str = process_date_single
        date_to_str = process_date_single

    if not date_from_str:
        yesterday = (datetime.now() - timedelta(days=1)).date()
        date_from = yesterday
        date_to = yesterday
    else:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else date_from

    all_staff = Staff.query.filter_by(is_active=True).all()
    marked_count = 0
    current_check = date_from

    while current_check <= date_to:
        # Skip Sundays
        if current_check.weekday() == 6:
            current_check += timedelta(days=1)
            continue

        for staff in all_staff:
            existing = Attendance.query.filter_by(staff_id=staff.id, date=current_check).first()

            # Already has a record with clock-in, holiday, or already absent → skip
            if existing and (existing.clock_in or existing.is_holiday or existing.is_absent):
                continue

            if not existing:
                existing = Attendance(staff_id=staff.id, date=current_check, notes='Auto-absent: no clock-in recorded')
                db.session.add(existing)

            existing.staff = staff
            existing.is_absent = True
            existing.is_holiday = False
            existing.clock_in = None
            existing.clock_out = None
            existing.hours_worked = 0
            existing.minutes_worked = 0
            existing.earned_amount = 0
            existing.calculate_hourly_rate()
            marked_count += 1

        current_check += timedelta(days=1)

    db.session.commit()
    log_activity('Attendance', f'Bulk processed absences', f'Range: {date_from} to {date_to}, Marked: {marked_count}')
    flash(f'Processed absences from {date_from} to {date_to}: {marked_count} total staff-days marked absent.', 'success')
    return redirect(url_for('attendance.index'))


# --- Attendance Management ---

@bp.route('/record/<int:attendance_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('attendance', action='edit')
def edit_attendance(attendance_id):
    """Edit attendance record (for manual corrections)"""
    attendance = Attendance.query.get_or_404(attendance_id)
    
    if request.method == 'POST':
        # Get values from form
        date_str = request.form.get('date')
        clock_in_str = request.form.get('clock_in')
        clock_out_str = request.form.get('clock_out')
        notes = request.form.get('notes')
        
        try:
            # Remember holiday state before edit so we know if the working-day divisor changes
            was_holiday = attendance.is_holiday

            # 1. Update the main date field
            if date_str:
                attendance.date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            # 2. Update clock in/out (ensuring they use the correct date)
            # Use provided date or keep existing date
            current_date_str = date_str or str(attendance.date)
            
            if clock_in_str:
                attendance.clock_in = datetime.strptime(f"{current_date_str} {clock_in_str}", '%Y-%m-%d %H:%M')
            else:
                attendance.clock_in = None
            
            if clock_out_str:
                attendance.clock_out = datetime.strptime(f"{current_date_str} {clock_out_str}", '%Y-%m-%d %H:%M')
            else:
                attendance.clock_out = None
            
            # 3. Update holiday/absent status, break, deductions, notes and recalculate
            is_holiday = True if request.form.get('is_holiday') else False
            is_absent = True if request.form.get('is_absent') else False

            # Mutual exclusivity: if marking holiday OR absent, clear the other
            if is_holiday:
                is_absent = False
            elif is_absent:
                is_holiday = False
                attendance.clock_in = None
                attendance.clock_out = None

            attendance.is_holiday = is_holiday
            attendance.is_absent = is_absent
            attendance.used_break = True if request.form.get('used_break') else False
            attendance.deduct_hours = float(request.form.get('deduct_hours') or 0)
            attendance.deduct_minutes = int(request.form.get('deduct_minutes') or 0)
            attendance.deduct_reason = request.form.get('deduct_reason')
            attendance.notes = notes
            
            # Recalculate based on new state
            attendance.calculate_hours_worked()
            attendance.calculate_earned_amount() # This also recalculates hourly rate based on the date

            # CRITICAL: If the holiday flag changed, the working-day divisor for this staff's
            # whole month changed too. Recalculate every other record in that month so they
            # don't stay stuck on the pre-edit rate (same fix already applied in mark_holiday()).
            if is_holiday != was_holiday:
                from sqlalchemy import extract
                all_month_records = Attendance.query.filter(
                    Attendance.staff_id == attendance.staff_id,
                    extract('year', Attendance.date) == attendance.date.year,
                    extract('month', Attendance.date) == attendance.date.month,
                    Attendance.id != attendance.id
                ).all()

                for record in all_month_records:
                    record.calculate_hourly_rate()
                    record.calculate_earned_amount()

            db.session.commit()
            log_activity('Attendance', f'Updated Attendance: {attendance.staff.name}', f'Date: {attendance.date}')
            flash('Attendance record updated!', 'success')
            return redirect(url_for('attendance.index'))
        except ValueError as e:
            flash(f'Invalid format: {e}', 'danger')
    
    return render_template('salary/edit_attendance.html', attendance=attendance)

@bp.route('/record/<int:attendance_id>/delete', methods=['POST'])
@login_required
@permission_required('attendance', action='delete')
def delete_attendance(attendance_id):
    """Delete attendance record"""
    attendance = Attendance.query.get_or_404(attendance_id)
    staff_name = attendance.staff.name
    att_date = attendance.date
    db.session.delete(attendance)
    db.session.commit()
    log_activity('Attendance', f'Deleted Attendance: {staff_name}', f'Date: {att_date}')
    flash(f'Attendance record for {staff_name} deleted.', 'info')
    return redirect(url_for('attendance.index'))

# --- API Endpoints for AJAX ---

@bp.route('/api/quick-clock/<int:staff_id>/<action>', methods=['POST'])
@login_required
@permission_required('attendance', action='edit')
def quick_clock(staff_id, action):
    """Quick clock in/out via API"""
    try:
        if action == 'in':
            result = clock_in(staff_id)
        elif action == 'out':
            result = clock_out(staff_id)
        else:
            return jsonify({'status': 'error', 'message': 'Invalid action'}), 400
        
        return jsonify({'status': 'success', 'message': f'Action: {action}'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

@bp.route('/api/staff/<int:staff_id>/today-status')
@login_required
def staff_today_status(staff_id):
    """Get staff's today status"""
    today = datetime.now().date()
    attendance = Attendance.query.filter_by(
        staff_id=staff_id,
        date=today
    ).first()
    
    status = {
        'staff_id': staff_id,
        'date': str(today),
        'clocked_in': False,
        'clocked_out': False,
        'time_summary': '0h 0m',
        'earned': 0
    }
    
    if attendance:
        status['clocked_in'] = attendance.clock_in is not None
        status['clocked_out'] = attendance.clock_out is not None
        status['time_summary'] = attendance.get_time_summary()
        status['earned'] = round(attendance.earned_amount, 2)
    
    return jsonify(status)

def parse_time(time_val, att_date):
    """Robustly parse time value from Excel into a datetime object."""
    from datetime import datetime, time as datetime_time
    if not time_val:
        return None
        
    if isinstance(time_val, datetime):
        # Already a datetime, just ensure it's on the correct date
        return datetime.combine(att_date, time_val.time())
    
    if isinstance(time_val, datetime_time):
        # It's a time object, combine with the attendance date
        return datetime.combine(att_date, time_val)
        
    # Handle string formats
    time_str = str(time_val).strip()
    formats = ['%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M:%S %p']
    
    for fmt in formats:
        try:
            parsed_t = datetime.strptime(time_str, fmt).time()
            return datetime.combine(att_date, parsed_t)
        except ValueError:
            continue
            
    raise ValueError(f"Invalid time format: {time_str}")

@bp.route('/bulk-upload', methods=['GET', 'POST'])
@login_required
@permission_required('attendance', action='add')
def bulk_upload():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(url_for('attendance.bulk_upload'))
            
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('attendance.bulk_upload'))
            
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('attendance.bulk_upload'))
            
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            from datetime import datetime
            file_content = file.read()
            wb = load_workbook(filename=BytesIO(file_content), data_only=True, read_only=True)
            ws = wb.active
            rows = list(ws.values)
            if not rows or len(rows) < 2:
                flash('File is empty or contains no data rows', 'error')
                return redirect(url_for('attendance.bulk_upload'))
                
            headers = [str(h).strip().lower() if h else '' for h in rows[0]]
            
            required_columns = ['staff_name', 'date', 'clock_in_time']
            missing = [col for col in required_columns if col not in headers]
            if missing:
                flash(f'Missing required columns: {", ".join(missing)}', 'error')
                return redirect(url_for('attendance.bulk_upload'))
                
            added = 0
            errors = []
            
            for idx, row in enumerate(rows[1:], start=2):
                # Skip truly empty rows
                if not any(row):
                    continue
                    
                try:
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = val
                            
                    staff_name = str(row_dict.get('staff_name', '')).strip()
                    if not staff_name:
                        errors.append(f'Row {idx}: Missing staff_name')
                        continue
                        
                    staff = Staff.query.filter_by(name=staff_name).first()
                    if not staff:
                        errors.append(f'Row {idx}: Staff "{staff_name}" not found')
                        continue

                        
                    date_val = row_dict.get('date')
                    if isinstance(date_val, datetime):
                        att_date = date_val.date()
                    elif date_val:
                        try:
                            # Split in case of "2024-05-01 00:00:00" strings
                            date_str = str(date_val).split()[0]
                            att_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except ValueError:
                            errors.append(f'Row {idx}: Invalid date format (expected YYYY-MM-DD)')
                            continue
                    else:
                        errors.append(f'Row {idx}: Missing date')
                        continue
                        
                    # Check for duplicate entry in same file/db
                    existing = Attendance.query.filter_by(staff_id=staff.id, date=att_date).first()
                    if existing:
                        errors.append(f'Row {idx}: Attendance for staff {staff.name} on {att_date} already exists')
                        continue
                        
                    try:
                        clock_in = parse_time(row_dict.get('clock_in_time'), att_date)
                        if not clock_in:
                            errors.append(f'Row {idx}: Missing clock_in_time')
                            continue
                            
                        clock_out = parse_time(row_dict.get('clock_out_time'), att_date)
                    except ValueError as ve:
                        errors.append(f'Row {idx}: {str(ve)}')
                        continue
                        
                    attendance = Attendance(
                        staff_id=staff.id,
                        date=att_date,
                        clock_in=clock_in,
                        clock_out=clock_out,
                        notes=str(row_dict.get('notes', '')).strip() if row_dict.get('notes') else None
                    )
                    
                    if clock_out:
                        attendance.calculate_hours_worked()
                        attendance.calculate_earned_amount()
                    else:
                        attendance.calculate_hourly_rate()
                        
                    db.session.add(attendance)
                    added += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
            
            if added > 0:
                db.session.commit()
                log_activity('Attendance', f'Bulk uploaded attendance', f'Records added: {added}')
                flash(f'Successfully added {added} attendance records!', 'success')
            
            if errors:
                flash(f'Encountered {len(errors)} issues. First 5 shown: {"; ".join(errors[:5])}', 'warning')
                
            if added == 0 and not errors:
                flash('No data entries found in file.', 'info')
                
            return redirect(url_for('attendance.index'))
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'error')
            return redirect(url_for('attendance.bulk_upload'))
            
    return render_template('salary/attendance_bulk_upload.html')


@bp.route('/download-sample')
@login_required
def download_sample():
    try:
        from openpyxl import Workbook
        from io import BytesIO
        from flask import send_file
        from datetime import datetime, date
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Attendance'
        
        headers = ['staff_name', 'date', 'clock_in_time', 'clock_out_time', 'notes']
        ws.append(headers)
        
        staff = Staff.query.first()
        staff_name = staff.name if staff else "John Doe"
        
        today = datetime.now().date()
        current_date_str = today.strftime('%Y-%m-%d')
        
        sample_data = [
            [staff_name, current_date_str, '09:00:00', '17:00:00', 'Regular day'],
            [staff_name, current_date_str, '09:15:00', '18:30:00', 'Overtime']
        ]
        
        for row in sample_data:
            ws.append(row)

            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='sample_attendance.xlsx', as_attachment=True)

    except Exception as e:
        flash(f'Error creating sample: {str(e)}', 'error')
        return redirect(url_for('attendance.bulk_upload'))
