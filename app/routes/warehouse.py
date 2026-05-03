from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from app.utils import permission_required
from flask_login import login_required, current_user
from app import db
from app.models import Warehouse, Product
from sqlalchemy import func
from app.routes.filters import apply_saved_filter_to_query

bp = Blueprint('warehouse', __name__)

@bp.route('/warehouses')
@login_required
def warehouses():
    search = request.args.get('search', '')
    query = Warehouse.query
    if search:
        query = query.filter(
            (Warehouse.name.ilike(f'%{search}%')) | 
            (Warehouse.code.ilike(f'%{search}%'))
        )
    
    query = apply_saved_filter_to_query(query, 'warehouse', request.args)
    warehouses = query.order_by(Warehouse.name).all()
    return render_template('warehouse/warehouses.html', warehouses=warehouses, active_module='warehouse')

@bp.route('/warehouse/add', methods=['GET', 'POST'])
@login_required
@permission_required('warehouse', action='add')
def add_warehouse():
    if request.method == 'POST':
        name = request.form.get('name')
        code = request.form.get('code')
        address = request.form.get('address')
        city = request.form.get('city')
        phone = request.form.get('phone')
        email = request.form.get('email')
        manager = request.form.get('manager')
        
        existing = Warehouse.query.filter_by(code=code).first()
        if existing:
            flash(f'Warehouse code "{code}" already exists.', 'error')
            return redirect(url_for('warehouse.add_warehouse'))
        
        warehouse = Warehouse(
            name=name,
            code=code,
            address=address,
            city=city,
            phone=phone,
            email=email,
            manager=manager
        )
        try:
            db.session.add(warehouse)
            db.session.commit()
            flash(f'Warehouse "{name}" created successfully!', 'success')
            return redirect(url_for('warehouse.warehouses'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating warehouse: {str(e)}', 'error')
    
    return render_template('warehouse/add_warehouse.html')

@bp.route('/warehouse/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('warehouse', action='edit')
def edit_warehouse(id):
    warehouse = Warehouse.query.get_or_404(id)
    
    if request.method == 'POST':
        name = request.form.get('name')
        code = request.form.get('code')
        address = request.form.get('address')
        city = request.form.get('city')
        phone = request.form.get('phone')
        email = request.form.get('email')
        manager = request.form.get('manager')
        is_active = request.form.get('is_active') == 'on'
        
        existing = Warehouse.query.filter(Warehouse.code == code, Warehouse.id != id).first()
        if existing:
            flash(f'Warehouse code "{code}" already exists.', 'error')
            return redirect(url_for('warehouse.edit_warehouse', id=id))
        
        warehouse.name = name
        warehouse.code = code
        warehouse.address = address
        warehouse.city = city
        warehouse.phone = phone
        warehouse.email = email
        warehouse.manager = manager
        warehouse.is_active = is_active
        
        try:
            db.session.commit()
            flash('Warehouse updated successfully!', 'success')
            return redirect(url_for('warehouse.warehouses'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating warehouse: {str(e)}', 'error')
    
    return render_template('warehouse/edit_warehouse.html', warehouse=warehouse)

@bp.route('/warehouse/<int:id>')
@login_required
def warehouse_detail(id):
    warehouse = Warehouse.query.get_or_404(id)
    products = Product.query.filter_by(warehouse_id=id).order_by(Product.name).all()
    
    total_quantity = sum(p.quantity for p in products)
    total_value = sum(p.quantity * p.cost_price for p in products)
    total_items = len(products)
    
    return render_template('warehouse/warehouse_detail.html', 
                         warehouse=warehouse, 
                         products=products,
                         total_quantity=total_quantity,
                         total_value=total_value,
                         total_items=total_items)

@bp.route('/warehouse/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('warehouse', action='delete')
def delete_warehouse(id):
    warehouse = Warehouse.query.get_or_404(id)
    
    if warehouse.products:
        flash('Cannot delete warehouse with associated products. Remove or reassign products first.', 'error')
        return redirect(url_for('warehouse.warehouse_detail', id=id))
    
    try:
        db.session.delete(warehouse)
        db.session.commit()
        flash('Warehouse deleted successfully!', 'success')
        return redirect(url_for('warehouse.warehouses'))
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting warehouse: {str(e)}', 'error')
        return redirect(url_for('warehouse.warehouse_detail', id=id))

@bp.route('/warehouse/bulk-delete', methods=['POST'])
@login_required
@permission_required('warehouse', action='delete')
def bulk_delete_warehouses():
    data = request.get_json()
    if not data or 'ids' not in data:
        return jsonify({'success': False, 'message': 'No warehouses selected'}), 400
        
    ids = data['ids']
    try:
        # Check if any selected warehouse has products
        warehouses_with_products = Warehouse.query.filter(
            Warehouse.id.in_(ids),
            Warehouse.products.any()
        ).count()
        
        if warehouses_with_products > 0:
            return jsonify({
                'success': False, 
                'message': f'Cannot delete {warehouses_with_products} warehouse(s) because they have associated products.'
            }), 400
            
        Warehouse.query.filter(Warehouse.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'Successfully deleted {len(ids)} warehouses.'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@bp.route('/api/warehouses')
@login_required
def api_warehouses():
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    return jsonify([{'id': w.id, 'name': w.name, 'code': w.code} for w in warehouses])

@bp.route('/api/warehouses/<int:id>/summary')
@login_required
def api_warehouse_summary(id):
    warehouse = Warehouse.query.get_or_404(id)
    products = Product.query.filter_by(warehouse_id=id).all()
    
    return jsonify({
        'id': warehouse.id,
        'name': warehouse.name,
        'code': warehouse.code,
        'total_products': len(products),
        'total_quantity': sum(p.quantity for p in products),
        'total_value': sum(p.quantity * p.cost_price for p in products)
    })

@bp.route('/bulk-upload', methods=['GET', 'POST'])
@login_required
@permission_required('warehouse', action='add')
def bulk_upload():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(url_for('warehouse.bulk_upload'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('warehouse.bulk_upload'))
            
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('warehouse.bulk_upload'))
            
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            file_content = file.read()
            wb = load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            rows = list(ws.values)
            if not rows:
                flash('File is empty', 'error')
                return redirect(url_for('warehouse.bulk_upload'))
            headers = [str(h).strip().lower() if h else '' for h in rows[0]]
            
            required_columns = ['name', 'code']
            missing = [col for col in required_columns if col not in headers]
            if missing:
                flash(f'Missing required columns: {", ".join(missing)}', 'error')
                return redirect(url_for('warehouse.bulk_upload'))
                
            added = 0
            errors = []
            
            for idx, row in enumerate(rows[1:], start=2):
                try:
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = val
                            
                    name = str(row_dict.get('name', '')).strip()
                    code = str(row_dict.get('code', '')).strip()
                    
                    if not name or not code:
                        errors.append(f'Row {idx}: Missing name or code')
                        continue
                        
                    existing = Warehouse.query.filter_by(code=code).first()
                    if existing:
                        errors.append(f'Row {idx}: Code "{code}" already exists')
                        continue
                        
                    warehouse = Warehouse(
                        name=name,
                        code=code,
                        city=str(row_dict.get('city', '')).strip() if row_dict.get('city') else None,
                        address=str(row_dict.get('address', '')).strip() if row_dict.get('address') else None,
                        is_active=bool(str(row_dict.get('is_active', '1')).strip().lower() in ['yes', '1', 'true', 't', 'active'])
                    )
                    db.session.add(warehouse)
                    added += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
            
            db.session.commit()
            if added > 0:
                flash(f'Successfully added {added} warehouses!', 'success')
            if errors:
                flash(f'Errors: {"; ".join(errors[:10])}', 'warning')
            return redirect(url_for('warehouse.warehouses'))
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'error')
            return redirect(url_for('warehouse.bulk_upload'))
            
    return render_template('warehouse/bulk_upload.html')

@bp.route('/download-sample')
@login_required
def download_sample():
    try:
        from openpyxl import Workbook
        from io import BytesIO
        from flask import send_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Warehouses'
        
        headers = ['name', 'code', 'city', 'address', 'is_active']
        ws.append(headers)
        
        sample_data = [
            ['Main Warehouse', 'WH-101', 'New York', '123 Main St', 'yes'],
            ['Secondary Hub', 'WH-102', 'Los Angeles', '456 West Blvd', 'yes']
        ]
        
        for row in sample_data:
            ws.append(row)
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='sample_warehouses.xlsx', as_attachment=True)
    except Exception as e:
        flash(f'Error creating sample: {str(e)}', 'error')
        return redirect(url_for('warehouse.bulk_upload'))