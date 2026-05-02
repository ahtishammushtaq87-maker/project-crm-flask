from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from app.utils import permission_required
from flask_login import login_required, current_user
from app import db
from app.models import Staff, Attendance
from datetime import datetime, timedelta
from sqlalchemy import func
import csv
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

bp = Blueprint('reports_attendance', __name__, url_prefix='/reports')

# --- Attendance Report Dashboard ---

@bp.route('/attendance/')
@login_required
def attendance_report():
    """Attendance report with comprehensive filtering and export options"""
    
    # Get filter parameters
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    staff_id_filter = request.args.get('staff_id')
    department_filter = request.args.get('department')
    report_type = request.args.get('report_type', 'summary')  # summary, detailed, daily
    
    # Set default date range (current month)
    if not date_from_str:
        today = datetime.now()
        date_from = today.replace(day=1)
    else:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
    
    if not date_to_str:
        date_to = datetime.now()
    else:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
    
    # Get attendance records
    query = Attendance.query.filter(
        Attendance.date >= date_from.date(),
        Attendance.date <= date_to.date()
    )
    
    if staff_id_filter:
        query = query.filter(Attendance.staff_id == staff_id_filter)
    
    attendance_records = query.order_by(Attendance.date.desc(), Attendance.staff_id).all()
    
    # Calculate summary statistics
    summary_stats = calculate_summary_stats(attendance_records)
    
    # Get all staff for filters
    all_staff = Staff.query.filter_by(is_active=True).all()
    departments = db.session.query(Staff.designation).distinct().all()
    departments = [d[0] for d in departments if d[0]]
    
    # Get report data based on type
    if report_type == 'detailed':
        report_data = get_detailed_report_data(attendance_records, date_from, date_to)
    elif report_type == 'daily':
        report_data = get_daily_report_data(attendance_records)
    else:  # summary
        report_data = get_summary_report_data(attendance_records, all_staff)
    
    return render_template('reports/attendance_report.html',
                         attendance_records=attendance_records,
                         all_staff=all_staff,
                         departments=departments,
                         date_from=date_from_str or date_from.strftime('%Y-%m-%d'),
                         date_to=date_to_str or date_to.strftime('%Y-%m-%d'),
                         staff_id_filter=staff_id_filter,
                         department_filter=department_filter,
                         report_type=report_type,
                         summary_stats=summary_stats,
                         report_data=report_data)


# --- Helper Functions for Report Calculations ---

def calculate_summary_stats(records):
    """Calculate summary statistics from attendance records"""
    total_hours = 0
    total_minutes = 0
    total_earned = 0
    total_records = len(records)
    
    for record in records:
        total_hours += record.hours_worked
        total_minutes += record.minutes_worked
        total_earned += record.earned_amount
    
    # Convert minutes to hours
    total_hours += total_minutes // 60
    total_minutes = total_minutes % 60
    
    # Calculate averages
    avg_hours = total_hours / total_records if total_records > 0 else 0
    avg_earned = total_earned / total_records if total_records > 0 else 0
    
    return {
        'total_hours': total_hours,
        'total_minutes': total_minutes,
        'total_earned': total_earned,
        'total_records': total_records,
        'avg_hours': avg_hours,
        'avg_earned': avg_earned,
        'days_worked': len(set(r.date for r in records))
    }


def get_summary_report_data(records, all_staff):
    """Get summary report data grouped by staff"""
    staff_summary = {}
    
    for staff in all_staff:
        staff_records = [r for r in records if r.staff_id == staff.id]
        if staff_records:
            total_hours = sum(r.hours_worked for r in staff_records)
            total_minutes = sum(r.minutes_worked for r in staff_records)
            total_earned = sum(r.earned_amount for r in staff_records)
            
            # Convert minutes
            total_hours += total_minutes // 60
            total_minutes = total_minutes % 60
            
            staff_summary[staff.id] = {
                'staff': staff,
                'total_hours': total_hours,
                'total_minutes': total_minutes,
                'total_earned': total_earned,
                'days_worked': len(set(r.date for r in staff_records)),
                'records_count': len(staff_records),
                'avg_hours': (total_hours + total_minutes/60) / len(staff_records) if staff_records else 0
            }
    
    return staff_summary


def get_detailed_report_data(records, date_from, date_to):
    """Get detailed report data for each day"""
    detailed_data = {}
    
    current_date = date_from.date()
    while current_date <= date_to.date():
        daily_records = [r for r in records if r.date == current_date]
        if daily_records:
            total_hours = sum(r.hours_worked for r in daily_records)
            total_minutes = sum(r.minutes_worked for r in daily_records)
            total_earned = sum(r.earned_amount for r in daily_records)
            
            detailed_data[current_date] = {
                'records': daily_records,
                'total_hours': total_hours,
                'total_minutes': total_minutes,
                'total_earned': total_earned,
                'staff_count': len(set(r.staff_id for r in daily_records))
            }
        
        current_date += timedelta(days=1)
    
    return detailed_data


def get_daily_report_data(records):
    """Get daily report data (each record on separate row)"""
    return records


# --- Export to CSV ---

@bp.route('/attendance/export/csv')
@login_required
def export_attendance_csv():
    """Export attendance report to CSV"""
    
    # Get filter parameters
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    staff_id_filter = request.args.get('staff_id')
    report_type = request.args.get('report_type', 'summary')
    
    # Set default date range
    if not date_from_str:
        today = datetime.now()
        date_from = today.replace(day=1)
    else:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
    
    if not date_to_str:
        date_to = datetime.now()
    else:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
    
    # Get attendance records
    query = Attendance.query.filter(
        Attendance.date >= date_from.date(),
        Attendance.date <= date_to.date()
    )
    
    if staff_id_filter:
        query = query.filter(Attendance.staff_id == staff_id_filter)
    
    attendance_records = query.order_by(Attendance.date.desc(), Attendance.staff_id).all()
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header based on report type
    if report_type == 'summary':
        writer.writerow(['Staff Name', 'Designation', 'Total Days', 'Total Hours', 'Total Minutes', 'Total Earned (PKR)'])
        
        # Group by staff
        staff_data = {}
        for record in attendance_records:
            if record.staff_id not in staff_data:
                staff_data[record.staff_id] = []
            staff_data[record.staff_id].append(record)
        
        # Write data
        for staff_id, records in staff_data.items():
            staff = Staff.query.get(staff_id)
            total_hours = sum(r.hours_worked for r in records)
            total_minutes = sum(r.minutes_worked for r in records)
            total_hours += total_minutes // 60
            total_minutes = total_minutes % 60
            total_earned = sum(r.earned_amount for r in records)
            
            writer.writerow([
                staff.name,
                staff.designation or '',
                len(records),
                total_hours,
                total_minutes,
                round(total_earned, 2)
            ])
    
    else:  # detailed
        writer.writerow(['Date', 'Staff Name', 'Designation', 'Clock In', 'Clock Out', 'Hours', 'Minutes', 'Hourly Rate (PKR)', 'Earned (PKR)', 'Notes'])
        
        for record in attendance_records:
            writer.writerow([
                record.date.strftime('%Y-%m-%d'),
                record.staff.name,
                record.staff.designation or '',
                record.clock_in.strftime('%H:%M:%S') if record.clock_in else '',
                record.clock_out.strftime('%H:%M:%S') if record.clock_out else '',
                record.hours_worked,
                record.minutes_worked,
                round(record.hourly_rate, 2),
                round(record.earned_amount, 2),
                record.notes or ''
            ])
    
    # Create response
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'attendance_report_{datetime.now().strftime("%Y-%m-%d")}.csv'
    )


# --- Export to Excel ---

@bp.route('/attendance/export/excel')
@login_required
def export_attendance_excel():
    """Export attendance report to Excel"""
    
    # Get filter parameters
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    staff_id_filter = request.args.get('staff_id')
    report_type = request.args.get('report_type', 'summary')
    
    # Set default date range
    if not date_from_str:
        today = datetime.now()
        date_from = today.replace(day=1)
    else:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
    
    if not date_to_str:
        date_to = datetime.now()
    else:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
    
    # Get attendance records
    query = Attendance.query.filter(
        Attendance.date >= date_from.date(),
        Attendance.date <= date_to.date()
    )
    
    if staff_id_filter:
        query = query.filter(Attendance.staff_id == staff_id_filter)
    
    attendance_records = query.order_by(Attendance.date.desc(), Attendance.staff_id).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    # Define styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:J1')
    title = ws['A1']
    title.value = f"Attendance Report - {date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')}"
    title.font = Font(bold=True, size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[1].height = 25
    
    # Add report date
    ws.merge_cells('A2:J2')
    report_date = ws['A2']
    report_date.value = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    report_date.font = Font(italic=True, size=10)
    report_date.alignment = Alignment(horizontal="center")
    
    # Start data from row 4
    current_row = 4
    
    if report_type == 'summary':
        # Header row
        headers = ['Staff Name', 'Designation', 'Total Days', 'Total Hours', 'Total Minutes', 'Total Earned (PKR)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        current_row += 1
        
        # Group by staff
        staff_data = {}
        for record in attendance_records:
            if record.staff_id not in staff_data:
                staff_data[record.staff_id] = []
            staff_data[record.staff_id].append(record)
        
        # Write data
        for staff_id, records in staff_data.items():
            staff = Staff.query.get(staff_id)
            total_hours = sum(r.hours_worked for r in records)
            total_minutes = sum(r.minutes_worked for r in records)
            total_hours += total_minutes // 60
            total_minutes = total_minutes % 60
            total_earned = sum(r.earned_amount for r in records)
            
            row_data = [
                staff.name,
                staff.designation or '',
                len(records),
                total_hours,
                total_minutes,
                round(total_earned, 2)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col > 2 else "left")
            
            current_row += 1
    
    else:  # detailed
        # Header row
        headers = ['Date', 'Staff Name', 'Designation', 'Clock In', 'Clock Out', 'Hours', 'Minutes', 'Hourly Rate (PKR)', 'Earned (PKR)', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        current_row += 1
        
        # Write data
        for record in attendance_records:
            row_data = [
                record.date.strftime('%Y-%m-%d'),
                record.staff.name,
                record.staff.designation or '',
                record.clock_in.strftime('%H:%M:%S') if record.clock_in else '',
                record.clock_out.strftime('%H:%M:%S') if record.clock_out else '',
                record.hours_worked,
                record.minutes_worked,
                round(record.hourly_rate, 2),
                round(record.earned_amount, 2),
                record.notes or ''
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = border
                if col == 1:
                    cell.alignment = Alignment(horizontal="center")
                elif col in [6, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
            
            current_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 20
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'attendance_report_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    )


# --- Export to PDF ---

@bp.route('/attendance/export/pdf')
@login_required
def export_attendance_pdf():
    """Export attendance report to PDF"""
    
    # Get filter parameters
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    staff_id_filter = request.args.get('staff_id')
    report_type = request.args.get('report_type', 'summary')
    
    # Set default date range
    if not date_from_str:
        today = datetime.now()
        date_from = today.replace(day=1)
    else:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
    
    if not date_to_str:
        date_to = datetime.now()
    else:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
    
    # Get attendance records
    query = Attendance.query.filter(
        Attendance.date >= date_from.date(),
        Attendance.date <= date_to.date()
    )
    
    if staff_id_filter:
        query = query.filter(Attendance.staff_id == staff_id_filter)
    
    attendance_records = query.order_by(Attendance.date.desc(), Attendance.staff_id).all()
    
    # Create PDF
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Define custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Oblique'
    )
    
    # Add title
    title = Paragraph("Attendance Report", title_style)
    elements.append(title)
    
    # Add date range
    date_range = Paragraph(
        f"Period: {date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        subtitle_style
    )
    elements.append(date_range)
    elements.append(Spacer(1, 0.2*inch))
    
    if report_type == 'summary':
        # Group by staff
        staff_data = {}
        for record in attendance_records:
            if record.staff_id not in staff_data:
                staff_data[record.staff_id] = []
            staff_data[record.staff_id].append(record)
        
        # Create table data
        table_data = [['Staff Name', 'Designation', 'Total Days', 'Total Hours', 'Total Earned (PKR)']]
        
        for staff_id, records in staff_data.items():
            staff = Staff.query.get(staff_id)
            total_hours = sum(r.hours_worked for r in records)
            total_minutes = sum(r.minutes_worked for r in records)
            total_hours += total_minutes // 60
            total_minutes = total_minutes % 60
            total_earned = sum(r.earned_amount for r in records)
            
            table_data.append([
                staff.name,
                staff.designation or '',
                str(len(records)),
                f"{int(total_hours)}h {int(total_minutes)}m",
                f"PKR {round(total_earned, 2):,.2f}"
            ])
    
    else:  # detailed
        # Create table data
        table_data = [['Date', 'Staff Name', 'Clock In', 'Clock Out', 'Hours', 'Earned (PKR)', 'Notes']]
        
        for record in attendance_records:
            table_data.append([
                record.date.strftime('%Y-%m-%d'),
                record.staff.name,
                record.clock_in.strftime('%H:%M') if record.clock_in else '',
                record.clock_out.strftime('%H:%M') if record.clock_out else '',
                f"{int(record.hours_worked)}h {int(record.minutes_worked)}m",
                f"PKR {round(record.earned_amount, 2):,.2f}",
                record.notes or ''
            ])
    
    # Create table
    table = Table(table_data, colWidths=[1.2*inch, 1.2*inch, 0.9*inch, 0.9*inch, 0.9*inch, 1*inch, 1.2*inch])
    
    # Style table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'attendance_report_{datetime.now().strftime("%Y-%m-%d")}.pdf'
    )
