from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response, current_app
from app.utils import permission_required, log_activity
from flask_login import login_required, current_user
from app import db
from app.models import PurchaseBill, PurchaseItem, Product, Vendor, Company, Currency, VendorAdvance, PurchaseOrder, PurchaseOrderItem, CostPriceHistory, PurchaseReturn, PurchaseReturnItem, PurchaseSettings, PurchaseReturnSettings, BillPayment, BillReceive, BillReceiveItem, ProductWarehouseStock, Warehouse, PaymentMethod
from app.forms import PurchaseForm, VendorForm, PurchaseReturnSettingsForm
from datetime import datetime, timedelta, date
from app.pdf_utils import generate_professional_pdf
from app.services.bom_versioning import BOMVersioningService
from werkzeug.utils import secure_filename
import os
from app.routes.filters import apply_saved_filter_to_query
import io
import csv
import json
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
from sqlalchemy import or_, and_

bp = Blueprint('purchase', __name__)

@bp.route('/bills')
@login_required
def bills():
    status = request.args.get('status', 'all')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    vendor_id = request.args.get('vendor_id')
    bill_number = request.args.get('bill_number')
    search = request.args.get('search', '')
    
    query = PurchaseBill.query
    
    if status == 'overdue':
        query = query.filter(
            PurchaseBill.status != 'paid',
            PurchaseBill.due_date < datetime.utcnow()
        )
    elif status != 'all':
        query = query.filter(PurchaseBill.status == status)
    
    if vendor_id:
        query = query.filter(PurchaseBill.vendor_id == vendor_id)
        
    if bill_number:
        query = query.filter(PurchaseBill.bill_number.ilike(f'%{bill_number}%'))
    
    if from_date:
        try:
            query = query.filter(PurchaseBill.date >= datetime.strptime(from_date, '%Y-%m-%d'))
        except ValueError:
            pass
    
    if to_date:
        try:
            to_date_dt = datetime.strptime(to_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(PurchaseBill.date <= to_date_dt)
        except ValueError:
            pass

    if search:
        search_filter = f'%{search}%'
        query = query.join(Vendor).filter(
            (PurchaseBill.bill_number.ilike(search_filter)) |
            (Vendor.name.ilike(search_filter)) |
            (PurchaseBill.notes.ilike(search_filter))
        )

    query = apply_saved_filter_to_query(query, 'purchase', request.args)

    bills = query.order_by(PurchaseBill.date.desc()).all()
    
    # Summary totals for the view
    total_amount = sum(bill.total for bill in bills)
    total_paid = sum(bill.paid_amount for bill in bills)
    total_balance = sum(bill.balance_due for bill in bills)
    
    # Get company date format
    company = Company.query.first()
    date_format = company.date_format if company and company.date_format else '%Y-%m-%d'
    
    vendors = Vendor.query.filter_by(is_active=True).all()
    # Fetch all bill numbers for the searchable dropdown
    all_bill_numbers = db.session.query(PurchaseBill.bill_number).distinct().order_by(PurchaseBill.bill_number).all()
    all_bill_numbers = [b[0] for b in all_bill_numbers]
    
    return render_template('purchase/bills.html', 
                         bills=bills, 
                         vendors=vendors,
                         all_bill_numbers=all_bill_numbers,
                         current_status=status,
                         from_date=from_date,
                         to_date=to_date,
                         search=search,
                         vendor_id=vendor_id,
                         bill_number=bill_number,
                         total_amount=total_amount,
                         total_paid=total_paid,
                         total_balance=total_balance,
                         date_format=date_format,
                         active_module='purchase',
                         filter_id=request.args.get('filter_id'))

@bp.route('/bill/create', methods=['GET', 'POST'])
@login_required
@permission_required('purchases', action='add')
def create_bill():
    form = PurchaseForm()
    vendors = Vendor.query.filter_by(is_active=True).all()
    products = Product.query.filter_by(is_active=True).all()
    currencies = Currency.query.filter_by(is_active=True).all()
    warehouses = Warehouse.query.all()
    
    form.vendor_id.choices = [(v.id, v.name) for v in vendors]
    
    if form.validate_on_submit():
        vendor_id = form.vendor_id.data
        date = form.date.data

        # Get items from form
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        warehouse_ids = request.form.getlist('warehouse_id[]')

        # Generate bill number using settings
        settings = PurchaseSettings.query.first()
        if not settings:
            settings = PurchaseSettings(bill_prefix='PB-', bill_suffix='', next_bill_number=1,
                                        po_prefix='PO-', po_suffix='', next_po_number=1)
            db.session.add(settings)
        bill_number = f"{settings.bill_prefix}{settings.next_bill_number}{settings.bill_suffix}"
        settings.next_bill_number += 1
        
        due_date_str = request.form.get('due_date')
        if due_date_str:
            try:
                due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
            except ValueError:
                due_date = date + timedelta(days=30)
        else:
            due_date = date + timedelta(days=30)

        bill = PurchaseBill(
            bill_number=bill_number,
            vendor_id=vendor_id,
            date=date,
            due_date=due_date,
            tax_rate=float(request.form.get('tax_rate', 0)),
            discount=float(request.form.get('discount', 0)),
            shipping_charge=float(request.form.get('shipping_charge', 0)),
            currency_id=request.form.get('currency_id'),
            exchange_rate=float(request.form.get('exchange_rate', 1)),
            notes=request.form.get('notes'),
            created_by=current_user.id,
            is_approved=(current_user.role == 'admin')
        )
        
        db.session.add(bill)
        
        for i in range(len(product_ids)):
            if product_ids[i] and quantities[i] and float(quantities[i]) > 0:
                item = PurchaseItem(
                    product_id=int(product_ids[i]),
                    quantity=float(quantities[i]),
                    unit_price=float(prices[i]),
                    warehouse_id=int(warehouse_ids[i]) if i < len(warehouse_ids) and warehouse_ids[i] else None,
                    total=float(quantities[i]) * float(prices[i])
                )
                bill.items.append(item)
        
        # IMPORTANT: Flush to session so items are linked before calculation
        db.session.flush()
        bill.calculate_totals()

        # Update paid amount from "Advance Paid" field in template
        try:
            advance = float(request.form.get('advance') or 0)
        except (ValueError, TypeError):
            advance = 0
        if advance > 0:
            vendor = Vendor.query.get(vendor_id)
            pending_advances = VendorAdvance.query.filter_by(
                vendor_id=vendor.id
            ).order_by(VendorAdvance.date).all()

            remaining_to_apply = advance
            for adv in pending_advances:
                if remaining_to_apply <= 0:
                    break
                can_apply = min(adv.remaining_balance, remaining_to_apply)
                if can_apply > 0:
                    adv.applied_amount += can_apply
                    bill.paid_amount += can_apply
                    # Track advance applied so it can be reversed if the bill is deleted
                    bill.advance_applied = (bill.advance_applied or 0) + can_apply
                    if adv.applied_amount >= adv.amount:
                        adv.is_adjusted = True
                        adv.adjusted_bill_id = bill.id
                    else:
                        adv.adjusted_bill_id = bill.id
                    remaining_to_apply -= can_apply

            if remaining_to_apply > 0:
                bill.paid_amount += remaining_to_apply
        else:
            bill.paid_amount = 0

        bill.update_status()

        # Handle bill image upload
        if 'bill_image' in request.files:
            file = request.files['bill_image']
            if file and file.filename:
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'webp'}
                ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                if ext in allowed_extensions:
                    upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'bills')
                    os.makedirs(upload_dir, exist_ok=True)
                    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                    filename = secure_filename(f"{bill_number}_{timestamp}.{ext}")
                    filepath = os.path.join(upload_dir, filename)
                    file.save(filepath)
                    bill.bill_image_path = os.path.join('uploads', 'bills', filename).replace('\\', '/')
                else:
                    flash('Invalid file type for bill image. Allowed: png, jpg, jpeg, gif, pdf, webp', 'warning')

        db.session.commit()

        # Auto-apply the vendor's existing advance credit to this new bill, but
        # only when the user didn't manually specify an advance amount above
        # (respect an explicit manual choice). Mirrors the customer-advance
        # auto-apply on sales invoice creation.
        advance_auto_applied = 0
        if advance <= 0:
            try:
                from app.utils import auto_apply_vendor_advance
                advance_auto_applied = auto_apply_vendor_advance(bill)
                if advance_auto_applied > 0:
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                current_app.logger.warning(f"Vendor advance auto-apply failed for bill {bill.bill_number}: {e}")

        log_activity('Purchase', f'Created Bill #{bill.bill_number}',
                    f'Vendor: {bill.vendor.name}, Total: {bill.total}')

        if advance_auto_applied > 0:
            flash(f'Purchase bill created successfully! PKR {advance_auto_applied:,.2f} of the vendor\'s existing advance was auto-applied to this bill.', 'success')
        else:
            flash('Purchase bill created successfully!', 'success')
        return redirect(url_for('purchase.bill_detail', id=bill.id))
    
    return render_template('purchase/create_bill.html', form=form, products=products, vendors=vendors, currencies=currencies, warehouses=warehouses)

@bp.route('/bill/<int:id>')
@login_required
def bill_detail(id):
    bill = PurchaseBill.query.get_or_404(id)
    company = Company.query.first()
    date_format = company.date_format if company and company.date_format else '%Y-%m-%d'
    # Build a dict of already-received quantity per purchase_item_id for the receive modal
    received_qty_map = {}
    for br in bill.bill_receives:
        for bri in br.receive_items:
            received_qty_map[bri.purchase_item_id] = received_qty_map.get(bri.purchase_item_id, 0) + bri.quantity_received
    
    warehouses = Warehouse.query.all()

    # All bills for this vendor (including the current one, badged "Current" in the
    # template). Mirrors the "Invoices for customer" table on the sales invoice page.
    other_bills = []
    if bill.vendor_id:
        other_bills = PurchaseBill.query.filter(
            PurchaseBill.vendor_id == bill.vendor_id,
            or_(PurchaseBill.id == bill.id, PurchaseBill.is_approved == True)
        ).order_by(PurchaseBill.date.desc()).all()

    # Custom payment methods (same source as the sales invoice pay popup)
    payment_methods = PaymentMethod.query.filter_by(is_active=True).order_by(PaymentMethod.name).all()

    return render_template('purchase/bill_detail.html', bill=bill, date_format=date_format,
                           received_qty_map=received_qty_map, warehouses=warehouses,
                           other_bills=other_bills, payment_methods=payment_methods)

@bp.route('/bill/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('purchases', action='edit')
def edit_bill(id):
    bill = PurchaseBill.query.get_or_404(id)
    form = PurchaseForm(obj=bill)
    vendors = Vendor.query.filter_by(is_active=True).all()
    products = Product.query.filter_by(is_active=True).all()
    currencies = Currency.query.filter_by(is_active=True).all()
    warehouses = Warehouse.query.all()

    form.vendor_id.choices = [(v.id, v.name) for v in vendors]

    if request.method == 'POST':
        # Only revert inventory if stock was already received for this bill
        if bill.inventory_received:
            if bill.bill_receives:
                for receive in bill.bill_receives:
                    for bri in receive.receive_items:
                        product = Product.query.get(bri.product_id)
                        if product:
                            product.update_quantity(-bri.quantity_received)
                            # Revert from the specific warehouse in the receive item
                            if bri.warehouse_id:
                                wh_stock = ProductWarehouseStock.query.filter_by(
                                    product_id=bri.product_id,
                                    warehouse_id=bri.warehouse_id
                                ).first()
                                if wh_stock:
                                    wh_stock.quantity -= bri.quantity_received
                                    if wh_stock.quantity < 0:
                                        wh_stock.quantity = 0
            else:
                # Fallback for legacy bills
                for item in bill.items:
                    product = Product.query.get(item.product_id)
                    if product:
                        product.update_quantity(-item.quantity)
                        # Revert from the specific warehouse used in the item
                        if item.warehouse_id:
                            wh_stock = ProductWarehouseStock.query.filter_by(
                                product_id=item.product_id,
                                warehouse_id=item.warehouse_id
                            ).first()
                            if wh_stock:
                                wh_stock.quantity -= item.quantity
                                if wh_stock.quantity < 0:
                                    wh_stock.quantity = 0

        # Safely remove old cost price history entries if not referenced by BOM items
        old_history_entries = CostPriceHistory.query.filter_by(purchase_bill_id=bill.id).all()
        for h in old_history_entries:
            from app.models import BOMItem
            referenced = BOMItem.query.filter_by(cost_price_history_id=h.id).first()
            if referenced:
                h.is_active = False
            else:
                db.session.delete(h)

        # Delete old receive records (since we are resetting the bill)
        if bill.inventory_received:
            BillReceive.query.filter_by(bill_id=bill.id).delete()
            bill.inventory_received = False

        # Delete old items
        PurchaseItem.query.filter_by(bill_id=bill.id).delete()

        # Rebuild from form
        bill.vendor_id = int(request.form.get('vendor_id', bill.vendor_id))
        date = request.form.get('date')
        if date:
            try:
                bill.date = datetime.strptime(date, '%Y-%m-%d')
            except ValueError:
                pass

        due_date_str = request.form.get('due_date')
        if due_date_str:
            try:
                bill.due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
            except ValueError:
                pass
        else:
            bill.due_date = bill.date + timedelta(days=30)

        bill.tax_rate = float(request.form.get('tax_rate', 0))
        bill.discount = float(request.form.get('discount', 0))
        bill.shipping_charge = float(request.form.get('shipping_charge', 0))
        bill.currency_id = request.form.get('currency_id') or None
        bill.exchange_rate = float(request.form.get('exchange_rate', 1))
        bill.notes = request.form.get('notes', '')

        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        warehouse_ids = request.form.getlist('warehouse_id[]')

        for i in range(len(product_ids)):
            if product_ids[i] and quantities[i] and float(quantities[i]) > 0:
                prod_id = int(product_ids[i])
                qty = float(quantities[i])
                price = float(prices[i])
                wh_id = int(warehouse_ids[i]) if i < len(warehouse_ids) and warehouse_ids[i] else None
                item_total = qty * price
                item = PurchaseItem(
                    product_id=prod_id,
                    quantity=qty,
                    unit_price=price,
                    warehouse_id=wh_id,
                    total=item_total
                )
                bill.items.append(item)
                # NOTE: inventory NOT updated on edit — user must re-receive

        # IMPORTANT: Flush to session so items are linked before calculation
        db.session.flush()
        bill.calculate_totals()
        
        # Cap paid amount if it now exceeds total
        if bill.paid_amount > bill.total:
            bill.paid_amount = bill.total
        bill.update_status()

        # Handle bill image upload on edit
        if 'bill_image' in request.files:
            file = request.files['bill_image']
            if file and file.filename:
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'webp'}
                ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                if ext in allowed_extensions:
                    upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'bills')
                    os.makedirs(upload_dir, exist_ok=True)
                    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                    filename = secure_filename(f"{bill.bill_number}_{timestamp}.{ext}")
                    filepath = os.path.join(upload_dir, filename)
                    file.save(filepath)
                    bill.bill_image_path = os.path.join('uploads', 'bills', filename).replace('\\', '/')

        db.session.commit()
        log_activity('Purchase', f'Updated Bill #{bill.bill_number}',
                    f'Vendor: {bill.vendor.name}, Total: {bill.total}')
        flash('Purchase bill updated successfully! Inventory and cost prices have been updated.', 'success')
        return redirect(url_for('purchase.bill_detail', id=bill.id))

    return render_template('purchase/edit_bill.html', form=form, bill=bill,
                           products=products, vendors=vendors, currencies=currencies, warehouses=warehouses)

@bp.route('/bill/<int:id>/update-shipping', methods=['POST'])
@login_required
@permission_required('purchases', action='edit')
def update_shipping(id):
    """Update shipping charge and recalculate product costs"""
    bill = PurchaseBill.query.get_or_404(id)
    
    try:
        new_shipping = float(request.form.get('shipping_charge', 0))
        old_shipping = bill.shipping_charge
        
        if new_shipping < 0:
            flash('Shipping charge cannot be negative', 'danger')
            return redirect(request.referrer or url_for('purchase.bill_detail', id=id))
        
        # Update shipping
        bill.shipping_charge = new_shipping
        bill.calculate_totals()
        
        # Recalculate product costs to include new shipping
        total_items_cost = sum(item.total for item in bill.items)
        if total_items_cost > 0:
            for item in bill.items:
                product = Product.query.get(item.product_id)
                if product:
                    # Calculate new cost including proportional shipping (based on item cost, not qty)
                    # Also include tax in allocation
                    taxable_amount = total_items_cost + new_shipping
                    tax_amount = (taxable_amount * bill.tax_rate) / 100 if bill.tax_rate > 0 else 0
                    total_additional = new_shipping + tax_amount
                    
                    # Allocate based on item cost ratio
                    allocation_ratio = item.total / total_items_cost if total_items_cost > 0 else 0
                    allocated_additional = total_additional * allocation_ratio
                    
                    new_cost_price = item.unit_price + (allocated_additional / item.quantity)
                    
                    if product.cost_price != new_cost_price:
                        old_price = product.cost_price
                        
                        # Create cost price history entry
                        cost_history = CostPriceHistory(
                            product_id=item.product_id,
                            purchase_bill_id=bill.id,
                            old_price=old_price if old_price > 0 else None,
                            new_price=new_cost_price,
                            quantity_at_old_price=product.quantity - item.quantity,
                            used_quantity=0,
                            reason=f"Shipping update - Old Shipping: PKR {old_shipping}, New Shipping: PKR {new_shipping}",
                            is_active=True,
                            created_by=current_user.id
                        )
                        db.session.add(cost_history)
                        
                        # Update product cost price
                        product.cost_price = new_cost_price
                        
                        # Trigger BOM versioning
                        try:
                            user_id = None
                            try:
                                if current_user and current_user.is_authenticated:
                                    user_id = current_user.id
                            except (AttributeError, TypeError):
                                pass
                            if user_id is None:
                                from app.models import User
                                admin_user = User.query.filter_by(username='admin').first()
                                user_id = admin_user.id if admin_user else 1
                            
                            BOMVersioningService.check_and_update_bom_for_cost_changes(
                                product_id=item.product_id,
                                created_by_id=user_id
                            )
                        except Exception as e:
                            print(f"Error updating BOM for shipping change: {e}")
        
        bill.update_status()
        db.session.commit()
        log_activity('Purchase', f'Updated Shipping on Bill #{bill.bill_number}',
                    f'Old: PKR {old_shipping:,.2f}, New: PKR {new_shipping:,.2f}')
        flash(f'Shipping charge updated to PKR {new_shipping:,.2f}. Product costs recalculated!', 'success')

    except ValueError:
        flash('Invalid shipping amount', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating shipping: {str(e)}', 'danger')
    
    return redirect(request.referrer or url_for('purchase.bill_detail', id=id))

@bp.route('/bill/<int:id>/update-tax', methods=['POST'])
@login_required
@permission_required('purchases', action='edit')
def update_tax(id):
    """Update tax rate and recalculate product costs"""
    bill = PurchaseBill.query.get_or_404(id)
    
    try:
        new_tax_rate = float(request.form.get('tax_rate', 0))
        old_tax_rate = bill.tax_rate
        
        if new_tax_rate < 0 or new_tax_rate > 100:
            flash('Tax rate must be between 0 and 100', 'danger')
            return redirect(request.referrer or url_for('purchase.bill_detail', id=id))
        
        # Update tax rate
        bill.tax_rate = new_tax_rate
        bill.calculate_totals()
        
        # Recalculate product costs to include new tax
        total_items_cost = sum(item.total for item in bill.items)
        if total_items_cost > 0:
            for item in bill.items:
                product = Product.query.get(item.product_id)
                if product:
                    # Calculate new cost including proportional tax (based on item cost, not qty)
                    # Also include shipping in calculation
                    taxable_amount = total_items_cost + bill.shipping_charge
                    new_tax_total = (taxable_amount * new_tax_rate) / 100 if new_tax_rate > 0 else 0
                    total_additional = bill.shipping_charge + new_tax_total
                    
                    # Allocate based on item cost ratio
                    allocation_ratio = item.total / total_items_cost if total_items_cost > 0 else 0
                    allocated_additional = total_additional * allocation_ratio
                    
                    new_cost_price = item.unit_price + (allocated_additional / item.quantity)
                    
                    if product.cost_price != new_cost_price:
                        old_price = product.cost_price
                        
                        # Create cost price history entry
                        cost_history = CostPriceHistory(
                            product_id=item.product_id,
                            purchase_bill_id=bill.id,
                            old_price=old_price if old_price > 0 else None,
                            new_price=new_cost_price,
                            quantity_at_old_price=product.quantity - item.quantity,
                            used_quantity=0,
                            reason=f"Tax rate update - Old Tax Rate: {old_tax_rate}%, New Tax Rate: {new_tax_rate}%",
                            is_active=True,
                            created_by=current_user.id
                        )
                        db.session.add(cost_history)
                        
                        # Update product cost price
                        product.cost_price = new_cost_price
                        
                        # Trigger BOM versioning
                        try:
                            user_id = None
                            try:
                                if current_user and current_user.is_authenticated:
                                    user_id = current_user.id
                            except (AttributeError, TypeError):
                                pass
                            if user_id is None:
                                from app.models import User
                                admin_user = User.query.filter_by(username='admin').first()
                                user_id = admin_user.id if admin_user else 1
                            
                            BOMVersioningService.check_and_update_bom_for_cost_changes(
                                product_id=item.product_id,
                                created_by_id=user_id
                            )
                        except Exception as e:
                            print(f"Error updating BOM for tax change: {e}")
        
        bill.update_status()
        db.session.commit()
        log_activity('Purchase', f'Updated Tax on Bill #{bill.bill_number}',
                    f'Old: {old_tax_rate}%, New: {new_tax_rate}%')
        flash(f'Tax rate updated to {new_tax_rate}%. Product costs recalculated!', 'success')

    except ValueError:
        flash('Invalid tax rate', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating tax: {str(e)}', 'danger')
    
    return redirect(request.referrer or url_for('purchase.bill_detail', id=id))

# ── Cancel / Reverse execution helpers (shared by direct-admin and approval paths) ──

def _apply_bill_cancellation(bill, qty_map):
    """Apply an itemized cancellation to the bill and mark it settled.

    qty_map: dict of {purchase_item_id (str or int): qty_to_cancel}. Mutates the
    bill in place; the CALLER is responsible for committing. Returns the total
    cancelled value. Quantities are re-validated against what is still cancellable.
    """
    total_cancelled_value = 0
    for item in bill.items:
        raw = qty_map.get(str(item.id), qty_map.get(item.id, 0))
        try:
            qty_to_cancel = float(raw or 0)
        except (ValueError, TypeError):
            qty_to_cancel = 0

        # Cannot cancel more than (Total - already received)
        received_qty = 0
        for br in bill.bill_receives:
            for bri in br.receive_items:
                if bri.purchase_item_id == item.id:
                    received_qty += bri.quantity_received

        max_cancellable = max(0, item.quantity - received_qty)
        if qty_to_cancel > max_cancellable:
            qty_to_cancel = max_cancellable

        if qty_to_cancel > 0:
            item.cancelled_quantity = qty_to_cancel
            item_value = (item.total / item.quantity) * qty_to_cancel
            total_cancelled_value += item_value

    bill.cancelled_amount = total_cancelled_value
    bill.cancelled_at = datetime.utcnow()
    bill.status = 'cancelled'

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    bill.notes = (bill.notes or "") + (
        f"\n[Cancelled on {now_str}] Itemized cancellation performed. "
        f"Total waived: PKR {total_cancelled_value:,.2f}"
    )
    return total_cancelled_value


def _apply_bill_reverse_cancellation(bill):
    """Undo a bill cancellation and restore the balance. Mutates bill; caller commits."""
    for item in bill.items:
        item.cancelled_quantity = 0
    bill.cancelled_amount = 0
    bill.cancelled_at = None
    # Reset status so update_status() doesn't return early on 'cancelled'
    bill.status = 'unpaid'
    bill.update_status()
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    bill.notes = (bill.notes or "") + f"\n[Rerversed on {now_str}] Cancellation was reversed. Balance restored."


def _clear_bill_pending_action(bill):
    """Clear any held cancel/reverse request from the bill."""
    bill.pending_action = None
    bill.pending_action_reason = None
    bill.pending_action_payload = None
    bill.pending_action_by = None
    bill.pending_action_at = None


@bp.route('/bill/<int:id>/cancel', methods=['POST'])
@login_required
@permission_required('purchases', action='edit')
def cancel_bill(id):
    """Cancel specific quantities of items and settle the bill.

    Admins apply the cancellation immediately. Non-admins (staff/manager) submit
    the request for admin approval — nothing on the bill changes until approved.
    """
    bill = PurchaseBill.query.get_or_404(id)

    if bill.status == 'cancelled':
        flash('Bill is already cancelled.', 'warning')
        return redirect(url_for('purchase.bill_detail', id=id))

    if bill.pending_action:
        flash('A request for this bill is already awaiting admin approval.', 'warning')
        return redirect(url_for('purchase.bill_detail', id=id))

    # Collect requested cancel quantities from the form
    qty_map = {}
    for item in bill.items:
        qty_map[str(item.id)] = request.form.get(f'cancel_qty_{item.id}', 0)
    reason = (request.form.get('cancel_reason') or '').strip()

    # Non-admins: hold the request for approval instead of applying it
    if not current_user.is_admin:
        try:
            bill.pending_action = 'cancel'
            bill.pending_action_reason = reason or None
            bill.pending_action_payload = json.dumps(qty_map)
            bill.pending_action_by = current_user.id
            bill.pending_action_at = datetime.utcnow()
            db.session.commit()
            log_activity('Purchase', f'Requested Cancellation of Bill #{bill.bill_number}',
                         f'Reason: {reason}' if reason else 'Awaiting admin approval')
            flash('Your cancellation request has been submitted for admin approval.', 'info')
        except Exception as e:
            db.session.rollback()
            flash(f'Error submitting request: {str(e)}', 'danger')
        return redirect(url_for('purchase.bill_detail', id=id))

    # Admin: apply immediately
    try:
        total_cancelled_value = _apply_bill_cancellation(bill, qty_map)
        db.session.commit()
        log_activity('Purchase', f'Cancelled Bill #{bill.bill_number}',
                     f'Cancelled Amount: PKR {total_cancelled_value:,.2f}')
        flash(f'Bill #{bill.bill_number} has been cancelled with itemized quantities.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error cancelling bill: {str(e)}', 'danger')

    return redirect(url_for('purchase.bill_detail', id=id))


@bp.route('/bill/<int:id>/reverse-cancel', methods=['POST'])
@login_required
@permission_required('purchases', action='edit')
def reverse_cancel_bill(id):
    """Undo the cancellation of a bill.

    Admins reverse immediately. Non-admins submit the request for admin approval —
    the bill stays cancelled until an admin approves the reversal.
    """
    bill = PurchaseBill.query.get_or_404(id)

    if bill.status != 'cancelled':
        flash('Bill is not cancelled.', 'warning')
        return redirect(url_for('purchase.bill_detail', id=id))

    if bill.pending_action:
        flash('A request for this bill is already awaiting admin approval.', 'warning')
        return redirect(url_for('purchase.bill_detail', id=id))

    reason = (request.form.get('reverse_reason') or '').strip()

    # Non-admins: hold the request for approval
    if not current_user.is_admin:
        try:
            bill.pending_action = 'reverse'
            bill.pending_action_reason = reason or None
            bill.pending_action_payload = None
            bill.pending_action_by = current_user.id
            bill.pending_action_at = datetime.utcnow()
            db.session.commit()
            log_activity('Purchase', f'Requested Reversal of Cancellation for Bill #{bill.bill_number}',
                         f'Reason: {reason}' if reason else 'Awaiting admin approval')
            flash('Your reversal request has been submitted for admin approval.', 'info')
        except Exception as e:
            db.session.rollback()
            flash(f'Error submitting request: {str(e)}', 'danger')
        return redirect(url_for('purchase.bill_detail', id=id))

    # Admin: reverse immediately
    try:
        _apply_bill_reverse_cancellation(bill)
        db.session.commit()
        log_activity('Purchase', f'Reversed Cancellation of Bill #{bill.bill_number}',
                     'Balance Restored')
        flash(f'Cancellation of Bill #{bill.bill_number} has been reversed.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error reversing cancellation: {str(e)}', 'danger')

    return redirect(url_for('purchase.bill_detail', id=id))


@bp.route('/bill/<int:id>/approve-action', methods=['POST'])
@login_required
def approve_bill_action(id):
    """Admin approves a held cancel/reverse request and applies it."""
    bill = PurchaseBill.query.get_or_404(id)

    if not current_user.is_admin:
        flash('Only an admin can approve this request.', 'danger')
        return redirect(url_for('purchase.bill_detail', id=id))

    action = bill.pending_action
    if not action:
        flash('There is no pending request for this bill.', 'warning')
        return redirect(url_for('purchase.bill_detail', id=id))

    label = bill.pending_action_label or action
    try:
        if action == 'cancel':
            if bill.status == 'cancelled':
                flash('Bill is already cancelled — clearing the pending request.', 'warning')
            else:
                qty_map = json.loads(bill.pending_action_payload or '{}')
                _apply_bill_cancellation(bill, qty_map)
        elif action == 'reverse':
            if bill.status != 'cancelled':
                flash('Bill is not cancelled — clearing the pending request.', 'warning')
            else:
                _apply_bill_reverse_cancellation(bill)

        _clear_bill_pending_action(bill)
        db.session.commit()
        log_activity('Purchase', f'Approved {label} for Bill #{bill.bill_number}',
                     'Request approved by admin')
        flash(f'Request approved: {label} applied to Bill #{bill.bill_number}.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error approving request: {str(e)}', 'danger')

    return redirect(url_for('purchase.bill_detail', id=id))


@bp.route('/bill/<int:id>/reject-action', methods=['POST'])
@login_required
def reject_bill_action(id):
    """Admin rejects a held cancel/reverse request. Nothing is applied to the bill."""
    bill = PurchaseBill.query.get_or_404(id)

    if not current_user.is_admin:
        flash('Only an admin can reject this request.', 'danger')
        return redirect(url_for('purchase.bill_detail', id=id))

    if not bill.pending_action:
        flash('There is no pending request for this bill.', 'warning')
        return redirect(url_for('purchase.bill_detail', id=id))

    label = bill.pending_action_label or bill.pending_action
    reason = (request.form.get('reject_reason') or '').strip()
    try:
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
        note = f"\n[Request Rejected on {now_str}] {label} request rejected by admin."
        if reason:
            note += f" Reason: {reason}"
        bill.notes = (bill.notes or "") + note

        _clear_bill_pending_action(bill)
        db.session.commit()
        log_activity('Purchase', f'Rejected {label} for Bill #{bill.bill_number}',
                     f'Reason: {reason}' if reason else 'Request rejected by admin')
        flash(f'{label} request has been rejected.', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Error rejecting request: {str(e)}', 'danger')

    return redirect(url_for('purchase.bill_detail', id=id))

@bp.route('/bill/<int:id>/pdf')
@login_required
def bill_pdf(id):
    bill = PurchaseBill.query.get_or_404(id)
    company = Company.query.first()
    from app.models import PurchaseSettings
    settings = PurchaseSettings.query.first()
    
    try:
        buffer = generate_professional_pdf('purchase', bill, company, settings)
        
        response = make_response(buffer)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename="bill_{bill.bill_number}.pdf"'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('purchase.bill_detail', id=id))

@bp.route('/bill/<int:id>/pdf/view')
@login_required
def bill_pdf_view(id):
    """View PDF inline for sharing"""
    bill = PurchaseBill.query.get_or_404(id)
    company = Company.query.first()
    from app.models import PurchaseSettings
    settings = PurchaseSettings.query.first()
    
    try:
        buffer = generate_professional_pdf('purchase', bill, company, settings)
        return send_file(buffer, mimetype='application/pdf')
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('purchase.bill_detail', id=id))

@bp.route('/bill/<int:id>/pdf/share')
@login_required
def bill_pdf_share(id):
    """Generate PDF, save to public folder, and return shareable link"""
    from flask import current_app
    import os
    import secrets
    
    bill = PurchaseBill.query.get_or_404(id)
    company = Company.query.first()
    from app.models import PurchaseSettings
    settings = PurchaseSettings.query.first()
    
    try:
        buffer = generate_professional_pdf('purchase', bill, company, settings)
        buffer.seek(0)
        pdf_data = buffer.getvalue()
        
        public_dir = os.path.join(current_app.root_path, 'static', 'shared_pdfs')
        os.makedirs(public_dir, exist_ok=True)
        
        token = secrets.token_urlsafe(8)
        filename = f"bill_{bill.bill_number}_{token}.pdf"
        filepath = os.path.join(public_dir, filename)
        
        with open(filepath, 'wb') as f:
            f.write(pdf_data)
        
        share_url = url_for('static', filename=f'shared_pdfs/{filename}')
        return {'share_url': share_url}
    except Exception as e:
        return {'error': str(e)}, 500


@bp.route('/settings', methods=['GET', 'POST'])
@login_required
def purchase_settings_page():
    from app.forms import PurchaseSettingsForm
    from app.models import PurchaseSettings
    settings = PurchaseSettings.query.first()
    
    form = PurchaseSettingsForm(obj=settings)
    
    if request.method == 'POST':
        if not settings:
            settings = PurchaseSettings()
            db.session.add(settings)
        
        settings.default_notes = form.default_notes.data
        settings.default_terms = form.default_terms.data
        settings.bill_prefix = form.bill_prefix.data or ''
        settings.bill_suffix = form.bill_suffix.data or ''
        settings.next_bill_number = form.next_bill_number.data or 1
        settings.po_prefix = form.po_prefix.data or ''
        settings.po_suffix = form.po_suffix.data or ''
        settings.next_po_number = form.next_po_number.data or 1
        
        db.session.commit()
        flash('Purchase settings saved!', 'success')
        return redirect(url_for('purchase.purchase_settings_page'))
    
    return render_template('purchase/purchase_settings.html', settings=settings, form=form)


@bp.route('/bill/<int:id>/pay', methods=['POST'])
@login_required
@permission_required('purchases', action='add')
def pay_bill(id):
    bill = PurchaseBill.query.get_or_404(id)
    payment_method = request.form.get('payment_method', 'cash')

    vendor = bill.vendor
    is_admin = (current_user.role == 'admin')
    total_payment = 0
    overpayment_credit = 0  # excess over the bill balance, credited to vendor advance
    # Keep the 3 built-in methods capitalised (Cash/Advance/Mixed); preserve the
    # exact label for custom/other methods (e.g. "Bank Transfer", "Cheque").
    if payment_method in ('cash', 'advance', 'mixed'):
        method_label = payment_method.capitalize()
    else:
        method_label = payment_method
    last_advance_id = None

    try:
        if payment_method == 'cash':
            amount = float(request.form.get('amount', 0))
            if amount > 0:
                total_payment = amount
                if is_admin:
                    # Cap at the bill balance; any excess becomes vendor advance credit
                    from app.utils import apply_bill_payment_with_credit
                    overpayment_credit += apply_bill_payment_with_credit(bill, amount, current_user.id)

        elif payment_method == 'advance':
            advance_id = request.form.get('selected_advance_id')
            if advance_id:
                advance = VendorAdvance.query.get(int(advance_id))
                if advance and advance.vendor_id == vendor.id:
                    can_apply = min(advance.remaining_balance, bill.balance_due)
                    if can_apply > 0:
                        total_payment = can_apply
                        last_advance_id = advance.id
                        if is_admin:
                            advance.applied_amount += can_apply
                            bill.paid_amount += can_apply
                            if advance.applied_amount >= advance.amount:
                                advance.is_adjusted = True
                                advance.adjusted_bill_id = bill.id
                            else:
                                advance.adjusted_bill_id = bill.id

        elif payment_method == 'mixed':
            advance_ids = request.form.getlist('value')
            cash_amount = float(request.form.get('cash_amount', 0))

            # Apply advances limited to vendor payable portion (balance_due)
            for advance_id in advance_ids:
                # Stop if vendor balance cleared
                if bill.balance_due <= 0:
                    break
                advance = VendorAdvance.query.get(int(advance_id))
                if advance and advance.vendor_id == vendor.id:
                    can_apply = min(advance.remaining_balance, bill.balance_due)
                    if can_apply > 0:
                        total_payment += can_apply
                        last_advance_id = advance.id # Stores the last one used for simple tracking
                        if is_admin:
                            advance.applied_amount += can_apply
                            bill.paid_amount += can_apply
                            if advance.applied_amount >= advance.amount:
                                advance.is_adjusted = True
                                advance.adjusted_bill_id = bill.id
                            else:
                                advance.adjusted_bill_id = bill.id

            # After advances, determine remaining total due (net of cancellations)
            remaining_total = bill.total - bill.paid_amount - (bill.cancelled_amount or 0)
            if cash_amount > 0 and remaining_total > 0:
                cash_to_apply = min(cash_amount, remaining_total)
                total_payment += cash_to_apply
                if is_admin:
                    bill.paid_amount += cash_to_apply

        else:
            # Direct non-cash methods (Bank Transfer, Cheque, Online, or any custom
            # payment method) — recorded like cash from the same Amount field, but
            # WITHOUT the cash-only auto-generated receipt PDF (handled further below).
            amount = float(request.form.get('amount', 0))
            if amount > 0:
                total_payment = amount
                if is_admin:
                    # Cap at the bill balance; any excess becomes vendor advance credit
                    from app.utils import apply_bill_payment_with_credit
                    overpayment_credit += apply_bill_payment_with_credit(bill, amount, current_user.id)

        if is_admin:
            bill.update_status()

        # ── Save BillPayment record ──────────────────────────────────────
        if total_payment > 0:
            bp_record = BillPayment(
                bill_id=bill.id,
                amount=total_payment,
                payment_method=method_label,
                reference_number=request.form.get('reference_number', '').strip() or None,
                notes=request.form.get('pay_notes', '').strip() or None,
                created_by=current_user.id,
                is_approved=is_admin,
                approved_by=current_user.id if is_admin else None,
                approved_at=datetime.utcnow() if is_admin else None,
                advance_id=last_advance_id
            )
            # Handle payment image upload
            if 'pay_image' in request.files:
                pay_file = request.files['pay_image']
                if pay_file and pay_file.filename:
                    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'webp'}
                    ext = pay_file.filename.rsplit('.', 1)[1].lower() if '.' in pay_file.filename else ''
                    if ext in allowed_extensions:
                        upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'payments')
                        os.makedirs(upload_dir, exist_ok=True)
                        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                        filename = secure_filename(f"pay_{bill.bill_number}_{timestamp}.{ext}")
                        filepath = os.path.join(upload_dir, filename)
                        pay_file.save(filepath)
                        bp_record.image_path = os.path.join('uploads', 'payments', filename).replace('\\', '/')
            db.session.add(bp_record)

            # Cash payments don't require the manual upload — auto-generate a
            # receipt PDF with the vendor's full details in its place, so the
            # Payment History "Receipt" column still has something to show.
            if payment_method == 'cash' and not bp_record.image_path:
                try:
                    db.session.flush()  # assign bp_record.id, used in the receipt filename/number
                    from app.pdf_utils import generate_bill_payment_receipt_pdf
                    company = Company.query.first()
                    pdf_buffer = generate_bill_payment_receipt_pdf(bp_record, bill, company)
                    import time, uuid
                    unique_prefix = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
                    filename = secure_filename(f"pay_{bill.bill_number}_{unique_prefix}_receipt.pdf")
                    upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'payments')
                    os.makedirs(upload_dir, exist_ok=True)
                    filepath = os.path.join(upload_dir, filename)
                    with open(filepath, 'wb') as f:
                        f.write(pdf_buffer.getvalue())
                    bp_record.image_path = os.path.join('uploads', 'payments', filename).replace('\\', '/')
                except Exception as e:
                    # Never block a cash payment from being recorded just because
                    # the auto-receipt failed to generate.
                    current_app.logger.warning(f"Auto-receipt generation failed for bill payment on {bill.bill_number}: {e}")

        db.session.commit()
        if total_payment > 0:
            log_activity('Purchase', f'Recorded Payment on Bill #{bill.bill_number}',
                        f'Amount: PKR {total_payment:,.2f}, Method: {method_label}')

        if total_payment > 0:
            msg = f'Payment of PKR {total_payment:,.2f} recorded successfully! ({method_label})'
            if overpayment_credit > 0:
                msg += f' PKR {overpayment_credit:,.2f} exceeded the balance due and has been credited to the vendor as a remaining advance.'
            flash(msg, 'success')
        else:
            flash('No payment was processed.', 'warning')

    except Exception as e:
        db.session.rollback()
        flash(f'Error processing payment: {str(e)}', 'danger')

    return redirect(request.referrer or url_for('purchase.bill_detail', id=bill.id))


@bp.route('/bill/<int:id>/payment/<int:pay_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('purchases', action='edit')
def edit_bill_payment(id, pay_id):
    """Edit existing BillPayment for purchase bill"""
    bill = PurchaseBill.query.get_or_404(id)
    payment = BillPayment.query.filter_by(id=pay_id, bill_id=bill.id).first_or_404()

    if request.method == 'POST':
        old_amount = payment.amount

        # Update payment details
        payment_date_str = request.form.get('payment_date')
        if payment_date_str:
            try:
                payment.date = datetime.strptime(payment_date_str, '%Y-%m-%d')
            except ValueError:
                payment.date = datetime.utcnow()
        else:
            payment.date = datetime.utcnow()

        payment.payment_method = request.form.get('payment_method', 'Cash')
        # Keep existing reference_number if not provided in form
        if 'reference_number' in request.form:
            payment.reference_number = request.form.get('reference_number', '')
        payment.notes = request.form.get('payment_notes', '')

        new_amount = float(request.form.get('amount', 0))

        # Handle image update
        if 'payment_image' in request.files:
            payment_file = request.files['payment_image']
            if payment_file and payment_file.filename:
                filename = secure_filename(payment_file.filename)
                upload_dir = os.path.join('app', 'static', 'uploads', 'payments')
                os.makedirs(upload_dir, exist_ok=True)
                file_path = os.path.join(upload_dir, filename)
                payment_file.save(file_path)
                payment.image_path = os.path.join('uploads', 'payments', filename).replace('\\', '/')

        # IMPORTANT: Update payment amount BEFORE calculating delta effect on bill
        payment.amount = new_amount

        # Safe paid_amount update: adjust bill's paid_amount by the difference
        delta = new_amount - old_amount
        bill.paid_amount += delta
        if bill.paid_amount < 0:
            bill.paid_amount = 0
        if bill.paid_amount > bill.total:
            bill.paid_amount = bill.total
        bill.update_status()

        db.session.commit()
        log_activity('Purchase', f'Updated Payment on Bill #{bill.bill_number}',
                    f'Old: PKR {old_amount:,.2f}, New: PKR {new_amount:,.2f}')
        flash(f'Payment updated: PKR{new_amount:,.2f} ({payment.payment_method})', 'success')
        return redirect(url_for('purchase.bill_detail', id=bill.id))

    # GET: render edit form (modal embedded in page)
    return render_template('purchase/edit_bill_payment.html', payment=payment, bill=bill)


_NO_IMAGE_PLACEHOLDER_SVG = (
    '<svg xmlns="http://www.w3.org/2000/svg" width="90" height="70" viewBox="0 0 90 70">'
    '<rect width="90" height="70" fill="#f1f3f5"/>'
    '<text x="45" y="39" font-family="sans-serif" font-size="10" fill="#adb5bd" '
    'text-anchor="middle">No Image</text></svg>'
)


@bp.route('/bill_payment/<int:pay_id>/image')
@login_required
def bill_payment_image(pay_id):
    """Serve bill payment receipt image"""
    from flask import current_app, Response
    payment = BillPayment.query.get_or_404(pay_id)
    if payment.image_path:
        # Stored path may be relative to static (uploads/...) or absolute (app/static/...)
        if os.path.isabs(payment.image_path):
            file_path = payment.image_path
        else:
            file_path = os.path.join(current_app.root_path, 'static', payment.image_path)
        if os.path.exists(file_path):
            # Explicit inline PDF disposition: without this, some browsers/proxies
            # fall back to a "Save As" download prompt instead of rendering the
            # PDF inline in the thumbnail iframe / preview modal.
            if file_path.lower().endswith('.pdf'):
                return send_file(file_path, mimetype='application/pdf', as_attachment=False)
            return send_file(file_path)
    # File missing on this server — this route is used directly as an <img>/<iframe>
    # src, so a flash+redirect to an HTML page breaks the preview; serve a small
    # placeholder graphic instead.
    return Response(_NO_IMAGE_PLACEHOLDER_SVG, mimetype='image/svg+xml')


@bp.route('/bill/<int:id>/payment/<int:pay_id>/delete', methods=['POST'])
@login_required
@permission_required('purchases', action='delete')
def delete_bill_payment(id, pay_id):
    """Safely delete BillPayment: reverse paid_amount, cleanup transactions"""
    bill = PurchaseBill.query.get_or_404(id)
    payment = BillPayment.query.filter_by(id=pay_id, bill_id=bill.id).first_or_404()
    
    # Reverse from bill
    bill.paid_amount -= payment.amount
    if bill.paid_amount < 0:
        bill.paid_amount = 0
    bill.update_status()
    
    # Delete linked accounting transactions
    from app.models import Transaction
    Transaction.query.filter(
        or_(
            and_(Transaction.reference_type == 'payment', Transaction.reference_id == payment.id),
            and_(Transaction.reference_type == 'purchase', Transaction.reference_id == bill.id, Transaction.amount == payment.amount)
        )
    ).delete()
    
    # Delete old image if exists
    if payment.image_path:
        try:
            image_path = os.path.join(current_app.root_path, 'static', payment.image_path)
            if os.path.exists(image_path):
                os.remove(image_path)
        except:
            pass
    
    db.session.delete(payment)
    db.session.commit()
    log_activity('Purchase', f'Deleted Payment on Bill #{bill.bill_number}',
                f'Amount: PKR {payment.amount:,.2f}')
    flash(f'Payment deleted. Balance updated.', 'success')
    return redirect(url_for('purchase.bill_detail', id=bill.id))



@bp.route('/bill/<int:id>/receive', methods=['POST'])
@login_required
def receive_quantity(id):
    """Receive stock from a purchase bill — updates inventory and cost price."""
    bill = PurchaseBill.query.get_or_404(id)

    try:
        # Gather posted quantities per purchase_item_id
        receive_notes = request.form.get('receive_notes', '').strip() or None
        received_any = False

        # Build shipping/tax data for cost allocation
        total_items_cost = sum(item.total for item in bill.items)
        shipping_charge = bill.shipping_charge
        tax_rate = bill.tax_rate
        taxable_amount = total_items_cost + shipping_charge
        tax_amount = (taxable_amount * tax_rate) / 100 if tax_rate > 0 else 0
        total_additional_cost = shipping_charge + tax_amount

        # Create the receive record
        receive_record = BillReceive(
            bill_id=bill.id,
            notes=receive_notes,
            created_by=current_user.id
        )
        db.session.add(receive_record)
        db.session.flush()  # get receive_record.id

        for item in bill.items:
            field_name = f'qty_{item.id}'
            qty_str = request.form.get(field_name, '').strip()
            if not qty_str:
                continue
            try:
                qty_to_receive = float(qty_str)
            except ValueError:
                continue
            if qty_to_receive <= 0:
                continue

            warehouse_id = request.form.get(f'warehouse_{item.id}')
            if not warehouse_id:
                # Fallback to the warehouse set on the bill item
                warehouse_id = item.warehouse_id

            # Save receive item
            bri = BillReceiveItem(
                receive_id=receive_record.id,
                purchase_item_id=item.id,
                product_id=item.product_id,
                quantity_received=qty_to_receive,
                warehouse_id=int(warehouse_id) if warehouse_id else None
            )
            db.session.add(bri)
            db.session.flush()  # Get bri.id for cost history link

            # Update inventory
            product = Product.query.get(item.product_id)
            if product:
                old_qty = product.quantity
                product.update_quantity(qty_to_receive)
                
                # Update per-warehouse stock in the bridge table
                if bri.warehouse_id:
                    wh_stock = ProductWarehouseStock.query.filter_by(
                        product_id=item.product_id,
                        warehouse_id=bri.warehouse_id
                    ).first()
                    if not wh_stock:
                        wh_stock = ProductWarehouseStock(
                            product_id=item.product_id,
                            warehouse_id=bri.warehouse_id,
                            quantity=0
                        )
                        db.session.add(wh_stock)
                    wh_stock.quantity += qty_to_receive
                elif product.warehouse_id:
                    # Legacy fallback
                    wh_stock = ProductWarehouseStock.query.filter_by(
                        product_id=item.product_id,
                        warehouse_id=product.warehouse_id
                    ).first()
                    if not wh_stock:
                        wh_stock = ProductWarehouseStock(
                            product_id=item.product_id,
                            warehouse_id=product.warehouse_id,
                            quantity=0
                        )
                        db.session.add(wh_stock)
                    wh_stock.quantity += qty_to_receive

                # Calculate proportional cost (shipping + tax allocated by item value)
                if total_items_cost > 0:
                    allocation_ratio = item.total / total_items_cost
                    allocated_additional = total_additional_cost * allocation_ratio
                else:
                    allocated_additional = 0
                new_unit_cost = item.unit_price + (allocated_additional / item.quantity) if item.quantity > 0 else item.unit_price

                # Record cost price history
                old_price = product.cost_price
                cost_history = CostPriceHistory(
                    product_id=item.product_id,
                    purchase_bill_id=bill.id,
                    bill_receive_item_id=bri.id,
                    old_price=old_price if old_price > 0 else None,
                    new_price=new_unit_cost,
                    quantity_at_old_price=old_qty,
                    used_quantity=0,
                    reason=f"Received {qty_to_receive} units from Bill {bill.bill_number} — Base: PKR {item.unit_price:.2f}, With shipping/tax: PKR {new_unit_cost:.2f}",
                    is_active=True,
                    created_by=current_user.id
                )
                db.session.add(cost_history)
                product.cost_price = new_unit_cost

                # Trigger BOM versioning
                try:
                    user_id = current_user.id if current_user.is_authenticated else 1
                    BOMVersioningService.check_and_update_bom_for_cost_changes(
                        product_id=item.product_id,
                        created_by_id=user_id
                    )
                except Exception as e:
                    print(f"BOM versioning error for product {item.product_id}: {e}")

                received_any = True

        if received_any:
            bill.inventory_received = True
            db.session.commit()
            log_activity('Purchase', f'Received Stock for Bill #{bill.bill_number}',
                        f'Vendor: {bill.vendor.name}')
            flash('Quantity received and inventory updated successfully!', 'success')
        else:
            db.session.rollback()
            flash('No quantities were entered. Nothing was received.', 'warning')

    except Exception as e:
        db.session.rollback()
        flash(f'Error receiving quantity: {str(e)}', 'danger')

    return redirect(url_for('purchase.bill_detail', id=bill.id))

@bp.route('/bill/<int:id>/discount', methods=['POST'])
@login_required
def apply_discount(id):
    bill = PurchaseBill.query.get_or_404(id)
    discount_amount = float(request.form.get('discount_amount', 0))
    
    if discount_amount > 0:
        bill.discount += discount_amount
        bill.calculate_totals()
        bill.update_status()
        
        db.session.commit()
        log_activity('Purchase', f'Applied Discount on Bill #{bill.bill_number}',
                    f'Discount: PKR {discount_amount:,.2f}')
        flash(f'Discount of PKR {discount_amount} applied successfully!', 'success')

    return redirect(request.referrer or url_for('purchase.bill_detail', id=bill.id))

@bp.route('/bill/receive/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('purchases', action='delete')
def delete_receive(id):
    """Reverse a receive entry: restore inventory and revert cost price if applicable"""
    receive = BillReceive.query.get_or_404(id)
    bill = receive.bill
    
    try:
        # Track affected products and their old costs
        affected_products = {}
        product_prev_cost = {}  # product_id -> old_price from deactivated entry
        
        for bri in receive.receive_items:
            product = Product.query.get(bri.product_id)
            if product:
                # Restore inventory (subtract received quantity)
                product.update_quantity(-bri.quantity_received)
                
                # Revert from the specific warehouse in bridge table
                if bri.warehouse_id:
                    wh_stock = ProductWarehouseStock.query.filter_by(
                        product_id=bri.product_id,
                        warehouse_id=bri.warehouse_id
                    ).first()
                    if wh_stock:
                        wh_stock.quantity -= bri.quantity_received
                        if wh_stock.quantity < 0:
                            wh_stock.quantity = 0
                
                affected_products[product.id] = product
            
            # Deactivate cost price history entries linked to this BillReceiveItem
            cost_entries = CostPriceHistory.query.filter_by(bill_receive_item_id=bri.id).all()
            for ch in cost_entries:
                ch.is_active = False
                ch.bill_receive_item_id = None  # break link to allow BillReceiveItem deletion
                # Capture old_price for product if not already stored
                if product and product.id not in product_prev_cost:
                    product_prev_cost[product.id] = ch.old_price
        
        # For each affected product, set cost price to latest active cost history (if any)
        for product in affected_products.values():
            latest_active = CostPriceHistory.query.filter_by(product_id=product.id, is_active=True)\
                .order_by(CostPriceHistory.change_date.desc()).first()
            if latest_active:
                product.cost_price = latest_active.new_price
            elif product.id in product_prev_cost and product_prev_cost[product.id] is not None:
                product.cost_price = product_prev_cost[product.id]
            # else: keep current cost
            
            # Trigger BOM versioning if cost changed
            try:
                user_id = current_user.id if current_user.is_authenticated else 1
                BOMVersioningService.check_and_update_bom_for_cost_changes(
                    product_id=product.id,
                    created_by_id=user_id
                )
            except Exception as e:
                print(f"BOM versioning error on reverse receive: {e}")
        
        # Delete the receive (cascades to BillReceiveItems)
        db.session.delete(receive)
        
        # If no more receives for this bill, update flag
        remaining = BillReceive.query.filter_by(bill_id=bill.id).count()
        if remaining == 0:
            bill.inventory_received = False
        
        db.session.commit()
        log_activity('Purchase', f'Reversed Receive for Bill #{bill.bill_number}',
                    f'Vendor: {bill.vendor.name}')
        flash('Receive entry reversed successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error reversing receive: {str(e)}', 'danger')

    return redirect(request.referrer or url_for('purchase.bill_detail', id=bill.id))

@bp.route('/bill/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('purchases', action='delete')
def delete_bill(id):
    bill = PurchaseBill.query.get_or_404(id)
    # Only revert inventory if stock was already received
    if bill.inventory_received:
        # Use BillReceive records for accurate reversal if they exist
        if bill.bill_receives:
            for receive in bill.bill_receives:
                for bri in receive.receive_items:
                    product = Product.query.get(bri.product_id)
                    if product:
                        product.update_quantity(-bri.quantity_received)
                        # Revert from the specific warehouse in the receive item
                        if bri.warehouse_id:
                            wh_stock = ProductWarehouseStock.query.filter_by(
                                product_id=bri.product_id,
                                warehouse_id=bri.warehouse_id
                            ).first()
                            if wh_stock:
                                wh_stock.quantity -= bri.quantity_received
                                if wh_stock.quantity < 0:
                                    wh_stock.quantity = 0
        else:
            # Fallback for legacy bills without receive records
            for item in bill.items:
                product = Product.query.get(item.product_id)
                if product:
                    product.update_quantity(-item.quantity)
                    # Revert from the specific warehouse in the bill item
                    if item.warehouse_id:
                        wh_stock = ProductWarehouseStock.query.filter_by(
                            product_id=item.product_id,
                            warehouse_id=item.warehouse_id
                        ).first()
                        if wh_stock:
                            wh_stock.quantity -= item.quantity
                            if wh_stock.quantity < 0:
                                wh_stock.quantity = 0

    # Handle cost adjustments / BOM versioning
    for item in bill.items:
        try:
            user_id = current_user.id if current_user.is_authenticated else 1
            BOMVersioningService.check_and_update_bom_for_cost_changes(
                product_id=item.product_id,
                created_by_id=user_id
            )
        except:
            pass

    # Delete associated bill image if exists
    if bill.bill_image_path:
        try:
            image_full_path = os.path.join(current_app.root_path, 'static', bill.bill_image_path)
            if os.path.exists(image_full_path):
                os.remove(image_full_path)
        except Exception as e:
            print(f"Warning: Could not delete bill image: {e}")

    # Give back any vendor advance credit that was applied to this bill (auto-apply
    # on creation or the manual "Advance Paid" field), so the vendor's remaining
    # advance balance is restored instead of staying permanently consumed.
    from app.utils import reverse_bill_advance_applied
    advance_reversed = reverse_bill_advance_applied(bill)

    bill_number = bill.bill_number
    vendor_name = bill.vendor.name
    db.session.delete(bill)
    db.session.commit()
    log_activity('Purchase', f'Deleted Bill #{bill_number}',
                f'Vendor: {vendor_name}, Advance restored: {advance_reversed:,.2f}')
    if advance_reversed > 0:
        flash(f'Purchase bill deleted. Inventory reversed and PKR {advance_reversed:,.2f} of vendor advance was restored.', 'success')
    else:
        flash('Purchase bill deleted successfully. Inventory and warehouse stocks have been reversed.', 'success')
    return redirect(url_for('purchase.bills'))

@bp.route('/bills/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_bills():
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No bills selected'}), 400
    
    deleted_count = 0
    errors = []
    
    for bill_id in ids:
        bill = PurchaseBill.query.get(bill_id)
        if not bill:
            continue
            
        try:
            # Revert inventory (Global + Warehouse)
            if bill.inventory_received:
                if bill.bill_receives:
                    for receive in bill.bill_receives:
                        for bri in receive.receive_items:
                            product = Product.query.get(bri.product_id)
                            if product:
                                product.update_quantity(-bri.quantity_received)
                                if bri.warehouse_id:
                                    wh_stock = ProductWarehouseStock.query.filter_by(
                                        product_id=bri.product_id,
                                        warehouse_id=bri.warehouse_id
                                    ).first()
                                    if wh_stock:
                                        wh_stock.quantity -= bri.quantity_received
                                        if wh_stock.quantity < 0:
                                            wh_stock.quantity = 0
                else:
                    for item in bill.items:
                        product = Product.query.get(item.product_id)
                        if product:
                            product.update_quantity(-item.quantity)
                            if item.warehouse_id:
                                wh_stock = ProductWarehouseStock.query.filter_by(
                                    product_id=item.product_id,
                                    warehouse_id=item.warehouse_id
                                ).first()
                                if wh_stock:
                                    wh_stock.quantity -= item.quantity
                                    if wh_stock.quantity < 0:
                                        wh_stock.quantity = 0

            # Restore any vendor advance applied to this bill before deleting it
            from app.utils import reverse_bill_advance_applied
            reverse_bill_advance_applied(bill)

            db.session.delete(bill)
            deleted_count += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f'Error deleting Bill {bill.bill_number}: {str(e)}')
            
    if deleted_count > 0:
        db.session.commit()
        
    message = f'Successfully deleted {deleted_count} bills.'
    if errors:
        return jsonify({'success': False, 'message': message, 'errors': errors}), 500
    return jsonify({'success': True, 'message': message})

@bp.route('/vendors')
@login_required
def vendors():
    status = request.args.get('status', 'all')
    search = request.args.get('search', '')
    vendor_id = request.args.get('vendor_id', type=int)
    city = request.args.get('city', '')
    state = request.args.get('state', '')
    country = request.args.get('country', '')
    postal_code = request.args.get('postal_code', '')
    
    query = Vendor.query
    
    if status == 'active':
        query = query.filter_by(is_active=True)
    elif status == 'inactive':
        query = query.filter_by(is_active=False)
    
    if city:
        query = query.filter(Vendor.city == city)
    if state:
        query = query.filter(Vendor.state == state)
    if country:
        query = query.filter(Vendor.country == country)
    if postal_code:
        query = query.filter(Vendor.postal_code == postal_code)
        
    if vendor_id:
        query = query.filter_by(id=vendor_id)
    elif search:
        search_filter = f"%{search}%"
        query = query.filter(
            (Vendor.name.ilike(search_filter)) |
            (Vendor.email.ilike(search_filter)) |
            (Vendor.phone.ilike(search_filter)) |
            (Vendor.city.ilike(search_filter)) |
            (Vendor.state.ilike(search_filter)) |
            (Vendor.country.ilike(search_filter)) |
            (Vendor.postal_code.ilike(search_filter))
        )
    
    query = apply_saved_filter_to_query(query, 'vendor', request.args)

    vendors = query.order_by(Vendor.name.asc()).all()
            
    # List of all vendors for the searchable dropdown
    all_vendors = Vendor.query.order_by(Vendor.name.asc()).all()
    
    # Get distinct cities, states, countries, and postal codes for filters
    all_cities = db.session.query(Vendor.city).filter(Vendor.city != None, Vendor.city != '').distinct().all()
    all_cities = [c[0] for c in all_cities]
    
    all_states = db.session.query(Vendor.state).filter(Vendor.state != None, Vendor.state != '').distinct().all()
    all_states = [s[0] for s in all_states]

    all_countries = db.session.query(Vendor.country).filter(Vendor.country != None, Vendor.country != '').distinct().all()
    all_countries = [c[0] for c in all_countries]

    all_postals = db.session.query(Vendor.postal_code).filter(Vendor.postal_code != None, Vendor.postal_code != '').distinct().all()
    all_postals = [p[0] for p in all_postals]
    
    return render_template('purchase/vendors.html', 
                         vendors=vendors,
                         all_vendors=all_vendors, 
                         all_cities=all_cities,
                         all_states=all_states,
                         all_countries=all_countries,
                         all_postals=all_postals,
                         current_status=status, 
                         search_query=search,
                         selected_vendor_id=vendor_id,
                         selected_city=city,
                         selected_state=state,
                         selected_country=country,
                         selected_postal=postal_code,
                         active_module='vendor',
                         filter_id=request.args.get('filter_id'))

@bp.route('/vendor/bulk-upload', methods=['GET', 'POST'])
@login_required
def bulk_upload_vendor():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('purchase.bulk_upload_vendor'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('purchase.bulk_upload_vendor'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('purchase.bulk_upload_vendor'))
        
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            wb = load_workbook(filename=BytesIO(file.read()), read_only=True)
            ws = wb.active
            rows = list(ws.values)
            if not rows:
                flash('File is empty', 'error')
                return redirect(url_for('purchase.bulk_upload_vendor'))
            headers = [str(h) if h else '' for h in rows[0]]
            df_data = []
            for row in rows[1:]:
                df_data.append({headers[i]: row[i] for i in range(len(headers))})
            import pandas as pd
            df = pd.DataFrame(df_data)
            
            required_columns = ['name']
            missing = [col for col in required_columns if col not in df.columns]
            if missing:
                flash(f'Missing required columns: {", ".join(missing)}', 'error')
                return redirect(url_for('purchase.bulk_upload_vendor'))
            
            added = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    name = str(row.get('name', '')).strip()
                    
                    if not name:
                        errors.append(f'Row {idx + 2}: Missing name')
                        continue
                    
                    vendor = Vendor(
                        name=name,
                        email=str(row.get('email', '')).strip() if pd.notna(row.get('email')) else None,
                        phone=str(row.get('phone', '')).strip() if pd.notna(row.get('phone')) else None,
                        address=str(row.get('address', '')).strip() if pd.notna(row.get('address')) else None,
                        shipping_address=str(row.get('shipping_address', '')).strip() if pd.notna(row.get('shipping_address')) else None,
                        gst_number=str(row.get('gst_number', '')).strip() if pd.notna(row.get('gst_number')) else None,
                        pan_number=str(row.get('pan_number', '')).strip() if pd.notna(row.get('pan_number')) else None,
                        contact_person=str(row.get('contact_person', '')).strip() if pd.notna(row.get('contact_person')) else None,
                        bank_name=str(row.get('bank_name', '')).strip() if pd.notna(row.get('bank_name')) else None,
                        account_holder_name=str(row.get('account_holder_name', '')).strip() if pd.notna(row.get('account_holder_name')) else None,
                        account_number=str(row.get('account_number', '')).strip() if pd.notna(row.get('account_number')) else None,
                        ifsc_code=str(row.get('ifsc_code', '')).strip() if pd.notna(row.get('ifsc_code')) else None,
                        swift_code=str(row.get('swift_code', '')).strip() if pd.notna(row.get('swift_code')) else None,
                    )
                    
                    db.session.add(vendor)
                    added += 1
                except Exception as e:
                    errors.append(f'Row {idx + 2}: {str(e)}')
            
            db.session.commit()
            if added > 0:
                log_activity('Vendors', f'Bulk Uploaded Vendors',
                            f'{added} vendors added')

            if added > 0:
                flash(f'Successfully added {added} vendors!', 'success')
            if errors:
                flash(f'Errors: {"; ".join(errors[:10])}', 'warning')

            return redirect(url_for('purchase.vendors'))
            
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'error')
            return redirect(url_for('purchase.bulk_upload_vendor'))
    
    return render_template('purchase/bulk_upload_vendor.html')

@bp.route('/vendor/download-sample')
@login_required
def download_vendor_sample():
    try:
        from openpyxl import Workbook
        from io import BytesIO
        from flask import send_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Vendors'
        
        headers = ['name', 'email', 'phone', 'address', 'shipping_address', 'gst_number', 'pan_number', 'contact_person', 'bank_name', 'account_holder_name', 'account_number', 'ifsc_code', 'swift_code']
        ws.append(headers)
        
        sample_data = [
            ['Vendor A', 'vendorA@example.com', '1234567890', '123 Factory St, City', '123 Factory St, City', 'GST123456789', 'ABCDE1234F', 'John Doe', 'State Bank of India', 'John Doe', '123456789012', 'SBIN0001234', 'SBININBB104'],
            ['Vendor B', 'vendorB@example.com', '2345678901', '456 Warehouse Ave, Town', '456 Warehouse Ave, Town', 'GST987654321', 'FGHI5678K', 'Jane Smith', 'HDFC Bank', 'Jane Smith', '987654321098', 'HDFC0005678', 'HDFCINBB'],
            ['Vendor C', 'vendorC@example.com', '3456789012', '789 Supply Rd, Village', '789 Supply Rd, Village', '', '', '', '', '', '', '', '']
        ]
        
        for row in sample_data:
            ws.append(row)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='sample_vendors.xlsx', as_attachment=True)
        
    except Exception as e:
        flash(f'Error creating sample: {str(e)}', 'error')
        return redirect(url_for('purchase.vendors'))

@bp.route('/vendor/<int:id>')
@login_required
def vendor_profile(id):
    from datetime import datetime as _dt
    vendor = Vendor.query.get_or_404(id)

    # ── Date filter ────────────────────────────────────────────────────────
    date_from_str = request.args.get('date_from', '').strip()
    date_to_str   = request.args.get('date_to',   '').strip()
    date_from = None
    date_to   = None
    try:
        if date_from_str:
            date_from = _dt.strptime(date_from_str, '%Y-%m-%d')
        if date_to_str:
            date_to = _dt.strptime(date_to_str, '%Y-%m-%d').replace(
                hour=23, minute=59, second=59)
    except ValueError:
        pass

    all_bills = sorted(vendor.bills, key=lambda b: b.date, reverse=True)
    if date_from or date_to:
        filtered_bills = [
            b for b in all_bills
            if (date_from is None or b.date >= date_from)
            and (date_to   is None or b.date <= date_to)
        ]
    else:
        filtered_bills = all_bills

    advances = sorted(vendor.advances, key=lambda a: a.date, reverse=True)

    return render_template('purchase/vendor_profile.html',
                           vendor=vendor,
                           bills=filtered_bills,
                           advances=advances,
                           date_from=date_from_str,
                           date_to=date_to_str)

@bp.route('/vendor/add', methods=['GET', 'POST'])
@login_required
@permission_required('vendors', action='add')
def add_vendor():
    form = VendorForm()
    if form.validate_on_submit():
        vendor = Vendor(
            name=form.name.data,
            email=form.email.data,
            phone=form.phone.data,
            address=form.address.data,
            city=form.city.data,
            state=form.state.data,
            country=form.country.data,
            postal_code=form.postal_code.data,
            gst_number=form.gst_number.data,
            payment_method=form.payment_method.data,
            bank_name=form.bank_name.data,
            account_holder_name=form.account_holder_name.data,
            account_number=form.account_number.data,
            swift_code=form.swift_code.data,
            ifsc_code=form.ifsc_code.data,
            sub_vendors=form.sub_vendors.data # JSON from JS
        )

        # Handle image upload
        if 'image' in request.files:
            image_file = request.files['image']
            if image_file and image_file.filename:
                filename = secure_filename(image_file.filename)
                image_path = os.path.join('app', 'static', 'uploads', 'vendors', filename)
                os.makedirs(os.path.dirname(image_path), exist_ok=True)
                image_file.save(image_path)
                vendor.image_path = image_path.replace('\\', '/')
        
        db.session.add(vendor)
        db.session.commit()
        
        log_activity('Vendors', f'Added Vendor: {vendor.name}', 
                    f'Company: {vendor.company_name or "N/A"}')
        
        flash('Vendor added successfully!', 'success')
        return redirect(url_for('purchase.vendors'))
    
    return render_template('purchase/add_vendor.html', form=form)


@bp.route('/vendor/<int:id>/advance', methods=['POST'])
@login_required
@permission_required('vendors', action='add')
def vendor_give_advance(id):
    """Record a quick advance payment given to a vendor against material."""
    vendor = Vendor.query.get_or_404(id)
    amount = float(request.form.get('amount', 0))
    description = request.form.get('description', '').strip()
    advance_date_str = request.form.get('advance_date', '')
    
    if amount <= 0:
        flash('Advance amount must be greater than zero.', 'danger')
        return redirect(url_for('purchase.vendor_profile', id=id))
    
    try:
        advance_date = datetime.strptime(advance_date_str, '%Y-%m-%d').date() if advance_date_str else date.today()
    except ValueError:
        advance_date = date.today()
    
    advance = VendorAdvance(
        vendor_id=vendor.id,
        amount=amount,
        date=advance_date,
        description=description or 'Advance against material',
        created_by=current_user.id
    )
    db.session.add(advance)
    db.session.commit()
    log_activity('Vendors', f'Recorded Advance for {vendor.name}',
                f'Amount: PKR {amount:,.2f}')
    flash(f'Advance of PKR {amount:,.2f} recorded for {vendor.name}.', 'success')
    return redirect(url_for('purchase.vendor_profile', id=id))


@bp.route('/vendor/<int:id>/advance/<int:adv_id>/adjust', methods=['POST'])
@login_required
@permission_required('vendors', action='edit')
def vendor_adjust_advance(id, adv_id):
    """Apply a vendor advance against a specific bill."""
    advance = VendorAdvance.query.get_or_404(adv_id)
    bill_id = request.form.get('bill_id')
    
    # Check if advance still has remaining balance
    if advance.remaining_balance <= 0:
        flash('This advance has no remaining balance to apply.', 'warning')
        return redirect(url_for('purchase.vendor_profile', id=id))
    
    if bill_id:
        bill = PurchaseBill.query.get(int(bill_id))
        if bill and bill.vendor_id == id:
            # Calculate how much of the advance can be applied to this bill
            bill_remaining = bill.total - bill.paid_amount
            apply_amount = min(advance.remaining_balance, bill_remaining)
            
            if apply_amount > 0:
                # Deduct the applied amount from the advance and bill
                advance.applied_amount += apply_amount
                bill.paid_amount += apply_amount
                bill.update_status()
                
                # Mark as fully adjusted only if the entire advance amount is applied
                if advance.applied_amount >= advance.amount:
                    advance.is_adjusted = True
                    advance.adjusted_bill_id = bill.id
                    db.session.commit()
                    log_activity('Vendors', f'Adjusted Advance Against Bill #{bill.bill_number}',
                                f'Amount: PKR {apply_amount:,.2f}, Fully Applied')
                    flash(f'Advance of PKR {apply_amount:,.2f} fully applied to bill {bill.bill_number}.', 'success')
                else:
                    db.session.commit()
                    log_activity('Vendors', f'Adjusted Advance Against Bill #{bill.bill_number}',
                                f'Amount: PKR {apply_amount:,.2f}, Remaining: PKR {advance.remaining_balance:,.2f}')
                    flash(f'Advance of PKR {apply_amount:,.2f} applied to bill {bill.bill_number}. Remaining PKR {advance.remaining_balance:,.2f} still available as advance.', 'success')
            return redirect(url_for('purchase.vendor_profile', id=id))
    return redirect(url_for('purchase.vendor_profile', id=id))


@bp.route('/vendor/<int:id>/advance/<int:adv_id>/delete', methods=['POST'])
@login_required
@permission_required('vendors', action='delete')
def vendor_delete_advance(id, adv_id):
    """Delete a vendor advance (reverses applied amount from bill if any was applied)."""
    advance = VendorAdvance.query.get_or_404(adv_id)
    if advance.vendor_id != id:
        flash('Invalid advance.', 'danger')
        return redirect(url_for('purchase.vendor_profile', id=id))
    
    # Reverse the applied amount from the adjusted bill
    if advance.applied_amount > 0 and advance.adjusted_bill_id:
        bill = PurchaseBill.query.get(advance.adjusted_bill_id)
        if bill:
            bill.paid_amount = max(0, bill.paid_amount - advance.applied_amount)
            bill.update_status()
    
    advance_amount = advance.amount
    db.session.delete(advance)
    db.session.commit()
    log_activity('Vendors', f'Deleted Vendor Advance',
                f'Amount: PKR {advance_amount:,.2f}')
    flash('Advance deleted and any applied amount has been reversed.', 'success')
    return redirect(url_for('purchase.vendor_profile', id=id))

@bp.route('/vendor/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('vendors', action='edit')
def edit_vendor(id):
    vendor = Vendor.query.get_or_404(id)
    form = VendorForm(obj=vendor)
    
    if form.validate_on_submit():
        vendor.name = form.name.data
        vendor.email = form.email.data
        vendor.phone = form.phone.data
        vendor.address = form.address.data
        vendor.city = form.city.data
        vendor.state = form.state.data
        vendor.country = form.country.data
        vendor.postal_code = form.postal_code.data
        vendor.gst_number = form.gst_number.data
        vendor.payment_method = form.payment_method.data
        vendor.bank_name = form.bank_name.data
        vendor.account_holder_name = form.account_holder_name.data
        vendor.account_number = form.account_number.data
        vendor.swift_code = form.swift_code.data
        vendor.ifsc_code = form.ifsc_code.data
        vendor.sub_vendors = form.sub_vendors.data # JSON from JS

        # Handle image upload
        if 'image' in request.files:
            image_file = request.files['image']
            if image_file and image_file.filename:
                filename = secure_filename(image_file.filename)
                image_path = os.path.join('app', 'static', 'uploads', 'vendors', filename)
                os.makedirs(os.path.dirname(image_path), exist_ok=True)
                image_file.save(image_path)
                vendor.image_path = image_path.replace('\\', '/')
        
        db.session.commit()
        log_activity('Vendors', f'Updated Vendor: {vendor.name}',
                    f'Vendor ID: {vendor.id}')
        flash('Vendor updated successfully!', 'success')
        return redirect(url_for('purchase.vendors'))

    return render_template('purchase/edit_vendor.html', form=form, vendor=vendor)

@bp.route('/vendor/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('vendors', action='delete')
def delete_vendor(id):
    vendor = Vendor.query.get_or_404(id)
    
    # Check if vendor has associated records (Purchases or Sales)
    if vendor.bills or vendor.sales or vendor.expenses:
        flash('Cannot delete vendor as they have associated transaction records.', 'danger')
        return redirect(url_for('purchase.vendors'))

    vendor_name = vendor.name
    db.session.delete(vendor)
    db.session.commit()
    log_activity('Vendors', f'Deleted Vendor: {vendor_name}',
                f'Vendor removed')
    flash('Vendor deleted successfully!', 'success')
    return redirect(url_for('purchase.vendors'))

@bp.route('/vendors/bulk-delete', methods=['POST'])
@login_required
@permission_required('vendors', action='delete')
def bulk_delete_vendors():
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No vendors selected'}), 400
    
    deleted_count = 0
    skipped_count = 0
    errors = []
    
    for vendor_id in ids:
        vendor = Vendor.query.get(vendor_id)
        if not vendor:
            continue
            
        # Check associations
        if vendor.bills or vendor.sales or vendor.expenses:
            skipped_count += 1
            continue
            
        try:
            db.session.delete(vendor)
            deleted_count += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f'Error deleting Vendor {vendor.name}: {str(e)}')
            
    if deleted_count > 0:
        db.session.commit()
        
    message = f'Successfully deleted {deleted_count} vendors.'
    if skipped_count > 0:
        message += f' Skipped {skipped_count} vendors with associated records.'
    
    if errors:
        return jsonify({'success': False, 'message': message, 'errors': errors}), 500
    return jsonify({'success': True, 'message': message})


# ---------------------------------------------------------------------------
# AJAX: Vendor advance info (used by create_bill and create_po forms)
# ---------------------------------------------------------------------------
@bp.route('/api/vendor/<int:vendor_id>/advances')
@login_required
def vendor_advance_info(vendor_id):
    """Return JSON with pending (unadjusted) advance balance for a vendor."""
    vendor = Vendor.query.get_or_404(vendor_id)
    pending = [a for a in vendor.advances if a.remaining_balance > 0]
    return jsonify({
        'vendor_id': vendor_id,
        'vendor_name': vendor.name,
        'total_advances': vendor.total_advances_given,
        'total_adjusted': vendor.total_advances_adjusted,
        'pending_balance': vendor.remaining_advance_balance,
        'advances': [
            {
                'id': a.id,
                'amount': a.amount,
                'applied_amount': a.applied_amount,
                'remaining': a.remaining_balance,
                'date': a.date.strftime('%d-%m-%Y'),
                'description': a.description or ''
            }
            for a in pending
        ]
    })


# ---------------------------------------------------------------------------
# AJAX: confirmed POs for a vendor (used by create_bill form)
# ---------------------------------------------------------------------------
@bp.route('/api/vendor/<int:vendor_id>/purchase-orders')
@login_required
def vendor_pos(vendor_id):
    """Return confirmed POs for a vendor (for the create-bill dropdown)."""
    pos = PurchaseOrder.query.filter_by(
        vendor_id=vendor_id, status='Confirmed'
    ).order_by(PurchaseOrder.date.desc()).all()
    return jsonify([
        {
            'id': po.id,
            'po_number': po.po_number,
            'date': po.date.strftime('%d-%m-%Y'),
            'total': po.total,
            'items': [
                {
                    'product_id': item.product_id,
                    'product_name': item.product.name,
                    'quantity': item.quantity,
                    'unit_price': item.unit_price,
                    'total': item.total
                }
                for item in po.items
            ],
            'shipping_charge': po.shipping_charge,
            'notes': po.notes or ''
        }
        for po in pos
    ])


# ---------------------------------------------------------------------------
# Purchase Orders
# ---------------------------------------------------------------------------
@bp.route('/orders')
@login_required
def purchase_orders():
    status = request.args.get('status', 'all')
    search = request.args.get('search', '')

    query = PurchaseOrder.query
    if status != 'all':
        query = query.filter_by(status=status)
    if search:
        query = query.join(Vendor).filter(
            PurchaseOrder.po_number.ilike(f'%{search}%') |
            Vendor.name.ilike(f'%{search}%')
        )
    query = apply_saved_filter_to_query(query, 'purchase_order', request.args)

    orders = query.order_by(PurchaseOrder.date.desc()).all()
    return render_template('purchase/purchase_orders.html',
                           orders=orders, current_status=status, search=search,
                           active_module='purchase_order',
                           filter_id=request.args.get('filter_id'))


def _parse_delivery_time(dt_str):
    """Parse datetime-local input string."""
    if not dt_str:
        return None
    try:
        return datetime.strptime(dt_str, '%Y-%m-%dT%H:%M')
    except ValueError:
        try:
            return datetime.strptime(dt_str, '%Y-%m-%d')
        except ValueError:
            return None


@bp.route('/order/create', methods=['GET', 'POST'])
@login_required
@permission_required('purchases', action='add')
def create_po():
    vendors = Vendor.query.filter_by(is_active=True).all()
    products = Product.query.filter_by(is_active=True).all()
    warehouses = Warehouse.query.all()

    if request.method == 'POST':
        vendor_id = int(request.form.get('vendor_id'))
        expected_date_str = request.form.get('expected_date', '')
        expected_date = datetime.strptime(expected_date_str, '%Y-%m-%d') if expected_date_str else None

        product_ids = request.form.getlist('product_id[]')
        quantities  = request.form.getlist('quantity[]')
        prices      = request.form.getlist('price[]')
        warehouse_ids = request.form.getlist('warehouse_id[]')

        # Auto-generate PO number using settings
        settings = PurchaseSettings.query.first()
        if not settings:
            settings = PurchaseSettings(bill_prefix='PB-', bill_suffix='', next_bill_number=1,
                                        po_prefix='PO-', po_suffix='', next_po_number=1)
            db.session.add(settings)
        po_number = f"{settings.po_prefix}{settings.next_po_number}{settings.po_suffix}"
        settings.next_po_number += 1

        po = PurchaseOrder(
            po_number=po_number,
            vendor_id=vendor_id,
            date=datetime.utcnow(),
            expected_date=expected_date,
            delivery_start=_parse_delivery_time(request.form.get('delivery_start')),
            delivery_end=_parse_delivery_time(request.form.get('delivery_end')),
            advance_amount=float(request.form.get('advance_amount', 0)),
            shipping_charge=float(request.form.get('shipping_charge', 0)),
            tax_rate=float(request.form.get('tax_rate', 0)),
            discount=float(request.form.get('discount', 0)),
            notes=request.form.get('notes', ''),
            status='Draft',
            created_by=current_user.id
        )
        db.session.add(po)

        for i in range(len(product_ids)):
            if product_ids[i] and quantities[i] and float(quantities[i]) > 0:
                qty   = float(quantities[i])
                price = float(prices[i])
                wh_id = int(warehouse_ids[i]) if i < len(warehouse_ids) and warehouse_ids[i] else None
                item  = PurchaseOrderItem(
                    product_id=int(product_ids[i]),
                    quantity=qty,
                    unit_price=price,
                    warehouse_id=wh_id,
                    total=qty * price
                )
                po.items.append(item)

        po.calculate_totals()
        
        # If advance amount is paid, also record it as vendor advance
        advance_amount = float(request.form.get('advance_amount', 0))
        if advance_amount > 0:
            is_admin = getattr(current_user, 'is_admin', False)
            vendor_advance = VendorAdvance(
                vendor_id=vendor_id,
                amount=advance_amount,
                date=datetime.utcnow().date(),
                description=f'Advance for PO {po_number}',
                is_approved=is_admin,
                approved_by=current_user.id if is_admin else None,
                approved_at=datetime.utcnow() if is_admin else None,
                created_by=current_user.id
            )
            db.session.add(vendor_advance)
        
        db.session.commit()
        log_activity('Purchase', f'Created PO #{po.po_number}',
                    f'Vendor: {Vendor.query.get(vendor_id).name}, Total: {po.total}')
        flash(f'Purchase Order {po.po_number} created!', 'success')
        return redirect(url_for('purchase.po_detail', id=po.id))

    return render_template('purchase/create_po.html', vendors=vendors, products=products, warehouses=warehouses)


@bp.route('/order/<int:id>')
@login_required
def po_detail(id):
    po = PurchaseOrder.query.get_or_404(id)
    start_ts = int(po.delivery_start.timestamp()) if po.delivery_start else None
    end_ts = int(po.delivery_end.timestamp()) if po.delivery_end else None
    return render_template('purchase/po_detail.html', po=po, start_ts=start_ts, end_ts=end_ts)


@bp.route('/order/<int:id>/pdf')
@login_required
def po_pdf(id):
    """Generate PDF for a Purchase Order."""
    po = PurchaseOrder.query.get_or_404(id)
    from app.models import Company
    company = Company.query.first()
    
    buffer = generate_professional_pdf('purchase_order', po, company, None)
    response = make_response(buffer)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename="{po.po_number}.pdf"'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@bp.route('/order/<int:id>/pdf/view')
@login_required
def po_pdf_view(id):
    """View PDF inline for sharing"""
    po = PurchaseOrder.query.get_or_404(id)
    from app.models import Company
    company = Company.query.first()
    
    buffer = generate_professional_pdf('purchase_order', po, company, None)
    return send_file(buffer, mimetype='application/pdf')

@bp.route('/order/<int:id>/pdf/share')
@login_required
def po_pdf_share(id):
    """Generate PDF, save to public folder, and return shareable link"""
    from flask import current_app
    import os
    import secrets
    
    po = PurchaseOrder.query.get_or_404(id)
    from app.models import Company
    company = Company.query.first()
    
    try:
        buffer = generate_professional_pdf('purchase_order', po, company, None)
        buffer.seek(0)
        pdf_data = buffer.getvalue()
        
        public_dir = os.path.join(current_app.root_path, 'static', 'shared_pdfs')
        os.makedirs(public_dir, exist_ok=True)
        
        token = secrets.token_urlsafe(8)
        filename = f"po_{po.po_number}_{token}.pdf"
        filepath = os.path.join(public_dir, filename)
        
        with open(filepath, 'wb') as f:
            f.write(pdf_data)
        
        share_url = url_for('static', filename=f'shared_pdfs/{filename}')
        return {'share_url': share_url}
    except Exception as e:
        return {'error': str(e)}, 500


@bp.route('/order/<int:id>/confirm', methods=['POST'])
@login_required
def confirm_po(id):
    po = PurchaseOrder.query.get_or_404(id)
    if po.status == 'Draft':
        po.status = 'Confirmed'
        db.session.commit()
        log_activity('Purchase', f'Confirmed PO #{po.po_number}',
                    f'Vendor: {po.vendor.name}')
        flash(f'Purchase Order {po.po_number} confirmed. You can now create a bill from it.', 'success')
    else:
        flash('Only Draft orders can be confirmed.', 'warning')
    return redirect(url_for('purchase.po_detail', id=id))


@bp.route('/order/<int:id>/cancel', methods=['POST'])
@login_required
def cancel_po(id):
    po = PurchaseOrder.query.get_or_404(id)
    if po.status in ('Draft', 'Confirmed'):
        po.status = 'Cancelled'
        db.session.commit()
        log_activity('Purchase', f'Cancelled PO #{po.po_number}',
                    f'Vendor: {po.vendor.name}')
        flash(f'Purchase Order {po.po_number} cancelled.', 'warning')
    else:
        flash('Cannot cancel a Converted or already Cancelled order.', 'danger')
    return redirect(url_for('purchase.po_detail', id=id))


@bp.route('/order/<int:id>/convert', methods=['POST'])
@login_required
def convert_po_to_bill(id):
    """Convert a confirmed PO into a Purchase Bill, pre-filling all items."""
    po = PurchaseOrder.query.get_or_404(id)
    if po.status != 'Confirmed':
        flash('Only Confirmed orders can be converted to a bill.', 'danger')
        return redirect(url_for('purchase.po_detail', id=id))

    last_bill = PurchaseBill.query.order_by(PurchaseBill.id.desc()).first()
    # Generate bill number using settings
    settings = PurchaseSettings.query.first()
    if not settings:
        settings = PurchaseSettings(bill_prefix='PB-', bill_suffix='', next_bill_number=1,
                                    po_prefix='PO-', po_suffix='', next_po_number=1)
        db.session.add(settings)
    bill_number = f"{settings.bill_prefix}{settings.next_bill_number}{settings.bill_suffix}"
    settings.next_bill_number += 1

    bill = PurchaseBill(
        bill_number=bill_number,
        vendor_id=po.vendor_id,
        po_id=po.id,
        date=datetime.utcnow(),
        due_date=datetime.utcnow() + timedelta(days=30),
        tax_rate=po.tax_rate,
        discount=po.discount,
        shipping_charge=po.shipping_charge,
        notes=po.notes,
        created_by=current_user.id
    )
    db.session.add(bill)

    for poi in po.items:
        item = PurchaseItem(
            product_id=poi.product_id,
            quantity=poi.quantity,
            unit_price=poi.unit_price,
            total=poi.total,
            warehouse_id=poi.warehouse_id
        )
        bill.items.append(item)
        # Update inventory
        product = Product.query.get(poi.product_id)
        if product:
            product.update_quantity(poi.quantity, warehouse_id=poi.warehouse_id)

    bill.calculate_totals()
    po.status = 'Converted'
    db.session.commit()
    log_activity('Purchase', f'Converted PO #{po.po_number} to Bill #{bill.bill_number}',
                f'Vendor: {po.vendor.name}, Total: {bill.total}')
    flash(f'Bill {bill.bill_number} created from PO {po.po_number}!', 'success')
    return redirect(url_for('purchase.bill_detail', id=bill.id))


@bp.route('/order/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('purchases', action='delete')
def delete_po(id):
    po = PurchaseOrder.query.get_or_404(id)
    po_number = po.po_number
    db.session.delete(po)
    db.session.commit()
    log_activity('Purchase', f'Deleted PO #{po_number}',
                f'Purchase Order removed')
    flash('Purchase Order deleted.', 'success')
    return redirect(url_for('purchase.purchase_orders'))

@bp.route('/purchase-orders/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_purchase_orders():
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No orders selected'}), 400
    
    deleted_count = 0
    errors = []
    
    for po_id in ids:
        po = PurchaseOrder.query.get(po_id)
        if not po:
            continue
            
        try:
            db.session.delete(po)
            deleted_count += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f'Error deleting Order {po.po_number}: {str(e)}')
            
    if deleted_count > 0:
        db.session.commit()
        
    message = f'Successfully deleted {deleted_count} orders.'
    if errors:
        return jsonify({'success': False, 'message': message, 'errors': errors}), 500
    return jsonify({'success': True, 'message': message})


# ---------------------------------------------------------------------------
# API: Get cost price history for a product
# ---------------------------------------------------------------------------
@bp.route('/api/product/<int:product_id>/cost-history')
@login_required
def product_cost_history(product_id):
    """Return JSON with cost price history for a product"""
    product = Product.query.get_or_404(product_id)
    history = CostPriceHistory.query.filter_by(product_id=product_id).order_by(CostPriceHistory.change_date.desc()).all()
    
    return jsonify({
        'product_id': product_id,
        'product_name': product.name,
        'current_cost_price': product.cost_price,
        'total_quantity': product.quantity,
        'history': [
            {
                'id': h.id,
                'old_price': h.old_price or 0,
                'new_price': h.new_price,
                'quantity_at_old_price': h.quantity_at_old_price,
                'remaining_at_old_price': h.remaining_at_old_price,
                'used_quantity': h.used_quantity,
                'change_date': h.change_date.strftime('%d-%m-%Y %H:%M'),
                'reason': h.reason,
                'is_active': h.is_active,
                'purchase_bill_id': h.purchase_bill_id
            }
            for h in history
        ]
    })


# ---------------------------------------------------------------------------
# API: Delete a cost price history entry (safe)
# ---------------------------------------------------------------------------
@bp.route('/api/cost-history/<int:history_id>/delete', methods=['POST'])
@login_required
@permission_required('purchases', action='delete')
def delete_cost_history(history_id):
    """Safely delete a cost price history entry if not referenced by BOM items"""
    history = CostPriceHistory.query.get_or_404(history_id)
    
    # Safety check: prevent deletion if referenced by BOM items
    from app.models import BOMItem
    referenced_bom_items = BOMItem.query.filter_by(cost_price_history_id=history_id).all()
    if referenced_bom_items:
        return jsonify({
            'success': False,
            'message': 'Cannot delete this history entry because it is referenced by BOM items. Deleting it would break cost tracking integrity.'
        }), 400
    
    product_id = history.product_id
    db.session.delete(history)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': 'Price history entry deleted successfully.',
        'product_id': product_id
    })


# ---------------------------------------------------------------------------
# Purchase Returns
# ---------------------------------------------------------------------------
@bp.route('/returns/')
@login_required
def purchase_return_list():
    """List all purchase returns"""
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    status = request.args.get('status', 'all')
    vendor_id = request.args.get('vendor_id')
    return_number = request.args.get('return_number')

    query = PurchaseReturn.query

    if from_date:
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d')
            query = query.filter(PurchaseReturn.date >= from_date_obj)
        except ValueError:
            pass

    if to_date:
        try:
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d')
            to_date_obj = to_date_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(PurchaseReturn.date <= to_date_obj)
        except ValueError:
            pass

    if status != 'all':
        query = query.filter(PurchaseReturn.status == status)
    
    if vendor_id:
        query = query.filter(PurchaseReturn.vendor_id == vendor_id)
    
    if return_number:
        query = query.filter(PurchaseReturn.return_number == return_number)

    query = apply_saved_filter_to_query(query, 'purchase_return', request.args)

    returns = query.order_by(PurchaseReturn.date.desc()).all()

    vendors = Vendor.query.filter_by(is_active=True).all()
    # Fetch all return numbers for the searchable dropdown
    all_return_numbers = db.session.query(PurchaseReturn.return_number).distinct().order_by(PurchaseReturn.return_number).all()
    all_return_numbers = [r[0] for r in all_return_numbers]
    # Fetch all bill numbers for the searchable dropdown
    all_bill_numbers = db.session.query(PurchaseBill.bill_number).distinct().order_by(PurchaseBill.bill_number).all()
    all_bill_numbers = [b[0] for b in all_bill_numbers]

    total_returns = sum(r.total for r in returns)
    total_count = len(returns)

    return render_template('purchase/returns.html',
                        returns=returns,
                        vendors=vendors,
                        all_return_numbers=all_return_numbers,
                        all_bill_numbers=all_bill_numbers,
                        from_date=from_date,
                        to_date=to_date,
                        current_status=status,
                        vendor_id=vendor_id,
                        return_number=return_number,
                        total_returns=total_returns,
                        total_count=total_count,
                        active_module='purchase_return',
                        filter_id=request.args.get('filter_id'))


@bp.route('/return/create', methods=['GET', 'POST'])
@login_required
@permission_required('purchases', action='add')
def create_purchase_return():
    """Create a new purchase return"""
    if request.method == 'GET':
        bill_id = request.args.get('bill_id')
        if bill_id:
            bill = PurchaseBill.query.get_or_404(int(bill_id))
            warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
            return render_template('purchase/create_return.html',
                               bill=bill,
                               products=Product.query.filter_by(is_active=True).all(),
                               warehouses=warehouses,
                               now=datetime.now())

        bills = PurchaseBill.query.order_by(PurchaseBill.date.desc()).all()
        warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
        return render_template('purchase/create_return.html',
                   bills=bills,
                   bill=None,
                   products=Product.query.filter_by(is_active=True).all(),
                   warehouses=warehouses,
                   now=datetime.now())

    # POST - process return
    bill_id = request.form.get('bill_id')
    bill = PurchaseBill.query.get_or_404(int(bill_id))

    product_ids = request.form.getlist('product_id[]')
    quantities = request.form.getlist('quantity[]')
    prices = request.form.getlist('price[]')
    warehouses = request.form.getlist('warehouse_id[]')

    subtotal = 0
    return_items = []

    for i in range(len(product_ids)):
        if product_ids[i] and quantities[i] and float(quantities[i]) > 0:
            product = Product.query.get(int(product_ids[i]))
            quantity = float(quantities[i])
            price = float(prices[i])
            total = quantity * price
            subtotal += total

            wh = None
            try:
                wh = int(warehouses[i]) if i < len(warehouses) and warehouses[i] else None
            except Exception:
                wh = None

            return_items.append({
                'product_id': product.id,
                'warehouse_id': wh,
                'quantity': quantity,
                'unit_price': price,
                'total': total
            })

    if not return_items:
        flash('Please add at least one item to return.', 'warning')
        return redirect(url_for('purchase.create_purchase_return', bill_id=bill_id))

    tax_rate = float(request.form.get('tax_rate', 0))
    tax = subtotal * (tax_rate / 100)
    discount = float(request.form.get('discount', 0))
    total = subtotal + tax - discount

    # Generate return number using settings
    settings = PurchaseReturnSettings.query.first()
    if not settings:
        settings = PurchaseReturnSettings(return_prefix='PRet-', return_suffix='', next_number=1)
        db.session.add(settings)

    # Sync next_number with actual highest return number in DB
    highest_return = PurchaseReturn.query.order_by(PurchaseReturn.id.desc()).first()
    if highest_return and highest_return.return_number:
        try:
            prefix_len = len(settings.return_prefix or '')
            suffix_len = len(settings.return_suffix or '')
            num_str = highest_return.return_number[prefix_len:]
            if suffix_len > 0:
                num_str = num_str[:-suffix_len]
            max_num = int(num_str)
            next_return_num = max(settings.next_number, max_num + 1)
        except (ValueError, IndexError):
            next_return_num = settings.next_number
    else:
        next_return_num = settings.next_number

    # Get unique return number
    prefix = settings.return_prefix or ''
    suffix = settings.return_suffix or ''
    while True:
        return_number = f"{prefix}{next_return_num}{suffix}"
        existing = PurchaseReturn.query.filter_by(return_number=return_number).first()
        if not existing:
            break
        next_return_num += 1

    is_admin = getattr(current_user, 'is_admin', False)
    purchase_return = PurchaseReturn(
        return_number=return_number,
        bill_id=bill.id,
        vendor_id=bill.vendor_id,
        date=datetime.strptime(request.form.get('date'), '%Y-%m-%d') if request.form.get('date') else datetime.now(),
        subtotal=subtotal,
        tax_rate=tax_rate,
        tax=tax,
        discount=discount,
        total=total,
        reason=request.form.get('reason', ''),
        status='approved' if is_admin else 'pending',
        is_approved=is_admin,
        approved_by=current_user.id if is_admin else None,
        approved_at=datetime.utcnow() if is_admin else None,
        returned_to_inventory=False,
        refund_status='none',
        created_by=current_user.id
    )

    db.session.add(purchase_return)
    db.session.flush()

    # Create return items
    for item in return_items:
        return_item = PurchaseReturnItem(
            return_id=purchase_return.id,
            product_id=item['product_id'],
            warehouse_id=item.get('warehouse_id'),
            quantity=item['quantity'],
            unit_price=item['unit_price'],
            total=item['total']
        )
        db.session.add(return_item)

    if purchase_return.is_approved:
        # Subtract from inventory
        for item in return_items:
            product = Product.query.get(item['product_id'])
            if product:
                # decrease global product quantity
                product.update_quantity(-item['quantity'])

                # decrease per-warehouse stock if specified
                wh_id = item.get('warehouse_id')
                if wh_id:
                    wh_stock = ProductWarehouseStock.query.filter_by(product_id=item['product_id'], warehouse_id=wh_id).first()
                    if not wh_stock:
                        wh_stock = ProductWarehouseStock(product_id=item['product_id'], warehouse_id=wh_id, quantity=0)
                        db.session.add(wh_stock)
                    wh_stock.quantity -= item['quantity']
                elif product.warehouse_id:
                    wh_stock = ProductWarehouseStock.query.filter_by(product_id=item['product_id'], warehouse_id=product.warehouse_id).first()
                    if not wh_stock:
                        wh_stock = ProductWarehouseStock(product_id=item['product_id'], warehouse_id=product.warehouse_id, quantity=0)
                        db.session.add(wh_stock)
                    wh_stock.quantity -= item['quantity']

        # Update bill totals - reduce bill total by return amount
        bill.subtotal -= subtotal
        bill.tax -= tax
        bill.total -= total
        if bill.total < 0:
            bill.total = 0

        # Adjust paid amount if needed
        if bill.paid_amount > 0:
            if bill.total < bill.paid_amount:
                bill.paid_amount = bill.total

        # Update bill status based on return
        original_total = bill.total + total  # Restore original to compare
        if original_total > 0:
            if bill.total == 0 or total >= original_total:
                bill.status = 'return'
            elif total > 0 and bill.total < original_total:
                bill.status = 'partial_return'
            else:
                bill.status = bill.status  # Keep existing status

    # Update return settings next number
    settings.next_number = next_return_num + 1

    db.session.commit()
    log_activity('Purchase', f'Created Return #{return_number}',
                f'Bill: #{bill.bill_number}, Total: {total}')
    flash(f'Purchase return {return_number} created successfully!', 'success')
    return redirect(url_for('purchase.purchase_return_detail', id=purchase_return.id))


@bp.route('/return/<int:id>')
@login_required
def purchase_return_detail(id):
    """View purchase return detail"""
    purchase_return = PurchaseReturn.query.get_or_404(id)
    return render_template('purchase/return_detail.html', return_obj=purchase_return)


@bp.route('/return/<int:id>/approve', methods=['POST'])
@login_required
@permission_required('purchases', action='edit')
def approve_purchase_return(id):
    if not getattr(current_user, 'is_admin', False):
        flash('Only admins can approve purchase returns.', 'danger')
        return redirect(url_for('purchase.purchase_return_detail', id=id))
        
    purchase_return = PurchaseReturn.query.get_or_404(id)
    
    if purchase_return.is_approved:
        flash('This return is already approved.', 'warning')
        return redirect(url_for('purchase.purchase_return_detail', id=id))

    # Set approval fields
    purchase_return.is_approved = True
    purchase_return.is_rejected = False
    purchase_return.status = 'approved'
    purchase_return.approved_by = current_user.id
    purchase_return.approved_at = datetime.utcnow()
    
    # Apply side effects (Inventory and Bill update)
    bill = purchase_return.bill
    
    # 1. Subtract from inventory
    for item in purchase_return.items:
        product = Product.query.get(item.product_id)
        if product:
            product.update_quantity(-item.quantity)
            wh_id = item.warehouse_id
            if wh_id:
                wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=wh_id).first()
                if not wh_stock:
                    wh_stock = ProductWarehouseStock(product_id=item.product_id, warehouse_id=wh_id, quantity=0)
                    db.session.add(wh_stock)
                wh_stock.quantity -= item.quantity
            elif product.warehouse_id:
                wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=product.warehouse_id).first()
                if not wh_stock:
                    wh_stock = ProductWarehouseStock(product_id=item.product_id, warehouse_id=product.warehouse_id, quantity=0)
                    db.session.add(wh_stock)
                wh_stock.quantity -= item.quantity

    # 2. Update bill totals
    bill.subtotal -= purchase_return.subtotal
    bill.tax -= purchase_return.tax
    bill.total -= purchase_return.total
    if bill.total < 0:
        bill.total = 0

    # 3. Adjust paid amount if needed
    if bill.paid_amount > 0:
        if bill.total < bill.paid_amount:
            bill.paid_amount = bill.total

    # 4. Update bill status
    original_total = bill.total + purchase_return.total
    if original_total > 0:
        if bill.total == 0 or purchase_return.total >= original_total:
            bill.status = 'return'
        elif purchase_return.total > 0 and bill.total < original_total:
            bill.status = 'partial_return'

    db.session.commit()
    log_activity('Purchase', f'Approved Return #{purchase_return.return_number}',
                f'Bill: #{bill.bill_number}')
    flash(f'Purchase return {purchase_return.return_number} approved successfully!', 'success')
    return redirect(url_for('purchase.purchase_return_detail', id=id))


@bp.route('/return/<int:id>/reject', methods=['POST'])
@login_required
@permission_required('purchases', action='edit')
def reject_purchase_return(id):
    if not getattr(current_user, 'is_admin', False):
        flash('Only admins can reject purchase returns.', 'danger')
        return redirect(url_for('purchase.purchase_return_detail', id=id))

    purchase_return = PurchaseReturn.query.get_or_404(id)
    reason = request.form.get('reason', '')
    
    purchase_return.is_approved = False
    purchase_return.is_rejected = True
    purchase_return.rejection_reason = reason
    purchase_return.status = 'rejected'
    
    db.session.commit()
    log_activity('Purchase', f'Rejected Return #{purchase_return.return_number}',
                f'Reason: {reason or "N/A"}')
    flash(f'Purchase return {purchase_return.return_number} has been rejected.', 'warning')
    return redirect(url_for('purchase.purchase_return_detail', id=id))


@bp.route('/return/<int:id>/mark_returned', methods=['POST'])
@login_required
def mark_purchase_returned(id):
    """Mark return as returned to inventory and actually restore stock levels"""
    purchase_return = PurchaseReturn.query.get_or_404(id)
    
    if not purchase_return.returned_to_inventory:
        # Restore inventory for each item
        for item in purchase_return.items:
            product = Product.query.get(item.product_id)
            if product:
                # increase global product quantity
                product.update_quantity(item.quantity)

                # increase per-warehouse stock
                wh_id = item.warehouse_id or product.warehouse_id
                if wh_id:
                    wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=wh_id).first()
                    if not wh_stock:
                        wh_stock = ProductWarehouseStock(product_id=item.product_id, warehouse_id=wh_id, quantity=0)
                        db.session.add(wh_stock)
                    wh_stock.quantity += item.quantity

        purchase_return.returned_to_inventory = True
        purchase_return.status = 'completed'
        db.session.commit()
        log_activity('Purchase', f'Marked Return #{purchase_return.return_number} as Returned',
                    f'Items returned to inventory')
        flash('Items successfully returned to inventory and stock levels updated.', 'success')
    else:
        flash('Items already returned to inventory.', 'info')
    
    return redirect(url_for('purchase.purchase_return_detail', id=id))


@bp.route('/return/<int:id>/refund', methods=['POST'])
@login_required
def process_purchase_refund(id):
    """Process refund to vendor"""
    purchase_return = PurchaseReturn.query.get_or_404(id)
    
    refund_amount = float(request.form.get('refund_amount', 0))
    
    if refund_amount <= 0:
        flash('Please enter a valid refund amount.', 'warning')
        return redirect(url_for('purchase.purchase_return_detail', id=id))
    
    if refund_amount > purchase_return.total:
        refund_amount = purchase_return.total
    
    purchase_return.refund_amount = refund_amount
    purchase_return.refund_status = 'paid'
    purchase_return.status = 'completed'
    
    # Create a vendor advance record for the refund (positive = refund credited as advance)
    if purchase_return.vendor:
        vendor_advance = VendorAdvance(
            vendor_id=purchase_return.vendor.id,
            amount=refund_amount,  # Positive amount: increases remaining advance
            date=datetime.now(),
            description=f'Refund for Return: {purchase_return.return_number}',
            created_by=current_user.id
        )
        db.session.add(vendor_advance)
    
    db.session.commit()
    log_activity('Purchase', f'Processed Refund for Return #{purchase_return.return_number}',
                f'Amount: PKR {refund_amount:,.2f}')
    flash(f'Refund of PKR{refund_amount:,.2f} processed to vendor.', 'success')
    return redirect(url_for('purchase.purchase_return_detail', id=id))


@bp.route('/returns/vendor/<int:vendor_id>')
@login_required
def vendor_purchase_returns(vendor_id):
    """View all purchase returns for a vendor"""
    vendor = Vendor.query.get_or_404(vendor_id)
    returns = PurchaseReturn.query.filter_by(vendor_id=vendor_id).order_by(PurchaseReturn.date.desc()).all()
    
    return render_template('purchase/vendor_returns.html',
                        vendor=vendor,
                        returns=returns)


@bp.route('/return/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('purchases', action='edit')
def edit_purchase_return(id):
    """Edit a purchase return"""
    purchase_return = PurchaseReturn.query.get_or_404(id)
    bill = purchase_return.bill
    
    if request.method == 'POST':
        purchase_return.reason = request.form.get('reason', '')
        purchase_return.date = datetime.strptime(request.form.get('date'), '%Y-%m-%d') if request.form.get('date') else purchase_return.date
        
        # Reverse inventory changes for old items ONLY if they weren't already returned to inventory
        if not purchase_return.returned_to_inventory:
            for item in purchase_return.items:
                product = Product.query.get(item.product_id)
                if product:
                    product.update_quantity(item.quantity)
                    
                    wh_id = item.warehouse_id or product.warehouse_id
                    if wh_id:
                        wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=wh_id).first()
                        if wh_stock:
                            wh_stock.quantity += item.quantity

        # Delete existing items
        for item in purchase_return.items:
            db.session.delete(item)
        
        # Get new data
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        warehouse_ids = request.form.getlist('warehouse_id[]')
        
        # Create new items
        for i in range(len(product_ids)):
            if product_ids[i] and quantities[i] and float(quantities[i]) > 0:
                product = Product.query.get(int(product_ids[i]))
                quantity = float(quantities[i])
                price = float(prices[i])
                total = quantity * price
                wh_id = int(warehouse_ids[i]) if i < len(warehouse_ids) and warehouse_ids[i] else (product.warehouse_id if product else None)
                
                return_item = PurchaseReturnItem(
                    return_id=purchase_return.id,
                    product_id=product.id,
                    quantity=quantity,
                    unit_price=price,
                    total=total,
                    warehouse_id=wh_id
                )
                db.session.add(return_item)
                
                # Subtract from inventory ONLY if not already returned to inventory
                if not purchase_return.returned_to_inventory:
                    if product:
                        product.update_quantity(-quantity)
                        
                        if wh_id:
                            wh_stock = ProductWarehouseStock.query.filter_by(product_id=product.id, warehouse_id=wh_id).first()
                            if not wh_stock:
                                wh_stock = ProductWarehouseStock(product_id=product.id, warehouse_id=wh_id, quantity=0)
                                db.session.add(wh_stock)
                            wh_stock.quantity -= quantity
        
        # Recalculate totals
        purchase_return.calculate_totals()
        
        # Update bill totals
        if bill:
            bill.subtotal = sum(item.total for item in bill.items) - sum(ret.total for ret in bill.purchase_returns)
            bill.calculate_totals()
            bill.update_status()

        db.session.commit()
        log_activity('Purchase', f'Updated Return #{purchase_return.return_number}',
                    f'Bill: #{bill.bill_number if bill else "N/A"}')
        flash(f'Return {purchase_return.return_number} updated successfully!', 'success')
        return redirect(url_for('purchase.purchase_return_detail', id=id))
    
    return render_template('purchase/edit_return.html',
                          return_obj=purchase_return,
                          products=Product.query.filter_by(is_active=True).all())


@bp.route('/return/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('purchases', action='delete')
def delete_purchase_return(id):
    """Delete a purchase return and restore stock if not already returned"""
    purchase_return = PurchaseReturn.query.get_or_404(id)
    bill = purchase_return.bill
    
    # Restore product quantities ONLY if they were still "out" (returned to vendor)
    # If returned_to_inventory is True, it means they already came back to shelf.
    if not purchase_return.returned_to_inventory:
        for item in purchase_return.items:
            product = Product.query.get(item.product_id)
            if product:
                # restore global product quantity
                product.update_quantity(item.quantity)

                # restore per-warehouse stock
                wh_id = item.warehouse_id or product.warehouse_id
                if wh_id:
                    wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=wh_id).first()
                    if not wh_stock:
                        wh_stock = ProductWarehouseStock(product_id=item.product_id, warehouse_id=wh_id, quantity=0)
                        db.session.add(wh_stock)
                    wh_stock.quantity += item.quantity
    
    # Reverse vendor refund if was paid
    if purchase_return.refund_status == 'paid' and purchase_return.vendor:
        # Find and delete the original positive advance created for this refund
        original_advance = VendorAdvance.query.filter_by(
            vendor_id=purchase_return.vendor.id,
            description=f'Refund for Return: {purchase_return.return_number}'
        ).first()

        if original_advance:
            db.session.delete(original_advance)
        else:
            # Fallback if original not found: Create a negative vendor advance to reverse the refund
            vendor_advance = VendorAdvance(
                vendor_id=purchase_return.vendor.id,
                amount=-purchase_return.refund_amount,
                date=datetime.now(),
                description=f'Reversal for Return: {purchase_return.return_number}',
                created_by=current_user.id
            )
            db.session.add(vendor_advance)
    
    # Restore bill totals and recalculate status
    if bill:
        bill.subtotal += purchase_return.subtotal
        bill.tax += purchase_return.tax
        bill.total += purchase_return.total
        
        bill.update_status()
    
    return_number = purchase_return.return_number
    db.session.delete(purchase_return)
    db.session.commit()
    log_activity('Purchase', f'Deleted Return #{return_number}',
                f'Return removed')
    flash(f'Return {return_number} deleted successfully. Inventory has been restored where applicable.', 'success')
    return redirect(url_for('purchase.purchase_return_list'))

@bp.route('/returns/bulk-delete', methods=['POST'])
@login_required
@permission_required('purchases', action='delete')
def bulk_delete_purchase_returns():
    """Bulk delete multiple purchase returns"""
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No returns selected'}), 400
    
    deleted_count = 0
    errors = []
    
    for return_id in ids:
        purchase_return = PurchaseReturn.query.get(return_id)
        if not purchase_return:
            continue
            
        try:
            bill = purchase_return.bill
            # Restore product quantities
            for item in purchase_return.items:
                if item.product:
                    item.product.quantity += item.quantity
            
            # Reverse vendor refund if was paid
            if purchase_return.refund_status == 'paid' and purchase_return.vendor:
                original_advance = VendorAdvance.query.filter_by(
                    vendor_id=purchase_return.vendor.id,
                    description=f'Refund for Return: {purchase_return.return_number}'
                ).first()
                if original_advance:
                    db.session.delete(original_advance)
            
            # Restore bill totals
            if bill:
                bill.subtotal += purchase_return.subtotal
                bill.tax += purchase_return.tax
                bill.total += purchase_return.total
                bill.update_status()
            
            db.session.delete(purchase_return)
            deleted_count += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f'Error deleting Return {purchase_return.return_number}: {str(e)}')
            
    if deleted_count > 0:
        db.session.commit()
        
    message = f'Successfully deleted {deleted_count} purchase returns.'
    if errors:
        return jsonify({'success': False, 'message': message, 'errors': errors}), 500
    return jsonify({'success': True, 'message': message})
@bp.route('/return/settings', methods=['GET', 'POST'])
@login_required
def purchase_return_settings():
    settings = PurchaseReturnSettings.query.first()
    form = PurchaseReturnSettingsForm(obj=settings)
    if form.validate_on_submit():
        if not settings:
            settings = PurchaseReturnSettings()
            db.session.add(settings)
        settings.return_prefix = form.return_prefix.data or 'PRet-'
        settings.return_suffix = form.return_suffix.data or ''
        settings.next_number = form.next_number.data or 1
        db.session.commit()
        flash('Purchase return settings updated successfully.', 'success')
        return redirect(url_for('purchase.purchase_return_settings'))
    return render_template('purchase/purchase_return_settings.html', settings=settings, form=form)

@bp.route("/vendor/<int:id>/export/csv")
@login_required
def vendor_export_csv(id):
    """Export vendor profile to CSV"""
    vendor = Vendor.query.get_or_404(id)
    bills = sorted(vendor.bills, key=lambda b: b.date, reverse=True)
    advances = sorted(vendor.advances, key=lambda a: a.date, reverse=True)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Vendor info
    writer.writerow(["Vendor Profile Export"])
    writer.writerow(["Name", vendor.name])
    writer.writerow(["Email", vendor.email or ""])
    writer.writerow(["Phone", vendor.phone or ""])
    writer.writerow(["Address", vendor.address or ""])
    writer.writerow(["Payment Method", vendor.payment_method or ""])
    writer.writerow(["Bank Name", vendor.bank_name or ""])
    writer.writerow(["Account Holder", vendor.account_holder_name or ""])
    writer.writerow(["Account Number", vendor.account_number or ""])
    writer.writerow(["IFSC Code", vendor.ifsc_code or ""])
    writer.writerow(["SWIFT Code", vendor.swift_code or ""])
    writer.writerow([])
    
    # Bills
    writer.writerow(["Purchase Bills"])
    writer.writerow(["Bill #", "Date", "Subtotal", "Tax", "Discount", "Total", "Paid", "Balance", "Status"])
    for bill in bills:
        writer.writerow([
            bill.bill_number,
            bill.date.strftime("%Y-%m-%d"),
            bill.subtotal,
            bill.tax,
            bill.discount,
            bill.total,
            bill.paid_amount,
            bill.balance_due,
            bill.status
        ])
    writer.writerow([])
    
    # Advances
    writer.writerow(["Advances"])
    writer.writerow(["Date", "Description", "Amount", "Status"])
    for adv in advances:
        writer.writerow([
            adv.date.strftime("%Y-%m-%d"),
            adv.description or "",
            adv.amount,
            "Adjusted" if adv.is_adjusted else "Pending"
        ])
    
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"vendor_{vendor.id}_{vendor.name}_profile.csv"
    )

@bp.route("/vendor/<int:id>/export/excel")
@login_required
def vendor_export_excel(id):
    """Export vendor profile to Excel"""
    vendor = Vendor.query.get_or_404(id)
    bills = sorted(vendor.bills, key=lambda b: b.date, reverse=True)
    advances = sorted(vendor.advances, key=lambda a: a.date, reverse=True)
    
    wb = openpyxl.Workbook()
    
    # Vendor Info sheet
    ws1 = wb.active
    ws1.title = "Vendor Info"
    ws1.append(["Vendor Profile Export"])
    ws1.append(["Name", vendor.name])
    ws1.append(["Email", vendor.email or ""])
    ws1.append(["Phone", vendor.phone or ""])
    ws1.append(["Address", vendor.address or ""])
    ws1.append(["Payment Method", vendor.payment_method or ""])
    ws1.append(["Bank Name", vendor.bank_name or ""])
    ws1.append(["Account Holder", vendor.account_holder_name or ""])
    ws1.append(["Account Number", vendor.account_number or ""])
    ws1.append(["IFSC Code", vendor.ifsc_code or ""])
    ws1.append(["SWIFT Code", vendor.swift_code or ""])
    
    # Bills sheet
    ws2 = wb.create_sheet("Purchase Bills")
    ws2.append(["Bill #", "Date", "Subtotal", "Tax", "Discount", "Total", "Paid", "Balance", "Status"])
    for bill in bills:
        ws2.append([
            bill.bill_number,
            bill.date.strftime("%Y-%m-%d"),
            bill.subtotal,
            bill.tax,
            bill.discount,
            bill.total,
            bill.paid_amount,
            bill.balance_due,
            bill.status
        ])
    
    # Advances sheet
    ws3 = wb.create_sheet("Advances")
    ws3.append(["Date", "Description", "Amount", "Status"])
    for adv in advances:
        ws3.append([
            adv.date.strftime("%Y-%m-%d"),
            adv.description or "",
            adv.amount,
            "Adjusted" if adv.is_adjusted else "Pending"
        ])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"vendor_{vendor.id}_{vendor.name}_profile.xlsx"
    )

@bp.route("/vendor/<int:id>/export/pdf")
@login_required
def vendor_export_pdf(id):
    """Export vendor profile to PDF - Mirroring Purchase Bill design"""
    from datetime import datetime as _dt
    from app.models import Company, PurchaseSettings
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image
    )
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    vendor = Vendor.query.get_or_404(id)
    company = Company.query.first() if 'Company' in globals() else None

    # Date filter from query params
    date_from_str = request.args.get('date_from', '').strip()
    date_to_str   = request.args.get('date_to',   '').strip()
    date_from_f = None
    date_to_f   = None
    try:
        if date_from_str:
            date_from_f = _dt.strptime(date_from_str, '%Y-%m-%d')
        if date_to_str:
            date_to_f = _dt.strptime(date_to_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except ValueError:
        pass

    all_bills = sorted(vendor.bills, key=lambda s: s.date, reverse=True)
    if date_from_f or date_to_f:
        bills = [
            s for s in all_bills
            if (date_from_f is None or s.date >= date_from_f)
            and (date_to_f   is None or s.date <= date_to_f)
        ]
    else:
        bills = all_bills

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=28, bottomMargin=72
    )

    PRIMARY_COLOR   = colors.HexColor("#c0392b")
    ACCENT_COLOR    = colors.HexColor("#f9f9f9")
    TEXT_COLOR      = colors.HexColor("#1a1a1a")
    MUTED_TEXT      = colors.HexColor("#6b6b6b")
    WHITE           = colors.white
    BLACK           = colors.black
    LIGHT_GREY      = colors.HexColor("#e0e0e0")
    BORDER_GREY     = colors.HexColor("#d0d0d0")
    HEADER_STRIPE   = colors.HexColor("#1a1a1a")
    PAGE_BG         = colors.HexColor("#f2f2f2")

    def _draw_footer(c, doc_obj):
        w, h = A4
        card_margin = 18
        lx = card_margin + 16
        rx = w - card_margin - 16
        divider_y = card_margin + 72 - 2
        y = divider_y - 11

        c.saveState()
        c.setStrokeColor(BORDER_GREY)
        c.setLineWidth(0.4)
        c.line(lx, divider_y, rx, divider_y)

        c.setFont('Helvetica-Bold', 7)
        c.setFillColor(TEXT_COLOR)
        c.drawString(lx, y, "PREPARED BY")

        c.setFont('Helvetica', 6.5)
        c.setFillColor(MUTED_TEXT)
        y -= 10
        c.drawString(lx, y, getattr(company, 'name', '') if company else "")

        sig_path = None
        if company:
            try:
                sig_path = company.signature_path
            except Exception:
                pass

        if sig_path and os.path.exists(sig_path):
            try:
                from reportlab.platypus import Image as RLImage
                sig_img = RLImage(sig_path)
                sig_img.drawWidth = 100
                sig_img.drawHeight = 100
                sig_img.drawOn(c, lx, y - 70)
            except Exception:
                pass
        
        if not sig_path or not os.path.exists(sig_path):
            y -= 14
            c.setStrokeColor(colors.HexColor("#555555"))
            c.setLineWidth(0.5)
            c.line(lx, y, lx + 110, y)
            y -= 8
            c.setFont('Helvetica', 6.5)
            c.setFillColor(MUTED_TEXT)
            c.drawString(lx, y, "Authorized Signature")

        msg_y = divider_y - 11 - 20
        c.setFont('Helvetica-Bold', 7.5)
        c.setFillColor(TEXT_COLOR)
        c.drawCentredString(w / 2, msg_y, "Thank you for your business!")

        cemail = getattr(company, 'email', '') if company else ""
        cphone = getattr(company, 'phone', '') if company else ""
        cwa    = getattr(company, 'whatsapp', '') if company else ""

        ry = divider_y - 11
        c.setFont('Helvetica-Bold', 7)
        c.drawRightString(rx, ry, "CUSTOMER SUPPORT")

        c.setFont('Helvetica', 6.5)
        c.setFillColor(MUTED_TEXT)
        for line in filter(None, [cemail, cphone, f"WhatsApp: {cwa}" if cwa else ""]):
            ry -= 9
            c.drawRightString(rx, ry, line)

        c.setFont('Helvetica', 7)
        c.setFillColor(MUTED_TEXT)
        c.drawCentredString(w / 2, card_margin + 6, f"Page {doc_obj.page}")

        c.restoreState()

    def _draw_page_decorations(c, doc_obj):
        w, h = A4
        c.saveState()
        c.setFillColor(PAGE_BG)
        c.rect(0, 0, w, h, fill=1, stroke=0)
        m = 18
        c.setFillColor(WHITE)
        c.roundRect(m, m, w - 2*m, h - 2*m, 6, fill=1, stroke=0)
        bh = 8
        c.setFillColor(HEADER_STRIPE)
        c.rect(m, h - m - bh, (w - 2*m) * 0.75, bh, fill=1, stroke=0)
        c.setFillColor(PRIMARY_COLOR)
        c.rect(m + (w - 2*m) * 0.75, h - m - bh, (w - 2*m) * 0.25, bh, fill=1, stroke=0)
        c.restoreState()
        
        _draw_footer(c, doc_obj)

    styles = getSampleStyleSheet()
    s_CompanyName = ParagraphStyle('CompanyName', parent=styles['Normal'], fontSize=11, textColor=TEXT_COLOR, fontName='Helvetica-Bold', spaceAfter=1)
    s_CompanyInfo = ParagraphStyle('CompanyInfo', parent=styles['Normal'], fontSize=7,  textColor=MUTED_TEXT, leading=10)
    s_Title       = ParagraphStyle('Title', parent=styles['Normal'], fontSize=20, textColor=BLACK, fontName='Helvetica-Bold', alignment=TA_RIGHT, spaceAfter=1)
    s_MetaLabel   = ParagraphStyle('MetaLabel', parent=styles['Normal'], fontSize=7.5, textColor=MUTED_TEXT, alignment=TA_RIGHT)
    s_MetaValue   = ParagraphStyle('MetaValue', parent=styles['Normal'], fontSize=8, textColor=TEXT_COLOR, fontName='Helvetica-Bold', alignment=TA_RIGHT)
    s_BoxTitle    = ParagraphStyle('BoxTitle', parent=styles['Normal'], fontSize=8, textColor=PRIMARY_COLOR, fontName='Helvetica-Bold', spaceAfter=3, leftIndent=0)
    s_BoxValue    = ParagraphStyle('BoxValue', parent=styles['Normal'], fontSize=7.5, textColor=TEXT_COLOR, leading=10, leftIndent=0)
    s_TblHeader   = ParagraphStyle('TblHeader', parent=styles['Normal'], fontSize=7.5, textColor=WHITE, fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_TblCell     = ParagraphStyle('TblCell', parent=styles['Normal'], fontSize=7.5, textColor=TEXT_COLOR, alignment=TA_RIGHT)
    s_TblCellC    = ParagraphStyle('TblCellC', parent=styles['Normal'], fontSize=7.5, textColor=TEXT_COLOR, alignment=TA_CENTER)
    s_TblCellL    = ParagraphStyle('TblCellL', parent=styles['Normal'], fontSize=7.5, textColor=TEXT_COLOR, alignment=TA_LEFT)
    s_TotalLabel  = ParagraphStyle('TotalLabel', parent=styles['Normal'], fontSize=8, textColor=TEXT_COLOR, alignment=TA_LEFT)
    s_TotalValue  = ParagraphStyle('TotalValue', parent=styles['Normal'], fontSize=8, textColor=TEXT_COLOR, alignment=TA_RIGHT)
    s_GrandLabel  = ParagraphStyle('GrandLabel', parent=styles['Normal'], fontSize=9, textColor=TEXT_COLOR, fontName='Helvetica-Bold', alignment=TA_LEFT)
    s_GrandValue  = ParagraphStyle('GrandValue', parent=styles['Normal'], fontSize=9, textColor=TEXT_COLOR, fontName='Helvetica-Bold', alignment=TA_RIGHT)

    elements = []

    # GET LOGO 
    logo = None
    if company and getattr(company, 'logo_path', None) and os.path.exists(company.logo_path):
        try:
            img = Image(company.logo_path)
            aspect = img.imageHeight / float(img.imageWidth)
            w = 1.0 * inch
            img.drawWidth  = w
            img.drawHeight = w * aspect
            logo = img
        except Exception:
            pass

    # HEADER LEFT
    left = []
    if logo:
        left.append(logo)
        left.append(Spacer(1, 6))
    
    cname = company.name if company else "COMPANY"
    left.append(Paragraph(cname, s_CompanyName))
    if company:
        for attr, label in [('email','<b>Email:</b> {}'), ('phone','<b>Phone:</b> {}'), ('whatsapp','<b>WhatsApp:</b> {}'), ('address','{}'), ('address2','{}')]:
            val = getattr(company, attr, None)
            if val:
                left.append(Paragraph(label.format(val), s_CompanyInfo))

    # HEADER RIGHT
    RIGHT_W      = 3.2 * inch
    META_LABEL_W = 1.1 * inch
    META_VAL_W   = RIGHT_W - META_LABEL_W

    meta_rows = [
        ["Report Date:", _dt.now().strftime('%d/%m/%Y')],
        ["Vendor ID:", str(vendor.id)],
    ]
    meta_tbl = Table(
        [[Paragraph(l, s_MetaLabel), Paragraph(v, s_MetaValue)] for l, v in meta_rows],
        colWidths=[META_LABEL_W, META_VAL_W]
    )
    meta_tbl.setStyle(TableStyle([
        ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 1),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
    ]))

    right_rows = [
        [Paragraph("VENDOR PROFILE", s_Title)],
        [Spacer(1, 22)],
        [meta_tbl],
    ]
    right_tbl = Table(right_rows, colWidths=[RIGHT_W])
    right_tbl.setStyle(TableStyle([
        ('ALIGN',  (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING',    (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
    ]))

    LEFT_W = (523 / 72) * inch - RIGHT_W - 0.1 * inch
    header_tbl = Table([[left, right_tbl]], colWidths=[LEFT_W, RIGHT_W])
    header_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN',  (0,0), (0,0),   'LEFT'),
        ('ALIGN',  (1,0), (1,0),   'RIGHT'),
        ('TOPPADDING',    (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
    ]))
    elements.append(header_tbl)
    elements.append(Spacer(1, 8))
    elements.append(HRFlowable(width="100%", thickness=0.6, color=BORDER_GREY, spaceAfter=8))

    # INFO BOX (VENDOR DETAILS)
    BOX_W = 3.5 * inch
    b1_rows = []
    b1_rows.append(f"<b>Name:</b> {vendor.company_name or vendor.name}")
    if vendor.email:   b1_rows.append(f"<b>Email:</b> {vendor.email}")
    if vendor.phone:   b1_rows.append(f"<b>Phone:</b> {vendor.phone}")
    if vendor.address: b1_rows.append(f"<b>Address:</b> {vendor.address}")
    if vendor.gst_number: b1_rows.append(f"<b>GST No:</b> {vendor.gst_number}")
    if getattr(vendor, 'contact_person', None): b1_rows.append(f"<b>Contact:</b> {vendor.contact_person}")

    data = [[Paragraph("VENDOR DETAILS", s_BoxTitle)]]
    for row in b1_rows:
        data.append([Paragraph(row, s_BoxValue)])
    box1 = Table(data, colWidths=[BOX_W])
    box1.setStyle(TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
        ('BACKGROUND',    (0,0), (-1,-1), WHITE),
        ('BOX',           (0,0), (-1,-1), 0.5, BORDER_GREY),
    ]))

    outer = Table([[box1]], colWidths=[7.1*inch])
    outer.setStyle(TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
        ('TOPPADDING',    (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(outer)
    elements.append(Spacer(1, 15))

    # INVOICES TABLE
    headers = ['Bill #', 'Date', 'Subtotal', 'Tax', 'Discount', 'Paid', 'Balance', 'Status']
    header_row = [Paragraph(f"<b>{h}</b>", s_TblHeader) for h in headers]
    table_data = [header_row]

    TABLE_WIDTH = 7.5 * inch
    # Weights for the 8 columns: 3, 2, 3, 3, 3, 3, 3, 2 (Total = 22)
    Cols = [(w/22.0)*TABLE_WIDTH for w in [3.2, 2.3, 3, 2.5, 3, 3, 3, 2]]

    def pkr(val):
        return f"PKR{val:,.0f}"

    def status_p(st):
        cmap = {
            "paid": colors.HexColor("#16a34a"),
            "partial": colors.HexColor("#d97706"),
            "unpaid": colors.HexColor("#dc2626")
        }
        sst = ParagraphStyle("sst", parent=styles["Normal"], fontSize=7,
                             leading=9, textColor=cmap.get(st, TEXT_COLOR),
                             alignment=TA_CENTER)
        return Paragraph(st.title(), sst)

    for bill in bills:
        table_data.append([
            Paragraph(bill.bill_number, s_TblCellL),
            Paragraph(bill.date.strftime("%d/%m/%Y"), s_TblCellC),
            Paragraph(pkr(bill.subtotal), s_TblCell),
            Paragraph(pkr(bill.tax), s_TblCell),
            Paragraph(pkr(bill.discount), s_TblCell),
            Paragraph(pkr(bill.paid_amount), s_TblCell),
            Paragraph(pkr(bill.balance_due), s_TblCell),
            status_p(bill.status)
        ])

    if not bills:
        empty = [Paragraph('No bills found', s_TblCellL)] + [Paragraph('-', s_TblCellC)] * 7
        table_data.append(empty)

    tbl = Table(table_data, colWidths=Cols, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  HEADER_STRIPE),
        ('TEXTCOLOR',     (0,0),  (-1,0),  WHITE),
        ('TOPPADDING',    (0,0),  (-1,0),  5),
        ('BOTTOMPADDING', (0,0),  (-1,0),  5),
        ('LEFTPADDING',   (0,0),  (-1,0),  6),
        ('RIGHTPADDING',  (0,0),  (-1,0),  6),
        ('TOPPADDING',    (0,1),  (-1,-1), 4),
        ('BOTTOMPADDING', (0,1),  (-1,-1), 4),
        ('LEFTPADDING',   (0,1),  (-1,-1), 6),
        ('RIGHTPADDING',  (0,1),  (-1,-1), 6),
        ('ROWBACKGROUNDS',(0,1),  (-1,-1), [WHITE, ACCENT_COLOR]),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
        ('ALIGN',         (0,0),  (-1,0),  'CENTER'),
        ('LINEBELOW',     (0,0),  (-1,0),  1.0, PRIMARY_COLOR),
        ('GRID',          (0,0),  (-1,-1), 0.3, LIGHT_GREY),
        ('LINEBELOW',     (0,-1), (-1,-1), 0.5, BORDER_GREY),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 10))

    # BOTTOM SECTION (Totals on the right)
    tot_bills = sum(b.total for b in bills)
    tot_disc = sum(b.discount for b in bills)
    out_bal = vendor.outstanding_balance
    tot_adv = vendor.total_advances_given
    adv_appl = vendor.total_advances_adjusted
    rem_adv = vendor.remaining_advance_balance

    totals = [
        ("Total Purchases",          pkr(tot_bills)),
        ("Total Discount",       pkr(tot_disc)),
        ("Outstanding Balance",  pkr(out_bal)),
        ("Total Advances",       pkr(tot_adv)),
        ("Advances Applied",     pkr(adv_appl)),
        ("Remaining Advance",    pkr(rem_adv)),
    ]

    # Outstanding Balance row index = 2
    DANGER_RED = colors.HexColor("#c0392b")
    s_GrandLabelRed = ParagraphStyle('GrandLabelRed', parent=styles['Normal'], fontSize=9,
                                      textColor=DANGER_RED, fontName='Helvetica-Bold', alignment=TA_LEFT)
    s_GrandValueRed = ParagraphStyle('GrandValueRed', parent=styles['Normal'], fontSize=9,
                                      textColor=DANGER_RED, fontName='Helvetica-Bold', alignment=TA_RIGHT)

    tot_rows = []
    for i, (label, value) in enumerate(totals):
        if label == "Outstanding Balance":
            lp = Paragraph(f"<b>{label}</b>", s_GrandLabelRed)
            vp = Paragraph(f"<b>{value}</b>", s_GrandValueRed)
        else:
            lp = Paragraph(f"<b>{label}</b>", s_GrandLabel)
            vp = Paragraph(f"<b>{value}</b>", s_GrandValue)
        tot_rows.append([lp, vp])

    tot_tbl = Table(tot_rows, colWidths=[2.0 * inch, 1.7 * inch])
    tot_style = [
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ('RIGHTPADDING',  (0,0), (-1,-1), 5),
        ('LINEBELOW',     (0,-1),(-1,-1), 0.5, BORDER_GREY),
    ]
    for i, (label, _) in enumerate(totals):
        bg = colors.HexColor("#fdecea") if label == "Outstanding Balance" else ACCENT_COLOR
        tot_style += [
            ('LINEABOVE',  (0,i), (-1,i), 0.8, BORDER_GREY),
            ('BACKGROUND', (0,i), (-1,i), bg)
        ]
    tot_tbl.setStyle(TableStyle(tot_style))

    # Build bank details box for the LEFT side (always shown)
    s_BankTitle = ParagraphStyle('BankTitle', parent=styles['Normal'], fontSize=8,
                                  textColor=PRIMARY_COLOR, fontName='Helvetica-Bold', spaceAfter=4)
    s_BankText  = ParagraphStyle('BankText',  parent=styles['Normal'], fontSize=7.5,
                                  textColor=TEXT_COLOR, leading=11)

    bank_fields = [
        ('bank_name',           'Bank'),
        ('account_holder_name', 'A/C Holder'),
        ('account_number',      'A/C No'),
        ('ifsc_code',           'IFSC'),
        ('swift_code',          'SWIFT'),
    ]

    bank_content = [[Paragraph("BANK DETAILS", s_BankTitle)]]
    for attr, label in bank_fields:
        val = getattr(vendor, attr, None) or '-'
        bank_content.append([Paragraph(f"<b>{label}:</b> {val}", s_BankText)])

    bank_tbl = Table(bank_content, colWidths=[3.5 * inch])
    bank_tbl.setStyle(TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
        ('BACKGROUND',    (0,0), (-1,-1), WHITE),
        ('BOX',           (0,0), (-1,-1), 0.5, BORDER_GREY),
    ]))
    left_cell = bank_tbl

    row1 = Table([[left_cell, tot_tbl]], colWidths=[3.75*inch, 3.75*inch])
    row1.setStyle(TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('ALIGN',         (1,0), (1,-1),  'RIGHT'),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
        ('TOPPADDING',    (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(row1)

    s_NotesTitle = ParagraphStyle('NotesTitle', parent=styles['Normal'], fontSize=8, textColor=PRIMARY_COLOR, fontName='Helvetica-Bold', spaceAfter=3)
    s_NotesText  = ParagraphStyle('NotesText', parent=styles['Normal'], fontSize=7, textColor=TEXT_COLOR, leading=10)
    s_ExtraThankYou = ParagraphStyle(
        'ExtraThankYou',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_LEFT,
        textColor=TEXT_COLOR,
        spaceAfter=6,
        leading=14,
        leftIndent=0
    )

    settings = PurchaseSettings.query.first()
    terms_text = settings.default_terms if settings and settings.default_terms else ""

    if terms_text:
        terms_content = [Paragraph("TERMS & CONDITIONS", s_NotesTitle)]
        for line in terms_text.split('\n'):
            if line.strip():
                terms_content.append(Paragraph(line.strip(), s_NotesText))
        
        terms_tbl = Table([[item] for item in terms_content], colWidths=[7.4 * inch])
        terms_tbl.setStyle(TableStyle([
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
            ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
            ('TOPPADDING',    (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ('LEFTPADDING',   (0,0), (-1,-1), 8),
            ('RIGHTPADDING',  (0,0), (-1,-1), 6),
            ('BOX',           (0,0), (-1,-1), 0.5, BORDER_GREY),
            ('BACKGROUND',    (0,0), (-1,-1), WHITE),
        ]))
        
        elements.append(Spacer(1, 10))
        elements.append(terms_tbl)

    doc.build(elements, onFirstPage=_draw_page_decorations, onLaterPages=_draw_page_decorations)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=f"vendor_{vendor.id}_{vendor.name.replace(' ', '_')}_profile.pdf"
    )

@bp.route('/public/purchase/<token>')
def public_purchase(token):
    from datetime import datetime
    from app.models import Company, PurchaseSettings
    from app.pdf_utils import generate_professional_pdf
    from flask import make_response
    
    bill = PurchaseBill.query.filter_by(access_token=token).first_or_404()
    
    # Check expiry
    if bill.token_expiry and bill.token_expiry < datetime.utcnow():
        return "Link expired", 403
        
    company = Company.query.first()
    settings = PurchaseSettings.query.first()
    
    try:
        buffer = generate_professional_pdf('purchase', bill, company, settings)
        
        response = make_response(buffer)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename="bill_{bill.bill_number or "unknown"}.pdf"'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        return str(e), 500


@bp.route('/vendor-advance/<int:id>/approve', methods=['POST'])
@login_required
def approve_vendor_advance(id):
    if not getattr(current_user, 'is_admin', False):
        flash('Only admins can approve advances.', 'danger')
        return redirect(request.referrer or url_for('purchase.vendor_list'))
        
    advance = VendorAdvance.query.get_or_404(id)
    if advance.is_approved:
        flash('Advance is already approved.', 'info')
        return redirect(request.referrer or url_for('purchase.vendor_detail', id=advance.vendor_id))
        
    advance.is_approved = True
    advance.is_rejected = False
    advance.approved_by = current_user.id
    advance.approved_at = datetime.utcnow()

    db.session.commit()
    log_activity('Vendors', f'Approved Vendor Advance',
                f'Amount: PKR {advance.amount:,.2f}')
    flash('Vendor advance approved successfully.', 'success')
    return redirect(request.referrer or url_for('purchase.vendor_detail', id=advance.vendor_id))


@bp.route('/vendor-advance/<int:id>/reject', methods=['POST'])
@login_required
def reject_vendor_advance(id):
    if not getattr(current_user, 'is_admin', False):
        flash('Only admins can reject advances.', 'danger')
        return redirect(request.referrer or url_for('purchase.vendor_list'))
        
    advance = VendorAdvance.query.get_or_404(id)
    reason = request.form.get('reason', '')
    
    advance.is_approved = False
    advance.is_rejected = True
    advance.rejection_reason = reason

    db.session.commit()
    log_activity('Vendors', f'Rejected Vendor Advance',
                f'Amount: PKR {advance.amount:,.2f}, Reason: {reason or "N/A"}')
    flash('Vendor advance rejected.', 'warning')
    return redirect(request.referrer or url_for('purchase.vendor_detail', id=advance.vendor_id))
