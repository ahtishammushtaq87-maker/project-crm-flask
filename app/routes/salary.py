from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from app.utils import permission_required
from flask_login import login_required, current_user
from app import db
from app.models import Staff, SalaryAdvance, SalaryPayment
from app.forms import StaffForm, SalaryAdvanceForm, SalaryPaymentForm
from datetime import datetime
from sqlalchemy import extract
import io

bp = Blueprint('salary', __name__)

from app.routes.filters import apply_saved_filter_to_query

# --- Staff Management ---

@bp.route('/staff')
@login_required
def staff_list():
    query = Staff.query
    query = apply_saved_filter_to_query(query, 'staff', request.args)
    staff_members = query.all()
    return render_template('salary/staff_list.html', staff_members=staff_members, active_module='staff')

@bp.route('/staff/add', methods=['GET', 'POST'])
@login_required
@permission_required('salary', action='add')
def add_staff():
    form = StaffForm()
    if form.validate_on_submit():
        staff = Staff(
            name=form.name.data,
            designation=form.designation.data,
            monthly_salary=form.monthly_salary.data,
            joining_date=form.joining_date.data or datetime.utcnow().date(),
            is_active=form.is_active.data
        )
        staff.calculate_daily_salary()  # Calculate daily salary
        db.session.add(staff)
        db.session.commit()
        flash('Staff member added successfully!', 'success')
        return redirect(url_for('salary.staff_list'))
    return render_template('salary/staff_form.html', form=form, title="Add Staff")

@bp.route('/staff/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('salary', action='edit')
def edit_staff(id):
    staff = Staff.query.get_or_404(id)
    form = StaffForm(obj=staff)
    if form.validate_on_submit():
        staff.name = form.name.data
        staff.designation = form.designation.data
        staff.monthly_salary = form.monthly_salary.data
        staff.joining_date = form.joining_date.data
        staff.is_active = form.is_active.data
        staff.calculate_daily_salary()  # Recalculate daily salary
        db.session.commit()
        flash('Staff details updated!', 'success')
        return redirect(url_for('salary.staff_list'))
    return render_template('salary/staff_form.html', form=form, title="Edit Staff")

@bp.route('/staff/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('salary', action='delete')
def delete_staff(id):
    staff = Staff.query.get_or_404(id)
    db.session.delete(staff)
    db.session.commit()
    flash('Staff member deleted.', 'info')
    return redirect(url_for('salary.staff_list'))

@bp.route('/staff/payments/<int:id>')
@login_required
def staff_payments(id):
    staff = Staff.query.get_or_404(id)
    payments = SalaryPayment.query.filter_by(staff_id=id).order_by(SalaryPayment.year.desc(), SalaryPayment.month.desc()).all()
    return render_template('salary/staff_payments.html', staff=staff, payments=payments)

@bp.route('/staff/bulk-upload', methods=['GET', 'POST'])
@login_required
@permission_required('salary', action='add')
def bulk_upload_staff():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(url_for('salary.bulk_upload_staff'))
            
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('salary.bulk_upload_staff'))
            
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('salary.bulk_upload_staff'))
            
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            import datetime
            file_content = file.read()
            wb = load_workbook(filename=BytesIO(file_content), data_only=True, read_only=True)
            ws = wb.active
            rows = list(ws.values)
            if not rows:
                flash('File is empty', 'error')
                return redirect(url_for('salary.bulk_upload_staff'))
            headers = [str(h).strip().lower() if h else '' for h in rows[0]]
            
            required_columns = ['name', 'monthly_salary']
            missing = [col for col in required_columns if col not in headers]
            if missing:
                flash(f'Missing required columns: {", ".join(missing)}', 'error')
                return redirect(url_for('salary.bulk_upload_staff'))
                
            added = 0
            errors = []
            
            for idx, row in enumerate(rows[1:], start=2):
                try:
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = val
                            
                    name = str(row_dict.get('name', '')).strip()
                    if not name:
                        errors.append(f'Row {idx}: Missing name')
                        continue
                        
                    monthly_salary = float(row_dict.get('monthly_salary', 0))
                    
                    staff = Staff(
                        name=name,
                        designation=str(row_dict.get('designation', '')).strip() if row_dict.get('designation') else None,
                        monthly_salary=monthly_salary,
                        is_active=bool(str(row_dict.get('is_active', '1')).strip().lower() in ['yes', '1', 'true', 't', 'active'])
                    )
                    staff.calculate_daily_salary()
                    db.session.add(staff)
                    added += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')

            
            db.session.commit()
            if added > 0:
                flash(f'Successfully added {added} staff members!', 'success')
            if errors:
                flash(f'Errors: {"; ".join(errors[:10])}', 'warning')
            return redirect(url_for('salary.staff_list'))
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'error')
            return redirect(url_for('salary.bulk_upload_staff'))
            
    return render_template('salary/staff_bulk_upload.html')

@bp.route('/staff/download-sample')
@login_required
def download_sample_staff():
    try:
        from openpyxl import Workbook
        from io import BytesIO
        from flask import send_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Staff'
        
        headers = ['name', 'designation', 'monthly_salary', 'is_active']
        ws.append(headers)
        
        sample_data = [
            ['John Doe', 'Manager', 50000, 'yes'],
            ['Jane Smith', 'Operator', 30000, 'yes']
        ]
        
        for row in sample_data:
            ws.append(row)

            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='sample_staff.xlsx', as_attachment=True)
    except Exception as e:
        flash(f'Error creating sample: {str(e)}', 'error')
        return redirect(url_for('salary.bulk_upload_staff'))

# --- Salary Advances ---

@bp.route('/advances')
@login_required
def advance_list():
    advances = SalaryAdvance.query.order_by(SalaryAdvance.date.desc()).all()
    return render_template('salary/advance_list.html', advances=advances)

@bp.route('/advances/add', methods=['GET', 'POST'])
@login_required
@permission_required('salary', action='add')
def add_advance():
    form = SalaryAdvanceForm()
    form.staff_id.choices = [(s.id, s.name) for s in Staff.query.filter_by(is_active=True).all()]
    if form.validate_on_submit():
        advance = SalaryAdvance(
            staff_id=form.staff_id.data,
            amount=form.amount.data,
            date=form.date.data,
            description=form.description.data
        )
        db.session.add(advance)
        db.session.commit()
        flash('Salary advance recorded!', 'success')
        return redirect(url_for('salary.advance_list'))
    return render_template('salary/advance_form.html', form=form, title="Record Advance")

@bp.route('/advances/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('salary', action='delete')
def delete_advance(id):
    advance = SalaryAdvance.query.get_or_404(id)
    if advance.is_deducted:
        flash('Cannot delete a deducted advance. Delete the salary payment first.', 'danger')
        return redirect(url_for('salary.advance_list'))
    db.session.delete(advance)
    db.session.commit()
    flash('Salary advance deleted.', 'info')
    return redirect(url_for('salary.advance_list'))

# --- Salary Payments ---

@bp.route('/payments')
@login_required
def payment_list():
    query = SalaryPayment.query
    query = apply_saved_filter_to_query(query, 'salary_payment', request.args)
    payments = query.order_by(SalaryPayment.year.desc(), SalaryPayment.month.desc()).all()
    return render_template('salary/payment_list.html', payments=payments, auto_download_id=None, active_module='salary_payment')

@bp.route('/payments/<int:paid>')
@login_required
def payment_list_with_download(paid):
    payments = SalaryPayment.query.order_by(SalaryPayment.year.desc(), SalaryPayment.month.desc()).all()
    return render_template('salary/payment_list.html', payments=payments, auto_download_id=paid)

@bp.route('/payments/process', methods=['GET', 'POST'])
@login_required
def process_salary_select():
    staff_members = Staff.query.filter_by(is_active=True).all()
    return render_template('salary/process_salary_select.html', staff_members=staff_members)

@bp.route('/payments/pay/<int:staff_id>', methods=['GET', 'POST'])
@login_required
def pay_salary(staff_id):
    staff = Staff.query.get_or_404(staff_id)
    form = SalaryPaymentForm()
    form.staff_id.choices = [(staff.id, staff.name)]
    
    # Pre-fill data
    if request.method == 'GET':
        form.staff_id.data = staff.id
        form.base_salary.data = staff.monthly_salary
        form.month.data = datetime.utcnow().month
        form.year.data = datetime.utcnow().year
        form.payment_date.data = datetime.utcnow().date()
        
        # Calculate outstanding advances
        outstanding_advances = SalaryAdvance.query.filter_by(staff_id=staff.id, is_deducted=False).all()
        form.advance_deduction.data = sum(a.amount for a in outstanding_advances)

    if form.validate_on_submit():
        print(f"Form validated for {staff.name}")
        # Check if already paid for this month/year
        existing = SalaryPayment.query.filter_by(staff_id=staff.id, month=form.month.data, year=form.year.data).first()
        if existing:
            flash(f'Salary already processed for {staff.name} for {form.month.data}/{form.year.data}', 'warning')
            return redirect(url_for('salary.payment_list'))

        net_salary = form.base_salary.data + form.bonus.data - form.advance_deduction.data - form.other_deductions.data
        
        payment = SalaryPayment(
            staff_id=staff.id,
            month=form.month.data,
            year=form.year.data,
            base_salary=form.base_salary.data,
            advance_deduction=form.advance_deduction.data,
            bonus=form.bonus.data,
            other_deductions=form.other_deductions.data,
            net_salary=net_salary,
            payment_date=form.payment_date.data,
            payment_method=form.payment_method.data,
            status=form.status.data,
            notes=form.notes.data
        )
        db.session.add(payment)
        db.session.flush() # Get payment.id

        # Mark advances as deducted
        if form.advance_deduction.data > 0:
            unpaid_advances = SalaryAdvance.query.filter_by(staff_id=staff.id, is_deducted=False).order_by(SalaryAdvance.date.asc()).all()
            deducted_so_far = 0
            for adv in unpaid_advances:
                if deducted_so_far < form.advance_deduction.data:
                    adv.is_deducted = True
                    adv.salary_payment_id = payment.id
                    deducted_so_far += adv.amount
        
        db.session.commit()
        flash(f'Salary processed for {staff.name}. Payment ID: {payment.id}', 'success')
        return redirect(url_for('salary.payment_list_with_download', paid=payment.id))

    return render_template('salary/pay_form.html', form=form, staff=staff)

@bp.route('/payments/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('salary', action='delete')
def delete_payment(id):
    payment = SalaryPayment.query.get_or_404(id)
    # Revert advances
    for advance in payment.deducted_advances:
        advance.is_deducted = False
        advance.salary_payment_id = None
    
    db.session.delete(payment)
    db.session.commit()
    flash('Salary payment deleted and advances reverted.', 'info')
    return redirect(url_for('salary.payment_list'))

@bp.route('/payments/download/<int:id>')
@login_required
def download_pay_stub(id):
    from app.models import Attendance, SalaryAdvance, Company
    from calendar import monthrange
    from datetime import date
    
    payment = SalaryPayment.query.get_or_404(id)
    staff = Staff.query.get_or_404(payment.staff_id)
    company = Company.query.first()
    
    year = payment.year
    month = payment.month
    _, last_day = monthrange(year, month)
    
    start_date = date(year, month, 1)
    end_date = date(year, month, last_day)
    
    attendance_records = Attendance.query.filter(
        Attendance.staff_id == staff.id,
        Attendance.date >= start_date,
        Attendance.date <= end_date
    ).all()
    
    total_hours = sum(a.hours_worked for a in attendance_records)
    total_minutes = sum(a.minutes_worked for a in attendance_records)
    total_hours += total_minutes / 60
    days_worked = len(attendance_records)
    hourly_rate = staff.monthly_salary / 30 / 8
    hourly_earnings = total_hours * hourly_rate
    
    attendance_data = {
        'total_hours': total_hours,
        'days_worked': days_worked,
        'hourly_rate': hourly_rate,
        'hourly_earnings': hourly_earnings
    }
    
    advances = SalaryAdvance.query.filter(
        SalaryAdvance.staff_id == staff.id
    ).all()
    
    from app.pdf_templates.pay_stub_template import generate_pay_stub_pdf_with_company
    
    pdf_buffer = generate_pay_stub_pdf_with_company(staff, payment, company, attendance_data, advances)
    
    filename = f"PaySlip_{staff.name.replace(' ', '_')}_{payment.month:02d}_{payment.year}.pdf"
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/payments/download/latest/<int:staff_id>')
@login_required
def download_latest_pay_stub(staff_id):
    from app.models import Attendance, SalaryAdvance, Company
    from calendar import monthrange
    from datetime import date
    
    payment = SalaryPayment.query.filter_by(staff_id=staff_id).order_by(SalaryPayment.year.desc(), SalaryPayment.month.desc()).first()
    if not payment:
        flash('No salary payment found for this staff.', 'warning')
        return redirect(url_for('salary.staff_list'))
    
    staff = Staff.query.get_or_404(staff_id)
    company = Company.query.first()
    
    year = payment.year
    month = payment.month
    _, last_day = monthrange(year, month)
    
    start_date = date(year, month, 1)
    end_date = date(year, month, last_day)
    
    attendance_records = Attendance.query.filter(
        Attendance.staff_id == staff.id,
        Attendance.date >= start_date,
        Attendance.date <= end_date
    ).all()
    
    total_hours = sum(a.hours_worked for a in attendance_records)
    total_minutes = sum(a.minutes_worked for a in attendance_records)
    total_hours += total_minutes / 60
    days_worked = len(attendance_records)
    hourly_rate = staff.monthly_salary / 30 / 8
    hourly_earnings = total_hours * hourly_rate
    
    attendance_data = {
        'total_hours': total_hours,
        'days_worked': days_worked,
        'hourly_rate': hourly_rate,
        'hourly_earnings': hourly_earnings
    }
    
    advances = SalaryAdvance.query.filter(
        SalaryAdvance.staff_id == staff.id
    ).all()
    
    from app.pdf_templates.pay_stub_template import generate_pay_stub_pdf_with_company
    
    pdf_buffer = generate_pay_stub_pdf_with_company(staff, payment, company, attendance_data, advances)
    
    filename = f"PaySlip_{staff.name.replace(' ', '_')}_{payment.month:02d}_{payment.year}.pdf"
    
    return send_file(
        pdf_buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@bp.route('/payments/bulk-upload', methods=['GET', 'POST'])
@login_required
@permission_required('salary', action='add')
def bulk_upload_payment():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(url_for('salary.bulk_upload_payment'))
            
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('salary.bulk_upload_payment'))
            
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('salary.bulk_upload_payment'))
            
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            from datetime import datetime
            file_content = file.read()
            wb = load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            rows = list(ws.values)
            if not rows:
                flash('File is empty', 'error')
                return redirect(url_for('salary.bulk_upload_payment'))
            headers = [str(h).strip().lower() if h else '' for h in rows[0]]
            
            required_columns = ['staff_id', 'month', 'year', 'base_salary', 'net_salary']
            missing = [col for col in required_columns if col not in headers]
            if missing:
                flash(f'Missing required columns: {", ".join(missing)}', 'error')
                return redirect(url_for('salary.bulk_upload_payment'))
                
            added = 0
            errors = []
            
            for idx, row in enumerate(rows[1:], start=2):
                try:
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = val
                            
                    staff_id = row_dict.get('staff_id')
                    try:
                        staff_id = int(str(staff_id).strip())
                    except ValueError:
                        errors.append(f'Row {idx}: staff_id must be a number')
                        continue
                        
                    staff = Staff.query.get(staff_id)
                    if not staff:
                        errors.append(f'Row {idx}: Staff ID {staff_id} not found')
                        continue
                        
                    month = int(row_dict.get('month', 1))
                    year = int(row_dict.get('year', 2024))
                    base_salary = float(row_dict.get('base_salary', 0))
                    net_salary = float(row_dict.get('net_salary', 0))
                    
                    existing = SalaryPayment.query.filter_by(staff_id=staff.id, month=month, year=year).first()
                    if existing:
                        errors.append(f'Row {idx}: Payment for staff {staff.name} for {month}/{year} already exists')
                        continue
                        
                    payment_date_val = row_dict.get('payment_date')
                    if isinstance(payment_date_val, datetime):
                        payment_date = payment_date_val.date()
                    elif payment_date_val:
                        payment_date = datetime.strptime(str(payment_date_val), '%Y-%m-%d').date()
                    else:
                        payment_date = datetime.utcnow().date()
                        
                    payment = SalaryPayment(
                        staff_id=staff.id,
                        month=month,
                        year=year,
                        base_salary=base_salary,
                        advance_deduction=float(row_dict.get('advance_deduction', 0) or 0),
                        bonus=float(row_dict.get('bonus', 0) or 0),
                        other_deductions=float(row_dict.get('other_deductions', 0) or 0),
                        net_salary=net_salary,
                        payment_date=payment_date,
                        payment_method=str(row_dict.get('payment_method', 'cash')).strip(),
                        status=str(row_dict.get('status', 'paid')).strip().lower(),
                        notes=str(row_dict.get('notes', '')).strip() if row_dict.get('notes') else None
                    )
                    db.session.add(payment)
                    added += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
            
            db.session.commit()
            if added > 0:
                flash(f'Successfully added {added} payments!', 'success')
            if errors:
                flash(f'Errors: {"; ".join(errors[:10])}', 'warning')
            return redirect(url_for('salary.payment_list'))
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'error')
            return redirect(url_for('salary.bulk_upload_payment'))
            
    return render_template('salary/payment_bulk_upload.html')

@bp.route('/payments/download-sample')
@login_required
def download_sample_payment():
    try:
        from openpyxl import Workbook
        from io import BytesIO
        from flask import send_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Payments'
        
        headers = ['staff_id', 'month', 'year', 'base_salary', 'advance_deduction', 'bonus', 'other_deductions', 'net_salary', 'payment_date', 'payment_method', 'status', 'notes']
        ws.append(headers)
        
        staff = Staff.query.first()
        staff_id = staff.id if staff else 1
        
        now = datetime.utcnow()
        current_month = now.month
        current_year = now.year
        current_date_str = now.strftime('%Y-%m-%d')
        
        sample_data = [
            [staff_id, current_month, current_year, 50000, 0, 2000, 0, 52000, current_date_str, 'Bank', 'paid', 'Monthly salary'],
            [staff_id, current_month, current_year, 50000, 5000, 0, 0, 45000, current_date_str, 'Cash', 'paid', 'Monthly salary with deduction']
        ]
        
        for row in sample_data:
            ws.append(row)

            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='sample_payments.xlsx', as_attachment=True)
    except Exception as e:
        flash(f'Error creating sample: {str(e)}', 'error')
        return redirect(url_for('salary.bulk_upload_payment'))

# --- API for dynamic updates ---
@bp.route('/api/get_staff_info/<int:staff_id>')
@login_required
def get_staff_info(staff_id):
    staff = Staff.query.get_or_404(staff_id)
    outstanding_advance = sum(a.amount for a in staff.advances if not a.is_deducted)
    return jsonify({
        'monthly_salary': staff.monthly_salary,
        'outstanding_advance': outstanding_advance
    })
