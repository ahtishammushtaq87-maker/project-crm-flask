from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response, current_app
from flask_login import login_required, current_user
from app.utils import permission_required, log_activity
from app import db
from app.models import Sale, SaleItem, Product, ProductWarehouseStock, Warehouse, Customer, Vendor, Company, InvoiceSettings, Currency, CustomerAdvance, SaleReturn, Salesman, CustomerGroup, Payment, PaymentMethod, SalesmanGroup, User
from app.forms import SaleForm, CustomerForm, InvoiceSettingsForm, SalesmanForm, CustomerGroupForm
from datetime import datetime, date
from sqlalchemy import func, or_, and_
from app.pdf_utils import generate_professional_pdf
import os
from werkzeug.utils import secure_filename
import json
from app.routes.filters import apply_saved_filter_to_query
import os
import io
import csv
import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

bp = Blueprint('sales', __name__)


def validate_item_discounts(sale_items):
    """
    Check each line item's OWN discount against that item's product-specific
    min/max discount rule (InvoiceSettings.product_discount_conditions).
    Returns a list of human-readable violation messages (empty if all pass).
    """
    violations = []
    settings = InvoiceSettings.query.first()
    if not settings or not settings.product_discount_conditions:
        return violations

    try:
        conditions = json.loads(settings.product_discount_conditions)
    except Exception:
        return violations

    rules_by_product = {}
    for cond in conditions:
        try:
            rules_by_product[cond['product_id']] = cond
        except Exception:
            continue

    for item in sale_items:
        cond = rules_by_product.get(item['product_id'])
        if not cond:
            continue
        item_discount = item.get('discount', 0) or 0
        min_allowed = float(cond.get('min_discount', 0))
        max_allowed = float(cond.get('max_discount', 0))
        product = Product.query.get(item['product_id'])
        product_label = product.name if product else f"Product #{item['product_id']}"

        if item_discount < min_allowed:
            violations.append(f"{product_label}: Discount PKR {item_discount} is less than min PKR {min_allowed}")
        if max_allowed > 0 and item_discount > max_allowed:
            violations.append(f"{product_label}: Discount PKR {item_discount} is more than max PKR {max_allowed}")

    return violations


@bp.route('/invoices')
@login_required
def invoices():
    status = request.args.get('status', 'all')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')
    salesman_id = request.args.get('salesman_id', type=int)
    invoice_number = request.args.get('invoice_number')
    customer_id = request.args.get('customer_id', type=int)
    customer_group_id = request.args.get('customer_group_id', type=int)
    search = request.args.get('search', '')
    
    query = Sale.query

    # ── Universal-status expressions (aligned with ApprovalService.get_status) ──
    # A Sale is "cancelled" when is_rejected + reason 'Cancelled by Admin' (the
    # Sale status enum has no 'cancelled' value, so cancel uses that fallback).
    CANCELLED_REASON = 'Cancelled by Admin'
    is_cancelled = (Sale.is_rejected == True) & (Sale.rejection_reason == CANCELLED_REASON)
    not_cancelled = (
        (Sale.is_rejected == False) |
        (Sale.rejection_reason.is_(None)) |
        (Sale.rejection_reason != CANCELLED_REASON)
    )
    is_rejected_only = (Sale.is_rejected == True) & (
        (Sale.rejection_reason.is_(None)) | (Sale.rejection_reason != CANCELLED_REASON)
    )
    is_draft_only = (Sale.is_rejected == False) & (Sale.is_approved == False) & (Sale.is_draft == True)
    is_pending = (Sale.is_rejected == False) & (Sale.is_approved == False) & (Sale.is_draft == False)

    if status == 'overdue':
        query = query.filter(
            not_cancelled,
            Sale.is_draft == False,
            Sale.status != 'paid',
            Sale.due_date < datetime.utcnow()
        )
    elif status == 'unapproved':
        # Pending invoices, OR invoices with PENDING (not rejected) payments/advances.
        # Draft/cancelled invoices are excluded entirely — their pending payment /
        # discount approval moves with them into the Draft/Cancelled tabs.
        query = query.filter(
            not_cancelled,
            Sale.is_draft == False,
            (
                is_pending |
                (Sale.payments.any((Payment.is_approved == False) & (Payment.is_rejected == False))) |
                (Sale.adjusted_advances.any((CustomerAdvance.is_approved == False) & (CustomerAdvance.is_rejected == False)))
            )
        )
    elif status == 'rejected_items':
        # Genuinely rejected invoices (excludes cancelled), OR any rejected payment/advance.
        query = query.filter(
            is_rejected_only |
            (Sale.payments.any(Payment.is_rejected == True)) |
            (Sale.adjusted_advances.any(CustomerAdvance.is_rejected == True))
        )
    elif status == 'cancelled':
        query = query.filter(is_cancelled)
    elif status == 'draft':
        query = query.filter(is_draft_only)
    elif status == 'all':
        # Active invoices only — cancelled and drafts live in their own tabs.
        query = query.filter(not_cancelled, Sale.is_draft == False)
    else:
        # paid / unpaid / partial payment-status tabs (exclude drafts & cancelled).
        query = query.filter(Sale.status == status, not_cancelled, Sale.is_draft == False)
    
    if salesman_id:
        query = query.filter(Sale.salesman_id == salesman_id)

    if invoice_number:
        query = query.filter(Sale.invoice_number == invoice_number)
    
    if customer_id:
        query = query.filter(Sale.customer_id == customer_id)
    
    if customer_group_id:
        query = query.join(Customer).filter(Customer.group_id == customer_group_id)
    
    if search:
        search_filter = f"%{search}%"
        query = query.outerjoin(Customer).filter(
            (Sale.invoice_number.ilike(search_filter)) |
            (Customer.name.ilike(search_filter))
        )
    
    if from_date:
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d')
            query = query.filter(Sale.date >= from_date_obj)
        except ValueError:
            pass  # Invalid date format, ignore filter
    
    if to_date:
        try:
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d')
            # Add one day to include the end date
            to_date_obj = to_date_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(Sale.date <= to_date_obj)
        except ValueError:
            pass  # Invalid date format, ignore filter
    
    query = apply_saved_filter_to_query(query, 'sale', request.args)
    
    sales = query.order_by(Sale.date.desc()).all()
    
    # Calculate totals
    total_subtotal = sum(sale.subtotal for sale in sales)
    total_tax = sum(sale.tax for sale in sales)
    total_amount = sum(sale.total for sale in sales)
    total_paid = sum(sale.paid_amount for sale in sales)
    total_balance = sum(sale.balance_due for sale in sales)
    
    # ── Unapproved summary (for alert box at top of page) ─────────────────
    # Always calculated regardless of which tab is active. EXCLUDE rejected items,
    # AND exclude anything belonging to a Draft or Cancelled invoice — those move
    # into the Draft/Cancelled tabs and must not raise a top notification.
    inv_draft_or_cancelled = (Sale.is_draft == True) | is_cancelled

    unapproved_invoices_all = Sale.query.filter(
        Sale.is_approved == False, Sale.is_rejected == False, Sale.is_draft == False
    ).all()
    unapproved_invoice_total = sum(s.total for s in unapproved_invoices_all)
    unapproved_invoice_count = len(unapproved_invoices_all)

    unapproved_payments_all = Payment.query.filter(
        Payment.is_approved == False, Payment.is_rejected == False,
        ~Payment.invoice.has(inv_draft_or_cancelled)
    ).all()
    unapproved_payment_total = sum(p.amount for p in unapproved_payments_all)
    unapproved_payment_count = len(unapproved_payments_all)

    unapproved_advances_all = CustomerAdvance.query.filter(
        CustomerAdvance.is_approved == False, CustomerAdvance.is_rejected == False,
        ~CustomerAdvance.adjusted_invoice.has(inv_draft_or_cancelled)
    ).all()
    unapproved_advance_total = sum(a.amount for a in unapproved_advances_all)
    unapproved_advance_count = len(unapproved_advances_all)
    
    # NEW: Calculate rejected items count for the new tab (Invoices OR Payments OR Advances)
    rejected_items_all = Sale.query.filter(
        is_rejected_only |
        (Sale.payments.any(Payment.is_rejected == True)) |
        (Sale.adjusted_advances.any(CustomerAdvance.is_rejected == True))
    ).all()
    rejected_item_count = len(rejected_items_all)

    # Cancelled & Draft counts for their dedicated tabs
    cancelled_count = Sale.query.filter(is_cancelled).count()
    draft_count = Sale.query.filter(is_draft_only).count()
    # ──────────────────────────────────────────────────────────────────────

    # Identify overdue invoices with suspended discounts (regardless of tab filter)
    # Use is_discount_restricted property which handles overdue status and customer group rules
    all_active_sales_with_discounts = Sale.query.filter(Sale.status != 'paid', Sale.discount > 0).all()
    suspended_discount_invoices = [s for s in all_active_sales_with_discounts if s.is_discount_restricted]

    # Get company date format
    company = Company.query.first()
    date_format = company.date_format if company and company.date_format else '%Y-%m-%d'
    
    # Load data for filter dropdowns
    salesmen = Salesman.query.filter_by(is_active=True).all()
    all_customers = Customer.query.filter_by(is_active=True).order_by(Customer.name).all()
    customer_groups = CustomerGroup.query.filter_by(is_active=True).order_by(CustomerGroup.name).all()
    # Limit number of invoices in dropdown to avoid performance issues if there are thousands
    recent_invoices = Sale.query.order_by(Sale.invoice_number.desc()).limit(1000).all()
    
    return render_template('sales/invoices.html', 
                         sales=sales, 
                         current_status=status,
                         from_date=from_date,
                         to_date=to_date,
                         salesman_id=salesman_id,
                         invoice_number=invoice_number,
                         customer_id=customer_id,
                         customer_group_id=customer_group_id,
                         search=search,
                         salesmen=salesmen,
                         all_customers=all_customers,
                         customer_groups=customer_groups,
                         recent_invoices=recent_invoices,
                         total_subtotal=total_subtotal,
                         total_tax=total_tax,
                         total_amount=total_amount,
                         total_paid=total_paid,
                         total_balance=total_balance,
                         date_format=date_format,
                         active_module='sale',
                         filter_id=request.args.get('filter_id'),
                         unapproved_invoice_total=unapproved_invoice_total,
                         unapproved_invoice_count=unapproved_invoice_count,
                         unapproved_payment_total=unapproved_payment_total,
                         unapproved_payment_count=unapproved_payment_count, 
                         unapproved_invoices_all=unapproved_invoices_all, 
                         unapproved_payments_all=unapproved_payments_all, 
                         unapproved_advance_total=unapproved_advance_total, 
                         unapproved_advance_count=unapproved_advance_count, 
                         unapproved_advances_all=unapproved_advances_all, 
                         rejected_item_count=rejected_item_count,
                         cancelled_count=cancelled_count,
                         draft_count=draft_count,
                         suspended_discount_invoices=suspended_discount_invoices)

@bp.route('/invoice/create', methods=['GET', 'POST'])
@login_required
@permission_required('sales', action='add')
def create_invoice():
    form = SaleForm()
    customers = Customer.query.all()
    vendors = Vendor.query.all()
    products = Product.query.all()
    currencies = Currency.query.filter_by(is_active=True).all()
    
    customer_advances = {c.id: float(c.remaining_advance_balance) for c in customers}
    customer_total_advances = {c.id: float(c.total_advances_received) for c in customers}
    
    form.customer_id.choices = [(0, 'Walk-in Customer')] + [(c.id, c.name) for c in customers]
    form.salesman_id.choices = [('', 'Select Salesman')] + [(s.id, s.name) for s in Salesman.query.filter_by(is_active=True).all()]
    
    if request.method == 'POST':
        customer_id = request.form.get('customer_id')
        selected_vendor_id = request.form.get('vendor_id')
        salesman_id = request.form.get('salesman_id')
        date = datetime.strptime(request.form.get('date'), '%Y-%m-%d')
        
        # Get items from form
        items = []
        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        deliveries = request.form.getlist('delivery[]')
        item_discounts = request.form.getlist('discount[]')
        warehouses = request.form.getlist('warehouse_id[]')

        subtotal = 0
        item_delivery_total = 0
        sale_items = []

        for i in range(len(product_ids)):
            if product_ids[i] and quantities[i] and float(quantities[i]) > 0:
                product = Product.query.get(int(product_ids[i]))
                quantity = float(quantities[i])
                price = float(prices[i])
                delivery_fee = float(deliveries[i]) if i < len(deliveries) else 0
                item_discount = float(item_discounts[i]) if i < len(item_discounts) and item_discounts[i] else 0
                item_subtotal = quantity * price
                total = item_subtotal + delivery_fee
                subtotal += total
                item_delivery_total += delivery_fee

                wh = None
                try:
                    wh = int(warehouses[i]) if i < len(warehouses) and warehouses[i] else None
                except Exception:
                    wh = None

                sale_items.append({
                    'product_id': product.id,
                    'warehouse_id': wh,
                    'quantity': quantity,
                    'unit_price': price,
                    'delivery_fee': delivery_fee,
                    'discount': item_discount,
                    'total': total
                })

        tax_rate = float(request.form.get('tax_rate', 0))
        tax = subtotal * (tax_rate / 100)
        # Discount is always the sum of the per-item discounts entered above.
        # Recomputed server-side (never trust the readonly header field the browser posts).
        discount = sum(item['discount'] for item in sale_items)
        delivery_charge = float(request.form.get('delivery_charge', 0))
        advance_applied = float(request.form.get('advance_applied', 0))

        total = subtotal + tax + delivery_charge - discount

        # Generate invoice number using settings
        settings = InvoiceSettings.query.first()
        if not settings:
            settings = InvoiceSettings(invoice_prefix='INV-', invoice_suffix='', next_number=1)
            db.session.add(settings)
        invoice_number = f"{settings.invoice_prefix}{settings.next_number}{settings.invoice_suffix}"
        settings.next_number += 1

        creator_is_admin = (current_user.role == 'admin')

        # Check discount conditions: each item's OWN discount is validated against
        # that item's product-specific min/max rule (rules are per-product, so with
        # per-item discounts now available each line must be checked individually).
        discount_violations = validate_item_discounts(sale_items)

        violation_found = len(discount_violations) > 0
        
        sale = Sale(
            invoice_number=invoice_number,
            customer_id=customer_id if customer_id != '0' else None,
            vendor_id=selected_vendor_id if selected_vendor_id and selected_vendor_id != '0' else None,
            date=date,
            subtotal=subtotal,
            tax_rate=tax_rate,
            tax=tax,
            discount=discount,
            discount_type='fixed',  # discount is always the summed PKR amount of per-item discounts
            delivery_charge=delivery_charge,
            advance_applied=advance_applied,
            total=total,
            status=request.form.get('status', 'unpaid'),
            currency_id=request.form.get('currency_id', None),
            exchange_rate=float(request.form.get('exchange_rate', 1)),
            salesman_id=salesman_id if salesman_id and salesman_id != '0' else None,
            paid_amount=advance_applied if not violation_found else 0, # Don't apply yet if violation
            created_by=current_user.id,
            is_approved=creator_is_admin and not violation_found,
            approved_by=current_user.id if (creator_is_admin and not violation_found) else None,
            approved_at=datetime.utcnow() if (creator_is_admin and not violation_found) else None,
            discount_violation="; ".join(discount_violations) if violation_found else None
        )

        if violation_found:
            flash(f"Notice: This invoice requires admin approval due to discount conditions: {'; '.join(discount_violations)}", 'warning')

        due_date_str = request.form.get('due_date')
        if due_date_str:
            try:
                sale.due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
                sale.overdue_date = sale.due_date
            except ValueError:
                sale.due_date = None
        
        db.session.add(sale)
        db.session.flush()
        
        # Handle advance application - link to customer advances
        if customer_id and customer_id != '0' and advance_applied > 0 and not violation_found:
            customer = Customer.query.get(int(customer_id))
            remaining_to_apply = advance_applied
            
            # Get advances sorted by date (oldest first)
            advances = CustomerAdvance.query.filter_by(customer_id=customer.id).filter(
                CustomerAdvance.is_adjusted == False
            ).order_by(CustomerAdvance.date).all()
            
            for adv in advances:
                if remaining_to_apply <= 0:
                    break
                    
                available = adv.remaining_balance
                if available > 0:
                    apply_amt = min(available, remaining_to_apply)
                    adv.applied_amount += apply_amt
                    remaining_to_apply -= apply_amt
                    
                    if adv.remaining_balance <= 0:
                        adv.is_adjusted = True
                        adv.adjusted_invoice_id = sale.id
        
        db.session.add(sale)
        db.session.flush()

        # Add items
        for item in sale_items:
            sale_item = SaleItem(
                sale_id=sale.id,
                product_id=item['product_id'],
                warehouse_id=item.get('warehouse_id'),
                quantity=item['quantity'],
                unit_price=item['unit_price'],
                delivery_fee=item.get('delivery_fee', 0),
                discount=item.get('discount', 0),
                total=item['total']
            )
            db.session.add(sale_item)

        if sale.is_approved:
            for item in sale_items:
                # Update inventory (global + per-warehouse)
                product = Product.query.get(item['product_id'])
                if product:
                    # reduce total product quantity
                    product.update_quantity(-item['quantity'])

                    # Update per-warehouse stock if warehouse selected
                    wh_id = item.get('warehouse_id')
                    if wh_id:
                        wh_stock = ProductWarehouseStock.query.filter_by(product_id=item['product_id'], warehouse_id=wh_id).first()
                        if not wh_stock:
                            wh_stock = ProductWarehouseStock(product_id=item['product_id'], warehouse_id=wh_id, quantity=0)
                            db.session.add(wh_stock)
                        wh_stock.quantity -= item['quantity']
                    elif product.warehouse_id:
                        # Legacy fallback
                        wh_stock = ProductWarehouseStock.query.filter_by(product_id=item['product_id'], warehouse_id=product.warehouse_id).first()
                        if not wh_stock:
                            wh_stock = ProductWarehouseStock(product_id=item['product_id'], warehouse_id=product.warehouse_id, quantity=0)
                            db.session.add(wh_stock)
                        wh_stock.quantity -= item['quantity']
            sale.stock_updated = True
        
        # IMPORTANT: Trigger the official total calculation from the model
        # This ensures the stored 'total' matches the model logic 100%
        sale.calculate_totals()
        sale.update_status()

        db.session.commit()
        
        log_activity('Sales', f'Created Invoice #{sale.invoice_number}', 
                    f'Customer: {sale.customer.name if sale.customer else "Walk-in Customer"}, Total: {sale.total}')
        
        flash('Invoice created successfully!', 'success')
        return redirect(url_for('sales.invoice_detail', id=sale.id))
    
    salesmen = Salesman.query.filter_by(is_active=True).all()
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    return render_template('sales/create_invoice.html', form=form, products=products, customers=customers, vendors=vendors, currencies=currencies, now=datetime.now(), customer_advances=customer_advances, customer_total_advances=customer_total_advances, salesmen=salesmen, warehouses=warehouses)

@bp.route('/invoice/<int:id>')
@login_required
def invoice_detail(id):
    sale = Sale.query.get_or_404(id)
    company = Company.query.first()
    date_format = company.date_format if company and company.date_format else '%Y-%m-%d'
    returns = SaleReturn.query.filter_by(sale_id=sale.id).order_by(SaleReturn.date.desc()).all()
    # Get payment methods for dropdown and existing payments for display
    payment_methods = PaymentMethod.query.filter_by(is_active=True).order_by(PaymentMethod.name).all()
    payments = Payment.query.filter_by(invoice_id=sale.id).order_by(Payment.date.desc()).all()
    
    other_invoices = []
    if sale.customer_id:
        other_invoices = Sale.query.filter(
            Sale.customer_id == sale.customer_id,
            Sale.id != sale.id,
            Sale.is_approved == True
        ).order_by(Sale.date.desc()).all()
        
    return render_template('sales/invoice_detail.html', 
                           sale=sale, 
                           returns=returns, 
                           date_format=date_format, 
                           payment_methods=payment_methods, 
                           payments=payments, 
                           today=datetime.utcnow().strftime('%Y-%m-%d'),
                           other_invoices=other_invoices)

@bp.route('/invoice/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('sales', action='edit')
def edit_invoice(id):
    sale = Sale.query.get_or_404(id)
    form = SaleForm(obj=sale)
    customers = Customer.query.all()
    vendors = Vendor.query.all()
    products = Product.query.all()
    currencies = Currency.query.filter_by(is_active=True).all()
    salesmen = Salesman.query.filter_by(is_active=True).all()

    form.customer_id.choices = [(0, 'Walk-in Customer')] + [(c.id, c.name) for c in customers]
    form.salesman_id.choices = [('', 'Select Salesman')] + [(s.id, s.name) for s in salesmen]

    if request.method == 'POST':
        # Revert old inventory ONLY if stock was previously deducted (i.e. sale was approved)
        if sale.stock_updated:
            for item in sale.items:
                product = Product.query.get(item.product_id)
                if product:
                    # restore total quantity
                    product.update_quantity(item.quantity)

                    # restore per-warehouse quantity if used
                    try:
                        wh_id = getattr(item, 'warehouse_id', None)
                    except Exception:
                        wh_id = None
                    if wh_id:
                        wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=wh_id).first()
                        if not wh_stock:
                            wh_stock = ProductWarehouseStock(product_id=item.product_id, warehouse_id=wh_id, quantity=0)
                            db.session.add(wh_stock)
                        wh_stock.quantity += item.quantity
                    elif product.warehouse_id:
                        wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=product.warehouse_id).first()
                        if not wh_stock:
                            wh_stock = ProductWarehouseStock(product_id=item.product_id, warehouse_id=product.warehouse_id, quantity=0)
                            db.session.add(wh_stock)
                        wh_stock.quantity += item.quantity
            sale.stock_updated = False

        # Delete old items
        SaleItem.query.filter_by(sale_id=sale.id).delete()

        # Rebuild from form
        customer_id = request.form.get('customer_id')
        selected_vendor_id = request.form.get('vendor_id')
        salesman_id = request.form.get('salesman_id')
        date = datetime.strptime(request.form.get('date'), '%Y-%m-%d')

        product_ids = request.form.getlist('product_id[]')
        quantities = request.form.getlist('quantity[]')
        prices = request.form.getlist('price[]')
        deliveries = request.form.getlist('delivery[]')
        item_discounts = request.form.getlist('discount[]')
        warehouses = request.form.getlist('warehouse_id[]')

        subtotal = 0
        sale_items = []

        for i in range(len(product_ids)):
            if product_ids[i] and quantities[i] and float(quantities[i]) > 0:
                product = Product.query.get(int(product_ids[i]))
                quantity = float(quantities[i])
                price = float(prices[i])
                delivery_fee = float(deliveries[i]) if i < len(deliveries) else 0
                item_discount = float(item_discounts[i]) if i < len(item_discounts) and item_discounts[i] else 0
                item_subtotal = quantity * price
                total = item_subtotal + delivery_fee
                subtotal += total

                wh = None
                try:
                    wh = int(warehouses[i]) if i < len(warehouses) and warehouses[i] else None
                except Exception:
                    wh = None

                sale_items.append({
                    'product_id': product.id,
                    'warehouse_id': wh,
                    'quantity': quantity,
                    'unit_price': price,
                    'delivery_fee': delivery_fee,
                    'discount': item_discount,
                    'total': total
                })

        tax_rate = float(request.form.get('tax_rate', 0))
        tax = subtotal * (tax_rate / 100)
        # Discount is always the sum of the per-item discounts entered above.
        discount = sum(item['discount'] for item in sale_items)
        delivery_charge = float(request.form.get('delivery_charge', 0))

        sale.customer_id = customer_id if customer_id != '0' else None
        sale.vendor_id = selected_vendor_id if selected_vendor_id and selected_vendor_id != '0' else None
        sale.salesman_id = salesman_id if salesman_id and salesman_id != '0' else None
        sale.date = date
        sale.subtotal = subtotal
        sale.tax_rate = tax_rate
        sale.tax = tax
        sale.discount = discount
        sale.discount_type = 'fixed'
        sale.delivery_charge = delivery_charge
        sale.advance_applied = float(request.form.get('advance_applied', sale.advance_applied or 0))
        sale.currency_id = request.form.get('currency_id') or None
        sale.exchange_rate = float(request.form.get('exchange_rate', 1))

        due_date_str = request.form.get('due_date')
        if due_date_str:
            try:
                sale.due_date = datetime.strptime(due_date_str, '%Y-%m-%d')
                sale.overdue_date = sale.due_date
            except ValueError:
                sale.due_date = None

        # Check discount conditions: each item's OWN discount is validated against
        # that item's product-specific min/max rule.
        discount_violations = validate_item_discounts(sale_items)

        if discount_violations:
            sale.is_approved = False
            sale.discount_violation = "; ".join(discount_violations)
            flash(f"Notice: This invoice requires admin approval due to discount conditions: {'; '.join(discount_violations)}", 'warning')
        else:
            if sale.discount_violation: # Violation fixed
                sale.discount_violation = None
                if current_user.role == 'admin':
                    sale.is_approved = True

        # Add items & update inventory (only deduct stock if sale is approved)
        for item in sale_items:
            sale_item = SaleItem(
                sale_id=sale.id,
                product_id=item['product_id'],
                warehouse_id=item.get('warehouse_id'),
                quantity=item['quantity'],
                unit_price=item['unit_price'],
                delivery_fee=item.get('delivery_fee', 0),
                discount=item.get('discount', 0),
                total=item['total']
            )
            db.session.add(sale_item)

        # Only deduct inventory if the sale is currently approved
        if sale.is_approved:
            for item in sale_items:
                product = Product.query.get(item['product_id'])
                if product:
                    product.update_quantity(-item['quantity'])

                    wh_id = item.get('warehouse_id')
                    if wh_id:
                        wh_stock = ProductWarehouseStock.query.filter_by(product_id=item['product_id'], warehouse_id=wh_id).first()
                        if not wh_stock:
                            wh_stock = ProductWarehouseStock(product_id=item['product_id'], warehouse_id=wh_id, quantity=0)
                            db.session.add(wh_stock)
                        wh_stock.quantity -= item['quantity']
                    elif product.warehouse_id:
                        wh_stock = ProductWarehouseStock.query.filter_by(product_id=item['product_id'], warehouse_id=product.warehouse_id).first()
                        if not wh_stock:
                            wh_stock = ProductWarehouseStock(product_id=item['product_id'], warehouse_id=product.warehouse_id, quantity=0)
                            db.session.add(wh_stock)
                        wh_stock.quantity -= item['quantity']
            sale.stock_updated = True

        sale.calculate_totals()
        # Preserve paid amount; update status accordingly
        sale.update_status()
        sale.updated_at = datetime.utcnow()

        db.session.commit()
        
        log_activity('Sales', f'Updated Invoice #{sale.invoice_number}', 
                    f'Customer: {sale.customer.name if sale.customer else "Walk-in Customer"}, New Total: {sale.total}')
        
        flash('Invoice updated successfully!', 'success')
        return redirect(url_for('sales.invoice_detail', id=sale.id))

    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    return render_template('sales/edit_invoice.html', form=form, sale=sale, products=products,
                           customers=customers, vendors=vendors, currencies=currencies,
                           salesmen=salesmen, now=datetime.now(), warehouses=warehouses)


@bp.route('/invoice/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('sales', action='delete')
def delete_invoice(id):
    sale = Sale.query.get_or_404(id)
    # Only restore inventory if stock was actually deducted (sale was approved at some point)
    if sale.stock_updated:
        for item in sale.items:
            product = Product.query.get(item.product_id)
            if product:
                # restore total quantity
                product.update_quantity(item.quantity)

                # restore per-warehouse quantity if sale item recorded a warehouse
                try:
                    wh_id = getattr(item, 'warehouse_id', None)
                except Exception:
                    wh_id = None
                if wh_id:
                    wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=wh_id).first()
                    if not wh_stock:
                        wh_stock = ProductWarehouseStock(product_id=item.product_id, warehouse_id=wh_id, quantity=0)
                        db.session.add(wh_stock)
                    wh_stock.quantity += item.quantity
                elif product.warehouse_id:
                    wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=product.warehouse_id).first()
                    if not wh_stock:
                        wh_stock = ProductWarehouseStock(product_id=item.product_id, warehouse_id=product.warehouse_id, quantity=0)
                        db.session.add(wh_stock)
                    wh_stock.quantity += item.quantity

    invoice_num = sale.invoice_number
    cust_name = sale.customer.name if sale.customer else "Walk-in Customer"
    total = sale.total
    
    db.session.delete(sale)
    db.session.commit()
    
    log_activity('Sales', f'Deleted Invoice #{invoice_num}', f'Customer: {cust_name}, Total: {total}')
    
    flash('Invoice deleted successfully.', 'success')
    return redirect(url_for('sales.invoices'))

@bp.route('/invoices/bulk-delete', methods=['POST'])
@login_required
@permission_required('sales', action='delete')
def bulk_delete_invoices():
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No invoices selected'}), 400
    
    deleted_count = 0
    errors = []
    
    for invoice_id in ids:
        sale = Sale.query.get(invoice_id)
        if not sale:
            continue
            
        try:
            # Only restore inventory if stock was actually deducted
            if sale.stock_updated:
                for item in sale.items:
                    product = Product.query.get(item.product_id)
                    if product:
                        product.update_quantity(item.quantity)
                        wh_id = getattr(item, 'warehouse_id', None)
                        if wh_id:
                            wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=wh_id).first()
                            if wh_stock:
                                wh_stock.quantity += item.quantity
                        elif product.warehouse_id:
                            wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=product.warehouse_id).first()
                            if wh_stock:
                                wh_stock.quantity += item.quantity
            
            db.session.delete(sale)
            deleted_count += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f'Error deleting Invoice {sale.invoice_number}: {str(e)}')
            
    if deleted_count > 0:
        db.session.commit()
        log_activity('Sales', f'Bulk Deleted Invoices', f'Deleted {deleted_count} invoices.')
        
    message = f'Successfully deleted {deleted_count} invoices.'
    if errors:
        return jsonify({'success': False, 'message': message, 'errors': errors}), 500
    return jsonify({'success': True, 'message': message})

@bp.route('/invoice/<int:id>/pay', methods=['POST'])
@login_required
@permission_required('sales', action='add')
def pay_invoice(id):
    sale = Sale.query.get_or_404(id)
    amount = float(request.form.get('amount', 0))
    
    if amount > 0:
        # Determine if this user's payment is auto-approved (admin) or needs approval (staff)
        payer_is_admin = (current_user.role == 'admin')

        # Only update paid_amount immediately for admin-approved payments
        # Staff payments remain pending until admin approves them
        if payer_is_admin:
            sale.paid_amount += amount
            sale.update_status()
        
        # Handle payment receipt image upload
        image_path = None
        if 'payment_image' in request.files:
            payment_file = request.files['payment_image']
            if payment_file and payment_file.filename:
                import time, uuid
                original_filename = secure_filename(payment_file.filename)
                # Generate a truly unique filename using timestamp and uuid
                unique_prefix = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
                filename = f"{unique_prefix}_{original_filename}"
                
                upload_dir = os.path.join('app', 'static', 'uploads', 'sale_payments')
                os.makedirs(upload_dir, exist_ok=True)
                file_path = os.path.join(upload_dir, filename)
                payment_file.save(file_path)
                # Store path relative to project root (required by payment_image route)
                image_path = f"app/static/uploads/sale_payments/{filename}"
        
        # Create Payment record for history tracking
        payment_date_str = request.form.get('payment_date', '')
        payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d') if payment_date_str else datetime.utcnow()
        payment_method = request.form.get('payment_method', 'Cash')
        payment_notes = request.form.get('payment_notes', '')
        
        # Generate payment number
        last_payment = Payment.query.order_by(Payment.id.desc()).first()
        payment_num = f"PAY-{last_payment.id + 1 if last_payment else 1}"
        
        payment = Payment(
            payment_number=payment_num,
            date=payment_date,
            amount=amount,
            method=payment_method,
            invoice_id=sale.id,
            notes=payment_notes,
            image_path=image_path,
            created_by=current_user.id,
            is_approved=payer_is_admin,
            approved_by=current_user.id if payer_is_admin else None,
            approved_at=datetime.utcnow() if payer_is_admin else None
        )
        db.session.add(payment)
        
        db.session.commit()
        log_activity('Sales', f'Recorded Payment on Invoice #{sale.invoice_number}',
                    f'Amount: PKR {amount:,.2f}, Method: {payment_method}, Approved: {payer_is_admin}')
        if payer_is_admin:
            flash(f'Payment of PKR {amount:,.2f} recorded and approved successfully!', 'success')
        else:
            flash(f'Payment of PKR {amount:,.2f} submitted — pending admin approval.', 'info')

    return redirect(url_for('sales.invoice_detail', id=sale.id))


# ── APPROVAL ROUTES (Admin only) ─────────────────────────────────────────────

@bp.route('/invoice/<int:id>/approve', methods=['POST'])
@login_required
def approve_invoice(id):
    """Admin approves a pending invoice — marks it approved so it counts in sales totals."""
    if current_user.role != 'admin':
        flash('Only admin can approve invoices.', 'danger')
        return redirect(url_for('sales.invoice_detail', id=id))

    sale = Sale.query.get_or_404(id)
    if sale.is_approved:
        flash('Invoice is already approved.', 'info')
        return redirect(url_for('sales.invoice_detail', id=id))

    sale.is_approved = True
    sale.approved_by = current_user.id
    sale.approved_at = datetime.utcnow()
    
    # Handle delayed advance application if present
    if sale.advance_applied > 0 and not sale.adjusted_advances:
        customer_id = sale.customer_id
        if customer_id:
            customer = Customer.query.get(customer_id)
            remaining_to_apply = sale.advance_applied
            
            # Get advances sorted by date (oldest first)
            advances = CustomerAdvance.query.filter_by(customer_id=customer_id, is_adjusted=False, is_approved=True).order_by(CustomerAdvance.date).all()
            
            for adv in advances:
                if remaining_to_apply <= 0:
                    break
                    
                available = adv.remaining_balance
                if available > 0:
                    apply_amt = min(available, remaining_to_apply)
                    adv.applied_amount += apply_amt
                    remaining_to_apply -= apply_amt
                    
                    if adv.remaining_balance <= 0:
                        adv.is_adjusted = True
                    adv.adjusted_invoice_id = sale.id
            
            # Update paid amount on sale now that advance is applied
            sale.paid_amount += (sale.advance_applied - remaining_to_apply)

    # Also auto-approve any pending payments on this invoice when the invoice itself is approved
    pending_payments = Payment.query.filter_by(invoice_id=sale.id, is_approved=False).all()
    for pmt in pending_payments:
        sale.paid_amount += pmt.amount
        pmt.is_approved = True
        pmt.approved_by = current_user.id
        pmt.approved_at = datetime.utcnow()
    sale.calculate_totals()
    sale.update_status()

    # Deduct inventory on approval if not already done (e.g. staff-created invoice)
    if not sale.stock_updated:
        for item in sale.items:
            product = Product.query.get(item.product_id)
            if product:
                product.update_quantity(-item.quantity)
                wh_id = getattr(item, 'warehouse_id', None)
                if wh_id:
                    wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=wh_id).first()
                    if not wh_stock:
                        wh_stock = ProductWarehouseStock(product_id=item.product_id, warehouse_id=wh_id, quantity=0)
                        db.session.add(wh_stock)
                    wh_stock.quantity -= item.quantity
                elif product.warehouse_id:
                    wh_stock = ProductWarehouseStock.query.filter_by(product_id=item.product_id, warehouse_id=product.warehouse_id).first()
                    if not wh_stock:
                        wh_stock = ProductWarehouseStock(product_id=item.product_id, warehouse_id=product.warehouse_id, quantity=0)
                        db.session.add(wh_stock)
                    wh_stock.quantity -= item.quantity
        sale.stock_updated = True

    db.session.commit()
    log_activity('Sales', f'Approved Invoice #{sale.invoice_number}',
                 f'Approved by {current_user.username}. Pending payments also approved: {len(pending_payments)}')
    flash(f'Invoice {sale.invoice_number} approved successfully!', 'success')
    return redirect(url_for('sales.invoice_detail', id=id))


@bp.route('/invoice/<int:id>/reject', methods=['POST'])
@login_required
def reject_invoice(id):
    """Admin rejects a pending invoice — mark it as rejected with a reason."""
    if current_user.role != 'admin':
        flash('Only admin can reject invoices.', 'danger')
        return redirect(url_for('sales.invoice_detail', id=id))

    sale = Sale.query.get_or_404(id)
    if sale.is_approved:
        flash('Cannot reject an already approved invoice.', 'warning')
        return redirect(url_for('sales.invoice_detail', id=id))

    reason = request.form.get('reason', '').strip()
    sale.is_rejected = True
    sale.is_approved = False
    sale.rejection_reason = reason
    db.session.commit()

    log_activity('Sales', f'Rejected Invoice #{sale.invoice_number}',
                 f'Rejected by {current_user.username}. Reason: {reason or "No reason provided"}')
    flash(f'Invoice {sale.invoice_number} has been rejected.', 'warning')
    return redirect(url_for('sales.invoice_detail', id=id))


@bp.route('/invoice/<int:id>/payment/<int:pay_id>/approve', methods=['POST'])
@login_required
def approve_payment(id, pay_id):
    """Admin approves a single pending payment — adds its amount to sale.paid_amount."""
    if current_user.role != 'admin':
        flash('Only admin can approve payments.', 'danger')
        return redirect(url_for('sales.invoice_detail', id=id))

    sale = Sale.query.get_or_404(id)
    payment = Payment.query.filter_by(id=pay_id, invoice_id=sale.id).first_or_404()

    if payment.is_approved:
        flash('Payment is already approved.', 'info')
        return redirect(url_for('sales.invoice_detail', id=id))

    payment.is_approved = True
    payment.approved_by = current_user.id
    payment.approved_at = datetime.utcnow()

    # Now add the payment amount to the sale's paid_amount
    sale.paid_amount += payment.amount
    sale.update_status()

    db.session.commit()
    log_activity('Sales', f'Approved Payment #{payment.payment_number} on Invoice #{sale.invoice_number}',
                 f'Amount: PKR {payment.amount:,.2f} approved by {current_user.username}')
    flash(f'Payment of PKR {payment.amount:,.2f} approved and applied to invoice!', 'success')
    return redirect(url_for('sales.invoice_detail', id=id))


@bp.route('/invoice/<int:id>/payment/<int:pay_id>/reject', methods=['POST'])
@login_required
def reject_payment(id, pay_id):
    """Admin rejects a pending payment."""
    if current_user.role != 'admin':
        flash('Only admin can reject payments.', 'danger')
        return redirect(url_for('sales.invoice_detail', id=id))

    sale = Sale.query.get_or_404(id)
    payment = Payment.query.filter_by(id=pay_id, invoice_id=sale.id).first_or_404()

    if payment.is_approved:
        flash('Cannot reject an already approved payment.', 'warning')
        return redirect(url_for('sales.invoice_detail', id=id))

    reason = request.form.get('reason', '').strip()
    payment.is_rejected = True
    payment.is_approved = False
    payment.rejection_reason = reason
    db.session.commit()

    log_activity('Sales', f'Rejected Payment on Invoice #{sale.invoice_number}',
                 f'Amount: PKR {payment.amount:,.2f} rejected by {current_user.username}. Reason: {reason}')
    flash(f'Payment of PKR {payment.amount:,.2f} rejected.', 'warning')
    return redirect(url_for('sales.invoice_detail', id=id))


@bp.route('/invoice/<int:id>/payment/<int:pay_id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('sales', action='edit')
def edit_payment(id, pay_id):
    """Edit existing payment for sales invoice"""
    sale = Sale.query.get_or_404(id)
    payment = Payment.query.filter_by(id=pay_id, invoice_id=sale.id).first_or_404()

    if request.method == 'POST':
        old_amount = payment.amount

        # Update payment details
        payment_date_str = request.form.get('payment_date')
        if payment_date_str:
            try:
                payment.date = datetime.strptime(payment_date_str, '%Y-%m-%d')
            except ValueError:
                payment.date = datetime.utcnow()
        else:
            payment.date = datetime.utcnow()

        payment.method = request.form.get('payment_method', 'Cash')
        # Only update reference_number if provided
        if 'reference_number' in request.form:
            payment.reference_number = request.form.get('reference_number', '')
        payment.notes = request.form.get('payment_notes', '')

        new_amount = float(request.form.get('amount', 0))

        # Handle image update
        if 'payment_image' in request.files:
            payment_file = request.files['payment_image']
            if payment_file and payment_file.filename:
                import time, uuid
                original_filename = secure_filename(payment_file.filename)
                unique_prefix = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
                filename = f"{unique_prefix}_{original_filename}"
                
                upload_dir = os.path.join('app', 'static', 'uploads', 'sale_payments')
                os.makedirs(upload_dir, exist_ok=True)
                file_path = os.path.join(upload_dir, filename)
                payment_file.save(file_path)
                # Store path relative to project root (required by payment_image route)
                payment.image_path = f"app/static/uploads/sale_payments/{filename}"

        # IMPORTANT: Update payment amount BEFORE calculating delta effect on sale
        payment.amount = new_amount

        # Safe paid_amount update: adjust sale's paid_amount by the difference
        delta = new_amount - old_amount
        from app.utils import safe_update_paid_amount
        safe_update_paid_amount(sale, delta)

        db.session.commit()
        log_activity('Sales', f'Updated Payment #{payment.payment_number} on Invoice #{sale.invoice_number}',
                    f'Old Amount: PKR {old_amount:,.2f}, New Amount: PKR {new_amount:,.2f}')
        flash(f'Payment updated: PKR{new_amount:,.2f} ({payment.method})', 'success')
        return redirect(url_for('sales.invoice_detail', id=sale.id))

    # GET: render edit form
    return render_template('sales/edit_payment.html', payment=payment, sale=sale)


@bp.route('/invoice/<int:id>/payment/<int:pay_id>/delete', methods=['POST'])
@login_required
@permission_required('sales', action='delete')
def delete_payment(id, pay_id):
    """Safely delete payment: reverse paid_amount, cleanup transactions"""
    sale = Sale.query.get_or_404(id)
    payment = Payment.query.filter_by(id=pay_id, invoice_id=sale.id).first_or_404()
    
    # Reverse from sale
    delta = -payment.amount
    from app.utils import safe_update_paid_amount
    safe_update_paid_amount(sale, delta)
    
    # Delete linked accounting transactions
    from app.utils import cleanup_linked_transactions
    cleanup_linked_transactions(payment)
    
    # Delete old image if exists
    if payment.image_path:
        try:
            image_path = os.path.join(current_app.root_path, 'static', payment.image_path)
            if os.path.exists(image_path):
                os.remove(image_path)
        except:
            pass
    
    pay_num = payment.payment_number
    pay_amount = payment.amount
    db.session.delete(payment)
    db.session.commit()
    log_activity('Sales', f'Deleted Payment #{pay_num} on Invoice #{sale.invoice_number}',
                f'Amount: PKR {pay_amount:,.2f}')
    flash(f'Payment deleted. Balance updated.', 'success')
    return redirect(url_for('sales.invoice_detail', id=sale.id))


_NO_IMAGE_PLACEHOLDER_SVG = (
    '<svg xmlns="http://www.w3.org/2000/svg" width="90" height="70" viewBox="0 0 90 70">'
    '<rect width="90" height="70" fill="#f1f3f5"/>'
    '<text x="45" y="39" font-family="sans-serif" font-size="10" fill="#adb5bd" '
    'text-anchor="middle">No Image</text></svg>'
)


@bp.route('/payment/<int:payment_id>/image')
@login_required
def payment_image(payment_id):
    """Serve payment receipt image"""
    from flask import current_app, Response
    payment = Payment.query.get_or_404(payment_id)
    if payment.image_path:
        # Stored path is relative to project root (e.g., 'app/static/uploads/sale_payments/xxx.png')
        if os.path.isabs(payment.image_path):
            file_path = payment.image_path
        else:
            project_root = os.path.dirname(current_app.root_path)
            file_path = os.path.join(project_root, payment.image_path)
        if os.path.exists(file_path):
            return send_file(file_path)
    # File missing on this server (e.g. uploaded on another environment and
    # never copied over) — this route is used directly as an <img> src, so a
    # flash+redirect to an HTML page breaks image rendering; serve a small
    # placeholder graphic instead.
    return Response(_NO_IMAGE_PLACEHOLDER_SVG, mimetype='image/svg+xml')

@bp.route('/invoice/<int:id>/discount', methods=['POST'])
@login_required
@permission_required('sales', action='edit')
def apply_discount(id):
    sale = Sale.query.get_or_404(id)

    # Block discount if invoice is overdue unless manually overridden
    ignore_overdue = request.form.get('ignore_overdue_discount') == 'on'
    
    # Synchronize the ignore_overdue_discount flag with the checkbox
    had_override = sale.ignore_overdue_discount
    sale.ignore_overdue_discount = ignore_overdue
    
    # Use 'or 0' to handle empty string inputs safely
    discount_amount_raw = request.form.get('discount_amount', '0')
    discount_amount = float(discount_amount_raw if discount_amount_raw.strip() else 0)

    # Optional: target specific item(s) on this invoice. Only ids that actually
    # belong to this sale are honored (never trust ids from the client blindly).
    requested_item_ids = {int(i) for i in request.form.getlist('item_ids[]') if i.strip().isdigit()}
    target_items = [item for item in sale.items if item.id in requested_item_ids] if requested_item_ids else []

    # Track if any significant change happened
    changes_made = False

    if discount_amount > 0:
        # Check if overdue rule applies to this customer's group
        rule_applies = False
        settings = InvoiceSettings.query.first()
        if settings:
            restricted_groups = settings.restricted_group_ids
            if sale.customer and sale.customer.group_id in restricted_groups:
                rule_applies = True

        # Check if overdue and NOT overridden BEFORE applying more discount
        if sale.is_overdue and not sale.ignore_overdue_discount and rule_applies:
            flash('Cannot add discount to an overdue invoice without "Override Restriction" selected.', 'warning')
            return redirect(request.referrer or url_for('sales.invoice_detail', id=sale.id))

        sale.discount += discount_amount

        if target_items:
            # Prorate the discount evenly across the selected items (bulk-safe:
            # any rounding remainder goes to the last item so shares sum exactly
            # to discount_amount), and reflect it on each item's Discount column.
            share = round(discount_amount / len(target_items), 2)
            allocated = 0
            for idx, item in enumerate(target_items):
                portion = share if idx < len(target_items) - 1 else round(discount_amount - allocated, 2)
                item.discount = (item.discount or 0) + portion
                allocated += portion
            item_names = ', '.join(item.product.name if item.product else f'Item #{item.id}' for item in target_items)
            flash(f'Discount of PKR {discount_amount:,.2f} applied successfully and prorated across: {item_names}.', 'success')
            log_activity('Sales', f'Applied Discount to Invoice #{sale.invoice_number}',
                        f'Amount: {discount_amount}, Items: {item_names}, Customer: {sale.customer.name if sale.customer else "Walk-in"}')
        else:
            flash(f'Discount of PKR {discount_amount:,.2f} applied successfully!', 'success')
            log_activity('Sales', f'Applied Discount to Invoice #{sale.invoice_number}',
                        f'Amount: {discount_amount}, Customer: {sale.customer.name if sale.customer else "Walk-in"}')
        changes_made = True
    
    if ignore_overdue != had_override:
        if ignore_overdue:
            flash('Overdue restriction overridden successfully.', 'info')
            log_activity('Sales', f'Overrode Overdue Restriction for Invoice #{sale.invoice_number}', 
                        f'Invoice was overdue but discount was allowed/updated by {current_user.username}')
        else:
            flash('Overdue restriction re-applied. Any existing discount is now suspended.', 'warning')
            log_activity('Sales', f'Re-applied Overdue Restriction for Invoice #{sale.invoice_number}', 
                        f'Restriction was restored by {current_user.username}. Effective discount is now 0.')
        changes_made = True

    if changes_made:
        sale.calculate_totals()
        sale.update_status()
        db.session.commit()
    else:
        # Only flash if something was actually attempted but nothing changed
        # (Though with the logic above, changes_made should be true if either occurred)
        flash('No changes were made.', 'secondary')

    return redirect(request.referrer or url_for('sales.invoice_detail', id=sale.id))


@bp.route('/invoice/<int:id>/apply-advance', methods=['POST'])
@login_required
@permission_required('sales', action='edit')
def apply_advance_to_overdue(id):
    """Apply available customer advance balance to an overdue invoice."""
    sale = Sale.query.get_or_404(id)

    if not sale.is_overdue:
        flash('This invoice is not overdue. Use the standard discount option.', 'warning')
        return redirect(url_for('sales.invoice_detail', id=sale.id))

    if not sale.customer_id:
        flash('No customer linked to this invoice.', 'error')
        return redirect(url_for('sales.invoice_detail', id=sale.id))

    customer = sale.customer
    available_advance = customer.remaining_advance_balance

    if available_advance <= 0:
        flash('No advance balance available for this customer.', 'warning')
        return redirect(url_for('sales.invoice_detail', id=sale.id))

    balance_due = sale.total - sale.paid_amount
    if balance_due <= 0:
        flash('Invoice is already fully paid.', 'info')
        return redirect(url_for('sales.invoice_detail', id=sale.id))

    # Apply whichever is smaller: available advance or balance due
    apply_amount = min(available_advance, balance_due)

    # Record advance application
    sale.advance_applied = (sale.advance_applied or 0) + apply_amount
    sale.paid_amount = (sale.paid_amount or 0) + apply_amount
    sale.update_status()

    # Deduct from customer advance records (oldest first)
    remaining_to_apply = apply_amount
    for adv in sorted(customer.advances, key=lambda a: a.date):
        if remaining_to_apply <= 0:
            break
        available_in_adv = adv.amount - (adv.applied_amount or 0)
        if available_in_adv <= 0:
            continue
        use = min(available_in_adv, remaining_to_apply)
        adv.applied_amount = (adv.applied_amount or 0) + use
        remaining_to_apply -= use

    db.session.commit()
    log_activity('Sales', f'Applied Advance to Overdue Invoice #{sale.invoice_number}',
                f'Amount: PKR {apply_amount:,.2f}, Customer: {customer.name}')
    flash(f'Advance of PKR {apply_amount:,.2f} applied to overdue invoice successfully!', 'success')
    return redirect(url_for('sales.invoice_detail', id=sale.id))


@bp.route('/customers')
@login_required
def customers():
    status = request.args.get('status', 'all')
    search = request.args.get('search', '')
    customer_id = request.args.get('customer_id', type=int)
    customer_group_id = request.args.get('customer_group_id', type=int)
    selected_city   = request.args.get('city', '')
    selected_state  = request.args.get('state', '')
    selected_country = request.args.get('country', '')
    selected_postal = request.args.get('postal_code', '')
    
    query = Customer.query
    
    if status == 'active':
        query = query.filter_by(is_active=True)
    elif status == 'inactive':
        query = query.filter_by(is_active=False)
    
    if customer_id:
        query = query.filter_by(id=customer_id)
    
    if customer_group_id:
        query = query.filter_by(group_id=customer_group_id)

    if selected_city:
        query = query.filter(Customer.city == selected_city)
    
    if selected_state:
        query = query.filter(Customer.state == selected_state)

    if selected_country:
        query = query.filter(Customer.country == selected_country)

    if selected_postal:
        query = query.filter(Customer.postal_code == selected_postal)

    if not customer_id and search:
        search_filter = f"%{search}%"
        query = query.filter(
            (Customer.name.ilike(search_filter)) |
            (Customer.email.ilike(search_filter)) |
            (Customer.phone.ilike(search_filter)) |
            (Customer.city.ilike(search_filter)) |
            (Customer.state.ilike(search_filter)) |
            (Customer.country.ilike(search_filter)) |
            (Customer.sub_customers.ilike(search_filter))
        )
    
    query = apply_saved_filter_to_query(query, 'customer', request.args)
    
    customers = query.order_by(Customer.name.asc()).all()
            
    # List of all customers for the searchable dropdown
    all_customers = Customer.query.order_by(Customer.name.asc()).all()
    customer_groups = CustomerGroup.query.filter_by(is_active=True).order_by(CustomerGroup.name).all()

    # Distinct city / state / country / postal lists for filter dropdowns (skip nulls/empty)
    all_cities = [r[0] for r in db.session.query(Customer.city).filter(
        Customer.city.isnot(None), Customer.city != ''
    ).distinct().order_by(Customer.city).all()]

    all_states = [r[0] for r in db.session.query(Customer.state).filter(
        Customer.state.isnot(None), Customer.state != ''
    ).distinct().order_by(Customer.state).all()]

    all_countries = [r[0] for r in db.session.query(Customer.country).filter(
        Customer.country.isnot(None), Customer.country != ''
    ).distinct().order_by(Customer.country).all()]

    all_postals = [r[0] for r in db.session.query(Customer.postal_code).filter(
        Customer.postal_code.isnot(None), Customer.postal_code != ''
    ).distinct().order_by(Customer.postal_code).all()]

    return render_template('sales/customers.html', 
                         customers=customers, 
                         all_customers=all_customers,
                         customer_groups=customer_groups,
                         current_status=status, 
                         search_query=search,
                         selected_customer_id=customer_id,
                         selected_group_id=customer_group_id,
                         all_cities=all_cities,
                         all_states=all_states,
                         all_countries=all_countries,
                         all_postals=all_postals,
                         selected_city=selected_city,
                         selected_state=selected_state,
                         selected_country=selected_country,
                         selected_postal=selected_postal,
                         active_module='customer',
                         filter_id=request.args.get('filter_id'))

@bp.route('/customer/bulk-upload', methods=['GET', 'POST'])
@login_required
@permission_required('customers', action='add')
def bulk_upload_customer():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('sales.bulk_upload_customer'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('sales.bulk_upload_customer'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('sales.bulk_upload_customer'))
        
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            file_content = file.read()
            wb = load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            rows = list(ws.values)
            if not rows:
                flash('File is empty', 'error')
                return redirect(url_for('sales.bulk_upload_customer'))
            headers = [str(h) if h else '' for h in rows[0]]
            df_data = []
            for row in rows[1:]:
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = val
                df_data.append(row_dict)
            import pandas as pd
            df = pd.DataFrame(df_data)
            
            required_columns = ['name']
            missing = [col for col in required_columns if col not in df.columns]
            if missing:
                flash(f'Missing required columns: {", ".join(missing)}', 'error')
                return redirect(url_for('sales.bulk_upload_customer'))
            
            added = 0
            errors = []
            
            for idx, row in df.iterrows():
                try:
                    name = str(row.get('name', '')).strip()
                    
                    if not name:
                        errors.append(f'Row {idx + 2}: Missing name')
                        continue
                    
                    customer = Customer(
                        name=name,
                        email=str(row.get('email', '')).strip() if pd.notna(row.get('email')) else None,
                        phone=str(row.get('phone', '')).strip() if pd.notna(row.get('phone')) else None,
                        address=str(row.get('address', '')).strip() if pd.notna(row.get('address')) else None,
                        gst_number=str(row.get('gst_number', '')).strip() if pd.notna(row.get('gst_number')) else None,
                        pan_number=str(row.get('pan_number', '')).strip() if pd.notna(row.get('pan_number')) else None,
                        contact_person=str(row.get('contact_person', '')).strip() if pd.notna(row.get('contact_person')) else None,
                    )
                    
                    db.session.add(customer)
                    added += 1
                except Exception as e:
                    errors.append(f'Row {idx + 2}: {str(e)}')
            
            db.session.commit()
            if added > 0:
                log_activity('Customers', f'Bulk Uploaded Customers', f'Added {added} customers via Excel upload')

            if added > 0:
                flash(f'Successfully added {added} customers!', 'success')
            if errors:
                flash(f'Errors: {"; ".join(errors[:10])}', 'warning')

            return redirect(url_for('sales.customers'))
            
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'error')
            return redirect(url_for('sales.bulk_upload_customer'))
    
    return render_template('sales/bulk_upload_customer.html')

@bp.route('/customer/download-sample')
@login_required
def download_customer_sample():
    try:
        from openpyxl import Workbook
        from io import BytesIO
        from flask import send_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Customers'
        
        headers = ['name', 'email', 'phone', 'address', 'gst_number', 'pan_number', 'contact_person']
        ws.append(headers)
        
        sample_data = [
            ['Customer A', 'customerA@example.com', '1234567890', '123 Main St, City', 'GST123456789', 'ABCDE1234F', 'John Doe'],
            ['Customer B', 'customerB@example.com', '2345678901', '456 Oak Ave, Town', 'GST987654321', 'FGHI5678K', 'Jane Smith'],
            ['Customer C', 'customerC@example.com', '3456789012', '789 Pine Rd, Village', '', '', '']
        ]
        
        for row in sample_data:
            ws.append(row)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='sample_customers.xlsx', as_attachment=True)
        
    except Exception as e:
        flash(f'Error creating sample: {str(e)}', 'error')
        return redirect(url_for('sales.customers'))


@bp.route('/customer/<int:id>')
@login_required
def customer_profile(id):
    from datetime import datetime as _dt
    customer = Customer.query.get_or_404(id)
    
    # ── Date filter ────────────────────────────────────────────────────────
    date_from_str = request.args.get('date_from', '').strip()
    date_to_str   = request.args.get('date_to',   '').strip()
    date_from = None
    date_to   = None
    try:
        if date_from_str:
            date_from = _dt.strptime(date_from_str, '%Y-%m-%d')
        if date_to_str:
            date_to = _dt.strptime(date_to_str, '%Y-%m-%d').replace(
                hour=23, minute=59, second=59)
    except ValueError:
        pass

    # Filter sales: only approved sales contribute to financial balance in profile
    all_sales = sorted([s for s in customer.sales if s.is_approved], key=lambda s: s.date, reverse=True)
    unapproved_sales = sorted([s for s in customer.sales if not s.is_approved], key=lambda s: s.date, reverse=True)

    if date_from or date_to:
        filtered_sales = [
            s for s in all_sales
            if (date_from is None or s.date >= date_from)
            and (date_to   is None or s.date <= date_to)
        ]
    else:
        filtered_sales = all_sales

    # Filter advances: only approved advances are listed for application
    all_advances = sorted([a for a in customer.advances if a.is_approved], key=lambda a: a.date, reverse=True)
    unapproved_advances = sorted([a for a in customer.advances if not a.is_approved], key=lambda a: a.date, reverse=True)
    
    return render_template('sales/customer_profile.html', 
                          customer=customer, 
                          sales=filtered_sales, 
                          unapproved_sales=unapproved_sales,
                          advances=all_advances,
                          unapproved_advances=unapproved_advances,
                          date_from=date_from_str,
                          date_to=date_to_str,
                          now=datetime.now())

@bp.route('/customer/<int:id>/ledger-pdf')
def customer_export_pdf(id, is_public=False):
    """Export customer detailed ledger to PDF"""
    from datetime import datetime as _dt, time as _time
    from app.models import Company, InvoiceSettings, Sale, CustomerAdvance, SaleReturn, Payment, Customer
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image
    )
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from flask_login import current_user

    if not is_public and not current_user.is_authenticated:
        return current_app.login_manager.unauthorized()

    customer = Customer.query.get_or_404(id)
    company  = Company.query.first()
    settings = InvoiceSettings.query.first()
    
    selected_ids = request.args.get('ids', '').split(',') if request.args.get('ids') else []

    # Opening balance
    opening_balance = float(customer.opening_balance or 0)
    
    # Only include approved entities in the official ledger PDF
    sales_q = Sale.query.filter_by(customer_id=id, is_approved=True).order_by(Sale.date.asc()).all()
    advances = CustomerAdvance.query.filter_by(customer_id=id, is_approved=True).order_by(CustomerAdvance.date.asc()).all()
    returns = SaleReturn.query.filter_by(customer_id=id).order_by(SaleReturn.date.asc()).all()
    
    sale_ids = [s.id for s in sales_q]
    # Only include approved payments
    payments = Payment.query.filter(Payment.invoice_id.in_(sale_ids), Payment.is_approved == True).order_by(Payment.date.asc()).all() if sale_ids else []
    
    # Build advances as a sorted list to interleave by date
    advance_events = []
    for a in advances:
        if selected_ids: continue  # Skip advances if filtering to specific invoices
        dt = _dt.combine(a.date, _time.min)
        advance_events.append((dt, {
            'date': dt,
            'type': 'advance',
            'desc': f"Advance Received: {a.description or ''} (PKR {float(a.amount or 0):,.2f})",
            'inv': '-',
            'debit': 0,
            'credit': 0
        }))
    advance_events.sort(key=lambda x: x[0])
    adv_idx = 0

    # Build events grouped per-invoice:
    # Invoice → its payments → its discount → its advance_applied → its returns
    # Standalone advances are interleaved before each invoice by date.
    events = []
    for s in sorted(sales_q, key=lambda x: x.date):
        if selected_ids and s.invoice_number not in selected_ids:
            continue

        # Insert any advances that occurred before or on this invoice's date
        while adv_idx < len(advance_events) and advance_events[adv_idx][0].date() <= s.date.date():
            events.append(advance_events[adv_idx][1])
            adv_idx += 1

        # 1. The invoice itself
        events.append({
            'date': s.date,
            'type': 'sale',
            'desc': f"Invoice #{s.invoice_number}",
            'inv': s.invoice_number,
            'debit': float(sum(item.total for item in s.items)) * (1 + float(s.tax_rate or 0) / 100) + float(s.delivery_charge or 0),
            'credit': 0,
            'sale_obj': s
        })

        # 2. Payments for this invoice (sorted by date)
        for p in sorted(s.payments, key=lambda x: x.date):
            if not getattr(p, 'is_approved', True):
                continue
            events.append({
                'date': p.date,
                'type': 'payment',
                'desc': f"Payment ({p.method or 'Cash'})",
                'inv': s.invoice_number,
                'debit': 0,
                'credit': float(p.amount or 0)
            })

        # 3. Discount for this invoice
        is_restricted = getattr(s, 'is_discount_restricted', False)
        effective_discount = getattr(s, 'effective_discount_amount', 0)
        if is_restricted and getattr(s, 'base_discount_amount', 0) > 0:
            events.append({
                'date': s.date, 'type': 'discount',
                'desc': f"Discount (Inv #{s.invoice_number}): Reversal due to overdue",
                'inv': s.invoice_number, 'debit': 0, 'credit': 0
            })
        elif effective_discount > 0:
            events.append({
                'date': s.date, 'type': 'discount',
                'desc': f"Discount (Inv #{s.invoice_number})",
                'inv': s.invoice_number, 'debit': 0,
                'credit': float(effective_discount or 0)
            })

        # 4. Advance Applied for this invoice
        if getattr(s, 'advance_applied', 0) > 0:
            events.append({
                'date': s.date, 'type': 'advance_applied',
                'desc': f"Advance Applied (Inv #{s.invoice_number})",
                'inv': s.invoice_number, 'debit': 0,
                'credit': float(s.advance_applied or 0)
            })

        # 5. Returns for this invoice (sorted by date)
        for r in sorted(s.returns, key=lambda x: x.date):
            if selected_ids and r.return_number not in selected_ids:
                continue
            events.append({
                'date': r.date, 'type': 'return',
                'desc': f"Sales Return #{r.return_number}",
                'inv': r.return_number, 'debit': 0,
                'credit': float(r.total or 0)
            })

    # Append any remaining advances after all invoices
    while adv_idx < len(advance_events):
        events.append(advance_events[adv_idx][1])
        adv_idx += 1
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=28, bottomMargin=72
    )

    PRIMARY_COLOR   = colors.HexColor("#c0392b")
    ACCENT_COLOR    = colors.HexColor("#f9f9f9")
    TEXT_COLOR      = colors.HexColor("#1a1a1a")
    MUTED_TEXT      = colors.HexColor("#6b6b6b")
    WHITE           = colors.white
    BLACK           = colors.black
    LIGHT_GREY      = colors.HexColor("#e0e0e0")
    BORDER_GREY     = colors.HexColor("#d0d0d0")
    HEADER_STRIPE   = colors.HexColor("#1a1a1a")
    PAGE_BG         = colors.HexColor("#f2f2f2")

    def _draw_footer(c, doc_obj):
        w, h = A4
        card_margin = 18
        lx = card_margin + 16
        rx = w - card_margin - 16
        divider_y = card_margin + 72 - 2
        y = divider_y - 11

        c.saveState()
        c.setStrokeColor(BORDER_GREY)
        c.setLineWidth(0.4)
        c.line(lx, divider_y, rx, divider_y)

        c.setFont('Helvetica-Bold', 7)
        c.setFillColor(TEXT_COLOR)
        c.drawString(lx, y, "PREPARED BY")

        c.setFont('Helvetica', 6.5)
        c.setFillColor(MUTED_TEXT)
        y -= 10
        c.drawString(lx, y, getattr(company, 'name', '') if company else "")

        sig_path = None
        if company:
            try:
                sig_path = company.signature_path
            except Exception:
                pass

        if sig_path and os.path.exists(sig_path):
            try:
                from reportlab.platypus import Image as RLImage
                sig_img = RLImage(sig_path)
                sig_img.drawWidth = 100
                sig_img.drawHeight = 100
                sig_img.drawOn(c, lx, y - 70)
            except Exception:
                pass
        
        if not sig_path or not os.path.exists(sig_path):
            y -= 14
            c.setStrokeColor(colors.HexColor("#555555"))
            c.setLineWidth(0.5)
            c.line(lx, y, lx + 110, y)
            y -= 8
            c.setFont('Helvetica', 6.5)
            c.setFillColor(MUTED_TEXT)
            c.drawString(lx, y, "Authorized Signature")

        msg_y = divider_y - 11 - 20
        c.setFont('Helvetica-Bold', 7.5)
        c.setFillColor(TEXT_COLOR)
        c.drawCentredString(w / 2, msg_y, "Thank you for your business!")

        cemail = getattr(company, 'email', '') if company else ""
        cphone = getattr(company, 'phone', '') if company else ""
        cwa    = getattr(company, 'whatsapp', '') if company else ""

        ry = divider_y - 11
        c.setFont('Helvetica-Bold', 7)
        c.drawRightString(rx, ry, "CUSTOMER SUPPORT")

        c.setFont('Helvetica', 6.5)
        c.setFillColor(MUTED_TEXT)
        for line in filter(None, [cemail, cphone, f"WhatsApp: {cwa}" if cwa else ""]):
            ry -= 9
            c.drawRightString(rx, ry, line)

        c.setFont('Helvetica', 7)
        c.setFillColor(MUTED_TEXT)
        c.drawCentredString(w / 2, card_margin + 6, f"Page {doc_obj.page}")

        c.restoreState()

    def _draw_page_decorations(c, doc_obj):
        w, h = A4
        c.saveState()
        c.setFillColor(PAGE_BG)
        c.rect(0, 0, w, h, fill=1, stroke=0)
        m = 18
        c.setFillColor(WHITE)
        c.roundRect(m, m, w - 2*m, h - 2*m, 6, fill=1, stroke=0)
        bh = 8
        c.setFillColor(HEADER_STRIPE)
        c.rect(m, h - m - bh, (w - 2*m) * 0.75, bh, fill=1, stroke=0)
        c.setFillColor(PRIMARY_COLOR)
        c.rect(m + (w - 2*m) * 0.75, h - m - bh, (w - 2*m) * 0.25, bh, fill=1, stroke=0)
        c.restoreState()
        
        _draw_footer(c, doc_obj)

    styles = getSampleStyleSheet()
    s_CompanyName = ParagraphStyle('CompanyName', parent=styles['Normal'], fontSize=11, textColor=TEXT_COLOR, fontName='Helvetica-Bold', spaceAfter=1)
    s_CompanyInfo = ParagraphStyle('CompanyInfo', parent=styles['Normal'], fontSize=7,  textColor=MUTED_TEXT, leading=10)
    s_Title       = ParagraphStyle('Title', parent=styles['Normal'], fontSize=20, textColor=BLACK, fontName='Helvetica-Bold', alignment=TA_RIGHT, spaceAfter=1)
    s_MetaLabel   = ParagraphStyle('MetaLabel', parent=styles['Normal'], fontSize=7.5, textColor=MUTED_TEXT, alignment=TA_RIGHT)
    s_MetaValue   = ParagraphStyle('MetaValue', parent=styles['Normal'], fontSize=8, textColor=TEXT_COLOR, fontName='Helvetica-Bold', alignment=TA_RIGHT)
    s_BoxTitle    = ParagraphStyle('BoxTitle', parent=styles['Normal'], fontSize=8, textColor=PRIMARY_COLOR, fontName='Helvetica-Bold', spaceAfter=3, leftIndent=0)
    s_BoxValue    = ParagraphStyle('BoxValue', parent=styles['Normal'], fontSize=7.5, textColor=TEXT_COLOR, leading=10, leftIndent=0)
    s_TblHeader   = ParagraphStyle('TblHeader', parent=styles['Normal'], fontSize=7.5, textColor=WHITE, fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_TblCell     = ParagraphStyle('TblCell', parent=styles['Normal'], fontSize=7.5, textColor=TEXT_COLOR, alignment=TA_RIGHT)
    s_TblCellC    = ParagraphStyle('TblCellC', parent=styles['Normal'], fontSize=7.5, textColor=TEXT_COLOR, alignment=TA_CENTER)
    s_TblCellL    = ParagraphStyle('TblCellL', parent=styles['Normal'], fontSize=7.5, textColor=TEXT_COLOR, alignment=TA_LEFT)
    s_TotalLabel  = ParagraphStyle('TotalLabel', parent=styles['Normal'], fontSize=8, textColor=TEXT_COLOR, alignment=TA_LEFT)
    s_TotalValue  = ParagraphStyle('TotalValue', parent=styles['Normal'], fontSize=8, textColor=TEXT_COLOR, alignment=TA_RIGHT)
    s_GrandLabel  = ParagraphStyle('GrandLabel', parent=styles['Normal'], fontSize=9, textColor=TEXT_COLOR, fontName='Helvetica-Bold', alignment=TA_LEFT)
    s_GrandValue  = ParagraphStyle('GrandValue', parent=styles['Normal'], fontSize=9, textColor=TEXT_COLOR, fontName='Helvetica-Bold', alignment=TA_RIGHT)

    elements = []

    # GET LOGO 
    logo = None
    if company and getattr(company, 'logo_path', None) and os.path.exists(company.logo_path):
        try:
            from reportlab.platypus import Image as RLImage
            img = RLImage(company.logo_path)
            aspect = img.imageHeight / float(img.imageWidth)
            w = 1.0 * inch
            img.drawWidth  = w
            img.drawHeight = w * aspect
            logo = img
        except Exception:
            pass

    # HEADER LEFT
    left = []
    if logo:
        left.append(logo)
        left.append(Spacer(1, 6))
    
    cname = company.name if company else "COMPANY"
    left.append(Paragraph(cname, s_CompanyName))
    if company:
        for attr, label in [('email','<b>Email:</b> {}'), ('phone','<b>Phone:</b> {}'), ('whatsapp','<b>WhatsApp:</b> {}'), ('address','{}'), ('address2','{}')]:
            val = getattr(company, attr, None)
            if val:
                left.append(Paragraph(label.format(val), s_CompanyInfo))

    # HEADER RIGHT
    RIGHT_W      = 3.2 * inch
    META_LABEL_W = 1.1 * inch
    META_VAL_W   = RIGHT_W - META_LABEL_W

    meta_rows = [
        ["Report Date:", _dt.now().strftime('%d/%m/%Y')],
        ["Customer ID:", str(customer.id)],
    ]
    meta_tbl = Table(
        [[Paragraph(l, s_MetaLabel), Paragraph(v, s_MetaValue)] for l, v in meta_rows],
        colWidths=[META_LABEL_W, META_VAL_W]
    )
    meta_tbl.setStyle(TableStyle([
        ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 1),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
    ]))

    right_rows = [
        [Paragraph("CUSTOMER LEDGER", s_Title)],
        [Spacer(1, 22)],
        [meta_tbl],
    ]
    right_tbl = Table(right_rows, colWidths=[RIGHT_W])
    right_tbl.setStyle(TableStyle([
        ('ALIGN',  (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING',    (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
    ]))

    LEFT_W = (523 / 72) * inch - RIGHT_W - 0.1 * inch
    header_tbl = Table([[left, right_tbl]], colWidths=[LEFT_W, RIGHT_W])
    header_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN',  (0,0), (0,0),   'LEFT'),
        ('ALIGN',  (1,0), (1,0),   'RIGHT'),
        ('TOPPADDING',    (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
    ]))
    elements.append(header_tbl)
    elements.append(Spacer(1, 8))
    elements.append(HRFlowable(width="100%", thickness=0.6, color=BORDER_GREY, spaceAfter=8))

    # INFO BOXES (CUSTOMER DETAILS & OUR LOCATIONS)
    BOX_W = 3.5 * inch
    GAP_W = 0.2 * inch
    
    # Box 1: Customer Details
    hs = customer.health_status
    if hs == 'danger':
        flag_label = '<font color="#e74a3b">● 30+ Days Overdue</font>'
    elif hs == 'warning':
        flag_label = '<font color="#f6c23e">● Overdue</font>'
    else:
        flag_label = '<font color="#1cc88a">● On Time</font>'

    b1_rows = []
    b1_rows.append(f"<b>Name:</b> {customer.company_name or customer.name}  {flag_label}")
    if customer.email:   b1_rows.append(f"<b>Email:</b> {customer.email}")
    if customer.phone:   b1_rows.append(f"<b>Phone:</b> {customer.phone}")
    if customer.address: b1_rows.append(f"<b>Address:</b> {customer.address}")
    if customer.gst_number: b1_rows.append(f"<b>GST No:</b> {customer.gst_number}")
    if customer.group:   b1_rows.append(f"<b>Group:</b> {customer.group.name}")

    data = [[Paragraph("CUSTOMER DETAILS", s_BoxTitle)]]
    for row in b1_rows:
        data.append([Paragraph(row, s_BoxValue)])
    box1 = Table(data, colWidths=[BOX_W])
    
    box_style = TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('ALIGN',         (0,0), (-1,-1), 'LEFT'),
        ('TOPPADDING',    (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING',   (0,0), (-1,-1), 8),
        ('RIGHTPADDING',  (0,0), (-1,-1), 6),
        ('BACKGROUND',    (0,0), (-1,-1), WHITE),
        ('BOX',           (0,0), (-1,-1), 0.5, BORDER_GREY),
    ])
    box1.setStyle(box_style)

    # Box 4: Our Locations (matching sales invoice layout)
    addr_rows = [
        "<b>Address 1:</b> No. 139 Tiyu West Road, Unit 15-A, Tianhe District, Guangzhou, Guangdong, 510620",
        "<b>Address 2:</b> LS-25, Super Auto Parts Market, Main Super Highway, Sohrab Goth, Karachi."
    ]
    data4 = [[Paragraph("OUR LOCATIONS", s_BoxTitle)]]
    for row in addr_rows:
        data4.append([Paragraph(row, s_BoxValue)])
    box4 = Table(data4, colWidths=[BOX_W])
    box4.setStyle(box_style)

    # Combine into a single row
    outer = Table([[box1, Spacer(GAP_W, 1), box4]], colWidths=[BOX_W, GAP_W, BOX_W])
    outer.setStyle(TableStyle([
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
        ('TOPPADDING',    (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(outer)
    elements.append(Spacer(1, 15))

    # LEDGER TABLE
    headers = ['Date', 'Description', 'Invoice #', 'Status', 'Debit', 'Credit', 'Balance']
    header_row = [Paragraph(f"<b>{h}</b>", s_TblHeader) for h in headers]
    table_data = [header_row]

    TABLE_WIDTH = 7.5 * inch
    # Weights: 1.8, 3.2, 1.3, 1.8, 1.9, 1.9, 2.1 (Total = 14)
    Cols = [(w/14.0)*TABLE_WIDTH for w in [1.8, 3.2, 1.3, 1.8, 1.9, 1.9, 2.1]]

    FLAG_GREEN  = colors.HexColor("#1cc88a")
    FLAG_YELLOW = colors.HexColor("#f6c23e")
    FLAG_RED    = colors.HexColor("#e74a3b")
    s_StatusGreen  = ParagraphStyle('StatusGreen',  parent=styles['Normal'], fontSize=7, textColor=FLAG_GREEN,  fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_StatusYellow = ParagraphStyle('StatusYellow', parent=styles['Normal'], fontSize=7, textColor=FLAG_YELLOW, fontName='Helvetica-Bold', alignment=TA_CENTER)
    s_StatusRed    = ParagraphStyle('StatusRed',    parent=styles['Normal'], fontSize=7, textColor=FLAG_RED,    fontName='Helvetica-Bold', alignment=TA_CENTER)

    def pkr(val):
        return f"PKR {val:,.2f}"

    # Opening Balance
    running_balance = opening_balance
    table_data.append([
        Paragraph("-", s_TblCellC),
        Paragraph("Opening Balance", s_TblCellL),
        Paragraph("-", s_TblCellC),
        Paragraph("-", s_TblCellC),
        Paragraph(pkr(opening_balance) if opening_balance > 0 else "0.00", s_TblCell),
        Paragraph(pkr(abs(opening_balance)) if opening_balance < 0 else "0.00", s_TblCell),
        Paragraph(pkr(running_balance), s_TblCell)
    ])

    total_sales = 0
    total_discounts = 0
    total_paid = 0
    total_advances_applied = 0

    for e in events:
        debit = e['debit']
        credit = e['credit']
        running_balance += (debit - credit)
        # If balance goes below zero, clamp to 0. The excess credit is treated as
        # a pending advance and must NOT carry over to subtract from future invoices.
        if running_balance < 0:
            running_balance = 0
        
        if e['type'] == 'sale': total_sales += debit
        elif e['type'] == 'discount': total_discounts += credit
        elif e['type'] == 'payment': total_paid += credit
        elif e['type'] == 'advance_applied': total_advances_applied += credit

        status_cell = Paragraph("-", s_TblCellC)
        sale_ref = e.get('sale_obj')
        if e['type'] == 'sale' and sale_ref:
            hs = sale_ref.invoice_health_status
            if hs == 'danger':
                status_cell = Paragraph(f"● Overdue ({sale_ref.days_overdue}d)", s_StatusRed)
            elif hs == 'warning':
                status_cell = Paragraph(f"● Overdue ({sale_ref.days_overdue}d)", s_StatusYellow)
            else:
                if sale_ref.status == 'paid':
                    status_cell = Paragraph("● Paid", s_StatusGreen)
                else:
                    status_cell = Paragraph("● On Time", s_StatusGreen)

        table_data.append([
            Paragraph(e['date'].strftime("%d/%m/%Y"), s_TblCellC),
            Paragraph(e['desc'], s_TblCellL),
            Paragraph(e['inv'], s_TblCellC),
            status_cell,
            Paragraph(pkr(debit), s_TblCell),
            Paragraph(pkr(credit), s_TblCell),
            Paragraph(pkr(max(0, running_balance)), s_TblCell)
        ])

    tbl = Table(table_data, colWidths=Cols, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),  HEADER_STRIPE),
        ('TEXTCOLOR',     (0,0),  (-1,0),  WHITE),
        ('TOPPADDING',    (0,0),  (-1,0),  5),
        ('BOTTOMPADDING', (0,0),  (-1,0),  5),
        ('GRID',          (0,0),  (-1,-1), 0.5, BORDER_GREY),
        ('VALIGN',        (0,0),  (-1,-1), 'MIDDLE'),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 20))

    # Prepare Bank Details Table
    bank_src = settings if settings else company
    bank_content = [[Paragraph("PAYMENT DETAILS", s_BoxTitle)]]
    has_bank = False
    for attr, label in [('bank_name','Bank'), ('account_holder_name','A/C Holder'), ('account_number','A/C No'), ('ifsc_code','IFSC')]:
        val = getattr(bank_src, attr, None)
        if val:
            has_bank = True
            bank_content.append([Paragraph(f"<b>{label}:</b> {val}", s_BoxValue)])

    bank_tbl = Spacer(1, 1)
    if has_bank:
        bank_tbl = Table(bank_content, colWidths=[3.2 * inch])
        bank_tbl.setStyle(TableStyle([
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
            ('BOX',           (0,0), (-1,-1), 0.5, BORDER_GREY),
            ('TOPPADDING',    (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING',   (0,0), (-1,-1), 8),
            ('BACKGROUND',    (0,0), (-1,-1), WHITE),
        ]))

    # Prepare Summary Table
    summary_data = [
        [Paragraph("<b>LEDGER SUMMARY</b>", s_BoxTitle), ""],
        [Paragraph("Total Sales (Invoices Gross)", s_TotalLabel), Paragraph(pkr(total_sales), s_TotalValue)],
        [Paragraph("Total Discounts Given", s_TotalLabel), Paragraph(pkr(total_discounts), s_TotalValue)],
        [Paragraph("Total Payments Received", s_TotalLabel), Paragraph(pkr(total_paid), s_TotalValue)],
        [Paragraph("Total Advances Received", s_TotalLabel), Paragraph(pkr(float(customer.total_advances_received or 0)), s_TotalValue)],
        [Paragraph("Total Advances Applied", s_TotalLabel), Paragraph(pkr(total_advances_applied), s_TotalValue)],
        [Paragraph("Remaining Advance Balance", s_TotalLabel), Paragraph(pkr(float(customer.remaining_advance_balance or 0)), s_TotalValue)],
        [Paragraph("<b>Closing Balance / Outstanding</b>", s_GrandLabel), Paragraph(pkr(max(0, running_balance)), s_GrandValue)],
    ]
    
    sum_w = 4.0 * inch
    summary_tbl = Table(summary_data, colWidths=[sum_w * 0.6, sum_w * 0.4])
    summary_tbl.setStyle(TableStyle([
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW', (0,-1), (-1,-1), 1, BLACK),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ('BACKGROUND', (0,-1), (-1,-1), LIGHT_GREY),
    ]))
    
    # Combined row for Bank (Left) and Summary (Right)
    combined_tbl = Table([[bank_tbl, summary_tbl]], colWidths=[TABLE_WIDTH - sum_w, sum_w])
    combined_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (0,0), (0,0),   'LEFT'),
        ('ALIGN', (1,0), (1,0),   'RIGHT'),
        ('LEFTPADDING',   (0,0), (-1,-1), 0),
        ('RIGHTPADDING',  (0,0), (-1,-1), 0),
    ]))
    
    elements.append(combined_tbl)
    elements.append(Spacer(1, 15))

    # ── Thank You Message Section ──────────────────────────────────────────
    # Calculate dynamic values
    latest_due_date = "Immediate"
    if sales_q:
        latest = sales_q[-1]
        if hasattr(latest, 'due_date') and latest.due_date:
            latest_due_date = latest.due_date.strftime('%d-%m-%Y')
    
    # Register Urdu font if not already
    import arabic_reshaper
    from bidi.algorithm import get_display
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    u_font = 'Helvetica'
    # Ensure we get the project root even when running from app/routes/
    base_dir = os.path.abspath(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
    possible_paths = [
        os.path.join(base_dir, 'app', 'static', 'fonts', 'arial.ttf'),
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/TTF/arial.ttf",
        "C:\\Windows\\Fonts\\arial.ttf"
    ]
    for fp in possible_paths:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont('UrduFont', fp))
                u_font = 'UrduFont'
                break
            except: continue
            
    if settings and settings.default_terms:
        terms_content = [Paragraph("TERMS & CONDITIONS", s_BoxTitle)]
        for line in settings.default_terms.split('\n'):
            if line.strip():
                terms_content.append(Paragraph(line.strip(), s_BoxValue))
        terms_tbl = Table([[c] for c in terms_content], colWidths=[7.3 * inch])
        terms_tbl.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 0.5, BORDER_GREY),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        elements.append(terms_tbl)
        elements.append(Spacer(1, 15))

    # English Message (Left)
    eng_style = ParagraphStyle('EngThank', parent=styles['Normal'], fontSize=8.5, textColor=TEXT_COLOR, leading=12)
    eng_text = (
        f"Thank you, {customer.name}, for your recent purchase.<br/>"
        f"Your total amount due is {pkr(running_balance)}, due by {latest_due_date}.<br/>"
        f"Any discount is valid only if paid by the due date.<br/>"
        f"If you have any questions, please let us know.<br/><br/>"
        f"Best regards, We look forward to serving you again in the future."
    )
    
    # Urdu Message (Shaped & Right Aligned) - Aligned with pdf_utils.py
    urdu_style = ParagraphStyle('UrduThank', parent=styles['Normal'], fontSize=9, fontName=u_font, alignment=TA_RIGHT, textColor=TEXT_COLOR, leading=14)
    u_text = (
        f"شکریہ،  {customer.name}  ، آپ کی حالیہ خریداری کے لیے۔\n"
        f"آپ کی قابلِ ادائیگی رقم {pkr(running_balance)} ہے، جس کی آخری تاریخ {latest_due_date} ہے۔"
        f"کوئی بھی رعایت صرف مقررہ تاریخ تک ادائیگی کی صورت میں ہی قابلِ قبول ہوگی۔\n"
        f"اگر آپ کے کوئی سوالات ہوں تو براہِ کرم ہمیں آگاہ کریں۔\n"
        f"نیک تمناؤں کے ساتھ،"
        f"ہم مستقبل میں دوبارہ آپ کی خدمت کرنے کے منتظر ہیں۔"
    )
    reshaped_text = arabic_reshaper.reshape(u_text)
    bidi_text = get_display(reshaped_text).replace('\n', '<br/>')
    
    thank_tbl = Table([[Paragraph(eng_text, eng_style), Paragraph(bidi_text, urdu_style)]], colWidths=[3.7*inch, 3.7*inch])
    thank_tbl.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    
    elements.append(thank_tbl)

    doc.build(elements, onFirstPage=_draw_page_decorations, onLaterPages=_draw_page_decorations)
    output.seek(0)
    
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=False,
        download_name=f"Ledger_{customer.name.replace(' ', '_')}_{_dt.now().strftime('%Y%m%d')}.pdf"
    )

@bp.route('/customer/<int:id>/advances-json')
@login_required
def customer_advances_json(id):
    customer = Customer.query.get_or_404(id)
    advances = CustomerAdvance.query.filter_by(customer_id=id).order_by(CustomerAdvance.date.desc()).all()
    return jsonify({
        'advances': [{
            'date': a.date.strftime('%d-%b-%Y'),
            'amount': a.amount,
            'applied': a.applied_amount,
            'balance': a.remaining_balance
        } for a in advances]
    })

@bp.route('/customer/add', methods=['GET', 'POST'])
@login_required
@permission_required('customers', action='add')
def add_customer():
    form = CustomerForm()
    groups = CustomerGroup.query.filter_by(is_active=True).all()
    form.group_id.choices = [(0, '- Select Group -')] + [(g.id, g.name) for g in groups]
    
    if form.validate_on_submit():
        # Handle image upload
        image_path = None
        if 'image' in request.files:
            image_file = request.files['image']
            if image_file and image_file.filename:
                import time, uuid
                original_filename = secure_filename(image_file.filename)
                unique_prefix = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
                filename = f"{unique_prefix}_{original_filename}"
                
                upload_dir = os.path.join('app', 'static', 'uploads', 'customers')
                os.makedirs(upload_dir, exist_ok=True)
                file_path = os.path.join(upload_dir, filename)
                image_file.save(file_path)
                image_path = f"uploads/customers/{filename}" # Store relative to static folder

        customer = Customer(
            name=form.name.data,
            company_name=form.company_name.data,
            group_id=form.group_id.data if form.group_id.data != 0 else None,
            email=form.email.data,
            phone=form.phone.data,
            address=form.address.data,
            city=form.city.data,
            state=form.state.data,
            country=form.country.data,
            postal_code=form.postal_code.data,
            gst_number=form.gst_number.data,
            payment_method=form.payment_method.data,
            sub_customers=form.sub_customers.data,
            image_path=image_path
        )
        db.session.add(customer)
        db.session.commit()
        
        log_activity('Customers', f'Added Customer: {customer.name}', 
                    f'Company: {customer.company_name or "N/A"}')
        
        flash('Customer added successfully!', 'success')
        return redirect(url_for('sales.customers'))
    
    return render_template('sales/add_customer.html', form=form)

@bp.route('/customer/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('customers', action='edit')
def edit_customer(id):
    customer = Customer.query.get_or_404(id)
    form = CustomerForm(obj=customer)
    groups = CustomerGroup.query.filter_by(is_active=True).all()
    form.group_id.choices = [(0, '- Select Group -')] + [(g.id, g.name) for g in groups]
    
    if form.validate_on_submit():
        customer.name = form.name.data
        customer.company_name = form.company_name.data
        customer.group_id = form.group_id.data if form.group_id.data != 0 else None
        customer.email = form.email.data
        customer.phone = form.phone.data
        customer.address = form.address.data
        customer.city = form.city.data
        customer.state = form.state.data
        customer.country = form.country.data
        customer.postal_code = form.postal_code.data
        customer.gst_number = form.gst_number.data
        customer.payment_method = form.payment_method.data
        customer.sub_customers = form.sub_customers.data

        # Handle image upload
        if 'image' in request.files:
            image_file = request.files['image']
            if image_file and image_file.filename:
                import time, uuid
                original_filename = secure_filename(image_file.filename)
                unique_prefix = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
                filename = f"{unique_prefix}_{original_filename}"
                
                upload_dir = os.path.join('app', 'static', 'uploads', 'customers')
                os.makedirs(upload_dir, exist_ok=True)
                file_path = os.path.join(upload_dir, filename)
                image_file.save(file_path)
                customer.image_path = f"uploads/customers/{filename}" # Store relative to static folder

        db.session.commit()
        log_activity('Customers', f'Updated Customer: {customer.name}', f'ID: {customer.id}')
        flash('Customer updated successfully!', 'success')
        return redirect(url_for('sales.customers'))
    
    return render_template('sales/edit_customer.html', form=form, customer=customer)

@bp.route('/customer/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('customers', action='delete')
def delete_customer(id):
    customer = Customer.query.get_or_404(id)
    
    # Check if customer has associated records (Sales)
    if customer.sales:
        flash('Cannot delete customer as they have associated sales records.', 'danger')
        return redirect(url_for('sales.customers'))

    cust_name = customer.name
    db.session.delete(customer)
    db.session.commit()
    log_activity('Customers', f'Deleted Customer: {cust_name}', f'ID: {id}')
    flash('Customer deleted successfully!', 'success')
    return redirect(url_for('sales.customers'))

@bp.route('/customers/bulk-delete', methods=['POST'])
@login_required
@permission_required('customers', action='delete')
def bulk_delete_customers():
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No customers selected'}), 400
    
    deleted_count = 0
    skipped_count = 0
    errors = []
    
    for customer_id in ids:
        customer = Customer.query.get(customer_id)
        if not customer:
            continue
            
        # Check associations
        if customer.sales:
            skipped_count += 1
            continue
            
        try:
            db.session.delete(customer)
            deleted_count += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f'Error deleting Customer {customer.name}: {str(e)}')
            
    if deleted_count > 0:
        db.session.commit()
        log_activity('Customers', f'Bulk Deleted Customers', f'Deleted {deleted_count} customers')

    message = f'Successfully deleted {deleted_count} customers.'
    if skipped_count > 0:
        message += f' Skipped {skipped_count} customers with associated records.'
    
    if errors:
        return jsonify({'success': False, 'message': message, 'errors': errors}), 500
    return jsonify({'success': True, 'message': message})


@bp.route('/invoice/<int:id>/pdf')
@login_required
def invoice_pdf(id):
    sale = Sale.query.get_or_404(id)
    company = Company.query.first()
    invoice_settings = InvoiceSettings.query.first()
    
    try:
        buffer = generate_professional_pdf('invoice', sale, company, invoice_settings)
        
        response = make_response(buffer)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename="invoice_{sale.invoice_number or "unknown"}.pdf"'
        return response
    except Exception as e:
        print(f"PDF generation error: {str(e)}")
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('sales.invoice_detail', id=id))

@bp.route('/invoice/<int:id>/pdf/view')
@login_required
def invoice_pdf_view(id):
    """View PDF inline (for sharing)"""
    sale = Sale.query.get_or_404(id)
    company = Company.query.first()
    invoice_settings = InvoiceSettings.query.first()
    
    try:
        buffer = generate_professional_pdf('invoice', sale, company, invoice_settings)
        
        response = make_response(buffer)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename="{sale.invoice_number}.pdf"'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('sales.invoice_detail', id=id))

@bp.route('/invoice/<int:id>/pdf/share')
@login_required
def invoice_pdf_share(id):
    """Generate PDF, save to public folder, and return shareable link"""
    from flask import current_app
    import os
    import secrets
    
    sale = Sale.query.get_or_404(id)
    company = Company.query.first()
    invoice_settings = InvoiceSettings.query.first()
    
    try:
        buffer = generate_professional_pdf('invoice', sale, company, invoice_settings)
        buffer.seek(0)
        pdf_data = buffer.getvalue()
        
        # Save to public folder
        public_dir = os.path.join(current_app.root_path, 'static', 'shared_pdfs')
        os.makedirs(public_dir, exist_ok=True)
        
        # Generate unique filename
        token = secrets.token_urlsafe(8)
        filename = f"invoice_{sale.invoice_number or 'unknown'}_{token}.pdf"
        filepath = os.path.join(public_dir, filename)
        
        # Save buffer to file
        with open(filepath, 'wb') as f:
            f.write(pdf_data)
        
        # Return the shareable URL
        share_url = url_for('static', filename=f'shared_pdfs/{filename}')
        return {'share_url': share_url, 'download_url': url_for('sales.invoice_pdf', id=id)}
    except Exception as e:
        return {'error': str(e)}, 500

@bp.route('/company', methods=['GET', 'POST'])
@login_required
def company_settings():
    company = Company.query.first()

    if request.method == 'POST':
        # Get name from form - required field
        company_name = request.form.get('name')
        if not company_name:
            flash('Company name is required.', 'error')
            return redirect(url_for('sales.company_settings'))

        if not company:
            company = Company(name=company_name)
            db.session.add(company)
        else:
            company.name = company_name

        company.address = request.form.get('address')
        company.phone = request.form.get('phone')
        company.email = request.form.get('email')
        company.gst_number = request.form.get('gst_number')
        company.pan_number = request.form.get('pan_number')
        company.website = request.form.get('website')
        company.date_format = request.form.get('date_format', '%Y-%m-%d')
        company.mo_prefix = request.form.get('mo_prefix', 'MO-')
        company.mo_suffix = request.form.get('mo_suffix', '')
        company.next_mo_number = request.form.get('next_mo_number', 1, type=int)
        company.bank_name = request.form.get('bank_name')
        company.account_number = request.form.get('account_number')
        company.ifsc_code = request.form.get('ifsc_code')
        company.account_holder_name = request.form.get('account_holder_name')

        # Handle logo upload
        if 'logo' in request.files:
            logo_file = request.files['logo']
            if logo_file and logo_file.filename:
                # Validate file type
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
                filename = secure_filename(logo_file.filename)
                ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
                if ext not in allowed_extensions:
                    flash('Invalid file type for logo. Only PNG, JPG, JPEG, GIF allowed.', 'error')
                    return redirect(url_for('sales.company_settings'))
                # Save file
                upload_dir = os.path.join('app', 'static', 'uploads')
                os.makedirs(upload_dir, exist_ok=True)
                logo_path = os.path.join(upload_dir, filename)
                logo_file.save(logo_path)
                # Store path
                company.logo_path = logo_path

        # Handle signature upload
        if 'signature' in request.files:
            sig_file = request.files['signature']
            if sig_file and sig_file.filename:
                # Validate file type
                allowed_extensions = {'png', 'jpg', 'jpeg'}
                filename = secure_filename(sig_file.filename)
                ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
                if ext not in allowed_extensions:
                    flash('Invalid file type for signature. Only PNG, JPG, JPEG allowed.', 'error')
                    return redirect(url_for('sales.company_settings'))
                # Save file
                upload_dir = os.path.join('app', 'static', 'uploads')
                os.makedirs(upload_dir, exist_ok=True)
                sig_path = os.path.join(upload_dir, filename)
                sig_file.save(sig_path)
                company.signature_path = sig_path

        # Handle logo removal
        if request.form.get('remove_logo') == 'on' and company.logo_path:
            try:
                logo_file_path = company.logo_path
                if os.path.exists(logo_file_path):
                    os.remove(logo_file_path)
            except Exception:
                pass  # Ignore deletion errors
            company.logo_path = None

        # Handle signature removal
        if request.form.get('remove_signature') == 'on' and company.signature_path:
            try:
                sig_file_path = company.signature_path
                if os.path.exists(sig_file_path):
                    os.remove(sig_file_path)
            except Exception:
                pass  # Ignore deletion errors
            company.signature_path = None

        db.session.commit()
        log_activity('Settings', f'Updated Company Settings', f'Company: {company.name}')
        flash('Company settings updated successfully!', 'success')
        return redirect(url_for('sales.company_settings'))

    return render_template('sales/company_settings.html', company=company)

@bp.route('/invoice/settings', methods=['GET', 'POST'])
@login_required
def invoice_settings():
    settings = InvoiceSettings.query.first()
    
    form = InvoiceSettingsForm(obj=settings)
    
    # Populate customer groups for the dropdown
    customer_groups = CustomerGroup.query.filter_by(is_active=True).order_by(CustomerGroup.name).all()
    form.overdue_restricted_groups.choices = [(g.id, g.name) for g in customer_groups]
    
    if request.method == 'GET' and settings and settings.overdue_restricted_groups:
        try:
            form.overdue_restricted_groups.data = json.loads(settings.overdue_restricted_groups)
        except:
            form.overdue_restricted_groups.data = []

    if form.validate_on_submit():
        if not settings:
            settings = InvoiceSettings()
            db.session.add(settings)
        
        settings.default_notes = form.default_notes.data
        settings.default_terms = form.default_terms.data
        settings.bank_name = form.bank_name.data
        settings.account_holder_name = form.account_holder_name.data
        settings.account_number = form.account_number.data
        settings.ifsc_code = form.ifsc_code.data
        settings.swift_code = form.swift_code.data
        settings.bank_address = form.bank_address.data
        settings.payment_instructions = form.payment_instructions.data
        settings.invoice_prefix = form.invoice_prefix.data or ''
        settings.invoice_suffix = form.invoice_suffix.data or ''
        settings.next_number = form.next_number.data or 1
        settings.tax_name = form.tax_name.data
        settings.tax_rate = form.tax_rate.data
        settings.payment_terms = form.payment_terms.data
        settings.notes = form.notes.data
        
        # Save restricted groups as JSON
        if form.overdue_restricted_groups.data:
            settings.overdue_restricted_groups = json.dumps(form.overdue_restricted_groups.data)
        else:
            settings.overdue_restricted_groups = None

        settings.product_discount_conditions = form.product_discount_conditions.data
            
        db.session.commit()
        log_activity('Settings', f'Updated Invoice Settings', f'Prefix: {settings.invoice_prefix}, Next #: {settings.next_number}')
        flash('Invoice settings updated successfully.', 'success')
        return redirect(url_for('sales.invoice_settings'))
    
    # Product Discount Conditions selector: only finished goods (manufactured items).
    products = Product.query.filter_by(is_active=True, is_manufactured=True).order_by(Product.name).all()

    # Lookup for ALL active products so any previously-saved condition (including
    # legacy non-finished-good entries) still renders its name/price correctly,
    # even though such products are no longer selectable in the dropdown.
    all_active = Product.query.filter_by(is_active=True).all()
    product_lookup = {
        p.id: {'name': p.name, 'price': p.unit_price or 0, 'cost': p.cost_price or 0}
        for p in all_active
    }

    return render_template('sales/invoice_settings.html', settings=settings, form=form,
                           products=products, product_lookup=product_lookup)


# ===== CUSTOMER ADVANCE ROUTES =====

@bp.route('/customer/<int:id>/advance', methods=['POST'])
@login_required
@permission_required('customers', action='add')
def customer_receive_advance(id):
    """Record a quick advance payment received from a customer against material."""
    customer = Customer.query.get_or_404(id)
    amount = float(request.form.get('amount', 0))
    description = request.form.get('description', '').strip()
    advance_date_str = request.form.get('advance_date', '')
    
    if amount <= 0:
        flash('Advance amount must be greater than zero.', 'danger')
        return redirect(url_for('sales.customer_profile', id=id))
    
    try:
        advance_date = datetime.strptime(advance_date_str, '%Y-%m-%d').date() if advance_date_str else date.today()
    except ValueError:
        advance_date = date.today()
    
    # Determine approval: admin-recorded advances are pre-approved; staff need admin approval
    creator_is_admin = (current_user.role == 'admin')

    advance = CustomerAdvance(
        customer_id=customer.id,
        amount=amount,
        date=advance_date,
        description=description or 'Advance for purchase',
        created_by=current_user.id,
        is_approved=creator_is_admin,
        approved_by=current_user.id if creator_is_admin else None,
        approved_at=datetime.utcnow() if creator_is_admin else None
    )
    db.session.add(advance)
    db.session.commit()
    log_activity('Customers', f'Received Advance from {customer.name}',
                f'Amount: PKR {amount:,.2f}, Approved: {creator_is_admin}')

    if creator_is_admin:
        flash(f'Advance of PKR {amount:,.2f} recorded and approved successfully.', 'success')
    else:
        flash(f'Advance of PKR {amount:,.2f} recorded — pending admin approval.', 'info')
    
    return redirect(url_for('sales.customer_profile', id=id))


@bp.route('/customer/<int:id>/advance/<int:adv_id>/approve', methods=['POST'])
@login_required
def approve_advance(id, adv_id):
    """Admin approves a pending customer advance."""
    if current_user.role != 'admin':
        flash('Only admin can approve advances.', 'danger')
        return redirect(url_for('sales.customer_profile', id=id))

    advance = CustomerAdvance.query.get_or_404(adv_id)
    if advance.is_approved:
        flash('Advance is already approved.', 'info')
        return redirect(url_for('sales.customer_profile', id=id))

    advance.is_approved = True
    advance.approved_by = current_user.id
    advance.approved_at = datetime.utcnow()
    db.session.commit()
    
    log_activity('Customers', f'Approved Advance from {advance.customer.name}',
                 f'Amount: PKR {advance.amount:,.2f} approved by {current_user.username}')
                 
    flash(f'Advance of PKR {advance.amount:,.2f} approved successfully!', 'success')
    return redirect(url_for('sales.customer_profile', id=id))


@bp.route('/customer/<int:id>/advance/<int:adv_id>/reject', methods=['POST'])
@login_required
def reject_advance(id, adv_id):
    """Admin rejects a pending customer advance."""
    if current_user.role != 'admin':
        flash('Only admin can reject advances.', 'danger')
        return redirect(url_for('sales.customer_profile', id=id))

    advance = CustomerAdvance.query.get_or_404(adv_id)
    if advance.is_approved:
        flash('Cannot reject an already approved advance.', 'warning')
        return redirect(url_for('sales.customer_profile', id=id))

    reason = request.form.get('reason', '').strip()
    advance.is_rejected = True
    advance.is_approved = False
    advance.rejection_reason = reason
    db.session.commit()

    log_activity('Customers', f'Rejected Advance from {advance.customer.name}',
                 f'Amount: PKR {advance.amount:,.2f} rejected by {current_user.username}. Reason: {reason}')
    flash(f'Advance of PKR {advance.amount:,.2f} has been rejected.', 'warning')
    return redirect(url_for('sales.customer_profile', id=id))


@bp.route('/customer/<int:customer_id>/advance/<int:adv_id>/apply', methods=['POST'])
@login_required
@permission_required('customers', action='edit')
def customer_apply_advance(customer_id, adv_id):
    """Apply a customer advance against a specific invoice."""
    advance = CustomerAdvance.query.get_or_404(adv_id)
    invoice_id = request.form.get('invoice_id')
    
    # Check if advance is approved
    if not advance.is_approved:
        flash('This advance is pending approval and cannot be applied yet.', 'warning')
        return redirect(url_for('sales.customer_profile', id=customer_id))

    # Check if advance still has remaining balance
    if advance.remaining_balance <= 0:
        flash('This advance has no remaining balance to apply.', 'warning')
        return redirect(url_for('sales.customer_profile', id=customer_id))
    
    invoice = Sale.query.get_or_404(invoice_id)
    
    # Ensure the invoice belongs to the customer
    if invoice.customer_id != customer_id:
        flash('Invalid invoice for this customer.', 'danger')
        return redirect(url_for('sales.customer_profile', id=customer_id))
    
    # Check invoice balance
    if invoice.balance_due <= 0:
        flash('This invoice has no balance to apply advance against.', 'warning')
        return redirect(url_for('sales.customer_profile', id=customer_id))
    
    # Calculate how much to apply (minimum of advance balance and invoice balance)
    apply_amount = min(advance.remaining_balance, invoice.balance_due)
    
    # Update advance
    advance.applied_amount += apply_amount
    if advance.remaining_balance <= 0:
        advance.is_adjusted = True
        advance.adjusted_invoice_id = invoice.id
    
    # Update invoice advance applied
    invoice.advance_applied = (invoice.advance_applied or 0) + apply_amount
    
    # Recalculate invoice total
    invoice.calculate_totals()
    invoice.paid_amount += apply_amount
    invoice.update_status()
    
    db.session.commit()
    log_activity('Customers', f'Applied Advance to Invoice #{invoice.invoice_number}',
                f'Amount: PKR {apply_amount:,.2f}, Customer ID: {customer_id}')
    flash(f'Advance of PKR {apply_amount:,.2f} applied to invoice {invoice.invoice_number}.', 'success')
    return redirect(url_for('sales.customer_profile', id=customer_id))


@bp.route('/customer/<int:customer_id>/advance/<int:adv_id>/delete', methods=['POST'])
@login_required
@permission_required('customers', action='delete')
def customer_delete_advance(customer_id, adv_id):
    """Delete a customer advance. If applied, first unapplies from invoices."""
    advance = CustomerAdvance.query.get_or_404(adv_id)
    
    if advance.customer_id != customer_id:
        flash('Invalid advance for this customer.', 'danger')
        return redirect(url_for('sales.customer_profile', id=customer_id))
    
    # If advance has been applied, unapply it first
    if advance.applied_amount > 0:
        from app.models import Sale
        invoices_with_advance = Sale.query.filter(
            Sale.customer_id == customer_id,
            Sale.advance_applied > 0
        ).all()
        
        reverse_amount = advance.applied_amount
        
        for invoice in invoices_with_advance:
            if reverse_amount <= 0:
                break
            if invoice.advance_applied > 0:
                reverse_amt = min(invoice.advance_applied, reverse_amount)
                invoice.advance_applied -= reverse_amt
                invoice.calculate_totals()
                invoice.paid_amount -= reverse_amt
                invoice.update_status()
                reverse_amount -= reverse_amt
        
        advance.applied_amount = 0
        advance.is_adjusted = False
        advance.adjusted_invoice_id = None
    
    customer_name = advance.customer.name
    amount = advance.amount
    
    db.session.delete(advance)
    db.session.commit()
    log_activity('Customers', f'Deleted Advance from {customer_name}',
                f'Amount: PKR {amount:,.2f}')
    flash(f'Advance of PKR {amount:,.2f} from {customer_name} has been deleted.', 'success')
    return redirect(url_for('sales.customer_profile', id=customer_id))


@bp.route('/customer/<int:customer_id>/advance/<int:adv_id>/unapply', methods=['POST'])
@login_required
@permission_required('customers', action='edit')
def customer_unapply_advance(customer_id, adv_id):
    """Unapply/reverse an advance that was applied to an invoice."""
    advance = CustomerAdvance.query.get_or_404(adv_id)
    
    if advance.customer_id != customer_id:
        flash('Invalid advance for this customer.', 'danger')
        return redirect(url_for('sales.customer_profile', id=customer_id))
    
    if advance.applied_amount <= 0:
        flash('This advance has not been applied to any invoice.', 'warning')
        return redirect(url_for('sales.customer_profile', id=customer_id))
    
    # Find invoices that have this advance applied
    from app.models import Sale
    invoices_with_advance = Sale.query.filter(
        Sale.customer_id == customer_id,
        Sale.advance_applied > 0
    ).all()
    
    # Reverse the applied amount from invoices
    reverse_amount = advance.applied_amount
    
    for invoice in invoices_with_advance:
        if reverse_amount <= 0:
            break
        if invoice.advance_applied > 0:
            reverse_amt = min(invoice.advance_applied, reverse_amount)
            invoice.advance_applied -= reverse_amt
            invoice.calculate_totals()
            invoice.paid_amount -= reverse_amt
            invoice.update_status()
            reverse_amount -= reverse_amt
    
    # Reset advance
    advance.applied_amount = 0
    advance.is_adjusted = False
    advance.adjusted_invoice_id = None
    
    db.session.commit()
    log_activity('Customers', f'Unapplied Advance for Customer ID {customer_id}',
                f'Amount: PKR {advance.amount:,.2f}')
    flash(f'Advance of PKR {advance.amount:,.2f} has been unapplied from invoices.', 'success')
    return redirect(url_for('sales.customer_profile', id=customer_id))


# --- Salesman Management ---

@bp.route('/salesmen')
@login_required
def salesmen_list():
    salesmen = Salesman.query.all()
    return render_template('sales/salesmen.html', salesmen=salesmen)

@bp.route('/salesman/add', methods=['GET', 'POST'])
@login_required
@permission_required('salesmen', action='add')
def add_salesman():
    form = SalesmanForm()
    groups = SalesmanGroup.query.filter_by(is_active=True).all()
    form.group_id.choices = [(0, '- Select Group -')] + [(g.id, g.name) for g in groups]
    users = User.query.filter_by(is_active=True).order_by(User.username).all()
    form.user_id.choices = [(0, '- Not Linked -')] + [(u.id, u.username) for u in users]

    if form.validate_on_submit():
        salesman = Salesman(
            name=form.name.data,
            email=form.email.data,
            phone=form.phone.data,
            address=form.address.data,
            commission_rate=form.commission_rate.data,
            group_id=form.group_id.data if form.group_id.data != 0 else None,
            user_id=form.user_id.data if form.user_id.data and form.user_id.data != 0 else None,
            is_active=form.is_active.data
        )
        db.session.add(salesman)
        db.session.commit()
        log_activity('Sales', f'Added Salesman: {salesman.name}', f'ID: {salesman.id}')
        flash('Salesman added successfully!', 'success')
        return redirect(url_for('sales.salesmen_list'))
    return render_template('sales/salesman_form.html', form=form, title='Add Salesman')

@bp.route('/salesman/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('salesmen', action='edit')
def edit_salesman(id):
    salesman = Salesman.query.get_or_404(id)
    form = SalesmanForm(obj=salesman)
    groups = SalesmanGroup.query.filter_by(is_active=True).all()
    form.group_id.choices = [(0, '- Select Group -')] + [(g.id, g.name) for g in groups]
    users = User.query.filter_by(is_active=True).order_by(User.username).all()
    form.user_id.choices = [(0, '- Not Linked -')] + [(u.id, u.username) for u in users]
    if request.method == 'GET':
        form.user_id.data = salesman.user_id or 0

    if form.validate_on_submit():
        salesman.name = form.name.data
        salesman.email = form.email.data
        salesman.phone = form.phone.data
        salesman.address = form.address.data
        salesman.commission_rate = form.commission_rate.data
        salesman.group_id = form.group_id.data if form.group_id.data != 0 else None
        salesman.user_id = form.user_id.data if form.user_id.data and form.user_id.data != 0 else None
        salesman.is_active = form.is_active.data
        db.session.commit()
        log_activity('Sales', f'Updated Salesman: {salesman.name}', f'ID: {salesman.id}')
        flash('Salesman updated successfully!', 'success')
        return redirect(url_for('sales.salesmen_list'))
    return render_template('sales/salesman_form.html', form=form, title='Edit Salesman', salesman=salesman)

@bp.route('/salesman/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('salesmen', action='delete')
def delete_salesman(id):
    salesman = Salesman.query.get_or_404(id)
    if salesman.sales:
        flash('Cannot delete salesman with associated sales records.', 'danger')
        return redirect(url_for('sales.salesmen_list'))
    sm_name = salesman.name
    db.session.delete(salesman)
    db.session.commit()
    log_activity('Sales', f'Deleted Salesman: {sm_name}', f'ID: {id}')
    flash('Salesman deleted successfully.', 'info')
    return redirect(url_for('sales.salesmen_list'))

@bp.route('/salesman/quick-add', methods=['POST'])
@login_required
@permission_required('salesmen', action='add')
def quick_add_salesman():
    name = request.form.get('name')
    if not name:
        return jsonify({'success': False, 'message': 'Name is required'}), 400
    
    group_id = request.form.get('group_id')
    
    salesman = Salesman(name=name, group_id=group_id if group_id and group_id != '0' else None)
    db.session.add(salesman)
    db.session.commit()
    log_activity('Sales', f'Quick Added Salesman: {salesman.name}', f'ID: {salesman.id}')

    return jsonify({
        'success': True, 
        'salesman': {
            'id': salesman.id,
            'name': salesman.name
        }
    })

@bp.route('/salesman/group/quick-add', methods=['POST'])
@login_required
@permission_required('salesmen', action='add')
def quick_add_salesman_group():
    name = request.form.get('name')
    if not name:
        return jsonify({'success': False, 'message': 'Name is required'}), 400
    
    if SalesmanGroup.query.filter_by(name=name).first():
        return jsonify({'success': False, 'message': 'Group already exists'}), 400

    group = SalesmanGroup(name=name)
    db.session.add(group)
    db.session.commit()
    log_activity('Sales', f'Quick Added Salesman Group: {group.name}', f'ID: {group.id}')

    return jsonify({
        'success': True,
        'group': {
            'id': group.id,
            'name': group.name
        }
    })

# --- Customer Group Management ---

@bp.route('/customer-groups')
@login_required
def customer_groups_list():
    groups = CustomerGroup.query.all()
    return render_template('sales/customer_groups.html', groups=groups)

@bp.route('/customer-group/add', methods=['GET', 'POST'])
@login_required
@permission_required('customer_groups', action='add')
def add_customer_group():
    form = CustomerGroupForm()
    if form.validate_on_submit():
        group = CustomerGroup(
            name=form.name.data,
            description=form.description.data
        )
        db.session.add(group)
        db.session.commit()
        log_activity('Customers', f'Added Customer Group: {group.name}', f'ID: {group.id}')
        flash('Customer Group added successfully!', 'success')
        return redirect(url_for('sales.customer_groups_list'))
    return render_template('sales/customer_group_form.html', form=form, title='Add Customer Group')

@bp.route('/customer-group/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('customer_groups', action='edit')
def edit_customer_group(id):
    group = CustomerGroup.query.get_or_404(id)
    form = CustomerGroupForm(obj=group)
    if form.validate_on_submit():
        group.name = form.name.data
        group.description = form.description.data
        db.session.commit()
        log_activity('Customers', f'Updated Customer Group: {group.name}', f'ID: {group.id}')
        flash('Customer Group updated successfully!', 'success')
        return redirect(url_for('sales.customer_groups_list'))
    return render_template('sales/customer_group_form.html', form=form, title='Edit Customer Group', group=group)

@bp.route('/customer-group/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('customer_groups', action='delete')
def delete_customer_group(id):
    group = CustomerGroup.query.get_or_404(id)
    if group.customers:
        flash('Cannot delete group with associated customers.', 'danger')
        return redirect(url_for('sales.customer_groups_list'))
    grp_name = group.name
    db.session.delete(group)
    db.session.commit()
    log_activity('Customers', f'Deleted Customer Group: {grp_name}', f'ID: {id}')
    flash('Customer Group deleted successfully.', 'info')
    return redirect(url_for('sales.customer_groups_list'))

@bp.route('/customer-group/quick-add', methods=['POST'])
@login_required
@permission_required('customer_groups', action='add')
def quick_add_customer_group():
    name = request.form.get('name')
    if not name:
        return jsonify({'success': False, 'message': 'Name is required'}), 400
    
    if CustomerGroup.query.filter_by(name=name).first():
        return jsonify({'success': False, 'message': 'Group already exists'}), 400

    group = CustomerGroup(name=name)
    db.session.add(group)
    db.session.commit()
    log_activity('Customers', f'Quick Added Customer Group: {group.name}', f'ID: {group.id}')

    return jsonify({
        'success': True,
        'group': {
            'id': group.id,
            'name': group.name
        }
    })

@bp.route('/public/invoice/<token>')
def public_invoice(token):
    from datetime import datetime
    from app.models import Company, InvoiceSettings
    sale = Sale.query.filter_by(access_token=token).first_or_404()
    
    # Check expiry
    if sale.token_expiry and sale.token_expiry < datetime.utcnow():
        return "Link expired", 403
        
    company = Company.query.first()
    invoice_settings = InvoiceSettings.query.first()
    
    try:
        buffer = generate_professional_pdf('invoice', sale, company, invoice_settings)
        
        response = make_response(buffer)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename="invoice_{sale.invoice_number or "unknown"}.pdf"'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        return f"Error generating PDF: {str(e)}", 500

@bp.route('/public/ledger/<token>')
def public_ledger(token):
    from datetime import datetime
    customer = Customer.query.filter_by(access_token=token).first_or_404()
    
    # Check expiry
    if customer.token_expiry and customer.token_expiry < datetime.utcnow():
        return "Link expired", 403
        
    # Reuse the existing PDF generation logic
    return customer_export_pdf(customer.id, is_public=True)

@bp.route('/customer/<int:id>/ledger-json')
@login_required
def customer_ledger_json(id):
    customer = Customer.query.get_or_404(id)
    
    # Opening balance
    opening_balance = float(customer.opening_balance or 0)
    
    # Get all components
    sales = Sale.query.filter_by(customer_id=id).order_by(Sale.date.asc()).all()
    advances = CustomerAdvance.query.filter_by(customer_id=id).order_by(CustomerAdvance.date.asc()).all()
    returns = SaleReturn.query.filter_by(customer_id=id).order_by(SaleReturn.date.asc()).all()
    
    # In this system, payments are linked to sales. 
    sale_ids = [s.id for s in sales]
    payments = Payment.query.filter(Payment.invoice_id.in_(sale_ids)).order_by(Payment.date.asc()).all() if sale_ids else []
    
    ledger_entries = []
    running_balance = opening_balance
    
    # Initial entry for Opening Balance
    ledger_entries.append({
        'date': '-',
        'invoice_number': 'Opening Balance',
        'customer_name': customer.name,
        'debit': opening_balance if opening_balance > 0 else 0,
        'credit': abs(opening_balance) if opening_balance < 0 else 0,
        'invoice_balance': 0,
        'running_balance': running_balance,
        'status': '-',
        'aging': '-'
    })
    
    today = datetime.utcnow().date()
    
    # Build advances as a sorted list to interleave by date
    advance_events_list = []
    for a in advances:
        dt = datetime.combine(a.date, datetime.min.time())
        advance_events_list.append((dt, {
            'date': dt,
            'type': 'advance',
            'invoice_number': '-',
            'desc': f"Advance Received: {a.description or ''} (PKR {float(a.amount or 0):,.2f})",
            'debit': 0,
            'credit': 0,
            'status': '-',
            'obj': a
        }))
    advance_events_list.sort(key=lambda x: x[0])
    adv_idx = 0

    # Build events grouped per-invoice:
    # Invoice → its payments → its discount → its advance_applied → its returns
    # Standalone advances are interleaved before each invoice by date.
    events = []
    for s in sales:
        # Determine status for ledger display
        display_status = (s.status or 'unpaid').capitalize()
        if s.is_overdue:
            days = s.days_overdue
            display_status = f'Overdue ({days}d)'

        # Insert any advances that occurred before or on this invoice's date
        while adv_idx < len(advance_events_list) and advance_events_list[adv_idx][0].date() <= s.date.date():
            events.append(advance_events_list[adv_idx][1])
            adv_idx += 1

        # 1. The invoice itself
        events.append({
            'date': s.date,
            'type': 'sale',
            'invoice_number': s.invoice_number,
            'desc': f"Invoice #{s.invoice_number}",
            'debit': float(sum(item.total for item in s.items)) * (1 + float(s.tax_rate or 0) / 100) + float(s.delivery_charge or 0),
            'credit': 0,
            'status': display_status,
            'flag_color': s.invoice_health_status,
            'days_overdue': s.days_overdue,
            'obj': s
        })

        # 2. Payments for this invoice (sorted by date)
        for p in sorted(s.payments, key=lambda x: x.date):
            if not getattr(p, 'is_approved', True):
                continue
            events.append({
                'date': p.date,
                'type': 'payment',
                'invoice_number': s.invoice_number,
                'desc': f"Payment Received ({p.method or 'Cash'})",
                'debit': 0,
                'credit': float(p.amount or 0),
                'status': '-',
                'obj': p
            })

        # 3. Discount for this invoice
        is_restricted = getattr(s, 'is_discount_restricted', False)
        effective_discount = getattr(s, 'effective_discount_amount', 0)
        if is_restricted and getattr(s, 'base_discount_amount', 0) > 0:
            events.append({
                'date': s.date, 'type': 'discount',
                'invoice_number': s.invoice_number,
                'desc': f"Discount (Inv #{s.invoice_number}): Reversal due to overdue",
                'debit': 0, 'credit': 0, 'status': '-', 'obj': s
            })
        elif effective_discount > 0:
            events.append({
                'date': s.date, 'type': 'discount',
                'invoice_number': s.invoice_number,
                'desc': f"Discount (Inv #{s.invoice_number})",
                'debit': 0, 'credit': float(effective_discount or 0),
                'status': '-', 'obj': s
            })

        # 4. Advance Applied for this invoice
        if hasattr(s, 'advance_applied') and s.advance_applied and s.advance_applied > 0:
            events.append({
                'date': s.date, 'type': 'advance_applied',
                'invoice_number': s.invoice_number,
                'desc': f"Advance Applied (Inv #{s.invoice_number})",
                'debit': 0, 'credit': float(s.advance_applied or 0),
                'status': '-', 'obj': s
            })

        # 5. Returns for this invoice (sorted by date)
        for r in sorted(s.returns, key=lambda x: x.date):
            events.append({
                'date': r.date, 'type': 'return',
                'invoice_number': r.return_number,
                'desc': f"Sales Return #{r.return_number}",
                'debit': 0, 'credit': float(r.total or 0),
                'status': (r.status or 'completed').capitalize(),
                'obj': r
            })

    # Append any remaining advances after all invoices
    while adv_idx < len(advance_events_list):
        events.append(advance_events_list[adv_idx][1])
        adv_idx += 1
    
    # Calculate Aging Buckets (on current unpaid invoices only)
    aging_buckets = {'0-30': 0, '31-60': 0, '61-90': 0, '91-180': 0, '>180': 0}
    for s in sales:
        if s.status != 'paid':
            invoice_date = s.date.date()
            days_old = (today - invoice_date).days
            bucket = '>180'
            if days_old <= 30: bucket = '0-30'
            elif days_old <= 60: bucket = '31-60'
            elif days_old <= 90: bucket = '61-90'
            elif days_old <= 180: bucket = '91-180'
            aging_buckets[bucket] += float(s.total - s.paid_amount)

    total_sales = 0
    total_discounts = 0
    total_paid = 0
    total_advances_applied = 0
    
    for e in events:
        debit = e['debit']
        credit = e['credit']
        running_balance += (debit - credit)
        # If balance goes below zero, clamp to 0. The excess credit is treated as
        # a pending advance and must NOT carry over to subtract from future invoices.
        if running_balance < 0:
            running_balance = 0
        
        if e['type'] == 'sale':
            total_sales += debit
        elif e['type'] == 'discount':
            total_discounts += credit
        elif e['type'] == 'payment':
            total_paid += credit
        elif e['type'] == 'advance_applied':
            total_advances_applied += credit
            
        ledger_entries.append({
            'date': e['date'].strftime('%d/%m/%Y'),
            'invoice_number': e['desc'],
            'customer_name': customer.name,
            'debit': debit,
            'credit': credit,
            'invoice_balance': 0,
            'running_balance': max(0, running_balance),
            'status': e['status'],
            'flag_color': e.get('flag_color', ''),
            'days_overdue': e.get('days_overdue', 0),
            'aging': '-' if e['type'] != 'sale' else '-',
            'entity_id': e['obj'].id if 'obj' in e and hasattr(e['obj'], 'id') else None,
            'entity_type': e['type'] if e['type'] in ['sale', 'return', 'advance', 'payment', 'advance_applied'] else None
        })

    summary = {
        'total_sales': total_sales,
        'total_discounts': total_discounts,
        'total_advances': float(customer.total_advances_received or 0),
        'total_advances_applied': total_advances_applied,
        'remaining_advance': float(customer.remaining_advance_balance or 0),
        'total_paid': total_paid,
        'total_outstanding': max(0, running_balance),
        'aging_buckets': aging_buckets
    }
    
    return jsonify({
        'customer_name': customer.name,
        'customer_phone': customer.phone,
        'customer_health': customer.health_status,
        'access_token': customer.valid_access_token,
        'entries': ledger_entries,
        'summary': summary
    })

@bp.route('/customer/<int:id>/ledger-excel')
@login_required
def customer_ledger_excel(id):
    customer = Customer.query.get_or_404(id)
    
    # Filter selected IDs if provided
    selected_ids = request.args.get('ids', '').split(',') if request.args.get('ids') else []
    
    # Opening balance
    opening_balance = float(customer.opening_balance or 0)
    
    sales = Sale.query.filter_by(customer_id=id).order_by(Sale.date.asc()).all()
    advances = CustomerAdvance.query.filter_by(customer_id=id).order_by(CustomerAdvance.date.asc()).all()
    returns = SaleReturn.query.filter_by(customer_id=id).order_by(SaleReturn.date.asc()).all()
    
    sale_ids = [s.id for s in sales]
    payments = Payment.query.filter(Payment.invoice_id.in_(sale_ids)).order_by(Payment.date.asc()).all() if sale_ids else []
    
    events = []
    for s in sales:
        events.append({
            'date': s.date,
            'type': 'sale',
            'desc': f"Invoice #{s.invoice_number}",
            'inv': s.invoice_number,
            'debit': float(s.subtotal or 0) + float(s.tax or 0) + float(s.delivery_charge or 0),
            'credit': 0,
            'status': (s.status or 'unpaid').capitalize()
        })
        # Apply overdue restriction logic
        is_restricted = getattr(s, 'is_discount_restricted', False)
        effective_discount = getattr(s, 'effective_discount_amount', 0)
        
        if is_restricted and getattr(s, 'base_discount_amount', 0) > 0:
            events.append({
                'date': s.date,
                'type': 'discount',
                'desc': f"Discount (Inv #{s.invoice_number}): Reversal due to overdue",
                'inv': s.invoice_number,
                'debit': 0,
                'credit': 0,
                'status': '-'
            })
        elif effective_discount > 0:
            events.append({
                'date': s.date,
                'type': 'discount',
                'desc': f"Discount (Inv #{s.invoice_number})",
                'inv': s.invoice_number,
                'debit': 0,
                'credit': float(effective_discount or 0),
                'status': '-'
            })
            
        # Advance Applied (Credit)
        if hasattr(s, 'advance_applied') and s.advance_applied and s.advance_applied > 0:
            events.append({
                'date': s.date,
                'type': 'advance_applied',
                'desc': f"Advance Applied (Inv #{s.invoice_number})",
                'inv': s.invoice_number,
                'debit': 0,
                'credit': float(s.advance_applied or 0),
                'status': '-'
            })
            
    for p in payments:
        events.append({
            'date': p.date,
            'type': 'payment',
            'desc': f"Payment ({p.method or 'Cash'})",
            'inv': p.invoice.invoice_number if p.invoice else '-',
            'debit': 0,
            'credit': float(p.amount or 0),
            'status': '-'
        })
        
    for a in advances:
        dt = datetime.combine(a.date, datetime.min.time())
        events.append({
            'date': dt,
            'type': 'advance',
            'desc': f"Advance Received: {a.description or ''}",
            'inv': '-',
            'debit': 0,
            # Informational only — does not affect balance
            'credit': 0,
            'status': '-'
        })
        
    for r in returns:
        events.append({
            'date': r.date,
            'type': 'return',
            'desc': f"Sales Return #{r.return_number}",
            'inv': r.return_number,
            'debit': 0,
            'credit': float(r.total or 0),
            'status': (r.status or 'completed').capitalize()
        })
        
    events.sort(key=lambda x: x['date'])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer Ledger"
    
    # Headers
    headers = ['Date', 'Description', 'Invoice #', 'Debit', 'Credit', 'Balance']
    ws.append(headers)
    
    # Styles
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
    bold_font = Font(bold=True)
    
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Opening Balance Row
    running_balance = opening_balance
    ws.append([
        '-', 
        'Opening Balance', 
        '-', 
        opening_balance if opening_balance > 0 else 0, 
        abs(opening_balance) if opening_balance < 0 else 0, 
        running_balance
    ])
    ws.cell(row=2, column=6).font = bold_font
    
    total_sales_sum = 0
    total_discounts_sum = 0
    total_paid_sum = 0
    total_advances_applied_sum = 0
    total_advances_sum = float(customer.total_advances_received or 0)

    # Data Rows
    row_idx = 3
    for e in events:
        if selected_ids and e['inv'] not in selected_ids and e['inv'] != '-':
            continue
            
        debit = e['debit']
        credit = e['credit']
        running_balance += (debit - credit)
        
        if e['type'] == 'sale': total_sales_sum += debit
        elif e['type'] == 'discount': total_discounts_sum += credit
        elif e['type'] == 'payment': total_paid_sum += credit
        elif e['type'] == 'advance_applied': 
            total_paid_sum += credit
            total_advances_applied_sum += credit
        
        d_str = e['date'].strftime('%d/%m/%Y') if hasattr(e['date'], 'strftime') else str(e['date'])
        
        ws.append([
            d_str,
            e['desc'],
            e['inv'],
            debit,
            credit,
            running_balance
        ])
        ws.cell(row=row_idx, column=6).font = bold_font
        row_idx += 1

    # Summary Section
    ws.append([]) # Empty row
    row_idx += 1
    
    summary_rows = [
        ("Total Sales", total_sales_sum),
        ("Total Discounts", total_discounts_sum),
        ("Total Advances Received", total_advances_sum),
        ("Total Advances Applied", total_advances_applied_sum),
        ("Remaining Advance Balance", float(customer.remaining_advance_balance or 0)),
        ("Total Paid/Applied", total_paid_sum),
        ("Outstanding Balance", running_balance)
    ]
    
    for label, val in summary_rows:
        ws.append(["", "", "", "", label, val])
        ws.cell(row=row_idx, column=5).font = bold_font
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=row_idx, column=6).font = bold_font
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=row_idx, column=6).number_format = '"PKR "#,##0.00'
        row_idx += 1
        
    # Column Widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Ledger_{customer.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename)

@bp.route('/recalculate-all', methods=['POST'])
@login_required
@permission_required('sales', action='edit')
def recalculate_all():
    """Safely recalculate all invoice totals, payments, and statuses without manual data entry"""
    try:
        # 1. Recalculate Item Totals to ensure consistency
        all_items = SaleItem.query.all()
        for item in all_items:
            item.total = (item.quantity * item.unit_price) + (item.delivery_fee or 0)
        
        # 2. Recalculate Sales Return Totals
        all_returns = SaleReturn.query.all()
        for ret in all_returns:
            ret.calculate_totals()
            
        # 3. Recalculate Invoices
        sales = Sale.query.all()
        recalculated_count = 0
        
        for sale in sales:
            # Recalculate Original Invoice Totals (Subtotal, Tax, Delivery, Calculated Discount)
            # This handles the overdue discount restriction logic internally
            sale.calculate_totals()
            
            # Subtract Return Totals
            returns_total = 0
            if sale.returns:
                # Include 'pending' returns as the system traditionally subtracts them immediately
                returns_total = sum(ret.total for ret in sale.returns if ret.status in ['approved', 'completed', 'pending'])
            
            sale.total = max(0, sale.total - returns_total)
            
            # Synchronize Paid Amount from approved Payments and manual Advances
            approved_payments_sum = db.session.query(func.sum(Payment.amount)).filter(
                Payment.invoice_id == sale.id,
                Payment.is_approved == True
            ).scalar() or 0
            
            sale.paid_amount = approved_payments_sum + (sale.advance_applied or 0)
            
            # Safety cap: if returns reduced the total below what was already paid
            if sale.paid_amount > sale.total:
                sale.paid_amount = sale.total
                
            # Update Payment Status (Paid, Partial, Unpaid)
            sale.update_status()
            
            recalculated_count += 1
            
        db.session.commit()
        
        log_activity('Sales', 'Bulk Recalculation', f'Successfully recalculated {recalculated_count} invoices and synced with returns/payments.')
        flash(f'Audit complete: {recalculated_count} invoices have been successfully recalculated and synchronized.', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Recalculation Error: {str(e)}")
        flash(f'Recalculation failed: {str(e)}', 'danger')
        
    return redirect(url_for('sales.invoices'))
