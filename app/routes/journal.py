"""Journal Entry module.

A small, self-contained double-entry style bookkeeping tool:
  • Create named accounts (Cash, Bank, an expense head, …), with an opening balance.
  • Add money to an account (a quick one-line entry).
  • Record a full journal entry with a date and one-or-more lines, each line
    posting a debit or credit against an account.
  • View each account's ledger with a running balance.

It uses its own tables (journal_accounts / journal_entries / journal_lines) and
does not touch any other module.
"""
from datetime import date, datetime

from flask import (
    Blueprint, render_template, request, redirect, url_for, flash, jsonify
)
from flask_login import login_required, current_user

from app import db
from app.models import JournalAccount, JournalEntry, JournalLine

bp = Blueprint('journal', __name__)


def _save_line_bill_image(file_storage):
    """Save an uploaded per-line bill image under app/static/uploads/bills/
    and return the project-root-relative path to store on JournalLine."""
    if not file_storage or not file_storage.filename:
        return None
    import os
    import time
    import uuid
    from werkzeug.utils import secure_filename

    original_filename = secure_filename(file_storage.filename)
    unique_prefix = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
    filename = f"{unique_prefix}_{original_filename}"

    full_path = os.path.join('app', 'static', 'uploads', 'bills', filename)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    file_storage.save(full_path)
    return f"app/static/uploads/bills/{filename}"


# ─── Permission helpers ────────────────────────────────────────────────────────

def _can_view():
    return current_user.is_admin or getattr(current_user, 'can_view_accounting', False)


def _guard():
    if not _can_view():
        flash('You do not have permission to access the Journal module.', 'danger')
        return redirect(url_for('dashboard.index'))
    return None


def _admin_guard():
    """Ensures the current user is an admin. Use for create/edit/delete routes."""
    if current_user.role != 'admin' and not getattr(current_user, 'is_admin', False):
        flash('Only admins can perform this action.', 'danger')
        return redirect(url_for('journal.entries'))
    return None


def _parse_date(s):
    if not s:
        return date.today()
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return date.today()


def _date_or_none(s):
    """Parse a YYYY-MM-DD string, returning None for blank/invalid input."""
    if not s:
        return None
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return None


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# ─── Accounts ──────────────────────────────────────────────────────────────────

@bp.route('/accounts')
@login_required
def accounts():
    guard = _guard()
    if guard:
        return guard
    admin_guard = _admin_guard()
    if admin_guard:
        return admin_guard
    accounts = JournalAccount.query.order_by(JournalAccount.name).all()
    total_balance = sum(a.balance for a in accounts)
    return render_template('journal/accounts.html', accounts=accounts,
                           total_balance=total_balance)


@bp.route('/accounts/create', methods=['GET', 'POST'])
@login_required
def create_account():
    guard = _guard()
    if guard:
        return guard
    admin_guard = _admin_guard()
    if admin_guard:
        return admin_guard

    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        if not name:
            flash('Account name is required.', 'warning')
            return redirect(url_for('journal.create_account'))
        if JournalAccount.query.filter(db.func.lower(JournalAccount.name) == name.lower()).first():
            flash(f'An account named "{name}" already exists.', 'warning')
            return redirect(url_for('journal.create_account'))

        acct = JournalAccount(
            name=name,
            account_type=(request.form.get('account_type') or '').strip() or None,
            opening_balance=_to_float(request.form.get('opening_balance')),
            notes=(request.form.get('notes') or '').strip() or None,
            created_by=current_user.id,
        )
        db.session.add(acct)
        db.session.commit()
        flash(f'Account "{name}" created.', 'success')
        return redirect(url_for('journal.accounts'))

    return render_template('journal/account_form.html', account=None)


@bp.route('/accounts/<int:account_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_account(account_id):
    guard = _guard()
    if guard:
        return guard
    admin_guard = _admin_guard()
    if admin_guard:
        return admin_guard
    acct = JournalAccount.query.get_or_404(account_id)

    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        if not name:
            flash('Account name is required.', 'warning')
            return redirect(url_for('journal.edit_account', account_id=acct.id))
        dupe = JournalAccount.query.filter(
            db.func.lower(JournalAccount.name) == name.lower(),
            JournalAccount.id != acct.id
        ).first()
        if dupe:
            flash(f'Another account named "{name}" already exists.', 'warning')
            return redirect(url_for('journal.edit_account', account_id=acct.id))

        acct.name = name
        acct.account_type = (request.form.get('account_type') or '').strip() or None
        acct.opening_balance = _to_float(request.form.get('opening_balance'))
        acct.notes = (request.form.get('notes') or '').strip() or None
        acct.is_active = bool(request.form.get('is_active'))
        db.session.commit()
        flash('Account updated.', 'success')
        return redirect(url_for('journal.accounts'))

    return render_template('journal/account_form.html', account=acct)


@bp.route('/accounts/<int:account_id>/delete', methods=['POST'])
@login_required
def delete_account(account_id):
    guard = _guard()
    if guard:
        return guard
    admin_guard = _admin_guard()
    if admin_guard:
        return admin_guard
    acct = JournalAccount.query.get_or_404(account_id)
    if acct.lines:
        flash('Cannot delete an account that has journal lines. Deactivate it instead.', 'warning')
        return redirect(url_for('journal.accounts'))
    db.session.delete(acct)
    db.session.commit()
    flash('Account deleted.', 'success')
    return redirect(url_for('journal.accounts'))


@bp.route('/accounts/<int:account_id>/ledger')
@login_required
def account_ledger(account_id):
    guard = _guard()
    if guard:
        return guard
    acct = JournalAccount.query.get_or_404(account_id)

    # All lines for this account, oldest first, with a running balance.
    lines = (JournalLine.query
             .filter_by(account_id=acct.id)
             .join(JournalEntry)
             .order_by(JournalEntry.date, JournalLine.id)
             .all())

    running = acct.opening_balance or 0
    rows = []
    for ln in lines:
        delta = (ln.amount or 0) if ln.entry_type == 'debit' else -(ln.amount or 0)
        running += delta
        rows.append({'line': ln, 'running': running})

    return render_template('journal/ledger.html', account=acct, rows=rows)


# ─── Quick "Add Money" to an account ───────────────────────────────────────────

def _create_add_money_entry(acct, amount, entry_type, date_str, reference, description, bill_file):
    """Shared by both the full add-money page and the quick modal on the
    entries list — creates a pending-approval JournalEntry + single line
    (auto-approved when an admin submits it), same as any other entry."""
    bill_path = _save_line_bill_image(bill_file)
    is_admin = (getattr(current_user, 'role', '') == 'admin') or getattr(current_user, 'is_admin', False)
    entry = JournalEntry(
        date=_parse_date(date_str),
        reference=(reference or '').strip() or None,
        description=(description or '').strip() or None,
        created_by=current_user.id,
        is_approved=is_admin,
        is_rejected=False,
        approved_by=current_user.id if is_admin else None,
        approved_at=datetime.utcnow() if is_admin else None,
    )
    entry.lines.append(JournalLine(
        account_id=acct.id,
        description=(description or '').strip() or None,
        entry_type=entry_type,
        amount=amount,
        bill_image_path=bill_path,
    ))
    db.session.add(entry)
    db.session.commit()
    return entry, is_admin


@bp.route('/accounts/<int:account_id>/add-money', methods=['GET', 'POST'])
@login_required
def add_money(account_id):
    guard = _guard()
    if guard:
        return guard
    acct = JournalAccount.query.get_or_404(account_id)

    if request.method == 'POST':
        amount = _to_float(request.form.get('amount'))
        if amount <= 0:
            flash('Enter an amount greater than zero.', 'warning')
            return redirect(url_for('journal.add_money', account_id=acct.id))

        entry_type = request.form.get('entry_type', 'debit')
        if entry_type not in ('debit', 'credit'):
            entry_type = 'debit'

        entry, is_admin = _create_add_money_entry(
            acct, amount, entry_type, request.form.get('date'),
            request.form.get('reference'), request.form.get('description'),
            request.files.get('bill_image'),
        )
        verb = 'added to' if entry_type == 'debit' else 'taken out of'
        if is_admin:
            flash(f'PKR {amount:,.0f} {verb} "{acct.name}".', 'success')
        else:
            flash(f'Entry of PKR {amount:,.0f} {verb} "{acct.name}" submitted for Admin approval.', 'info')
        return redirect(url_for('journal.account_ledger', account_id=acct.id))

    return render_template('journal/add_money.html', account=acct,
                           today=date.today())


@bp.route('/accounts/add-money-quick', methods=['POST'])
@login_required
def add_money_quick():
    """AJAX endpoint for the "Add Money" modal on the Journal Entries list
    page — same logic as add_money() but takes the account from the posted
    form (a dropdown in the modal) and replies with JSON instead of a
    redirect, so the staff member never leaves /journal/entries."""
    if not _can_view():
        return jsonify({'success': False, 'message': 'You do not have permission to access the Journal module.'}), 403

    acct = JournalAccount.query.get(request.form.get('account_id', type=int))
    if not acct:
        return jsonify({'success': False, 'message': 'Please select a valid account.'})

    amount = _to_float(request.form.get('amount'))
    if amount <= 0:
        return jsonify({'success': False, 'message': 'Enter an amount greater than zero.'})

    entry_type = request.form.get('entry_type', 'debit')
    if entry_type not in ('debit', 'credit'):
        entry_type = 'debit'

    entry, is_admin = _create_add_money_entry(
        acct, amount, entry_type, request.form.get('date'),
        request.form.get('reference'), request.form.get('description'),
        request.files.get('bill_image'),
    )
    verb = 'added to' if entry_type == 'debit' else 'taken out of'
    if is_admin:
        message = f'PKR {amount:,.0f} {verb} "{acct.name}".'
    else:
        message = f'Entry of PKR {amount:,.0f} {verb} "{acct.name}" submitted for Admin approval.'

    return jsonify({'success': True, 'message': message, 'entry_id': entry.id})


# ─── Journal Entries ───────────────────────────────────────────────────────────

def _filtered_entries(args, newest_first=True):
    """Apply the account/date/category filters (shared by the list page and the
    exports) and return (entries, account_id, date_from, date_to, category_id)."""
    account_id  = args.get('account_id',  type=int)
    category_id = args.get('category_id', type=int)
    date_from   = _date_or_none(args.get('date_from'))
    date_to     = _date_or_none(args.get('date_to'))

    query = JournalEntry.query
    if account_id:
        query = query.filter(JournalEntry.lines.any(JournalLine.account_id == account_id))
    if category_id:
        query = query.filter(JournalEntry.lines.any(JournalLine.category_id == category_id))
    if date_from:
        query = query.filter(JournalEntry.date >= date_from)
    if date_to:
        query = query.filter(JournalEntry.date <= date_to)

    if newest_first:
        query = query.order_by(JournalEntry.date.desc(), JournalEntry.id.desc())
    else:
        query = query.order_by(JournalEntry.date.asc(), JournalEntry.id.asc())
    return query.all(), account_id, date_from, date_to, category_id


def _grouped_by_account(entries, account_id=None):
    """Group the lines of the given entries by account (oldest first within
    each account). When account_id is set only that account's lines are kept.
    Returns a list of dicts sorted by account name."""
    from collections import OrderedDict
    groups = OrderedDict()
    for e in sorted(entries, key=lambda x: (x.date, x.id)):
        for l in e.lines:
            if account_id and l.account_id != account_id:
                continue
            if l.account_id not in groups:
                groups[l.account_id] = {'account': l.account, 'lines': []}
            groups[l.account_id]['lines'].append((e, l))
    result = list(groups.values())
    result.sort(key=lambda g: (g['account'].name.lower() if g['account'] else ''))
    return result


@bp.route('/')
@bp.route('/entries')
@login_required
def entries():
    guard = _guard()
    if guard:
        return guard

    all_entries, account_id, date_from, date_to, category_id = _filtered_entries(request.args)

    accounts = JournalAccount.query.order_by(JournalAccount.name).all()
    from app.models import ExpenseCategory
    expense_categories = (ExpenseCategory.query
                          .filter_by(is_active=True)
                          .order_by(ExpenseCategory.name).all())

    # ── Summary stats for the top cards ────────────────────────────────────────
    grand_debit  = sum(e.total_debit  for e in all_entries)
    grand_credit = sum(e.total_credit for e in all_entries)
    # Total remaining bank balance = sum of all active account balances
    bank_balance = sum(a.balance for a in accounts if a.is_active)

    return render_template('journal/entries.html', entries=all_entries,
                           accounts=accounts,
                           expense_categories=expense_categories,
                           f_account_id=account_id,
                           f_category_id=category_id,
                           f_date_from=request.args.get('date_from', ''),
                           f_date_to=request.args.get('date_to', ''),
                           grand_debit=grand_debit,
                           grand_credit=grand_credit,
                           bank_balance=bank_balance)


# ─── Account-wise exports (Excel / PDF), grouped by account ────────────────────

def _filter_caption(date_from, date_to):
    if date_from and date_to:
        return f"{date_from.strftime('%d-%m-%Y')} to {date_to.strftime('%d-%m-%Y')}"
    if date_from:
        return f"From {date_from.strftime('%d-%m-%Y')}"
    if date_to:
        return f"Up to {date_to.strftime('%d-%m-%Y')}"
    return "All dates"


@bp.route('/entries/export/excel')
@login_required
def export_excel():
    guard = _guard()
    if guard:
        return guard

    import io
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    entries, account_id, date_from, date_to, category_id = _filtered_entries(request.args, newest_first=False)
    groups = _grouped_by_account(entries, account_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Journal by Account'

    # Styles
    title_font = Font(bold=True, size=14)
    acct_font = Font(bold=True, size=12, color='FFFFFF')
    acct_fill = PatternFill('solid', fgColor='0D6EFD')
    head_font = Font(bold=True)
    head_fill = PatternFill('solid', fgColor='E9ECEF')
    total_font = Font(bold=True)
    total_fill = PatternFill('solid', fgColor='F1F3F5')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal='right')

    headers = ['Date', 'Entry #', 'Reference', 'Description', 'Debit (In)', 'Credit (Out)', 'Balance']
    NCOL = len(headers)

    ws['A1'] = 'Journal Entries — Account-wise Ledger'
    ws['A1'].font = title_font
    ws['A2'] = f'Filter: {_filter_caption(date_from, date_to)}'
    ws['A2'].font = Font(italic=True, color='6C757D')
    r = 4

    grand_debit = grand_credit = grand_balance = 0
    for g in groups:
        acct = g['account']
        opening = acct.opening_balance or 0
        # Account header row (name) — spans the table width.
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
        cell = ws.cell(row=r, column=1,
                       value=f"{acct.name}"
                             + (f"  ({acct.account_type})" if acct and acct.account_type else ''))
        cell.font = acct_font
        cell.fill = acct_fill
        r += 1

        # Column headers
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=r, column=ci, value=h)
            c.font = head_font
            c.fill = head_fill
            c.border = border
            if ci >= 5:
                c.alignment = right
        r += 1

        # Opening balance row (amount shown on top of the account).
        running = opening
        c = ws.cell(row=r, column=4, value='Opening Balance')
        c.font = total_font; c.alignment = right; c.border = border
        c = ws.cell(row=r, column=7, value=running)
        c.font = total_font; c.alignment = right; c.border = border; c.number_format = '#,##0'
        for ci in (1, 2, 3, 5, 6):
            ws.cell(row=r, column=ci).border = border
        r += 1

        sub_debit = sub_credit = 0
        for e, l in g['lines']:
            debit = l.amount if l.entry_type == 'debit' else 0
            credit = l.amount if l.entry_type == 'credit' else 0
            # Debit adds (+), credit subtracts (−) from the running balance.
            running += debit - credit
            sub_debit += debit
            sub_credit += credit
            row_vals = [
                e.date.strftime('%d-%m-%Y') if e.date else '',
                f'#{e.id}',
                e.reference or '',
                l.description or (e.description or ''),
                debit or '',
                credit or '',
                running,
            ]
            for ci, v in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=ci, value=v)
                c.border = border
                if ci >= 5:
                    c.alignment = right
                    if v != '':
                        c.number_format = '#,##0'
            r += 1

        # Remaining balance row for this account (running == closing balance).
        c = ws.cell(row=r, column=4, value='Remaining Balance')
        c.font = total_font; c.fill = total_fill; c.alignment = right; c.border = border
        for ci, v in [(5, sub_debit), (6, sub_credit), (7, running)]:
            c = ws.cell(row=r, column=ci, value=v)
            c.font = total_font; c.fill = total_fill; c.alignment = right
            c.border = border; c.number_format = '#,##0'
        for ci in (1, 2, 3):
            ws.cell(row=r, column=ci).fill = total_fill
            ws.cell(row=r, column=ci).border = border
        grand_debit += sub_debit
        grand_credit += sub_credit
        grand_balance += running
        r += 2  # blank separator row between accounts

    if not groups:
        ws.cell(row=r, column=1, value='No entries for the selected filter.')
    else:
        # Grand total (total remaining across all accounts shown)
        c = ws.cell(row=r, column=4, value='GRAND TOTAL')
        c.font = Font(bold=True, size=12); c.alignment = right
        for ci, v in [(5, grand_debit), (6, grand_credit), (7, grand_balance)]:
            c = ws.cell(row=r, column=ci, value=v)
            c.font = Font(bold=True, size=12); c.alignment = right; c.number_format = '#,##0'

    widths = [14, 10, 18, 38, 14, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='journal_by_account.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.route('/entries/export/pdf')
@login_required
def export_pdf():
    guard = _guard()
    if guard:
        return guard

    import io
    from flask import send_file
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)

    entries, account_id, date_from, date_to, category_id = _filtered_entries(request.args, newest_first=False)
    groups = _grouped_by_account(entries, account_id)

    styles = getSampleStyleSheet()
    st_title = ParagraphStyle('t', parent=styles['Normal'], fontSize=15, leading=18, fontName='Helvetica-Bold')
    st_sub = ParagraphStyle('s', parent=styles['Normal'], fontSize=9, leading=12, textColor=colors.HexColor('#6c757d'))
    st_acct = ParagraphStyle('a', parent=styles['Normal'], fontSize=11, leading=14,
                             fontName='Helvetica-Bold', textColor=colors.white)
    st_cell = ParagraphStyle('c', parent=styles['Normal'], fontSize=8.5, leading=11)

    CONTENT_W = A4[0] - 72
    col_w = [CONTENT_W * x for x in (0.12, 0.08, 0.13, 0.29, 0.125, 0.125, 0.13)]

    def money(v):
        return f'{v:,.0f}' if v else ''

    def net_money(v):
        return f'{v:,.0f}' if v else '0'

    elements = [Paragraph('Journal Entries — Account-wise Ledger', st_title),
                Paragraph(f'Filter: {_filter_caption(date_from, date_to)}', st_sub),
                Spacer(1, 8)]

    grand_debit = grand_credit = grand_balance = 0
    for g in groups:
        acct = g['account']
        opening = acct.opening_balance or 0
        # Account banner (name only)
        banner = Table([[Paragraph(
            f"{acct.name}" + (f" ({acct.account_type})" if acct.account_type else ''), st_acct)]],
            colWidths=[CONTENT_W])
        banner.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#0d6efd')),
            ('LEFTPADDING', (0, 0), (-1, -1), 8), ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(banner)

        data = [[Paragraph(f'<b>{h}</b>', st_cell) for h in
                 ['Date', 'Entry #', 'Reference', 'Description', 'Debit (In)', 'Credit (Out)', 'Balance']]]
        # Opening balance row
        running = opening
        data.append([Paragraph('', st_cell), Paragraph('', st_cell), Paragraph('', st_cell),
                     Paragraph('<b>Opening Balance</b>', st_cell), Paragraph('', st_cell),
                     Paragraph('', st_cell), Paragraph(f'<b>{net_money(running)}</b>', st_cell)])
        sub_debit = sub_credit = 0
        for e, l in g['lines']:
            debit = l.amount if l.entry_type == 'debit' else 0
            credit = l.amount if l.entry_type == 'credit' else 0
            running += debit - credit  # debit adds, credit subtracts
            sub_debit += debit; sub_credit += credit
            data.append([
                Paragraph(e.date.strftime('%d-%m-%Y') if e.date else '', st_cell),
                Paragraph(f'#{e.id}', st_cell),
                Paragraph(e.reference or '', st_cell),
                Paragraph(l.description or (e.description or ''), st_cell),
                Paragraph(money(debit), st_cell),
                Paragraph(money(credit), st_cell),
                Paragraph(net_money(running), st_cell),
            ])
        data.append([Paragraph('', st_cell), Paragraph('', st_cell), Paragraph('', st_cell),
                     Paragraph('<b>Remaining Balance</b>', st_cell),
                     Paragraph(f'<b>{money(sub_debit)}</b>', st_cell),
                     Paragraph(f'<b>{money(sub_credit)}</b>', st_cell),
                     Paragraph(f'<b>{net_money(running)}</b>', st_cell)])
        grand_debit += sub_debit; grand_credit += sub_credit; grand_balance += running

        t = Table(data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e9ecef')),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f8f9fa')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f1f3f5')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (4, 0), (6, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 5), ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 10))

    if not groups:
        elements.append(Paragraph('No entries for the selected filter.', st_sub))
    else:
        gt = Table([[Paragraph('<b>GRAND TOTAL</b>', st_cell),
                     Paragraph(f'<b>{money(grand_debit)}</b>', st_cell),
                     Paragraph(f'<b>{money(grand_credit)}</b>', st_cell),
                     Paragraph(f'<b>{net_money(grand_balance)}</b>', st_cell)]],
                   colWidths=[CONTENT_W * 0.49, CONTENT_W * 0.17, CONTENT_W * 0.17, CONTENT_W * 0.17])
        gt.setStyle(TableStyle([
            ('ALIGN', (1, 0), (3, 0), 'RIGHT'),
            ('LINEABOVE', (0, 0), (-1, 0), 1, colors.HexColor('#0d6efd')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(gt)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=36, bottomMargin=36,
                            leftMargin=36, rightMargin=36, title='Journal by Account')
    doc.build(elements)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='journal_by_account.pdf', mimetype='application/pdf')


@bp.route('/entries/create', methods=['GET', 'POST'])
@login_required
def create_entry():
    guard = _guard()
    if guard:
        return guard

    accounts = JournalAccount.query.filter_by(is_active=True).order_by(JournalAccount.name).all()
    from app.models import ExpenseCategory
    expense_categories = ExpenseCategory.query.filter_by(is_active=True).order_by(ExpenseCategory.name).all()

    if request.method == 'POST':
        account_ids = request.form.getlist('line_account_id[]')
        category_ids = request.form.getlist('line_category_id[]')
        descriptions = request.form.getlist('line_description[]')
        types = request.form.getlist('line_type[]')
        amounts = request.form.getlist('line_amount[]')
        bill_images = request.files.getlist('line_bill_image[]')

        is_admin = (getattr(current_user, 'role', '') == 'admin') or getattr(current_user, 'is_admin', False)
        entry = JournalEntry(
            date=_parse_date(request.form.get('date')),
            reference=(request.form.get('reference') or '').strip() or None,
            description=(request.form.get('description') or '').strip() or None,
            created_by=current_user.id,
            is_approved=is_admin,
            is_rejected=False,
            approved_by=current_user.id if is_admin else None,
            approved_at=datetime.utcnow() if is_admin else None,
        )

        added = 0
        for i in range(len(account_ids)):
            acct_id = account_ids[i]
            amount = _to_float(amounts[i]) if i < len(amounts) else 0
            if not acct_id or amount <= 0:
                continue
            etype = types[i] if i < len(types) else 'debit'
            if etype not in ('debit', 'credit'):
                etype = 'debit'
            cat_id = category_ids[i] if i < len(category_ids) else None
            cat_id_int = int(cat_id) if (cat_id and str(cat_id).isdigit()) else None
            bill_file = bill_images[i] if i < len(bill_images) else None
            bill_path = _save_line_bill_image(bill_file)

            entry.lines.append(JournalLine(
                account_id=int(acct_id),
                category_id=cat_id_int,
                description=(descriptions[i].strip() if i < len(descriptions) and descriptions[i] else None),
                entry_type=etype,
                amount=amount,
                bill_image_path=bill_path,
            ))
            added += 1

        if added == 0:
            flash('Add at least one line with an account and amount.', 'warning')
            return render_template('journal/entry_form.html', accounts=accounts,
                                   expense_categories=expense_categories,
                                   entry=None, today=date.today())

        db.session.add(entry)
        db.session.commit()
        if is_admin:
            flash(f'Journal entry saved with {added} line(s).', 'success')
        else:
            flash(f'Journal entry with {added} line(s) submitted for Admin approval.', 'info')
        return redirect(url_for('journal.entry_detail', entry_id=entry.id))

    return render_template('journal/entry_form.html', accounts=accounts,
                           expense_categories=expense_categories,
                           entry=None, today=date.today())


@bp.route('/entries/<int:entry_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_entry(entry_id):
    guard = _guard()
    if guard:
        return guard
    if not (current_user.is_admin or getattr(current_user, 'can_edit_accounting', False)):
        flash('You do not have permission to edit journal entries.', 'danger')
        return redirect(url_for('journal.entries'))

    entry = JournalEntry.query.get_or_404(entry_id)
    accounts = JournalAccount.query.filter_by(is_active=True).order_by(JournalAccount.name).all()
    from app.models import ExpenseCategory
    expense_categories = ExpenseCategory.query.filter_by(is_active=True).order_by(ExpenseCategory.name).all()

    if request.method == 'POST':
        account_ids = request.form.getlist('line_account_id[]')
        category_ids = request.form.getlist('line_category_id[]')
        descriptions = request.form.getlist('line_description[]')
        types = request.form.getlist('line_type[]')
        amounts = request.form.getlist('line_amount[]')
        bill_images = request.files.getlist('line_bill_image[]')
        bill_images_existing = request.form.getlist('line_bill_image_existing[]')

        entry.date = _parse_date(request.form.get('date'))
        entry.reference = (request.form.get('reference') or '').strip() or None
        entry.description = (request.form.get('description') or '').strip() or None

        # Rebuild lines
        JournalLine.query.filter_by(entry_id=entry.id).delete()

        added = 0
        for i in range(len(account_ids)):
            acct_id = account_ids[i]
            amount = _to_float(amounts[i]) if i < len(amounts) else 0
            if not acct_id or amount <= 0:
                continue
            etype = types[i] if i < len(types) else 'debit'
            if etype not in ('debit', 'credit'):
                etype = 'debit'
            cat_id = category_ids[i] if i < len(category_ids) else None
            cat_id_int = int(cat_id) if (cat_id and str(cat_id).isdigit()) else None
            bill_file = bill_images[i] if i < len(bill_images) else None
            bill_path = _save_line_bill_image(bill_file)
            if not bill_path:
                # No new upload for this row — keep whatever image it already had.
                bill_path = bill_images_existing[i].strip() if i < len(bill_images_existing) and bill_images_existing[i] else None

            entry.lines.append(JournalLine(
                account_id=int(acct_id),
                category_id=cat_id_int,
                description=(descriptions[i].strip() if i < len(descriptions) and descriptions[i] else None),
                entry_type=etype,
                amount=amount,
                bill_image_path=bill_path,
            ))
            added += 1

        if added == 0:
            flash('Add at least one line with an account and amount.', 'warning')
            return render_template('journal/entry_form.html', accounts=accounts,
                                   expense_categories=expense_categories,
                                   entry=entry, today=entry.date or date.today())

        db.session.commit()
        flash('Journal entry updated successfully.', 'success')
        return redirect(url_for('journal.entry_detail', entry_id=entry.id))

    return render_template('journal/entry_form.html', accounts=accounts,
                           expense_categories=expense_categories,
                           entry=entry, today=entry.date or date.today())


@bp.route('/entries/<int:entry_id>')
@login_required
def entry_detail(entry_id):
    guard = _guard()
    if guard:
        return guard
    entry = JournalEntry.query.get_or_404(entry_id)
    return render_template('journal/entry_detail.html', entry=entry)


@bp.route('/entries/<int:entry_id>/send-to-expense', methods=['POST'])
@login_required
def send_entry_to_expense(entry_id):
    """Create an Expense record from a journal entry's credit (money-out) total.
    Mirrors the Delivering Voucher → Expense flow."""
    if not _can_view():
        return jsonify({'success': False, 'message': 'Permission denied.'}), 403

    entry = JournalEntry.query.get_or_404(entry_id)

    if not getattr(entry, 'is_approved', True):
        return jsonify({'success': False, 'message': 'Only approved entries can be sent to expense.'}), 400

    if entry.expense_id:
        return jsonify({'success': False, 'message': 'This entry has already been sent to expense.'})

    amount = entry.total_credit
    if amount <= 0:
        return jsonify({'success': False, 'message': 'Only entries with a credit (money-out) amount can be sent to expense.'})

    category_id = request.form.get('category_id')
    if not category_id:
        return jsonify({'success': False, 'message': 'Please select an expense category.'})

    from app.models import Expense, ExpenseCategory, ExpenseSettings
    from app.utils import log_activity

    category = ExpenseCategory.query.get(category_id)
    if not category:
        return jsonify({'success': False, 'message': 'Invalid expense category.'})

    try:
        from app.routes.accounting import get_unique_expense_number
        acc_settings = ExpenseSettings.query.first()
        if not acc_settings:
            acc_settings = ExpenseSettings()
            db.session.add(acc_settings)
            db.session.flush()

        exp_num, next_num = get_unique_expense_number(acc_settings, acc_settings.next_number)

        # Build a readable description from the entry's credit lines / narration.
        credit_accounts = ', '.join(sorted({
            l.account.name for l in entry.lines if l.entry_type == 'credit' and l.account
        }))
        desc = entry.description or credit_accounts or f'Journal Entry #{entry.id}'

        expense = Expense()
        expense.expense_number = exp_num
        expense.date = datetime.combine(entry.date, datetime.min.time()) if entry.date else datetime.utcnow()
        expense.category_id = int(category_id)
        expense.description = f'Journal Entry #{entry.id}: {desc}'
        expense.amount = amount
        expense.payment_method = 'Cash'
        expense.reference = entry.reference or f'JE-{entry.id}'
        expense.status = 'confirmed'
        expense.created_by = current_user.id

        # Carry over a bill image from the journal entry's lines so the user
        # doesn't have to upload it again on the Expense side.
        bill_image_path = next(
            (l.bill_image_path for l in entry.lines if l.entry_type == 'credit' and l.bill_image_path),
            None
        ) or next((l.bill_image_path for l in entry.lines if l.bill_image_path), None)
        expense.bill_image_path = bill_image_path

        db.session.add(expense)
        db.session.flush()

        entry.expense_id = expense.id
        acc_settings.next_number = next_num
        db.session.commit()

        log_activity('Journal', f'Sent Journal Entry #{entry.id} to Expense',
                     f'Category: {category.name}, Amount: {amount}')

        return jsonify({
            'success': True,
            'message': f'Expense #{exp_num} created under "{category.name}".',
            'expense_number': exp_num,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@bp.route('/entries/<int:entry_id>/delete', methods=['POST'])
@login_required
def delete_entry(entry_id):
    guard = _guard()
    if guard:
        return guard
    if not (current_user.is_admin or getattr(current_user, 'can_delete_accounting', False)):
        flash('You do not have permission to delete journal entries.', 'danger')
        return redirect(url_for('journal.entries'))
    entry = JournalEntry.query.get_or_404(entry_id)
    db.session.delete(entry)
    db.session.commit()
    flash('Journal entry deleted.', 'success')
    return redirect(url_for('journal.entries'))
