from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from app import db
from app.models import Product, User, Warehouse, ProductCategory
from app.forms import ProductForm
from sqlalchemy import func, inspect
from io import BytesIO
import os
from werkzeug.utils import secure_filename
from app.routes.filters import apply_saved_filter_to_query

bp = Blueprint('inventory', __name__)

def has_column(table_name, column_name):
    try:
        inspector = inspect(db.engine)
        return column_name in [c['name'] for c in inspector.get_columns(table_name)]
    except:
        return False

@bp.route('/products')
@login_required
def products():
    category = request.args.get('category', 'all')
    warehouse_id = request.args.get('warehouse_id', '', type=str)
    qty_min = request.args.get('qty_min', '', type=str)
    qty_max = request.args.get('qty_max', '', type=str)
    search_query = request.args.get('search', '', type=str).strip()
    
    db.session.expire_all()
    query = Product.query
    
    # Search filter
    if search_query:
        query = query.filter(
            (Product.name.ilike(f'%{search_query}%')) | 
            (Product.sku.ilike(f'%{search_query}%'))
        )
    
    if category != 'all':
        query = query.filter(Product.category_id == int(category))
    
    if warehouse_id and warehouse_id.isdigit():
        query = query.filter(Product.warehouse_id == int(warehouse_id))
    
    # Filter by quantity range
    if qty_min and qty_min.isdigit():
        query = query.filter(Product.quantity >= int(qty_min))
    
    if qty_max and qty_max.isdigit():
        query = query.filter(Product.quantity <= int(qty_max))
    
    query = apply_saved_filter_to_query(query, 'product', request.args)

    products = query.order_by(Product.name).all()
    
    # Get all categories for filter
    categories = ProductCategory.query.filter_by(is_active=True).order_by(ProductCategory.name).all()
    
    # Get all warehouses for filter
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    
    return render_template('inventory/products.html', 
                         products=products, 
                         categories=categories,
                         warehouses=warehouses,
                         current_category=category,
                         current_warehouse_id=warehouse_id,
                         qty_min=qty_min,
                         qty_max=qty_max,
                         search_query=search_query,
                         active_module='product',
                         filter_id=request.args.get('filter_id'))

@bp.route('/product/add', methods=['GET', 'POST'])
@login_required
def add_product():
    db.session.expire_all()
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    form = ProductForm()
    
    if request.method == 'POST':
        # Get form data directly from request
        name = request.form.get('name')
        sku = request.form.get('sku')
        description = request.form.get('description')
        unit_price = request.form.get('unit_price')
        cost_price = request.form.get('cost_price')
        quantity = request.form.get('quantity')
        reorder_level = request.form.get('reorder_level')
        category_id = request.form.get('category_id')
        is_manufactured = 'is_manufactured' in request.form
        finished_good_price = request.form.get('finished_good_price')
        
        # Validate required fields
        if name and sku and unit_price is not None:
            # Check if SKU already exists
            existing_product = Product.query.filter_by(sku=sku).first()
            if existing_product:
                flash(f'SKU "{sku}" already exists. Please use a different SKU.', 'error')
                return redirect(url_for('inventory.add_product'))
            
            warehouse_id = request.form.get('warehouse_id')
            # If finished good, use finished_good_price as the selling price
            if is_manufactured and finished_good_price:
                final_unit_price = float(finished_good_price)
            else:
                final_unit_price = float(unit_price)
            
            product = Product(
                name=name,
                sku=sku,
                description=description,
                unit_price=final_unit_price,
                cost_price=float(cost_price) if cost_price else 0.0,
                quantity=float(quantity) if quantity else 0,
                reorder_level=float(reorder_level) if reorder_level else 0,
                category_id=int(category_id) if category_id and category_id != '0' else None,
                warehouse_id=int(warehouse_id) if warehouse_id and warehouse_id != '0' else None
            )
            
            product.is_manufactured = is_manufactured
            product.finished_good_price = float(finished_good_price) if finished_good_price else None
            
            # Handle image upload
            if 'image' in request.files:
                image_file = request.files['image']
                if image_file and image_file.filename:
                    filename = secure_filename(image_file.filename)
                    image_path = os.path.join('app', 'static', 'uploads', 'products', filename)
                    os.makedirs(os.path.dirname(image_path), exist_ok=True)
                    image_file.save(image_path)
                    product.image_path = image_path.replace('\\', '/')
            
            try:
                db.session.add(product)
                db.session.commit()
                flash('Product added successfully!', 'success')
                return redirect(url_for('inventory.products'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error adding product: {str(e)}', 'error')
                return redirect(url_for('inventory.add_product'))
        else:
            flash('Please fill in all required fields.', 'error')
    
    # Fetch categories for the dropdown
    categories = ProductCategory.query.filter_by(is_active=True).order_by(ProductCategory.name).all()
    
    return render_template('inventory/add_product.html', form=form, categories=categories, warehouses=warehouses)


@bp.route('/product/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_product(id):
    product = Product.query.get_or_404(id)
    db.session.expire_all()
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    form = ProductForm(obj=product)
    
    if form.validate_on_submit():
        # Check if SKU is being changed and if the new SKU already exists
        if form.sku.data != product.sku:
            existing_product = Product.query.filter_by(sku=form.sku.data).first()
            if existing_product:
                flash(f'SKU "{form.sku.data}" already exists. Please use a different SKU.', 'error')
                return redirect(url_for('inventory.edit_product', id=product.id))
        
        # Store old cost for versioning check
        old_cost = product.cost_price
        
        warehouse_id = request.form.get('warehouse_id')
        product.name = form.name.data
        product.sku = form.sku.data
        product.description = form.description.data
        
        product.is_manufactured = form.is_manufactured.data if form.is_manufactured.data else False
        
        # Handle finished_good_price
        finished_good_price = request.form.get('finished_good_price')
        product.finished_good_price = float(finished_good_price) if finished_good_price else None
        
        # If finished good, use finished_good_price as the selling price
        if product.is_manufactured and product.finished_good_price:
            product.unit_price = product.finished_good_price
        else:
            product.unit_price = form.unit_price.data
        
        product.cost_price = form.cost_price.data if form.cost_price.data is not None else 0.0
        product.reorder_level = form.reorder_level.data
        category_id = request.form.get('category_id')
        product.category_id = int(category_id) if category_id and category_id != '0' else None
        product.warehouse_id = int(warehouse_id) if warehouse_id and warehouse_id != '0' else None
        
        # Handle quantity update
        quantity = request.form.get('quantity')
        if quantity is not None:
            try:
                product.quantity = float(quantity)
            except (ValueError, TypeError):
                pass
        
        # Handle image upload
        if 'image' in request.files:
            image_file = request.files['image']
            if image_file and image_file.filename:
                filename = secure_filename(image_file.filename)
                image_path = os.path.join('app', 'static', 'uploads', 'products', filename)
                os.makedirs(os.path.dirname(image_path), exist_ok=True)
                image_file.save(image_path)
                # Normalize path to use forward slashes for consistency
                product.image_path = image_path.replace('\\', '/')
        
        try:
            db.session.commit()
            print(f"\n[DEBUG] Product {product.id} ({product.name}) updated")
            print(f"[DEBUG] Old cost: {old_cost}, New cost: {product.cost_price}")
            print(f"[DEBUG] Costs equal? {old_cost == product.cost_price}")
            
            # Trigger BOM versioning if cost price changed
            if old_cost != product.cost_price:
                print(f"[DEBUG] Cost changed! Triggering BOM versioning...")
                from app.services.bom_versioning import BOMVersioningService
                try:
                    print(f"[DEBUG] Calling check_and_update_bom_for_cost_changes...")
                    # Use current_user.id if available, fallback to admin user
                    user_id = None
                    try:
                        if current_user and current_user.is_authenticated:
                            user_id = current_user.id
                    except (AttributeError, TypeError):
                        pass
                    
                    if user_id is None:
                        # Fallback to admin user
                        admin_user = User.query.filter_by(username='admin').first()
                        user_id = admin_user.id if admin_user else 1
                    
                    updated_boms = BOMVersioningService.check_and_update_bom_for_cost_changes(
                        product_id=product.id,
                        created_by_id=user_id
                    )
                    print(f"[DEBUG] Updated {len(updated_boms)} BOM(s)")
                    if updated_boms:
                        version_count = len(updated_boms)
                        flash(f'Product updated! BOM versions updated for {version_count} BOM(s).', 'info')
                    else:
                        flash('Product updated successfully!', 'success')
                except Exception as e:
                    flash(f'Product updated, but error updating BOM versions: {str(e)}', 'warning')
                    print(f"[DEBUG] Error updating BOM versions: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print(f"[DEBUG] Cost did not change, no BOM versioning needed")
                flash('Product updated successfully!', 'success')
            return redirect(url_for('inventory.products'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating product: {str(e)}', 'error')
            return redirect(url_for('inventory.edit_product', id=product.id))
    
    # Fetch categories for the dropdown
    categories = ProductCategory.query.filter_by(is_active=True).order_by(ProductCategory.name).all()
    
    return render_template('inventory/edit_product.html', form=form, product=product, categories=categories, warehouses=warehouses)

@bp.route('/product/<int:id>/delete', methods=['GET', 'POST'])
@login_required
def delete_product(id):
    product = Product.query.get_or_404(id)
    
    # Check if product is associated with any sales, purchases, or stock movements
    if product.sale_items or product.purchase_items or product.stock_movements:
        flash(f'Cannot delete product "{product.name}" because it has associated transaction history (sales, purchases, or stock movements). Try marking it as inactive instead.', 'danger')
        return redirect(url_for('inventory.products'))
        
    try:
        db.session.delete(product)
        db.session.commit()
        flash(f'Product "{product.name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting product: {str(e)}', 'error')
        
    return redirect(url_for('inventory.products'))

@bp.route('/products/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_products():
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No products selected'}), 400
    
    deleted_count = 0
    skipped_count = 0
    errors = []
    
    for product_id in ids:
        product = Product.query.get(product_id)
        if not product:
            continue
            
        # Check if product is associated with any sales, purchases, or stock movements
        if product.sale_items or product.purchase_items or product.stock_movements:
            skipped_count += 1
            continue
            
        try:
            db.session.delete(product)
            deleted_count += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f'Error deleting {product.name}: {str(e)}')
            
    if deleted_count > 0:
        db.session.commit()
        
    message = f'Successfully deleted {deleted_count} products.'
    if skipped_count > 0:
        message += f' Skipped {skipped_count} products with transaction history.'
    
    if errors:
        return jsonify({'success': False, 'message': message, 'errors': errors}), 500
        
    return jsonify({'success': True, 'message': message})

@bp.route('/products/bulk-assign-warehouse', methods=['POST'])
@login_required
def bulk_assign_warehouse():
    ids = request.json.get('ids', [])
    warehouse_id = request.json.get('warehouse_id')
    
    if not ids:
        return jsonify({'success': False, 'message': 'No products selected'}), 400
    
    if not warehouse_id:
        return jsonify({'success': False, 'message': 'No warehouse selected'}), 400
    
    # Verify warehouse exists
    warehouse = Warehouse.query.get(warehouse_id)
    if not warehouse:
        return jsonify({'success': False, 'message': 'Invalid warehouse'}), 400
    
    updated_count = 0
    skipped_count = 0
    errors = []
    
    for product_id in ids:
        product = Product.query.get(product_id)
        if not product:
            skipped_count += 1
            continue
        
        # Safety check: skip products with transaction history
        if product.sale_items or product.purchase_items or product.stock_movements:
            skipped_count += 1
            continue
        
        try:
            product.warehouse_id = int(warehouse_id)
            updated_count += 1
        except Exception as e:
            errors.append(f'Error updating {product.name}: {str(e)}')
    
    if updated_count > 0:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500
    
    message = f'Successfully assigned warehouse to {updated_count} products.'
    if skipped_count > 0:
        message += f' Skipped {skipped_count} products with transaction history (cannot reassign).'
    
    if errors:
        return jsonify({'success': False, 'message': message, 'errors': errors}), 500
    
    return jsonify({'success': True, 'message': message})

@bp.route('/stock-report')
@login_required
def stock_report():
    query = Product.query
    query = apply_saved_filter_to_query(query, 'product', request.args)
    products = query.all()
    
    # Calculate statistics
    total_products = len(products)
    total_value = sum(p.quantity * p.cost_price for p in products)
    low_stock_count = sum(1 for p in products if p.quantity <= p.reorder_level)
    out_of_stock = sum(1 for p in products if p.quantity == 0)
    
    return render_template('inventory/stock_report.html',
                         products=products,
                         total_products=total_products,
                         total_value=total_value,
                         low_stock_count=low_stock_count,
                         out_of_stock=out_of_stock,
                         active_module='product',
                         filter_id=request.args.get('filter_id'))

@bp.route('/product/bulk-upload', methods=['GET', 'POST'])
@login_required
def bulk_upload():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'error')
            return redirect(url_for('inventory.bulk_upload'))
        
        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(url_for('inventory.bulk_upload'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(url_for('inventory.bulk_upload'))
        
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            file_content = file.read()
            wb = load_workbook(filename=BytesIO(file_content), read_only=True)
            ws = wb.active
            rows = list(ws.values)
            if not rows:
                flash('File is empty', 'error')
                return redirect(url_for('inventory.bulk_upload'))
            headers = [str(h) if h else '' for h in rows[0]]
            
            required_columns = ['name', 'sku']
            missing = [col for col in required_columns if col not in headers]
            if missing:
                flash(f'Missing required columns: {", ".join(missing)}', 'error')
                return redirect(url_for('inventory.bulk_upload'))
            
            added = 0
            errors = []
            
            for idx, row in enumerate(rows[1:], start=2):
                try:
                    row_dict = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = val
                    
                    name = str(row_dict.get('name', '')).strip()
                    sku = str(row_dict.get('sku', '')).strip()
                    
                    if not name or not sku:
                        errors.append(f'Row {idx}: Missing name or SKU')
                        continue
                    
                    existing = Product.query.filter_by(sku=sku).first()
                    if existing:
                        errors.append(f'Row {idx}: SKU "{sku}" already exists')
                        continue
                    
                    product = Product(
                        name=name,
                        sku=sku,
                        description=str(row_dict.get('description', '')).strip() if row_dict.get('description') else None,
                        category_id=int(row_dict.get('category_id')) if row_dict.get('category_id') else None,
                        unit_price=float(row_dict.get('unit_price', 0)) if row_dict.get('unit_price') else 0,
                        cost_price=float(row_dict.get('cost_price', 0)) if row_dict.get('cost_price') else 0,
                        quantity=float(row_dict.get('quantity', 0)) if row_dict.get('quantity') else 0,
                        reorder_level=float(row_dict.get('reorder_level', 0)) if row_dict.get('reorder_level') else 0,
                    )
                    
                    db.session.add(product)
                    added += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
            
            db.session.commit()
            
            if added > 0:
                flash(f'Successfully added {added} products!', 'success')
            if errors:
                flash(f'Errors: {"; ".join(errors[:10])}', 'warning')
            
            return redirect(url_for('inventory.products'))
            
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'error')
            return redirect(url_for('inventory.bulk_upload'))
    
    return render_template('inventory/bulk_upload.html')

@bp.route('/product/download-sample')
@login_required
def download_sample():
    try:
        from openpyxl import Workbook
        from io import BytesIO
        from flask import send_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Products'
        
        headers = ['name', 'sku', 'description', 'unit_price', 'cost_price', 'quantity', 'reorder_level', 'category_id']
        ws.append(headers)
        
        sample_data = [
            ['Product A', 'SKU-001', 'Description for Product A', 100.00, 50.00, 10, 5, ''],
            ['Product B', 'SKU-002', 'Description for Product B', 200.00, 100.00, 20, 10, ''],
            ['Product C', 'SKU-003', 'Description for Product C', 50.00, 25.00, 100, 20, '']
        ]
        
        for row in sample_data:
            ws.append(row)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, download_name='sample_products.xlsx', as_attachment=True)
        
    except Exception as e:
        flash(f'Error creating sample: {str(e)}', 'error')
        return redirect(url_for('inventory.bulk_upload'))

@bp.route('/api/product/<int:id>')
@login_required
def get_product(id):
    product = Product.query.get_or_404(id)
    return jsonify({
        'id': product.id,
        'name': product.name,
        'sku': product.sku,
        'unit_price': product.unit_price,
        'cost_price': product.cost_price,
        'quantity': product.quantity,
        'reorder_level': product.reorder_level,
        'category': product.category
    })