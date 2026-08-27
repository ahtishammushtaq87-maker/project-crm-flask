from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash
from app.utils import permission_required, log_activity
from flask_login import login_required, current_user
from app import db
from app.models import Sale, PurchaseBill, Transaction, Expense, ExpenseCategory, Vendor, Account, Payment, BillPayment, TaxRate, Currency, RecurringExpense, Staff, Attendance, ExpenseSettings, Customer
from app.forms import ExpenseForm, ExpenseCategoryForm
from datetime import datetime, timedelta
from sqlalchemy import func, and_, inspect
from app.routes.filters import apply_saved_filter_to_query

bp = Blueprint('accounting', __name__)

def has_column(table_name, column_name):
    try:
        inspector = inspect(db.engine)
        return column_name in [c['name'] for c in inspector.get_columns(table_name)]
    except:
        return False

def get_unique_expense_number(settings, next_num):
    """Generate a unique expense number that doesn't exist in the database."""
    prefix = settings.expense_prefix or ''
    suffix = settings.expense_suffix or ''
    while True:
        expense_number = f"{prefix}{next_num}{suffix}"
        existing = Expense.query.filter_by(expense_number=expense_number).first()
        if not existing:
            return expense_number, next_num + 1


# ─── Expense's own accounts (ExpenseAccount / ExpenseAccountTransaction) ────
# Fully independent of the Journal module — Expense owns and manages these
# accounts end to end (create/edit/delete, debit/credit, balance) rather than
# sharing state with Journal's own separate accounts.

def _sync_expense_account_transaction(expense, account_id, is_confirmed):
    """Create/refresh the ExpenseAccountTransaction that mirrors this
    expense's effect on `account_id` (money out, i.e. 'credit'). Removes the
    transaction if account_id is falsy or invalid. `expense` must already
    have a flushed id."""
    from app.models import ExpenseAccount, ExpenseAccountTransaction

    existing = ExpenseAccountTransaction.query.filter_by(expense_id=expense.id).first()

    if not account_id or not ExpenseAccount.query.get(account_id):
        if existing:
            db.session.delete(existing)
        return

    expense_date = expense.date.date() if hasattr(expense.date, 'date') else expense.date

    if existing:
        txn = existing
    else:
        txn = ExpenseAccountTransaction(expense_id=expense.id, entry_type='credit',
                                        created_by=expense.created_by)
        db.session.add(txn)

    txn.account_id = account_id
    txn.date = expense_date
    txn.amount = expense.amount
    txn.description = expense.description
    txn.reference = expense.reference
    txn.bill_image_path = expense.bill_image_path
    txn.is_approved = is_confirmed
    txn.is_rejected = False
    txn.approved_by = expense.created_by if is_confirmed else None
    txn.approved_at = datetime.utcnow() if is_confirmed else None


def _delete_linked_expense_account_transaction(expense_id):
    """Remove the ExpenseAccountTransaction (if any) that
    _sync_expense_account_transaction created for this expense, so deleting
    the expense doesn't leave a dangling transaction still affecting an
    account's balance."""
    from app.models import ExpenseAccountTransaction
    txn = ExpenseAccountTransaction.query.filter_by(expense_id=expense_id).first()
    if txn:
        db.session.delete(txn)


# ─── Expense as a Sale/Purchase payment ("Add this to Invoice/Purchase
# Payment") ────────────────────────────────────────────────────────────────
# Checking the box on Add/Edit Expense applies the expense's amount as a real
# Payment (sale) or BillPayment (purchase bill) — found/reversed via
# expense_id, the same idempotent-bridge pattern ExpenseAccountTransaction and
# the (now-removed) JournalEntry link used. The Payment/BillPayment keeps its
# own independent approval workflow exactly like one created from the
# invoice/bill's own "Record Payment" — the linked Expense is just the record
# of where the money came from and never re-applies/reverses money based on
# the Expense's own approve/reject/draft status.

def _reverse_and_delete_sale_payment(payment, user_id):
    """Undo a Payment created from an Expense and delete it. Only reverses
    money if it was actually applied (is_approved) — mirrors
    sales.delete_payment."""
    from app.utils import adjust_sale_payment
    sale = Sale.query.get(payment.invoice_id)
    if sale and payment.is_approved:
        adjust_sale_payment(sale, -payment.amount, user_id)
    db.session.delete(payment)


def _reverse_and_delete_bill_payment(bill_payment):
    """Undo a BillPayment created from an Expense and delete it. Only
    reverses money if it was actually applied (is_approved) — mirrors
    purchase.delete_bill_payment."""
    bill = PurchaseBill.query.get(bill_payment.bill_id)
    if bill and bill_payment.is_approved:
        bill.paid_amount = max(0.0, (bill.paid_amount or 0) - bill_payment.amount)
        bill.update_status()
    db.session.delete(bill_payment)


def _reverse_expense_payment_transfer(expense, user_id):
    """Reverse and remove any Payment/BillPayment linked to this expense, and
    clear its transfer flags. Used when the expense is deleted, or when the
    checkbox is unchecked / target changed on edit. `expense` must already
    have a flushed id."""
    existing_payment = Payment.query.filter_by(expense_id=expense.id).first()
    if existing_payment:
        _reverse_and_delete_sale_payment(existing_payment, user_id)
    existing_bill_payment = BillPayment.query.filter_by(expense_id=expense.id).first()
    if existing_bill_payment:
        _reverse_and_delete_bill_payment(existing_bill_payment)
    expense.linked_sale_id = None
    expense.linked_bill_id = None
    expense.is_payment_transfer = False


def _sync_expense_payment_transfer(expense, target_type, target_id, amount, user_id, is_admin):
    """Keep the Payment/BillPayment tied to this expense in sync with the
    "Add this to Invoice/Purchase Payment" checkbox + dropdown + amount.
    Idempotent: finds-or-creates-or-reverses by expense_id, same convention as
    _sync_expense_account_transaction. `expense` must already have a flushed
    id. `target_type` is 'sale', 'bill', or falsy (box unchecked/no target)."""
    from app.utils import adjust_sale_payment, apply_sale_payment_with_credit, apply_bill_payment_with_credit

    if target_type not in ('sale', 'bill') or not target_id:
        _reverse_expense_payment_transfer(expense, user_id)
        return

    amount = round(float(amount or 0), 2)
    if amount <= 0:
        _reverse_expense_payment_transfer(expense, user_id)
        return

    existing_payment = Payment.query.filter_by(expense_id=expense.id).first()
    existing_bill_payment = BillPayment.query.filter_by(expense_id=expense.id).first()

    if target_type == 'sale':
        if existing_bill_payment:
            _reverse_and_delete_bill_payment(existing_bill_payment)

        sale = Sale.query.get(target_id)
        if not sale:
            _reverse_expense_payment_transfer(expense, user_id)
            return

        if existing_payment and existing_payment.invoice_id == target_id:
            delta = round(amount - existing_payment.amount, 2)
            existing_payment.amount = amount
            existing_payment.date = expense.date
            existing_payment.method = expense.payment_method or existing_payment.method
            existing_payment.notes = f'Transferred from Expense #{expense.expense_number}'
            # Payment.image_path uses the same project-root-relative format as
            # Expense.bill_image_path (e.g. "app/static/uploads/..."), so the
            # bill image the expense carries can be copied over as-is.
            if expense.bill_image_path:
                existing_payment.image_path = expense.bill_image_path
            if abs(delta) > 0.009 and existing_payment.is_approved:
                adjust_sale_payment(sale, delta, user_id)
        else:
            if existing_payment:
                _reverse_and_delete_sale_payment(existing_payment, user_id)
            apply_sale_payment_with_credit(sale, amount, user_id)
            last_payment = Payment.query.order_by(Payment.id.desc()).first()
            payment_num = f"PAY-{last_payment.id + 1 if last_payment else 1}"
            new_payment = Payment(
                payment_number=payment_num,
                date=expense.date,
                amount=amount,
                method=expense.payment_method or 'Cash',
                invoice_id=sale.id,
                expense_id=expense.id,
                notes=f'Transferred from Expense #{expense.expense_number}',
                image_path=expense.bill_image_path,
                created_by=user_id,
                is_approved=True,
                needs_approval=not is_admin,
                approved_by=user_id if is_admin else None,
                approved_at=datetime.utcnow() if is_admin else None
            )
            db.session.add(new_payment)

        expense.linked_sale_id = target_id
        expense.linked_bill_id = None
        expense.is_payment_transfer = True

    else:  # target_type == 'bill'
        if existing_payment:
            _reverse_and_delete_sale_payment(existing_payment, user_id)

        bill = PurchaseBill.query.get(target_id)
        if not bill:
            _reverse_expense_payment_transfer(expense, user_id)
            return

        # BillPayment.image_path is relative to the static folder (e.g.
        # "uploads/payments/xxx.png"), unlike Expense.bill_image_path which is
        # project-root-relative (e.g. "app/static/uploads/bills/xxx.png") — so
        # the "app/static/" prefix has to be stripped off when copying it over.
        bp_image_path = None
        if expense.bill_image_path:
            bp_image_path = expense.bill_image_path.replace('app/static/', '').replace('app\\static\\', '')

        if existing_bill_payment and existing_bill_payment.bill_id == target_id:
            delta = round(amount - existing_bill_payment.amount, 2)
            existing_bill_payment.amount = amount
            existing_bill_payment.date = expense.date
            existing_bill_payment.payment_method = expense.payment_method or existing_bill_payment.payment_method
            existing_bill_payment.notes = f'Transferred from Expense #{expense.expense_number}'
            if bp_image_path:
                existing_bill_payment.image_path = bp_image_path
            if abs(delta) > 0.009 and existing_bill_payment.is_approved:
                bill.paid_amount = max(0.0, min(bill.total, (bill.paid_amount or 0) + delta))
                bill.update_status()
        else:
            if existing_bill_payment:
                _reverse_and_delete_bill_payment(existing_bill_payment)
            if is_admin:
                apply_bill_payment_with_credit(bill, amount, user_id)
                bill.update_status()
            new_bp = BillPayment(
                bill_id=bill.id,
                date=expense.date,
                amount=amount,
                payment_method=expense.payment_method or 'Cash',
                notes=f'Transferred from Expense #{expense.expense_number}',
                image_path=bp_image_path,
                created_by=user_id,
                expense_id=expense.id,
                is_approved=is_admin,
                approved_by=user_id if is_admin else None,
                approved_at=datetime.utcnow() if is_admin else None
            )
            db.session.add(new_bp)

        expense.linked_bill_id = target_id
        expense.linked_sale_id = None
        expense.is_payment_transfer = True


def _sync_add_money_sale_transfer(txn, account, sale_id, amount, user_id, is_admin):
    """Mirror of _sync_expense_payment_transfer's 'sale' branch, but for a
    standalone debit ('Add Money') ExpenseAccountTransaction instead of an
    Expense — Add Money entries never create an Expense row (see
    _add_money_from_expense_form), so the Payment is created directly with no
    expense_id (Payment.expense_id is nullable, same as for Payments recorded
    straight from the Sales module); `txn.linked_payment_id` is the link back
    instead. Only ever creates — called once, right after the entry is
    created. Editing an already-transferred entry's amount is handled inline
    in edit_expense_account_debit_entry() instead (it only adjusts the
    existing Payment's amount via adjust_sale_payment, it never re-targets
    which Sale a transfer points at)."""
    from app.utils import apply_sale_payment_with_credit

    sale = Sale.query.get(sale_id)
    if not sale:
        return

    apply_sale_payment_with_credit(sale, amount, user_id)
    last_payment = Payment.query.order_by(Payment.id.desc()).first()
    payment_num = f"PAY-{last_payment.id + 1 if last_payment else 1}"
    notes = f'Transferred from Add Money entry on "{account.name}"'
    if txn.reference:
        notes += f' (Ref: {txn.reference})'
    new_payment = Payment(
        payment_number=payment_num,
        date=txn.date,
        amount=amount,
        method='Cash',
        invoice_id=sale.id,
        notes=notes,
        image_path=txn.bill_image_path,
        created_by=user_id,
        is_approved=True,
        needs_approval=not is_admin,
        approved_by=user_id if is_admin else None,
        approved_at=datetime.utcnow() if is_admin else None,
    )
    db.session.add(new_payment)
    db.session.flush()
    txn.linked_payment_id = new_payment.id


def _add_money_from_expense_form():
    """Handle the 'Debit' branch of the Add Expense form: the user picked an
    account + Debit, meaning money is going INTO that account. This is NOT an
    expense, so no Expense row is created — just a debit ExpenseAccountTransaction
    against the Expense module's own account. Optionally, that incoming money
    can instead be recorded as a customer's payment against a Sale invoice
    ("Add this to Invoice Payment" — see _sync_add_money_sale_transfer) since
    a Debit is money coming IN, same direction as a customer payment. The
    mirror-image "Add this to Purchase Payment" (paying a vendor bill) only
    makes sense for outgoing Credit expenses, so it's handled in add_expense()
    instead, not here."""
    from app.models import ExpenseAccount, ExpenseAccountTransaction

    account_id = request.form.get('account_id', type=int)
    acct = ExpenseAccount.query.get(account_id) if account_id else None
    if not acct:
        flash('Please select an account to add money to.', 'warning')
        return redirect(url_for('accounting.add_expense'))

    try:
        amount = float(request.form.get('amount') or 0)
    except (TypeError, ValueError):
        amount = 0
    if amount <= 0:
        flash('Enter an amount greater than zero.', 'warning')
        return redirect(url_for('accounting.add_expense'))

    bill_path = _save_fixed_expense_bill_image(request.files.get('bill_image'))
    is_admin = getattr(current_user, 'is_admin', False)
    txn = ExpenseAccountTransaction(
        account_id=acct.id,
        entry_type='debit',
        transaction_type='add_money',
        date=_parse_expense_date(request.form.get('date')) or datetime.utcnow().date(),
        amount=amount,
        description=(request.form.get('description') or '').strip() or None,
        reference=(request.form.get('reference') or '').strip() or None,
        bill_image_path=bill_path,
        customer_id=request.form.get('customer_id', type=int) or None,
        warehouse_id=request.form.get('warehouse_id', type=int) or None,
        is_approved=is_admin,
        approved_by=current_user.id if is_admin else None,
        approved_at=datetime.utcnow() if is_admin else None,
        created_by=current_user.id,
    )
    db.session.add(txn)
    db.session.commit()

    # "Add this to Invoice Payment" — only meaningful for Debit (money in),
    # so this is the only place that ever honors payment_transfer_type='sale'.
    transfer_type = request.form.get('payment_transfer_type')
    transfer_target_id = request.form.get('payment_transfer_target_id', type=int)
    if transfer_type == 'sale' and transfer_target_id:
        _sync_add_money_sale_transfer(txn, acct, transfer_target_id, amount, current_user.id, is_admin)
        db.session.commit()

    if is_admin:
        flash(f'PKR {amount:,.0f} added to "{acct.name}". Recorded as an account transaction, not an Expense.', 'success')
    else:
        flash(f'Entry of PKR {amount:,.0f} added to "{acct.name}" submitted for Admin approval.', 'info')
    return redirect(url_for('accounting.account_activity', account_id=acct.id))


def _parse_expense_date(raw):
    """Parse a bulk-upload sheet's Date cell — openpyxl already hands back a
    datetime/date object for date-formatted cells, otherwise fall back to the
    common text formats. Returns a datetime, or None if unparseable."""
    if raw is None or str(raw).strip() in ('', '-'):
        return None
    if isinstance(raw, datetime):
        return raw
    if hasattr(raw, 'year') and hasattr(raw, 'month') and hasattr(raw, 'day'):
        return datetime.combine(raw, datetime.min.time())
    s = str(raw).strip()
    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%y'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


@bp.route('/ledger')
def ledger():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    account = request.args.get('account', 'all')
    
    # Build query
    query = db.session.query(Transaction)
    
    if start_date:
        query = query.filter(Transaction.date >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Transaction.date <= datetime.strptime(end_date, '%Y-%m-%d'))
    if account != 'all':
        query = query.filter(Transaction.account == account)
    
    transactions = query.order_by(Transaction.date.desc()).all()
    
    # Get unique accounts for filter
    accounts = db.session.query(Transaction.account).distinct().all()
    accounts = [a[0] for a in accounts if a[0]]
    
    # Calculate balances
    total_debit = sum(t.debit for t in transactions)
    total_credit = sum(t.credit for t in transactions)
    
    return render_template('accounting/ledger.html',
                         transactions=transactions,
                         accounts=accounts,
                         total_debit=total_debit,
                         total_credit=total_credit,
                         start_date=start_date,
                         end_date=end_date,
                         current_account=account)


@bp.route('/')
@login_required
def dashboard():
    from flask import request
    from datetime import datetime, timedelta

    # Get date filters from request
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')

    # Set default date range (last 30 days) if not provided
    if not date_from_str:
        date_from = datetime.utcnow() - timedelta(days=30)
    else:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')

    if not date_to_str:
        date_to = datetime.utcnow()
    else:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d')

    # Apply date filters to all queries
    sales_query = Sale.query.filter(Sale.date >= date_from, Sale.date <= date_to)
    purchases_query = PurchaseBill.query.filter(PurchaseBill.date >= date_from, PurchaseBill.date <= date_to)
    expenses_query = Expense.query.filter(Expense.date >= date_from, Expense.date <= date_to, Expense.status == 'confirmed')

    from app.models import SaleItem, Product
    total_sales = db.session.query(func.sum(Sale.total)).filter(Sale.date >= date_from, Sale.date <= date_to).scalar() or 0
    
    # Calculate COGS (Cost of Goods Sold)
    total_cogs = db.session.query(func.sum(SaleItem.quantity * Product.cost_price))\
        .join(Sale, SaleItem.sale_id == Sale.id)\
        .join(Product, SaleItem.product_id == Product.id)\
        .filter(Sale.date >= date_from, Sale.date <= date_to)\
        .scalar() or 0
        
    total_purchases = db.session.query(func.sum(PurchaseBill.total)).filter(PurchaseBill.date >= date_from, PurchaseBill.date <= date_to).scalar() or 0
    
    # Calculate operating expenses - handle regular and divided expenses separately
    # IMPORTANT: Always filter out monthly-divided expenses from regular operating totals
    
    # Build filters for operating expenses
    operating_filter = [
        Expense.is_bom_overhead == False,
        Expense.is_shifted == False,
        Expense.is_inventory_shifted == False,
        Expense.date >= date_from,
        Expense.date <= date_to
    ]
    # Only filter divided expenses if column exists
    if has_column('expenses', 'is_monthly_divided'):
        operating_filter.append(Expense.is_monthly_divided == False)
    if has_column('expenses', 'is_payment_transfer'):
        operating_filter.append(Expense.is_payment_transfer == False)
    operating_filter.append(Expense.status == 'confirmed')

    operating_expenses = db.session.query(func.sum(Expense.amount)).filter(*operating_filter).scalar() or 0

    # Build filters for manufacturing overhead
    bom_filter = [
        Expense.is_bom_overhead == True,
        Expense.is_shifted == False,
        Expense.date >= date_from,
        Expense.date <= date_to
    ]
    # Only filter divided expenses if column exists
    if has_column('expenses', 'is_monthly_divided'):
        bom_filter.append(Expense.is_monthly_divided == False)
    if has_column('expenses', 'is_payment_transfer'):
        bom_filter.append(Expense.is_payment_transfer == False)
    bom_filter.append(Expense.status == 'confirmed')
    
    manufacturing_overhead = db.session.query(func.sum(Expense.amount)).filter(*bom_filter).scalar() or 0
    
    # Calculate today's daily expenses and handle monthly divided expenses for the period
    today = datetime.utcnow().date()
    today_daily_expenses = 0
    daily_expense_breakdown = []
    divided_expenses_for_period = 0  # Divided expenses applicable to the date range
    
    if has_column('expenses', 'is_monthly_divided'):
        # Get all monthly divided expenses (excluding shifted)
        all_monthly_expenses = Expense.query.filter(
            Expense.is_monthly_divided == True,
            Expense.is_shifted == False,
            Expense.status == 'confirmed'
        ).all()
        
        # Calculate today's divided expenses and period total
        for exp in all_monthly_expenses:
            daily_amount = exp.get_today_expense()
            
            # Add to today's total
            today_daily_expenses += daily_amount
            if daily_amount > 0:
                daily_expense_breakdown.append({
                    'description': exp.description or f"Expense {exp.expense_number}",
                    'daily_amount': daily_amount,
                    'category': exp.expense_category.name if exp.expense_category else 'Uncategorized'
                })
            
            # Calculate divided expense applicable to the date range
            if exp.monthly_start_date and exp.monthly_end_date:
                # Find overlap between expense period and filter period
                overlap_start = max(exp.monthly_start_date, date_from.date())
                overlap_end = min(exp.monthly_end_date, date_to.date())
                
                if overlap_start <= overlap_end:
                    # Calculate days in overlap
                    overlap_days = (overlap_end - overlap_start).days + 1
                    # Add proportional amount
                    divided_expenses_for_period += exp.daily_amount * overlap_days
    
    # Total expenses for the period (non-divided + proportional divided)
    total_expenses = operating_expenses + manufacturing_overhead + divided_expenses_for_period

    # Gross Profit = Sales - COGS
    gross_profit = total_sales - total_cogs

    # Net Profit = Gross Profit - operating_expenses (use divided amounts where applicable)
    # (BOM overhead is already in COGS, so we only subtract operating expenses here)
    net_profit = gross_profit - total_expenses

    outstanding_invoices = db.session.query(func.sum(Sale.total - Sale.paid_amount)).filter(Sale.status != 'paid', Sale.date >= date_from, Sale.date <= date_to).scalar() or 0
    paid_invoices = sales_query.filter(Sale.status == 'paid').count()
    unpaid_or_partial_invoices = sales_query.filter(Sale.status != 'paid').count()

    account_summary = {}
    transactions = Transaction.query.filter(Transaction.date >= date_from, Transaction.date <= date_to).all()
    for t in transactions:
        account_summary.setdefault(t.account, {'debit': 0, 'credit': 0})
        account_summary[t.account]['debit'] += t.debit
        account_summary[t.account]['credit'] += t.credit

    # Monthly trends within date range
    monthly_sales = []
    monthly_expenses = []
    monthly_labels = []

    # Generate monthly data for the selected date range
    current_date = date_from.replace(day=1)
    while current_date <= date_to:
        month_sales = db.session.query(func.sum(Sale.total)).filter(
            func.extract('year', Sale.date) == current_date.year,
            func.extract('month', Sale.date) == current_date.month
        ).scalar() or 0

        month_expenses = db.session.query(func.sum(Expense.amount)).filter(
            func.extract('year', Expense.date) == current_date.year,
            func.extract('month', Expense.date) == current_date.month,
            Expense.status == 'confirmed',
            Expense.is_shifted == False,
            Expense.is_inventory_shifted == False,
            Expense.is_payment_transfer == False
        ).scalar() or 0

        monthly_sales.append(float(month_sales))
        monthly_expenses.append(float(month_expenses))
        monthly_labels.append(current_date.strftime('%b %Y'))

        # Move to next month
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)

    # Yearly summary within date range
    yearly = []
    start_year = date_from.year
    end_year = date_to.year

    for y in range(start_year, end_year + 1):
        y_sales = db.session.query(func.sum(Sale.total)).filter(
            func.extract('year', Sale.date) == y
        ).scalar() or 0
        y_exp = db.session.query(func.sum(Expense.amount)).filter(
            func.extract('year', Expense.date) == y,
            Expense.status == 'confirmed',
            Expense.is_shifted == False,
            Expense.is_inventory_shifted == False,
            Expense.is_payment_transfer == False
        ).scalar() or 0
        yearly.append({'year': y, 'sales': float(y_sales), 'expenses': float(y_exp)})

    return render_template('accounting/dashboard.html',
                         total_sales=total_sales,
                         total_purchases=total_purchases,
                         total_cogs=total_cogs,
                         total_expenses=total_expenses,
                         operating_expenses=operating_expenses,
                         manufacturing_overhead=manufacturing_overhead,
                         divided_expenses_for_period=divided_expenses_for_period,
                         gross_profit=gross_profit,
                         net_profit=net_profit,
                         outstanding_invoices=outstanding_invoices,
                         paid_invoices=paid_invoices,
                         unpaid_or_partial_invoices=unpaid_or_partial_invoices,
                         account_summary=account_summary,
                         monthly_sales=monthly_sales,
                         monthly_expenses=monthly_expenses,
                         monthly_labels=monthly_labels,
                         yearly=yearly,
                         today_daily_expenses=today_daily_expenses,
                         daily_expense_breakdown=daily_expense_breakdown,
                         date_from=date_from.strftime('%Y-%m-%d'),
                         date_to=date_to.strftime('%Y-%m-%d'))


@bp.route('/accounts')
@login_required
def accounts():
    accounts = Account.query.order_by(Account.code.nullslast(), Account.name).all()
    return render_template('accounting/accounts.html', accounts=accounts)


@bp.route('/account/add', methods=['GET', 'POST'])
@login_required
@permission_required('accounting', action='add')
def add_account():
    if request.method == 'POST':
        name = request.form.get('name')
        code = request.form.get('code')
        typ = request.form.get('type')
        parent_id = request.form.get('parent_id') or None
        description = request.form.get('description')
        if name and typ:
            account = Account(name=name, code=code, type=typ, parent_id=parent_id if parent_id else None, description=description)
            db.session.add(account)
            db.session.commit()
            log_activity('Accounting', f'Created Account: {name}', f'Type: {typ}, Code: {code or "N/A"}')
            flash('Account created successfully', 'success')
            return redirect(url_for('accounting.accounts'))
        flash('Please provide required fields', 'danger')
    parents = Account.query.filter_by(parent_id=None).order_by(Account.name).all()
    return render_template('accounting/add_account.html', parents=parents)


@bp.route('/account/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('accounting', action='edit')
def edit_account(id):
    account = Account.query.get_or_404(id)
    if request.method == 'POST':
        account.name = request.form.get('name')
        account.code = request.form.get('code')
        account.type = request.form.get('type')
        account.parent_id = request.form.get('parent_id') or None
        account.description = request.form.get('description')
        db.session.commit()
        log_activity('Accounting', f'Updated Account: {account.name}', f'Type: {account.type}, Code: {account.code or "N/A"}')
        flash('Account updated successfully', 'success')
        return redirect(url_for('accounting.accounts'))
    parents = Account.query.filter(Account.id != account.id, Account.parent_id == None).order_by(Account.name).all()
    return render_template('accounting/edit_account.html', account=account, parents=parents)


@bp.route('/account/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('accounting', action='delete')
def delete_account(id):
    account = Account.query.get_or_404(id)
    acct_name = account.name
    db.session.delete(account)
    db.session.commit()
    log_activity('Accounting', f'Deleted Account: {acct_name}', f'ID: {id}')
    flash('Account deleted successfully', 'success')
    return redirect(url_for('accounting.accounts'))


@bp.route('/journal')
@login_required
def journal():
    return redirect(url_for('accounting.ledger'))


@bp.route('/chart-of-accounts')
@login_required
def chart_of_accounts():
    account_balances = {}
    for t in Transaction.query.all():
        account_balances.setdefault(t.account, {'debit': 0, 'credit': 0})
        account_balances[t.account]['debit'] += t.debit
        account_balances[t.account]['credit'] += t.credit

    return render_template('accounting/chart_of_accounts.html', account_balances=account_balances)


@bp.route('/transactions')
@login_required
def transactions():
    accounts = Account.query.order_by(Account.name).all()
    invoices = Sale.query.order_by(Sale.date.desc()).all()

    # Filters
    invoice_id = request.args.get('invoice_id', type=int)
    payment_mode = request.args.get('payment_mode')
    status = request.args.get('status')
    type_filter = request.args.get('type', 'all')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    search = request.args.get('search', '').strip()
    per_page = request.args.get('per_page', 25, type=int)
    page = request.args.get('page', 1, type=int)

    # Collect all payments from different modules
    payments = []

    # Sales payments
    sales_payments = Payment.query.filter(Payment.invoice_id != None).all()
    for p in sales_payments:
        txn = Transaction.query.filter(Transaction.reference_type == 'sale', Transaction.reference_id == p.invoice_id).first()
        payments.append({
            'id': p.id,
            'date': p.date,
            'amount': p.amount,
            'payment_mode': p.method,
            'reference_type': 'sale',
            'reference_id': p.invoice_id,
            'invoice': p.invoice,
            'debit_account': txn.debit_account if txn else None,
            'credit_account': txn.credit_account if txn else None,
            'status': txn.status if txn else 'Pending',
            'is_mapped': txn.is_mapped if txn else False,
            'description': p.notes,
            'transaction_id': txn.id if txn else None
        })

    # Expense payments
    expense_payments = Payment.query.filter(Payment.expense_id != None).all()
    for p in expense_payments:
        txn = Transaction.query.filter(Transaction.reference_type == 'expense', Transaction.reference_id == p.expense_id).first()
        payments.append({
            'id': p.id,
            'date': p.date,
            'amount': p.amount,
            'payment_mode': p.method,
            'reference_type': 'expense',
            'reference_id': p.expense_id,
            'invoice': None,
            'debit_account': txn.debit_account if txn else None,
            'credit_account': txn.credit_account if txn else None,
            'status': txn.status if txn else 'Pending',
            'is_mapped': txn.is_mapped if txn else False,
            'description': p.notes,
            'transaction_id': txn.id if txn else None
        })

    # Purchase payments (from bills with paid_amount)
    purchase_bills = PurchaseBill.query.filter(PurchaseBill.paid_amount > 0).all()
    for bill in purchase_bills:
        txn = Transaction.query.filter(Transaction.reference_type == 'purchase', Transaction.reference_id == bill.id).first()
        payments.append({
            'id': bill.id,
            'date': bill.date,
            'amount': bill.paid_amount,
            'payment_mode': 'Various',
            'reference_type': 'purchase',
            'reference_id': bill.id,
            'invoice': None,
            'debit_account': txn.debit_account if txn else None,
            'credit_account': txn.credit_account if txn else None,
            'status': txn.status if txn else 'Pending',
            'is_mapped': txn.is_mapped if txn else False,
            'description': f'Payment for bill {bill.bill_number}',
            'transaction_id': txn.id if txn else None
        })

    # General payments
    general_payments = Payment.query.filter(Payment.invoice_id.is_(None), Payment.expense_id.is_(None)).all()
    for p in general_payments:
        txn = Transaction.query.filter(Transaction.reference_type == 'payment', Transaction.reference_id == p.id).first()
        payments.append({
            'id': p.id,
            'date': p.date,
            'amount': p.amount,
            'payment_mode': p.method,
            'reference_type': 'payment',
            'reference_id': p.id,
            'invoice': None,
            'debit_account': txn.debit_account if txn else None,
            'credit_account': txn.credit_account if txn else None,
            'status': txn.status if txn else 'Pending',
            'is_mapped': txn.is_mapped if txn else False,
            'description': p.notes,
            'transaction_id': txn.id if txn else None
        })

    # Apply filters
    if type_filter and type_filter != 'all':
        payments = [p for p in payments if p['reference_type'] == type_filter]

    if invoice_id:
        payments = [p for p in payments if p['reference_type'] == 'sale' and p['reference_id'] == invoice_id]

    if payment_mode:
        payments = [p for p in payments if p['payment_mode'] == payment_mode]

    if status:
        payments = [p for p in payments if p['status'] == status]

    if date_from:
        date_from_dt = datetime.strptime(date_from, '%Y-%m-%d')
        payments = [p for p in payments if p['date'] >= date_from_dt]

    if date_to:
        date_to_dt = datetime.strptime(date_to, '%Y-%m-%d')
        payments = [p for p in payments if p['date'] <= date_to_dt]

    if search:
        payments = [p for p in payments if search.lower() in (p['description'] or '').lower() or search.lower() in (p['payment_mode'] or '').lower()]

    # Sort by date desc
    payments.sort(key=lambda x: x['date'], reverse=True)

    # Paginate
    start = (page - 1) * per_page
    end = start + per_page
    paginated_items = payments[start:end]

    # Create a simple pagination object
    total = len(payments)
    pages = (total + per_page - 1) // per_page
    pagination = type('Pagination', (), {
        'items': paginated_items,
        'page': page,
        'pages': pages,
        'per_page': per_page,
        'total': total,
        'has_prev': page > 1,
        'has_next': page < pages,
        'prev_num': page - 1 if page > 1 else None,
        'next_num': page + 1 if page < pages else None
    })()

    return render_template('accounting/transactions.html',
                         transactions=paginated_items,
                         accounts=accounts,
                         invoices=invoices,
                         pagination=pagination,
                         invoice_id=invoice_id,
                         payment_mode=payment_mode,
                         status=status,
                         type=type_filter,
                         date_from=date_from,
                         date_to=date_to,
                         search=search,
                         per_page=per_page)


@bp.route('/reports/profit-loss')
@login_required
def report_profit_loss():
    # reuse existing profit_loss logic
    return redirect(url_for('accounting.profit_loss'))


@bp.route('/reports/balance-sheet')
@login_required
def report_balance_sheet():
    # simple balance sheet from account totals
    accounts = Account.query.all()
    asset = liability = equity = income = expense = 0
    for a in accounts:
        t = Transaction.query.filter(Transaction.account == a.name).all()
        balance = sum(x.debit - x.credit for x in t)
        if a.type == 'Asset':
            asset += balance
        elif a.type == 'Liability':
            liability += balance
        elif a.type == 'Equity':
            equity += balance
        elif a.type == 'Income':
            income += balance
        elif a.type == 'Expense':
            expense += balance
    return render_template('accounting/report_balance_sheet.html', asset=asset, liability=liability, equity=equity, income=income, expense=expense)


@bp.route('/reports/cash-flow')
@login_required
def report_cash_flow():
    # placeholder for cash flow
    return render_template('accounting/report_cash_flow.html')


@bp.route('/reports/expense')
@login_required
def report_expense():
    expenses = Expense.query.order_by(Expense.date.desc()).all()
    return render_template('accounting/report_expense.html', expenses=expenses)


@bp.route('/reports/tax-summary')
@login_required
def report_tax_summary():
    taxes = TaxRate.query.all()
    return render_template('accounting/report_tax_summary.html', taxes=taxes)


@bp.route('/transaction/add', methods=['GET', 'POST'])
@login_required
@permission_required('accounting', action='add')
def add_transaction():
    accounts = Account.query.order_by(Account.name).all()
    invoices = Sale.query.order_by(Sale.invoice_number).all()
    now = datetime.utcnow().strftime('%Y-%m-%d')
    
    # Pre-fill from query params
    prefill = {
        'reference_type': request.args.get('reference_type'),
        'reference_id': request.args.get('reference_id'),
        'amount': request.args.get('amount'),
        'payment_mode': request.args.get('payment_mode'),
        'description': request.args.get('description')
    }
    
    if request.method == 'POST':
        tn = request.form.get('transaction_number')
        date = request.form.get('date')
        amount = float(request.form.get('amount') or 0)
        reference_type = request.form.get('reference_type')
        reference_id = request.form.get('reference_id') or None
        status = request.form.get('status', 'Pending')
        payment_mode = request.form.get('payment_mode', 'Cash')
        description = request.form.get('description')
        debit_account = request.form.get('debit_account')
        credit_account = request.form.get('credit_account')
        account = request.form.get('account')
        is_mapped = bool(reference_id)
        debit = float(request.form.get('debit') or 0)
        credit = float(request.form.get('credit') or 0)

        transaction = Transaction(
            transaction_number=tn,
            date=datetime.strptime(date, '%Y-%m-%d') if date else datetime.utcnow(),
            amount=amount,
            payment_mode=payment_mode,
            invoice_id=int(reference_id) if reference_type == 'sale' and reference_id else None,
            status=status,
            is_mapped=is_mapped,
            reference_type=reference_type,
            reference_id=int(reference_id) if reference_id else None,
            debit_account=debit_account,
            credit_account=credit_account,
            description=description,
            account=account,
            debit=debit,
            credit=credit
        )
        db.session.add(transaction)
        db.session.commit()
        log_activity('Accounting', f'Created Transaction: {tn}', f'Amount: {amount}, Type: {reference_type or "manual"}')
        flash('Transaction mapped', 'success')
        return redirect(url_for('accounting.transactions'))
    return render_template('accounting/add_transaction.html', accounts=accounts, invoices=invoices, now=now, prefill=prefill)
    
@bp.route('/transaction/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('accounting', action='delete')
def delete_transaction(id):
    transaction = Transaction.query.get_or_404(id)
    try:
        txn_num = transaction.transaction_number
        db.session.delete(transaction)
        db.session.commit()
        log_activity('Accounting', f'Deleted Transaction: {txn_num}', f'ID: {id}')
        flash('Transaction mapping removed successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting transaction: {str(e)}', 'error')
    return redirect(url_for('accounting.transactions'))


@bp.route('/map_transaction/<int:payment_id>/<ref_type>')
@login_required
def map_transaction(payment_id, ref_type):
    # Redirect to add_transaction with pre-filled data
    if ref_type == 'sale':
        payment = Payment.query.get_or_404(payment_id)
        invoice = Sale.query.get(payment.invoice_id)
        return redirect(url_for('accounting.add_transaction', 
                               reference_type='sale', 
                               reference_id=payment.invoice_id,
                               amount=payment.amount,
                               payment_mode=payment.method,
                               description=f'Payment for invoice {invoice.invoice_number}' if invoice else payment.notes))
    elif ref_type == 'expense':
        payment = Payment.query.get_or_404(payment_id)
        expense = Expense.query.get(payment.expense_id)
        return redirect(url_for('accounting.add_transaction', 
                               reference_type='expense', 
                               reference_id=payment.expense_id,
                               amount=payment.amount,
                               payment_mode=payment.method,
                               description=f'Payment for expense {expense.expense_number}' if expense else payment.notes))
    elif ref_type == 'purchase':
        bill = PurchaseBill.query.get_or_404(payment_id)
        return redirect(url_for('accounting.add_transaction', 
                               reference_type='purchase', 
                               reference_id=bill.id,
                               amount=bill.paid_amount,
                               payment_mode='Various',
                               description=f'Payment for bill {bill.bill_number}'))
    elif ref_type == 'payment':
        payment = Payment.query.get_or_404(payment_id)
        return redirect(url_for('accounting.add_transaction', 
                               reference_type='payment', 
                               reference_id=payment.id,
                               amount=payment.amount,
                               payment_mode=payment.method,
                               description=payment.notes))
    return redirect(url_for('accounting.transactions'))
    return redirect(url_for('accounting.transactions'))


@bp.route('/payments')
@login_required
def payments():
    payments = Payment.query.order_by(Payment.date.desc()).all()
    return render_template('accounting/payments.html', payments=payments)


@bp.route('/payment/add', methods=['GET', 'POST'])
@login_required
@permission_required('accounting', action='add')
def add_payment():
    invoices = Sale.query.filter(Sale.status != 'paid').all()
    if request.method == 'POST':
        payment_number = request.form.get('payment_number')
        date = request.form.get('date')
        amount = float(request.form.get('amount') or 0)
        method = request.form.get('method')
        invoice_id = request.form.get('invoice_id') or None
        reference_number = request.form.get('reference_number')
        notes = request.form.get('notes')

        payment = Payment(
            payment_number=payment_number,
            date=datetime.strptime(date, '%Y-%m-%d') if date else datetime.utcnow(),
            amount=amount,
            method=method,
            invoice_id=invoice_id if invoice_id else None,
            reference_number=reference_number,
            notes=notes
        )
        db.session.add(payment)

        if invoice_id:
            sale = Sale.query.get(int(invoice_id))
            if sale:
                sale.paid_amount += amount
                sale.update_status()

        db.session.commit()
        log_activity('Accounting', f'Created Payment: {payment_number}', f'Amount: {amount}, Method: {method}')
        flash('Payment recorded', 'success')
        return redirect(url_for('accounting.payments'))
    return render_template('accounting/add_payment.html', invoices=invoices, date_today=datetime.utcnow().strftime('%Y-%m-%d'))

@bp.route('/payment/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('accounting', action='delete')
def delete_payment(id):
    payment = Payment.query.get_or_404(id)
    
    # If it's linked to an invoice, subtract the amount
    if payment.invoice_id:
        sale = Sale.query.get(payment.invoice_id)
        if sale:
            sale.paid_amount -= payment.amount
            if sale.paid_amount < 0:
                sale.paid_amount = 0
            sale.update_status()
            
    # Also delete any mapped transaction for this payment
    # We search by reference_type and ID
    associated_txns = Transaction.query.filter_by(reference_type='payment', reference_id=payment.id).all()
    # Or if it was a sale/expense payment, find by those refs
    if payment.invoice_id:
        associated_txns += Transaction.query.filter_by(reference_type='sale', reference_id=payment.invoice_id, amount=payment.amount).all()
    elif payment.expense_id:
        associated_txns += Transaction.query.filter_by(reference_type='expense', reference_id=payment.expense_id, amount=payment.amount).all()
        
    for txn in associated_txns:
        db.session.delete(txn)
        
    pay_num = payment.payment_number
    pay_amount = payment.amount
    try:
        db.session.delete(payment)
        db.session.commit()
        log_activity('Accounting', f'Deleted Payment: {pay_num}', f'Amount: {pay_amount}')
        flash('Payment deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting payment: {str(e)}', 'error')

    return redirect(request.referrer or url_for('accounting.payments'))

@bp.route('/trial-balance')
def trial_balance():
    as_of_date = request.args.get('as_of_date', datetime.now().strftime('%Y-%m-%d'))
    as_of_datetime = datetime.strptime(as_of_date, '%Y-%m-%d')
    
    # Get all transactions up to date
    transactions = Transaction.query.filter(Transaction.date <= as_of_datetime).all()
    
    # Calculate balances per account
    account_balances = {}
    for t in transactions:
        if t.account not in account_balances:
            account_balances[t.account] = {'debit': 0, 'credit': 0}
        account_balances[t.account]['debit'] += t.debit
        account_balances[t.account]['credit'] += t.credit
    
    trial_balance_items = []
    total_debit = 0
    total_credit = 0
    
    for account, balances in account_balances.items():
        balance = balances['debit'] - balances['credit']
        if balance > 0:
            trial_balance_items.append({
                'account': account,
                'debit': balance,
                'credit': 0
            })
            total_debit += balance
        elif balance < 0:
            trial_balance_items.append({
                'account': account,
                'debit': 0,
                'credit': abs(balance)
            })
            total_credit += abs(balance)
    
    return render_template('accounting/trial_balance.html',
                         trial_balance=trial_balance_items,
                         total_debit=total_debit,
                         total_credit=total_credit,
                         as_of_date=as_of_date)

@bp.route('/profit-loss')
def profit_loss():
    start_date = request.args.get('start_date', 
                                  (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))
    
    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
    
    # Sales revenue
    total_sales = db.session.query(func.sum(Sale.total)).filter(
        and_(Sale.date >= start_datetime, Sale.date <= end_datetime)
    ).scalar() or 0
    
    # Purchase costs
    total_purchases = db.session.query(func.sum(PurchaseBill.total)).filter(
        and_(PurchaseBill.date >= start_datetime, PurchaseBill.date <= end_datetime)
    ).scalar() or 0
    
    # Gross profit
    gross_profit = total_sales - total_purchases
    
    # Operating Expenses - Simple/Daily
    operating_expenses = Expense.query.filter(
        Expense.date >= start_datetime, 
        Expense.date <= end_datetime,
        Expense.is_bom_overhead == False,
        Expense.is_monthly_divided == False,
        Expense.status == 'confirmed'
    ).all()
    
    expense_categories = {}
    for e in operating_expenses:
        cat_name = e.expense_category.name if e.expense_category else 'Other'
        expense_categories[cat_name] = expense_categories.get(cat_name, 0) + e.amount
    
    # Divided Expenses
    divided_expenses = Expense.query.filter(
        Expense.date >= start_datetime, 
        Expense.date <= end_datetime,
        Expense.is_bom_overhead == False,
        Expense.is_monthly_divided == True
    ).all()
    
    divided_expense_categories = {}
    for e in divided_expenses:
        cat_name = e.expense_category.name if e.expense_category else 'Other'
        if e.monthly_start_date and e.monthly_end_date:
            exp_start = datetime.combine(e.monthly_start_date, datetime.min.time())
            exp_end = datetime.combine(e.monthly_end_date, datetime.min.time())
            period_start = max(start_datetime, exp_start)
            period_end = min(end_datetime, exp_end)
            if period_end >= period_start:
                total_days = (e.monthly_end_date - e.monthly_start_date).days + 1
                active_days = (period_end.date() - period_start.date()).days + 1
                if total_days > 0:
                    daily_amount = e.amount / total_days
                    pro_rata_amount = daily_amount * active_days
                    divided_expense_categories[cat_name] = divided_expense_categories.get(cat_name, 0) + pro_rata_amount
        else:
            divided_expense_categories[cat_name] = divided_expense_categories.get(cat_name, 0) + e.amount
    
    total_divided_expenses = sum(divided_expense_categories.values())
    
    # Calculate Daily Payroll (same as Dashboard)
    from calendar import monthrange
    from datetime import timedelta
    
    attendance_records_by_date = {}
    attendance_records = Attendance.query.filter(
        Attendance.date >= start_datetime.date(),
        Attendance.date <= end_datetime.date()
    ).all()
    for record in attendance_records:
        if record.date not in attendance_records_by_date:
            attendance_records_by_date[record.date] = []
        attendance_records_by_date[record.date].append(record)
    
    attendance_payroll = sum(record.earned_amount for record in attendance_records)
    
    active_staff = Staff.query.filter_by(is_active=True).all()
    period_start = start_datetime.date()
    period_end = end_datetime.date()
    daily_payroll_for_period = attendance_payroll
    
    for staff in active_staff:
        if period_start.month == period_end.month and period_start.year == period_end.year:
            days_in_period = (period_end - period_start).days + 1
            _, days_in_month = monthrange(period_start.year, period_start.month)
            daily_rate = staff.monthly_salary / float(days_in_month)
            days_without_attendance = 0
            current_date = period_start
            while current_date <= period_end:
                if current_date not in attendance_records_by_date:
                    days_without_attendance += 1
                current_date += timedelta(days=1)
            daily_payroll_for_period += daily_rate * days_without_attendance
        else:
            current_date = period_start
            while current_date <= period_end:
                _, days_in_month = monthrange(current_date.year, current_date.month)
                month_end = datetime(current_date.year, current_date.month, days_in_month).date()
                actual_end = min(month_end, period_end)
                daily_rate = staff.monthly_salary / float(days_in_month)
                days_without_attendance = 0
                check_date = current_date
                while check_date <= actual_end:
                    if check_date not in attendance_records_by_date:
                        days_without_attendance += 1
                    check_date += timedelta(days=1)
                daily_payroll_for_period += daily_rate * days_without_attendance
                if actual_end == month_end:
                    current_date = datetime(
                        current_date.year if current_date.month < 12 else current_date.year + 1,
                        (current_date.month % 12) + 1,
                        1
                    ).date()
                else:
                    break
    
    total_payroll = daily_payroll_for_period
    
    total_expenses = sum(expense_categories.values()) + total_divided_expenses + total_payroll
    net_profit = gross_profit - total_expenses
    
    return render_template('accounting/profit_loss.html',
                         total_sales=total_sales,
                         total_purchases=total_purchases,
                         gross_profit=gross_profit,
                         expense_categories=expense_categories,
                         divided_expense_categories=divided_expense_categories,
                         total_divided_expenses=total_divided_expenses,
                         total_payroll=total_payroll,
                         total_expenses=total_expenses,
                         net_profit=net_profit,
                         start_date=start_date,
                         end_date=end_date)

@bp.route('/expenses')
@login_required
def expenses():
    from flask import session
    
    # Check for reset trigger
    if request.args.get('reset'):
        session.pop('expense_filters', None)
        return redirect(url_for('accounting.expenses'))

    # Helper to get either from request.args or from session
    def get_filter(name, type_func=None):
        val = request.args.get(name)
        if val is not None:
            # Update session if value is present in request
            if 'expense_filters' not in session:
                session['expense_filters'] = {}
            session['expense_filters'][name] = val
            
            if val == '':
                return None
            return type_func(val) if type_func else val
        
        # If not in request, check session
        saved_filters = session.get('expense_filters', {})
        val = saved_filters.get(name)
        if val and val != '':
            return type_func(val) if type_func else val
        return None

    # Write any fixed-expense cycle that has started into the book before the
    # list is built, so a new cycle shows up without any manual step.
    ensure_fixed_expense_rows()

    # Get filter parameters with persistence
    vendor_id = get_filter('vendor_id', int)
    customer_id = get_filter('customer_id', int)
    warehouse_id = get_filter('warehouse_id', int)
    category_id = get_filter('category_id', int)
    mo_id = get_filter('mo_id', int)
    start_date = get_filter('start_date')
    end_date = get_filter('end_date')

    # Credit (real Expenses, money out) vs Debit (standalone "Add Money"
    # entries, money in — see _add_money_from_expense_form) view toggle.
    entry_view = (request.args.get('entry_view') or 'credit').strip().lower()
    if entry_view not in ('credit', 'debit'):
        entry_view = 'credit'

    # Build query
    query = Expense.query
    
    if vendor_id:
        query = query.filter(Expense.vendor_id == vendor_id)
    if customer_id:
        query = query.filter(Expense.customer_id == customer_id)
    if warehouse_id:
        query = query.filter(Expense.warehouse_id == warehouse_id)
    if category_id:
        query = query.filter(Expense.category_id == category_id)
    if mo_id:
        query = query.filter(Expense.mo_id == mo_id)
    if start_date:
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        query = query.filter(Expense.date >= start_datetime)
    if end_date:
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
        # Add one day to include the end date fully
        end_datetime = end_datetime + timedelta(days=1)
        query = query.filter(Expense.date < end_datetime)
    
    query = apply_saved_filter_to_query(query, 'expense', request.args)

    # ── Approval status tabs ──────────────────────────────────────────────────
    # Derived from the same flags ApprovalService.get_status() reads, in the same
    # precedence order (rejected > approved > draft). A "cancelled" expense is a
    # rejection carrying the reason the cancel action writes.
    CANCEL_REASON = 'Cancelled by Admin'
    # NULL-safe: a row with a NULL status/flag must still land in exactly one tab,
    # never disappear from all of them.
    not_cancelled_status = ((Expense.status == None) | (Expense.status != 'cancelled'))
    is_cancelled = (((Expense.is_rejected == True) & (Expense.rejection_reason == CANCEL_REASON))
                    | (Expense.status == 'cancelled'))
    is_rejected_only = ((Expense.is_rejected == True)
                        & ((Expense.rejection_reason == None)
                           | (Expense.rejection_reason != CANCEL_REASON))
                        & not_cancelled_status)
    is_draft_only = ((Expense.is_draft == True)
                     & ((Expense.is_rejected == False) | (Expense.is_rejected == None))
                     & ((Expense.is_approved == False) | (Expense.is_approved == None)))
    is_unapproved = (((Expense.is_approved == False) | (Expense.is_approved == None))
                     & ((Expense.is_rejected == False) | (Expense.is_rejected == None))
                     & ((Expense.is_draft == False) | (Expense.is_draft == None))
                     & not_cancelled_status)

    exp_status = (request.args.get('status') or 'all').strip().lower()
    if exp_status not in ('all', 'approved', 'unapproved', 'rejected_items', 'draft', 'cancelled'):
        exp_status = 'all'

    if exp_status == 'unapproved':
        query = query.filter(is_unapproved)
    elif exp_status == 'rejected_items':
        query = query.filter(is_rejected_only)
    elif exp_status == 'draft':
        query = query.filter(is_draft_only)
    elif exp_status == 'cancelled':
        query = query.filter(is_cancelled)
    elif exp_status == 'approved':
        query = query.filter(Expense.is_approved == True, ~is_cancelled)
    else:
        # 'all' — the working view: cancelled and drafts live in their own tabs.
        query = query.filter(~is_cancelled,
                             (Expense.is_draft == False) | (Expense.is_draft == None))

    # Counts span the whole book, not the current tab.
    unapproved_count = Expense.query.filter(is_unapproved).count()
    rejected_item_count = Expense.query.filter(is_rejected_only).count()
    draft_count = Expense.query.filter(is_draft_only).count()
    cancelled_count = Expense.query.filter(is_cancelled).count()

    # Standalone debit ("Add Money") entries — never tied to an Expense row,
    # so otherwise invisible on this page. Transfer Funds legs are excluded:
    # deleting only one side of a transfer would break its double-entry pair,
    # so those stay visible only on Account Activity.
    from app.models import ExpenseAccountTransaction
    debit_base_query = ExpenseAccountTransaction.query.filter(
        ExpenseAccountTransaction.expense_id.is_(None),
        ExpenseAccountTransaction.entry_type == 'debit',
        ExpenseAccountTransaction.transaction_type != 'transfer',
    )
    debit_count = debit_base_query.count()

    debit_entries = []
    if entry_view == 'debit':
        debit_query = debit_base_query
        if start_date:
            debit_query = debit_query.filter(
                ExpenseAccountTransaction.date >= datetime.strptime(start_date, '%Y-%m-%d').date())
        if end_date:
            debit_query = debit_query.filter(
                ExpenseAccountTransaction.date < (datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)).date())
        debit_entries = (debit_query
                          .order_by(ExpenseAccountTransaction.date.desc(), ExpenseAccountTransaction.id.desc())
                          .all())

    expenses = query.order_by(Expense.date.desc()).all()
    # Calculate confirmed expenses, excluding those shifted to PD projects
    total_expense = sum(e.amount for e in expenses if e.status == 'confirmed' and not getattr(e, 'is_shifted', False) and not getattr(e, 'is_inventory_shifted', False) and not getattr(e, 'is_payment_transfer', False))
    total_pd_shifted_expense = sum(e.amount for e in expenses if e.status == 'confirmed' and getattr(e, 'is_shifted', False))
    
    # Get filter options
    categories = ExpenseCategory.query.filter_by(is_active=True).order_by(ExpenseCategory.name).all()
    vendors = Vendor.query.filter_by(is_active=True).order_by(Vendor.name).all()
    from app.models import Customer, Warehouse
    customers = Customer.query.filter_by(is_active=True).order_by(Customer.name).all()
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    from app.models import ManufacturingOrder, ExpenseAccount, PaymentMethod
    manufacturing_orders = ManufacturingOrder.query.order_by(ManufacturingOrder.order_number.desc()).all()
    expense_accounts = ExpenseAccount.query.filter_by(is_active=True).order_by(ExpenseAccount.name).all()
    payment_methods = PaymentMethod.query.filter_by(is_active=True).order_by(PaymentMethod.name).all()
    
    # Get active/draft PD projects for the shift modal
    from app.models import PDProject
    active_pd_projects = PDProject.query.filter(PDProject.status.in_(['Draft', 'Active'])).all()

    # Get active inventory items for the "shift to inventory cost" modal
    from app.models import Product
    inventory_items = Product.query.filter_by(is_active=True).order_by(Product.name).all()
    
    # Get PD expense categories for the shift modal
    pd_expense_categories = [
        'Sample Purchase', 'Reverse Engineering', 'Measurement', 'CAD', 'Prototype', 'Testing',
        'Mold', 'Die', 'Fixture', 'Pattern', 'Jig', 'Gauge',
        'Raw Material', 'Purchased Components', 'Machining', 'Casting',
        'Electricity', 'Maintenance', 'Factory Wages',
        'Office Rent', 'Salaries', 'Marketing', 'Travel',
        'Scrap', 'Prototype Failure', 'Warranty'
    ]

    # Get date format from company settings
    from app.models import Company
    company = Company.query.first()
    date_format = company.date_format if company and company.date_format else '%Y-%m-%d'
    
    return render_template('accounting/expenses.html', 
                         expenses=expenses, 
                         total_expense=total_expense,
                         total_pd_shifted_expense=total_pd_shifted_expense,
                         categories=categories,
                         vendors=vendors,
                         customers=customers,
                         warehouses=warehouses,
                         expense_accounts=expense_accounts,
                         payment_methods=payment_methods,
                         manufacturing_orders=manufacturing_orders,
                         active_pd_projects=active_pd_projects,
                         inventory_items=inventory_items,
                         pd_expense_categories=pd_expense_categories,
                         selected_vendor=vendor_id,
                         selected_customer=customer_id,
                         selected_warehouse=warehouse_id,
                         selected_category=category_id,
                         selected_mo_id=mo_id,
                         selected_start_date=start_date,
                         selected_end_date=end_date,
                         date_format=date_format,
                         today_date=datetime.utcnow().date(),
                         active_module='expense',
                         current_status=exp_status,
                         unapproved_count=unapproved_count,
                         rejected_item_count=rejected_item_count,
                         draft_count=draft_count,
                         cancelled_count=cancelled_count,
                         entry_view=entry_view,
                         debit_entries=debit_entries,
                         debit_count=debit_count,
                         filter_id=request.args.get('filter_id'))


@bp.route('/expenses/search-sales-json')
@login_required
def search_sales_json():
    """Searchable dropdown source for "Add this to Invoice Payment" on
    Add/Edit Expense — Sale invoices with a balance still due, matched by
    invoice number or customer name. Mirrors sales.customers_list_json."""
    q = request.args.get('q', '').strip()
    query = Sale.query.filter(Sale.is_draft == False, Sale.status != 'paid')
    if q:
        like = f"%{q}%"
        query = query.join(Sale.customer, isouter=True).filter(
            (Sale.invoice_number.ilike(like)) | (Customer.name.ilike(like))
        )
    sales = query.order_by(Sale.date.desc()).limit(50).all()
    results = []
    for s in sales:
        if s.balance_due <= 0.009:
            continue
        label = f"{s.invoice_number} — {s.customer.name if s.customer else 'No Customer'} (Due: PKR {s.balance_due:,.2f})"
        results.append({'id': s.id, 'text': label})
    return jsonify({'results': results})


@bp.route('/expenses/search-bills-json')
@login_required
def search_bills_json():
    """Searchable dropdown source for "Add this to Purchase Payment" on
    Add/Edit Expense — Purchase bills with a balance still due, matched by
    bill number or vendor name."""
    q = request.args.get('q', '').strip()
    query = PurchaseBill.query.filter(PurchaseBill.status.notin_(['paid', 'cancelled']))
    if q:
        like = f"%{q}%"
        query = query.join(PurchaseBill.vendor, isouter=True).filter(
            (PurchaseBill.bill_number.ilike(like)) | (Vendor.name.ilike(like))
        )
    bills = query.order_by(PurchaseBill.date.desc()).limit(50).all()
    results = []
    for b in bills:
        balance = max(0.0, b.total - b.paid_amount - (b.cancelled_amount or 0))
        if balance <= 0.009:
            continue
        label = f"{b.bill_number} — {b.vendor.name if b.vendor else 'No Vendor'} (Due: PKR {balance:,.2f})"
        results.append({'id': b.id, 'text': label})
    return jsonify({'results': results})


@bp.route('/expenses/account-activity')
@login_required
def account_activity():
    """Approved money-in, money-out, transfers, running balance, daily
    confirmation, and reconciliation for one of Expense's own accounts.
    Covers real Expenses (credit), plain 'Add Money' transactions (debit —
    see _add_money_from_expense_form, which never creates an Expense row),
    and Transfer Funds movements (a credit/debit pair across two accounts,
    see transfer_funds_quick). Read-only: no ledger data is created here.
    Entirely independent of the Journal module's own accounts."""
    from app.models import ExpenseAccount, ExpenseAccountTransaction, AccountDailyClose

    all_accounts = ExpenseAccount.query.order_by(ExpenseAccount.name).all()

    account_type = (request.args.get('account_type') or '').strip()
    custodian = (request.args.get('custodian') or '').strip()
    location = (request.args.get('location') or '').strip()

    filtered_accounts = [
        a for a in all_accounts
        if (not account_type or a.account_type == account_type)
        and (not custodian or a.custodian_name == custodian)
        and (not location or a.location == location)
    ]

    account_type_options = sorted({a.account_type for a in all_accounts if a.account_type})
    custodian_options = sorted({a.custodian_name for a in all_accounts if a.custodian_name})
    location_options = sorted({a.location for a in all_accounts if a.location})

    account_id = request.args.get('account_id', type=int)
    selected_account = ExpenseAccount.query.get(account_id) if account_id else None
    if not selected_account and filtered_accounts:
        selected_account = filtered_accounts[0]
        account_id = selected_account.id

    today = datetime.utcnow().date()

    def _parse_date(raw, default):
        if not raw:
            return default
        try:
            return datetime.strptime(raw, '%Y-%m-%d').date()
        except ValueError:
            return default

    # Default to the account's FULL history rather than an arbitrary lookback
    # window — a fixed "last 7 days" default silently hid older (or
    # future-dated) transactions from the ledger, summary cards, and running
    # balance, making a transfer look "missing" or the totals look wrong even
    # though it was recorded correctly. Explicit date_from/date_to in the URL
    # still narrow the view same as before.
    if selected_account:
        earliest_txn = (ExpenseAccountTransaction.query
                         .filter_by(account_id=selected_account.id)
                         .order_by(ExpenseAccountTransaction.date.asc())
                         .first())
        latest_txn = (ExpenseAccountTransaction.query
                       .filter_by(account_id=selected_account.id)
                       .order_by(ExpenseAccountTransaction.date.desc())
                       .first())
        default_date_from = earliest_txn.date if earliest_txn else (
            selected_account.created_at.date() if selected_account.created_at else today)
        default_date_to = max(today, latest_txn.date) if latest_txn else today
    else:
        default_date_from = today - timedelta(days=30)
        default_date_to = today

    date_from = _parse_date(request.args.get('date_from'), default_date_from)
    date_to = _parse_date(request.args.get('date_to'), default_date_to)
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    row_status = request.args.get('row_status', 'all')

    rows = []
    status_counts = {'all': 0, 'approved': 0, 'pending': 0, 'rejected': 0, 'draft': 0, 'reversed': 0}
    range_opening_balance = 0.0
    money_in = money_out = 0.0
    expected_balance = actual_balance = variance = 0.0
    latest_close = None
    daily_closes = []

    def _status_of(t):
        if t.is_reversed:
            return 'reversed'
        if t.is_draft:
            return 'draft'
        if t.is_rejected:
            return 'rejected'
        if t.is_approved:
            return 'approved'
        return 'pending'

    if selected_account:
        range_opening_balance = selected_account.opening_balance or 0
        prior_txns = (ExpenseAccountTransaction.query
                      .filter(ExpenseAccountTransaction.account_id == selected_account.id,
                              ExpenseAccountTransaction.is_approved.is_(True),
                              ExpenseAccountTransaction.date < date_from)
                      .all())
        for t in prior_txns:
            range_opening_balance += (t.amount or 0) if t.entry_type == 'debit' else -(t.amount or 0)

        range_txns = (ExpenseAccountTransaction.query
                      .filter(ExpenseAccountTransaction.account_id == selected_account.id,
                              ExpenseAccountTransaction.date >= date_from,
                              ExpenseAccountTransaction.date <= date_to)
                      .order_by(ExpenseAccountTransaction.date.desc(), ExpenseAccountTransaction.id.desc())
                      .all())

        for t in range_txns:
            status = _status_of(t)
            status_counts['all'] += 1
            status_counts[status] += 1
            if t.is_approved and not t.is_rejected:
                if t.entry_type == 'debit':
                    money_in += t.amount or 0
                else:
                    money_out += t.amount or 0

        expected_balance = range_opening_balance + money_in - money_out

        latest_close = (AccountDailyClose.query
                         .filter(AccountDailyClose.account_id == selected_account.id,
                                 AccountDailyClose.close_date <= date_to)
                         .order_by(AccountDailyClose.close_date.desc(), AccountDailyClose.closed_at.desc())
                         .first())
        actual_balance = latest_close.actual_balance if latest_close else expected_balance
        variance = actual_balance - expected_balance

        running = range_opening_balance
        for t in range_txns:
            if row_status != 'all' and _status_of(t) != row_status:
                continue
            delta = (t.amount or 0) if t.entry_type == 'debit' else -(t.amount or 0)
            running += delta

            if t.entry_type == 'debit':
                from_label = t.counterparty_account.name if t.counterparty_account else (t.payee or 'External')
                to_label = selected_account.name
            else:
                from_label = selected_account.name
                if t.counterparty_account:
                    to_label = t.counterparty_account.name
                elif t.payee:
                    to_label = t.payee
                elif t.expense:
                    to_label = t.expense.expense_number
                else:
                    to_label = 'Expense Payment'

            rows.append({'txn': t, 'running': running, 'expense': t.expense,
                        'from_label': from_label, 'to_label': to_label})

        daily_closes = (AccountDailyClose.query
                        .filter_by(account_id=selected_account.id)
                        .order_by(AccountDailyClose.close_date.desc())
                        .all())

    # No active_module passed here on purpose — that's what turns on the
    # floating "Advanced Filter" button (base.html), and this page has its
    # own dedicated filter bar, so it isn't needed.
    return render_template('accounting/account_activity.html',
                           accounts=filtered_accounts,
                           all_accounts=all_accounts,
                           account_type_options=account_type_options,
                           custodian_options=custodian_options,
                           location_options=location_options,
                           selected_account=selected_account,
                           selected_account_id=account_id,
                           selected_account_type=account_type,
                           selected_custodian=custodian,
                           selected_location=location,
                           date_from=date_from,
                           date_to=date_to,
                           row_status=row_status,
                           status_counts=status_counts,
                           rows=rows,
                           range_opening_balance=range_opening_balance,
                           money_in=money_in,
                           money_out=money_out,
                           expected_balance=expected_balance,
                           actual_balance=actual_balance,
                           variance=variance,
                           latest_close=latest_close,
                           daily_closes=daily_closes,
                           today=today,
                           current_status='account_activity')


@bp.route('/expenses/account-activity/export/excel')
@login_required
def account_activity_export_excel():
    """Excel download of one account's activity — same 'account-wise ledger'
    layout as the Journal module's own export (journal.export_excel), applied
    to Expense's own independent accounts."""
    import io
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from app.models import ExpenseAccount, ExpenseAccountTransaction

    account_id = request.args.get('account_id', type=int)
    selected_account = ExpenseAccount.query.get(account_id) if account_id else None
    if not selected_account:
        flash('Select an account first.', 'warning')
        return redirect(url_for('accounting.account_activity'))

    txns = (ExpenseAccountTransaction.query
            .filter_by(account_id=selected_account.id)
            .order_by(ExpenseAccountTransaction.date.asc(), ExpenseAccountTransaction.id.asc())
            .all())

    wb = Workbook()
    ws = wb.active
    ws.title = 'Account Activity'

    title_font = Font(bold=True, size=14)
    acct_font = Font(bold=True, size=12, color='FFFFFF')
    acct_fill = PatternFill('solid', fgColor='0D6EFD')
    head_font = Font(bold=True)
    head_fill = PatternFill('solid', fgColor='E9ECEF')
    total_font = Font(bold=True)
    total_fill = PatternFill('solid', fgColor='F1F3F5')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal='right')

    headers = ['Date', 'Entry #', 'Reference', 'Description', 'Debit (In)', 'Credit (Out)', 'Balance']
    NCOL = len(headers)

    ws['A1'] = 'Expense Account Activity'
    ws['A1'].font = title_font
    ws['A2'] = f'Account: {selected_account.name}'
    ws['A2'].font = Font(italic=True, color='6C757D')
    r = 4

    opening = selected_account.opening_balance or 0
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    cell = ws.cell(row=r, column=1,
                   value=selected_account.name
                         + (f"  ({selected_account.account_type})" if selected_account.account_type else ''))
    cell.font = acct_font
    cell.fill = acct_fill
    r += 1

    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = head_font
        c.fill = head_fill
        c.border = border
        if ci >= 5:
            c.alignment = right
    r += 1

    running = opening
    c = ws.cell(row=r, column=4, value='Opening Balance')
    c.font = total_font; c.alignment = right; c.border = border
    c = ws.cell(row=r, column=7, value=running)
    c.font = total_font; c.alignment = right; c.border = border; c.number_format = '#,##0'
    for ci in (1, 2, 3, 5, 6):
        ws.cell(row=r, column=ci).border = border
    r += 1

    sub_debit = sub_credit = 0
    for t in txns:
        debit = t.amount if t.entry_type == 'debit' else 0
        credit = t.amount if t.entry_type == 'credit' else 0
        # Debit adds (+), credit subtracts (−) from the running balance.
        running += debit - credit
        sub_debit += debit
        sub_credit += credit
        row_vals = [
            t.date.strftime('%d-%m-%Y') if t.date else '',
            f'#{t.id}',
            t.reference or '',
            t.description or '',
            debit or '',
            credit or '',
            running,
        ]
        for ci, v in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = border
            if ci >= 5:
                c.alignment = right
                if v != '':
                    c.number_format = '#,##0'
        r += 1

    # Remaining balance row for this account (running == closing balance).
    c = ws.cell(row=r, column=4, value='Remaining Balance')
    c.font = total_font; c.fill = total_fill; c.alignment = right; c.border = border
    for ci, v in [(5, sub_debit), (6, sub_credit), (7, running)]:
        c = ws.cell(row=r, column=ci, value=v)
        c.font = total_font; c.fill = total_fill; c.alignment = right
        c.border = border; c.number_format = '#,##0'
    for ci in (1, 2, 3):
        ws.cell(row=r, column=ci).fill = total_fill
        ws.cell(row=r, column=ci).border = border
    r += 2

    if not txns:
        ws.cell(row=r, column=1, value='No transactions for this account.')
    else:
        c = ws.cell(row=r, column=4, value='GRAND TOTAL')
        c.font = Font(bold=True, size=12); c.alignment = right
        for ci, v in [(5, sub_debit), (6, sub_credit), (7, running)]:
            c = ws.cell(row=r, column=ci, value=v)
            c.font = Font(bold=True, size=12); c.alignment = right; c.number_format = '#,##0'

    widths = [14, 10, 18, 38, 14, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    import re
    safe_name = re.sub(r'[^A-Za-z0-9_-]+', '_', selected_account.name).strip('_') or 'account'
    return send_file(buf, as_attachment=True,
                     download_name=f'account_activity_{safe_name}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── Expense's own account management (create/edit/delete) — AJAX/JSON ─────
# Entirely independent of the Journal module's equivalent routes. Read
# access only needs login (matches how Add/Edit Expense already shows the
# account dropdown to any user with accounting access); create/edit/delete
# are admin-only, matching the "Manage Accounts" button which is only
# rendered for admins in the first place.

@bp.route('/expense-accounts/list-json')
@login_required
def list_expense_accounts_json():
    from app.models import ExpenseAccount
    accounts = ExpenseAccount.query.order_by(ExpenseAccount.name).all()
    return jsonify({'success': True, 'accounts': [{
        'id': a.id,
        'name': a.name,
        'account_type': a.account_type or '',
        'custodian_name': a.custodian_name or '',
        'location': a.location or '',
        'opening_balance': round(a.opening_balance or 0, 2),
        'balance': round(a.balance, 2),
        'is_active': bool(a.is_active),
        'has_lines': bool(a.transactions),
        'image_url': (url_for('static', filename=a.image_path.replace('app/static/', '')) if a.image_path else None),
    } for a in accounts]})


@bp.route('/expense-accounts/create-quick', methods=['POST'])
@login_required
def create_expense_account_quick():
    from app.models import ExpenseAccount
    if not getattr(current_user, 'is_admin', False):
        return jsonify({'success': False, 'message': 'Only admins can create accounts.'}), 403

    name = (request.form.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Account name is required.'})
    if ExpenseAccount.query.filter(db.func.lower(ExpenseAccount.name) == name.lower()).first():
        return jsonify({'success': False, 'message': f'An account named "{name}" already exists.'})

    try:
        opening_balance = float(request.form.get('opening_balance') or 0)
    except (TypeError, ValueError):
        opening_balance = 0.0

    image_path = _save_fixed_expense_bill_image(request.files.get('image'))

    acct = ExpenseAccount(
        name=name,
        account_type=(request.form.get('account_type') or '').strip() or None,
        opening_balance=opening_balance,
        notes=(request.form.get('notes') or '').strip() or None,
        custodian_name=(request.form.get('custodian_name') or '').strip() or None,
        location=(request.form.get('location') or '').strip() or None,
        linked_funding_account_id=request.form.get('linked_funding_account_id', type=int) or None,
        image_path=image_path,
        created_by=current_user.id,
    )
    db.session.add(acct)
    db.session.commit()
    return jsonify({'success': True, 'account': {
        'id': acct.id, 'name': acct.name, 'account_type': acct.account_type or '',
        'custodian_name': acct.custodian_name or '', 'location': acct.location or '',
        'image_url': (url_for('static', filename=image_path.replace('app/static/', '')) if image_path else None),
    }})


@bp.route('/expense-accounts/<int:account_id>/edit-quick', methods=['POST'])
@login_required
def edit_expense_account_quick(account_id):
    from app.models import ExpenseAccount
    if not getattr(current_user, 'is_admin', False):
        return jsonify({'success': False, 'message': 'Only admins can edit accounts.'}), 403

    acct = ExpenseAccount.query.get_or_404(account_id)
    name = (request.form.get('name') or '').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Account name is required.'})
    dupe = ExpenseAccount.query.filter(
        db.func.lower(ExpenseAccount.name) == name.lower(),
        ExpenseAccount.id != acct.id
    ).first()
    if dupe:
        return jsonify({'success': False, 'message': f'Another account named "{name}" already exists.'})

    try:
        opening_balance = float(request.form.get('opening_balance') or 0)
    except (TypeError, ValueError):
        opening_balance = 0.0

    acct.name = name
    acct.account_type = (request.form.get('account_type') or '').strip() or None
    acct.opening_balance = opening_balance
    acct.custodian_name = (request.form.get('custodian_name') or '').strip() or None
    acct.location = (request.form.get('location') or '').strip() or None
    acct.linked_funding_account_id = request.form.get('linked_funding_account_id', type=int) or None
    if 'is_active' in request.form:
        acct.is_active = str(request.form.get('is_active')).lower() in ('1', 'true', 'on', 'yes')
    db.session.commit()
    return jsonify({'success': True, 'account': {
        'id': acct.id, 'name': acct.name, 'account_type': acct.account_type or '',
        'custodian_name': acct.custodian_name or '', 'location': acct.location or '',
        'opening_balance': round(acct.opening_balance or 0, 2),
        'balance': round(acct.balance, 2),
        'is_active': bool(acct.is_active),
        'has_lines': bool(acct.transactions),
    }})


@bp.route('/expense-accounts/<int:account_id>/delete-quick', methods=['POST'])
@login_required
def delete_expense_account_quick(account_id):
    from app.models import ExpenseAccount
    if not getattr(current_user, 'is_admin', False):
        return jsonify({'success': False, 'message': 'Only admins can delete accounts.'}), 403

    acct = ExpenseAccount.query.get_or_404(account_id)
    if acct.transactions:
        return jsonify({'success': False,
                        'message': 'Cannot delete an account that has transactions. Deactivate it instead.'})
    db.session.delete(acct)
    db.session.commit()
    return jsonify({'success': True})


# ─── Transfer Funds / Daily Cash Close / Reconciliation — AJAX/JSON ────────
# A transfer moves money between two of Expense's own accounts as a single
# atomic pair: a Credit (money out) on the source account and a matching
# Debit (money in) on the destination account, linked by transfer_group so
# they always show up together on both accounts' ledgers. Admin-only, same
# as creating/editing/deleting accounts.

@bp.route('/expense-accounts/transfer-quick', methods=['POST'])
@login_required
def transfer_funds_quick():
    from app.models import ExpenseAccount, ExpenseAccountTransaction
    import uuid

    if not getattr(current_user, 'is_admin', False):
        return jsonify({'success': False, 'message': 'Only admins can transfer funds between accounts.'}), 403

    from_id = request.form.get('from_account_id', type=int)
    to_id = request.form.get('to_account_id', type=int)
    if not from_id or not to_id:
        return jsonify({'success': False, 'message': 'Select both a From Account and a To Account.'})
    if from_id == to_id:
        return jsonify({'success': False, 'message': 'From Account and To Account must be different.'})

    from_acct = ExpenseAccount.query.get(from_id)
    to_acct = ExpenseAccount.query.get(to_id)
    if not from_acct or not to_acct:
        return jsonify({'success': False, 'message': 'One of the selected accounts no longer exists.'})
    if not from_acct.is_active or not to_acct.is_active:
        return jsonify({'success': False, 'message': 'Both accounts must be active to transfer funds.'})

    try:
        amount = float(request.form.get('amount') or 0)
    except (TypeError, ValueError):
        amount = 0
    if amount <= 0:
        return jsonify({'success': False, 'message': 'Enter a transfer amount greater than zero.'})

    date_str = request.form.get('date')
    try:
        txn_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else datetime.utcnow().date()
    except ValueError:
        txn_date = datetime.utcnow().date()

    reference = (request.form.get('reference') or '').strip() or None
    description = (request.form.get('description') or '').strip() or None
    bill_image_path = _save_fixed_expense_bill_image(request.files.get('bill_image'))

    transfer_group = uuid.uuid4().hex
    now = datetime.utcnow()

    out_txn = ExpenseAccountTransaction(
        account_id=from_acct.id, date=txn_date, entry_type='credit', amount=amount,
        description=description or f'Funds transferred to {to_acct.name}',
        reference=reference, transaction_type='transfer', bill_image_path=bill_image_path,
        counterparty_account_id=to_acct.id, transfer_group=transfer_group, payee=to_acct.name,
        is_approved=True, approved_by=current_user.id, approved_at=now, created_by=current_user.id,
    )
    in_txn = ExpenseAccountTransaction(
        account_id=to_acct.id, date=txn_date, entry_type='debit', amount=amount,
        description=description or f'Funds received from {from_acct.name}',
        reference=reference, transaction_type='transfer', bill_image_path=bill_image_path,
        counterparty_account_id=from_acct.id, transfer_group=transfer_group, payee=from_acct.name,
        is_approved=True, approved_by=current_user.id, approved_at=now, created_by=current_user.id,
    )
    db.session.add(out_txn)
    db.session.add(in_txn)
    db.session.commit()

    return jsonify({'success': True,
                    'message': f'PKR {amount:,.2f} transferred from {from_acct.name} to {to_acct.name}.'})


@bp.route('/expense-accounts/transfers')
@login_required
def all_transfers():
    """Every Transfer Funds movement across every account in one flat,
    searchable list — one row per transfer, keyed on its 'out' (credit) leg;
    the matching 'in' (debit) leg shares the same transfer_group and mirrors
    it, so showing just one side avoids listing each transfer twice."""
    from app.models import ExpenseAccountTransaction

    date_from_raw = (request.args.get('date_from') or '').strip()
    date_to_raw = (request.args.get('date_to') or '').strip()
    q = (request.args.get('q') or '').strip()

    def _parse_date(raw):
        if not raw:
            return None
        try:
            return datetime.strptime(raw, '%Y-%m-%d').date()
        except ValueError:
            return None

    date_from = _parse_date(date_from_raw)
    date_to = _parse_date(date_to_raw)

    query = ExpenseAccountTransaction.query.filter(
        ExpenseAccountTransaction.transaction_type == 'transfer',
        ExpenseAccountTransaction.entry_type == 'credit',
    )
    if date_from:
        query = query.filter(ExpenseAccountTransaction.date >= date_from)
    if date_to:
        query = query.filter(ExpenseAccountTransaction.date <= date_to)

    transfers = query.order_by(ExpenseAccountTransaction.date.desc(), ExpenseAccountTransaction.id.desc()).all()

    if q:
        ql = q.lower()

        def _matches(t):
            haystack = ' '.join(filter(None, [
                t.reference, t.description,
                t.account.name if t.account else '',
                t.counterparty_account.name if t.counterparty_account else '',
            ])).lower()
            return ql in haystack

        transfers = [t for t in transfers if _matches(t)]

    total_amount = sum(t.amount or 0 for t in transfers)

    return render_template('accounting/all_transfers.html',
                           transfers=transfers, total_amount=total_amount,
                           date_from=date_from_raw, date_to=date_to_raw, q=q)


@bp.route('/expense-accounts/transfer/<transfer_group>/delete', methods=['POST'])
@login_required
def delete_transfer(transfer_group):
    """Deletes both legs of a transfer (matched by transfer_group). Removing
    the rows IS the reversal — same live-computed-balance convention as
    delete_expense_account_debit_entry. Admin-only: an account-to-account
    transfer moves real money, same gating as creating one."""
    from app.models import ExpenseAccountTransaction
    if not getattr(current_user, 'is_admin', False):
        flash('Only admins can delete a transfer.', 'danger')
        return redirect(url_for('accounting.all_transfers'))

    legs = ExpenseAccountTransaction.query.filter_by(
        transfer_group=transfer_group, transaction_type='transfer').all()
    if not legs:
        flash('Transfer not found.', 'warning')
        return redirect(url_for('accounting.all_transfers'))

    for leg in legs:
        db.session.delete(leg)
    db.session.commit()
    flash('Transfer deleted — both accounts\' balances have been reversed.', 'success')
    return redirect(url_for('accounting.all_transfers'))


@bp.route('/expense-accounts/transfer/<transfer_group>/edit-quick', methods=['POST'])
@login_required
def edit_transfer_quick(transfer_group):
    """Edits both legs of a transfer (matched by transfer_group) — amount,
    date, reference, description, and optionally a new bill image. Never
    re-targets which two accounts were involved (that would mean undoing one
    transfer and creating a different one, not editing this one) — use
    Delete + a fresh Transfer Funds for that instead. Admin-only, same
    gating as creating or deleting a transfer."""
    from app.models import ExpenseAccountTransaction
    if not getattr(current_user, 'is_admin', False):
        return jsonify({'success': False, 'message': 'Only admins can edit a transfer.'}), 403

    legs = ExpenseAccountTransaction.query.filter_by(
        transfer_group=transfer_group, transaction_type='transfer').all()
    if len(legs) != 2:
        return jsonify({'success': False, 'message': 'Transfer not found.'})

    try:
        amount = float(request.form.get('amount') or 0)
    except (TypeError, ValueError):
        amount = 0
    if amount <= 0:
        return jsonify({'success': False, 'message': 'Enter an amount greater than zero.'})

    date_str = request.form.get('date')
    try:
        txn_date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else None
    except ValueError:
        txn_date = None

    reference = (request.form.get('reference') or '').strip() or None
    description = (request.form.get('description') or '').strip() or None
    new_bill = _save_fixed_expense_bill_image(request.files.get('bill_image'))

    for leg in legs:
        leg.amount = amount
        if txn_date:
            leg.date = txn_date
        leg.reference = reference
        if description:
            leg.description = description
        if new_bill:
            leg.bill_image_path = new_bill

    db.session.commit()
    return jsonify({'success': True, 'message': 'Transfer updated.'})


@bp.route('/expense-accounts/<int:account_id>/daily-close-quick', methods=['POST'])
@login_required
def daily_close_quick(account_id):
    from app.models import ExpenseAccount, AccountDailyClose

    account = ExpenseAccount.query.get_or_404(account_id)
    try:
        actual_balance = float(request.form.get('actual_balance'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Enter the counted actual balance.'})

    notes = (request.form.get('notes') or '').strip() or None
    expected_balance = account.balance
    today = datetime.utcnow().date()

    close = AccountDailyClose.query.filter_by(account_id=account.id, close_date=today).first()
    if not close:
        close = AccountDailyClose(account_id=account.id, close_date=today)
        db.session.add(close)

    close.expected_balance = expected_balance
    close.actual_balance = actual_balance
    close.notes = notes
    close.closed_by = current_user.id
    close.closed_at = datetime.utcnow()
    # Re-closing the same day resets any prior reconciliation — finance needs to re-verify.
    close.is_reconciled = False
    close.reconciled_by = None
    close.reconciled_at = None

    db.session.commit()

    return jsonify({'success': True, 'message': f'Daily cash close saved for {account.name}.',
                    'expected_balance': round(expected_balance, 2),
                    'actual_balance': round(actual_balance, 2),
                    'variance': round(close.variance, 2)})


@bp.route('/account-daily-closes/<int:close_id>/reconcile-quick', methods=['POST'])
@login_required
def reconcile_daily_close_quick(close_id):
    from app.models import AccountDailyClose
    if not getattr(current_user, 'is_admin', False):
        return jsonify({'success': False, 'message': 'Only admins can mark a close as reconciled.'}), 403

    close = AccountDailyClose.query.get_or_404(close_id)
    close.is_reconciled = True
    close.reconciled_by = current_user.id
    close.reconciled_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'success': True, 'message': 'Marked as reconciled.'})


def _parse_shifted_product_costs(raw, expense_amount):
    """Parse `Expense.shifted_to_product_ids` into {product_id: ('set', old_cost) | ('add', amount)}
    describing how to reverse the shift on that item.

    Current format is 'pid:new_cost:old_cost,...' — the item's cost is SET to
    new_cost when shifted, so undoing it restores old_cost directly ('set').
    Two older formats are also understood, both additive (the item's cost was
    increased by an amount rather than replaced, so undoing subtracts it back
    off, 'add'): the per-item 'pid:amount,...' format used before shifts
    replaced cost outright, and the original equal-split plain-id format
    ('pid,pid,...') from before per-item costs existed at all."""
    if not raw:
        return {}
    tokens = [t.strip() for t in raw.split(',') if t.strip()]
    if not tokens:
        return {}
    first_parts = tokens[0].split(':')
    if len(first_parts) >= 3:
        result = {}
        for t in tokens:
            parts = t.split(':')
            if len(parts) < 3:
                continue
            try:
                pid, old_cost = int(parts[0]), float(parts[2])
            except (ValueError, IndexError):
                continue
            result[pid] = ('set', old_cost)
        return result
    if len(first_parts) == 2:
        result = {}
        for t in tokens:
            try:
                pid_str, amt_str = t.split(':', 1)
                result[int(pid_str)] = ('add', float(amt_str))
            except (ValueError, IndexError):
                continue
        return result
    ids = [int(t) for t in tokens if t.isdigit()]
    if not ids:
        return {}
    per = expense_amount / len(ids)
    return {pid: ('add', per) for pid in ids}


@bp.route('/expense/shift-to-inventory', methods=['POST'])
@login_required
def shift_expense_to_inventory():
    """Shift an operating expense onto one or more inventory items' cost.

    Each selected item's share defaults to an equal split of the expense
    amount but is editable in the popup, so the amounts actually applied may
    differ per item — they must still add up to the expense's full amount.
    Each item's cost_price is REPLACED with its share (not added on top of
    whatever it already was) — the popup's "Cost to Add" value becomes the
    item's new cost outright. Item quantities are left unchanged.
    """
    if not getattr(current_user, 'is_admin', False):
        return jsonify({'success': False, 'message': 'Admin access required.'}), 403

    expense_id = request.form.get('expense_id', type=int)
    # Accept both `product_ids` and `product_ids[]` from the multi-select
    raw_ids = request.form.getlist('product_ids') or request.form.getlist('product_ids[]')

    if not expense_id or not raw_ids:
        return jsonify({'success': False, 'message': 'Please select at least one inventory item.'}), 400

    from app.models import Product

    expense = Expense.query.get_or_404(expense_id)

    # Only plain operating expenses can be shifted to inventory cost.
    if getattr(expense, 'is_inventory_shifted', False):
        return jsonify({'success': False, 'message': 'This expense is already shifted to inventory.'}), 400
    if getattr(expense, 'is_shifted', False):
        return jsonify({'success': False, 'message': 'This expense is already shifted to a PD project.'}), 400
    if getattr(expense, 'is_bom_overhead', False):
        return jsonify({'success': False, 'message': 'Overhead expenses cannot be shifted to inventory.'}), 400
    if getattr(expense, 'is_monthly_divided', False):
        return jsonify({'success': False, 'message': 'Monthly divided expenses cannot be shifted to inventory.'}), 400
    if getattr(expense, 'is_payment_transfer', False):
        return jsonify({'success': False, 'message': 'This expense was transferred to a Sale/Purchase payment and cannot be shifted to inventory.'}), 400

    # Resolve unique, valid product ids
    seen = set()
    product_ids = []
    for rid in raw_ids:
        try:
            pid = int(rid)
        except (TypeError, ValueError):
            continue
        if pid and pid not in seen:
            seen.add(pid)
            product_ids.append(pid)

    products = Product.query.filter(Product.id.in_(product_ids)).all()
    if not products:
        return jsonify({'success': False, 'message': 'No valid inventory items found.'}), 400

    # Per-item new cost — the "Shift to Inventory Cost" popup pre-fills each
    # item's share as an equal split of the expense amount but leaves it
    # editable, so the values actually submitted may differ per item. They
    # must still add up to the expense's full amount. Each value REPLACES
    # that item's cost_price outright — it is not added on top of it.
    item_costs = {}
    total_entered = 0.0
    for product in products:
        raw_cost = (request.form.get(f'item_cost_{product.id}') or '').strip()
        try:
            cost = round(float(raw_cost), 2) if raw_cost else None
        except (TypeError, ValueError):
            cost = None
        if cost is None or cost < 0:
            return jsonify({'success': False, 'message': f'Enter a valid cost amount for {product.name}.'}), 400
        item_costs[product.id] = cost
        total_entered += cost

    if abs(round(total_entered - expense.amount, 2)) > 0.01:
        return jsonify({'success': False,
                        'message': f'The item costs must add up to the expense amount (PKR {expense.amount:,.2f}). '
                                   f'They currently total PKR {total_entered:,.2f}.'}), 400

    try:
        applied_names = []
        shifted_tokens = []
        for product in products:
            old_cost = product.cost_price or 0
            new_cost = item_costs[product.id]
            product.cost_price = new_cost
            applied_names.append(f'{product.name} (PKR {old_cost:,.2f} -> PKR {new_cost:,.2f})')
            shifted_tokens.append(f'{product.id}:{new_cost}:{old_cost}')

        expense.is_inventory_shifted = True
        expense.shifted_to_product_ids = ','.join(shifted_tokens)

        db.session.commit()

        log_activity('Accounting',
                     f'Shifted expense {expense.expense_number} to inventory cost',
                     f'PKR {expense.amount} applied across {len(products)} item(s): {", ".join(applied_names)}')

        if len(products) == 1:
            msg = f'{products[0].name} cost set to PKR {item_costs[products[0].id]:,.2f}.'
        else:
            msg = f'Cost updated on {len(products)} items (PKR {expense.amount:,.2f} total).'
        return jsonify({'success': True, 'message': msg})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/expense/unshift-from-inventory', methods=['POST'])
@login_required
def unshift_expense_from_inventory():
    """Reverse an inventory cost shift — restores each item's cost back to
    what it was before the shift (or, for a shift made before costs were
    replaced outright, subtracts the amount that had been added)."""
    if not getattr(current_user, 'is_admin', False):
        return jsonify({'success': False, 'message': 'Admin access required.'}), 403

    from app.models import Product

    expense_id = request.form.get('expense_id', type=int)
    expense = Expense.query.get_or_404(expense_id)

    if not getattr(expense, 'is_inventory_shifted', False):
        return jsonify({'success': False, 'message': 'This expense is not shifted to inventory.'}), 400

    try:
        item_specs = _parse_shifted_product_costs(expense.shifted_to_product_ids, expense.amount)
        products = Product.query.filter(Product.id.in_(item_specs.keys())).all() if item_specs else []
        for product in products:
            mode, value = item_specs.get(product.id, ('add', 0))
            if mode == 'set':
                product.cost_price = max(0, value)
            else:
                product.cost_price = max(0, (product.cost_price or 0) - value)

        expense.is_inventory_shifted = False
        expense.shifted_to_product_ids = None

        db.session.commit()
        log_activity('Accounting',
                     f'Reversed inventory shift for expense {expense.expense_number}',
                     f'PKR {expense.amount} removed from {len(products)} item(s)')
        return jsonify({'success': True, 'message': 'Inventory shift reversed.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/expense/add', methods=['GET', 'POST'])
@login_required
@permission_required('accounting', action='add')
def add_expense():
    from app.models import ExpenseCategory, Vendor, BOM, ExpenseAccount
    from werkzeug.utils import secure_filename
    import os

    if request.method == 'POST' and request.form.get('txn_type') == 'debit':
        return _add_money_from_expense_form()

    form = ExpenseForm()
    
    # Populate category choices
    categories = ExpenseCategory.query.filter_by(is_active=True).order_by(ExpenseCategory.name).all()
    form.category_id.choices = [(cat.id, cat.name) for cat in categories]
    
    # Populate vendor choices
    vendors = Vendor.query.filter_by(is_active=True).order_by(Vendor.name).all()
    form.vendor_id.choices = [(0, 'Select Vendor (Optional)')] + [(v.id, v.name) for v in vendors]

    # Populate customer / warehouse choices
    from app.models import Customer, Warehouse
    customers = Customer.query.filter_by(is_active=True).order_by(Customer.name).all()
    form.customer_id.choices = [(0, 'Select Customer (Optional)')] + [(c.id, c.name) for c in customers]
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    form.warehouse_id.choices = [(0, 'Select Warehouse (Optional)')] + [(w.id, w.name) for w in warehouses]

    # Populate Payment Method choices
    from app.models import PaymentMethod
    methods = PaymentMethod.query.filter_by(is_active=True).order_by(PaymentMethod.name).all()
    form.payment_method.choices = [(m.name, m.name) for m in methods]
    
    # Populate manufactured product choices (if column exists)
    from app.models import Product
    if has_column('products', 'is_manufactured'):
        manufactured_products = Product.query.filter_by(is_manufactured=True, is_active=True).order_by(Product.name).all()
    else:
        manufactured_products = []
    form.product_id.choices = [(0, 'Select Finished Product (Optional)')] + [(p.id, p.name) for p in manufactured_products]
    
    # Populate BOM choices
    boms = BOM.query.filter_by(is_active=True).order_by(BOM.name).all()
    form.bom_id.choices = [(0, 'Select BOM (Optional)')] + [(b.id, b.name) for b in boms]
    
    # Populate In Progress Manufacturing Order choices (no placeholder; Select2 shows placeholder text)
    from app.models import ManufacturingOrder
    in_progress_mos = (
        ManufacturingOrder.query
        .filter_by(status='In Progress')
        .order_by(ManufacturingOrder.order_number)
        .all()
    )
    form.mo_id.choices = [(mo.id, f"{mo.order_number} — {mo.bom.product.name}") for mo in in_progress_mos]
    
    if form.validate_on_submit():
        # Get selected targets
        base_amount = form.amount.data
        is_overhead = form.is_bom_overhead.data
        selected_mo_id = (form.mo_id.data or 0) if is_overhead else 0
        
        # Determine the active mode
        mode = request.form.get('overhead_mode', 'bulk')
        is_admin = getattr(current_user, 'is_admin', False)
        target_status = 'confirmed' if is_admin else 'pending'
        
        # Handle bill image upload
        bill_path = None
        if 'bill_image' in request.files:
            bill_file = request.files['bill_image']
            if bill_file and bill_file.filename:
                import time, uuid
                original_filename = secure_filename(bill_file.filename)
                unique_prefix = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
                filename = f"{unique_prefix}_{original_filename}"
                
                full_bill_path = os.path.join('app', 'static', 'uploads', 'bills', filename)
                os.makedirs(os.path.dirname(full_bill_path), exist_ok=True)
                bill_file.save(full_bill_path)
                # Store path relative to project root
                bill_path = f"app/static/uploads/bills/{filename}"
        
        common_kwargs = dict(
            date=form.date.data,
            category_id=form.category_id.data,
            vendor_id=form.vendor_id.data if form.vendor_id.data != 0 else None,
            customer_id=form.customer_id.data if form.customer_id.data != 0 else None,
            warehouse_id=form.warehouse_id.data if form.warehouse_id.data != 0 else None,
            description=form.description.data,
            payment_method=form.payment_method.data,
            reference=form.reference.data,
            notes=form.notes.data,
            is_bom_overhead=is_overhead,
            bill_image_path=bill_path,
            is_monthly_divided=form.is_monthly_divided.data,
            monthly_start_date=form.monthly_start_date.data if form.is_monthly_divided.data else None,
            monthly_end_date=form.monthly_end_date.data if form.is_monthly_divided.data else None,
            is_approved=is_admin,
            approved_by=current_user.id if is_admin else None,
            approved_at=datetime.utcnow() if is_admin else None,
            is_rejected=False
        )
        
        # Get expense number settings
        settings = ExpenseSettings.query.first()
        if not settings:
            settings = ExpenseSettings(expense_prefix='EXP-', expense_suffix='', next_number=1)
            db.session.add(settings)

        # Sync next_expense_num with actual highest expense number in DB
        highest_expense = Expense.query.order_by(Expense.id.desc()).first()
        if highest_expense and highest_expense.expense_number:
            try:
                prefix_len = len(settings.expense_prefix or '')
                suffix_len = len(settings.expense_suffix or '')
                num_str = highest_expense.expense_number[prefix_len:]
                if suffix_len > 0:
                    num_str = num_str[:-suffix_len]
                max_num = int(num_str)
                next_expense_num = max(settings.next_number, max_num + 1)
            except (ValueError, IndexError):
                next_expense_num = settings.next_number
        else:
            next_expense_num = settings.next_number
        
        created_expenses = []   # track all new expense objects
        
        # ── MODE 1: Direct MO link (Single or Bulk) ────────────────────────
        if is_overhead and mode == 'mo':
            selected_mo_ids = form.mo_id.data if form.mo_id.data else []
            from app.models import ManufacturingOrder
            valid_mos = []
            for mo_id in selected_mo_ids:
                if mo_id != 0:
                    target_mo = ManufacturingOrder.query.get(mo_id)
                    if target_mo and target_mo.status == 'In Progress':
                        valid_mos.append(target_mo)
            
            if not valid_mos and any(m != 0 for m in selected_mo_ids):
                flash('Invalid or completed Manufacturing Order(s) selected.', 'danger')
                return redirect(url_for('accounting.add_expense'))

            num_mos = len(valid_mos)
            
            if num_mos == 0:
                # No specific MO selected
                expense_number, next_expense_num = get_unique_expense_number(settings, next_expense_num)
                exp = Expense(
                    expense_number=expense_number,
                    amount=base_amount,
                    status=target_status,
                    created_by=current_user.id,
                    **common_kwargs
                )
                if exp.is_monthly_divided:
                    exp.calculate_daily_amount()
                db.session.add(exp)
                created_expenses.append(exp)
                if target_status == 'confirmed':
                    flash_msg = f'Overhead expense PKR {base_amount} added (Unassigned).'
                else:
                    flash_msg = f'Overhead expense PKR {base_amount} created (Unassigned) and waiting for admin confirmation.'
            elif num_mos == 1:
                target_mo = valid_mos[0]
                expense_number, next_expense_num = get_unique_expense_number(settings, next_expense_num)
                exp = Expense(
                    expense_number=expense_number,
                    amount=base_amount,
                    status=target_status,
                    mo_id=target_mo.id,
                    created_by=current_user.id,
                    **common_kwargs
                )
                if exp.is_monthly_divided:
                    exp.calculate_daily_amount()
                db.session.add(exp)
                created_expenses.append(exp)
                
                # Update MO overhead only if confirmed
                if target_status == 'confirmed':
                    target_mo.actual_overhead_cost = (target_mo.actual_overhead_cost or 0) + base_amount
                    target_mo.total_cost = (target_mo.actual_material_cost or 0) + (target_mo.actual_labor_cost or 0) + target_mo.actual_overhead_cost
                    flash_msg = f'Overhead expense PKR {base_amount} added and linked to {target_mo.order_number}.'
                else:
                    flash_msg = f'Overhead expense PKR {base_amount} created for {target_mo.order_number} and is waiting for admin confirmation.'
            else:
                amount_per_mo = base_amount / num_mos
                for i, target_mo in enumerate(valid_mos):
                    kwargs = dict(common_kwargs)
                    kwargs['expense_number'], next_expense_num = get_unique_expense_number(settings, next_expense_num)
                    kwargs['description'] = f"{form.description.data} (Allocation {i+1}/{num_mos})"
                    kwargs['amount'] = amount_per_mo
                    kwargs['mo_id'] = target_mo.id

                    exp = Expense(
                        status=target_status,
                        created_by=current_user.id,
                        **kwargs
                    )
                    if exp.is_monthly_divided:
                        exp.calculate_daily_amount()
                    db.session.add(exp)
                    created_expenses.append(exp)

                    if target_status == 'confirmed':
                        target_mo.actual_overhead_cost = (target_mo.actual_overhead_cost or 0) + amount_per_mo
                        target_mo.total_cost = (target_mo.actual_material_cost or 0) + (target_mo.actual_labor_cost or 0) + target_mo.actual_overhead_cost
                
                if target_status == 'confirmed':
                    flash_msg = f'Expense(s) added. PKR {base_amount} divided into {num_mos} Manufacturing Orders.'
                else:
                    flash_msg = f'Expense(s) created and waiting for admin confirmation. PKR {base_amount} allocated to {num_mos} Manufacturing Orders.'
        
        # ── MODE 2: Bulk Product/BOM allocation ───────────────────────────
        else:
            selected_product_ids = form.product_id.data if is_overhead else []
            selected_bom_ids = form.bom_id.data if is_overhead else []
            targets = []
            for pid in selected_product_ids:
                if pid and pid != 0: targets.append(('product', pid))
            for bid in selected_bom_ids:
                if bid and bid != 0: targets.append(('bom', bid))
            
            num_targets = len(targets)
            amount_per = base_amount / num_targets if num_targets > 0 else base_amount
            
            if num_targets == 0:
                expense_number, next_expense_num = get_unique_expense_number(settings, next_expense_num)
                # If unassigned but specific IDs provided, attach them
                pid_list = [p for p in (form.product_id.data or []) if p and p != 0]
                bid_list = [b for b in (form.bom_id.data or []) if b and b != 0]
                
                exp = Expense(
                    expense_number=expense_number,
                    amount=base_amount,
                    status=target_status,
                    created_by=current_user.id,
                    product_id=pid_list[0] if pid_list else None,
                    bom_id=bid_list[0] if bid_list else None,
                    **common_kwargs
                )
                if exp.is_monthly_divided:
                    exp.calculate_daily_amount()
                db.session.add(exp)
                created_expenses.append(exp)
            else:
                for i, (target_type, target_id) in enumerate(targets):
                    kwargs = dict(common_kwargs)
                    kwargs['expense_number'], next_expense_num = get_unique_expense_number(settings, next_expense_num)
                    kwargs['description'] = f"{form.description.data} (Allocation {i+1}/{num_targets})"
                    kwargs['amount'] = amount_per
                    kwargs['is_bom_overhead'] = True
                    if target_type == 'product':
                        kwargs['product_id'] = target_id
                    else:
                        kwargs['bom_id'] = target_id
                    exp = Expense(**kwargs)
                    if exp.is_monthly_divided:
                        exp.calculate_daily_amount()
                    db.session.add(exp)
                    created_expenses.append(exp)
            
            if target_status == 'confirmed':
                flash_msg = f'Expense(s) added. PKR {base_amount} divided into {max(1, num_targets)} record(s).'
            else:
                flash_msg = f'Expense(s) created and waiting for admin confirmation. PKR {base_amount} divided into {max(1, num_targets)} record(s).'
        # ─────────────────────────────────────────────────────────────────
        
        # Update expense settings next number
        settings.next_number = next_expense_num

        try:
            account_id = request.form.get('account_id', type=int)
            if account_id:
                db.session.flush()  # assign ids to the new Expense row(s) first
                for exp in created_expenses:
                    _sync_expense_account_transaction(exp, account_id, target_status == 'confirmed')

            # "Add this to Invoice/Purchase Payment" — only meaningful for a
            # single plain expense; a BOM-overhead allocation split across
            # several MOs/products/rows has no one row to hang a real
            # Sale/Bill payment off, so the checkbox is ignored if the form
            # ended up creating more than one Expense row. "Add to Invoice
            # Payment" (sale) never applies here — this is the Credit/expense
            # path (money OUT), so only "Add to Purchase Payment" (bill,
            # paying a vendor) makes sense; sale transfers only happen from
            # the Debit path (_add_money_from_expense_form /
            # _sync_add_money_sale_transfer), money coming IN.
            transfer_type = request.form.get('payment_transfer_type')
            if transfer_type == 'sale':
                transfer_type = None
            transfer_target_id = request.form.get('payment_transfer_target_id', type=int)
            if transfer_type in ('sale', 'bill') and transfer_target_id and len(created_expenses) == 1:
                db.session.flush()
                _sync_expense_payment_transfer(created_expenses[0], transfer_type, transfer_target_id,
                                               created_expenses[0].amount, current_user.id, is_admin)

            db.session.commit()

            log_activity('Accounting', f'Added Expense: {flash_msg}',
                        f'Total: {base_amount}, Mode: {mode}, Status: {target_status}')
            
            flash(flash_msg, 'success')
            return redirect(url_for('accounting.expenses'))
        except Exception as e:
            db.session.rollback()
            if 'UNIQUE constraint failed' in str(e) and 'expense_number' in str(e):
                flash('Failed to create expense due to a numbering conflict. Please try again.', 'danger')
            else:
                flash(f'Error creating expense: {str(e)}', 'danger')
            return redirect(url_for('accounting.add_expense'))

    expense_accounts = ExpenseAccount.query.filter_by(is_active=True).order_by(ExpenseAccount.name).all()
    return render_template('accounting/add_expense.html', form=form, expense_accounts=expense_accounts)


@bp.route('/expense/bulk-upload', methods=['GET', 'POST'])
@login_required
@permission_required('accounting', action='add')
def bulk_upload_expense():
    """Bulk-create/update Expenses from an uploaded Excel sheet (same shape as
    the app's own Expense report export). Deliberately out of scope: BOM
    Overhead linkage, Manufacturing Order allocation, monthly division, and
    the Account/Debit-Credit link — those stay individual-only via
    Add/Edit Expense so a spreadsheet can never silently mis-allocate cost."""
    from app.models import ExpenseCategory, Vendor, ExpenseAccountTransaction
    import re

    if request.method == 'POST':
        if 'file' not in request.files or not request.files['file'].filename:
            flash('No file selected.', 'error')
            return redirect(url_for('accounting.bulk_upload_expense'))

        file = request.files['file']
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls).', 'error')
            return redirect(url_for('accounting.bulk_upload_expense'))

        try:
            from openpyxl import load_workbook
            from io import BytesIO
            wb = load_workbook(filename=BytesIO(file.read()), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.values)
        except Exception as e:
            flash(f'Error reading file: {str(e)}', 'error')
            return redirect(url_for('accounting.bulk_upload_expense'))

        if not rows:
            flash('File is empty.', 'error')
            return redirect(url_for('accounting.bulk_upload_expense'))

        # Some exports (like the app's own) put a title row above the real
        # header — scan the first few rows for one that has both Date and Amount.
        header_idx = None
        for i, r in enumerate(rows[:5]):
            cells = [str(c).strip().lower() if c is not None else '' for c in r]
            if 'date' in cells and 'amount' in cells:
                header_idx = i
                break
        if header_idx is None:
            flash('Could not find a header row with Date and Amount columns.', 'error')
            return redirect(url_for('accounting.bulk_upload_expense'))

        headers = [str(c).strip().lower() if c is not None else '' for c in rows[header_idx]]

        def col(*names):
            for name in names:
                if name in headers:
                    return headers.index(name)
            return None

        idx_num = col('expense #', 'expense#', 'expense number')
        idx_date = col('date')
        idx_category = col('category')
        idx_vendor = col('vendor')
        idx_desc = col('description')
        idx_amount = col('amount')
        idx_approval = col('approval', 'status')

        if None in (idx_date, idx_category, idx_desc, idx_amount):
            flash('The sheet must have Date, Category, Description and Amount columns.', 'error')
            return redirect(url_for('accounting.bulk_upload_expense'))

        settings = ExpenseSettings.query.first()
        if not settings:
            settings = ExpenseSettings(expense_prefix='EXP-', expense_suffix='', next_number=1)
            db.session.add(settings)
            db.session.commit()

        is_admin = getattr(current_user, 'is_admin', False)
        category_cache = {c.name.lower(): c for c in ExpenseCategory.query.all()}
        vendor_cache = {v.name.lower(): v for v in Vendor.query.all()}

        created = 0
        updated = 0
        errors = []

        for i, row in enumerate(rows[header_idx + 1:]):
            row_num = header_idx + 2 + i  # Excel row number, matching what the user sees

            def cell(idx):
                return row[idx] if idx is not None and idx < len(row) else None

            # Skip fully blank rows silently (common at the end of exported sheets)
            if all(c is None or str(c).strip() in ('', '-') for c in row):
                continue

            try:
                raw_amount = cell(idx_amount)
                if raw_amount is None or str(raw_amount).strip() in ('', '-'):
                    errors.append(f'Row {row_num}: Missing amount')
                    continue
                amount_str = re.sub(r'[^\d.\-]', '', str(raw_amount))
                try:
                    amount = float(amount_str)
                except ValueError:
                    errors.append(f'Row {row_num}: Invalid amount "{raw_amount}"')
                    continue
                if amount <= 0:
                    errors.append(f'Row {row_num}: Amount must be greater than zero')
                    continue

                exp_date = _parse_expense_date(cell(idx_date))
                if not exp_date:
                    errors.append(f'Row {row_num}: Invalid or missing date "{cell(idx_date)}"')
                    continue

                cat_name = str(cell(idx_category) or '').strip()
                if not cat_name or cat_name == '-':
                    errors.append(f'Row {row_num}: Missing category')
                    continue
                category = category_cache.get(cat_name.lower())
                if not category:
                    category = ExpenseCategory(name=cat_name)
                    db.session.add(category)
                    db.session.flush()
                    category_cache[cat_name.lower()] = category

                vendor = None
                if idx_vendor is not None:
                    vendor_name = str(cell(idx_vendor) or '').strip()
                    if vendor_name and vendor_name != '-':
                        vendor = vendor_cache.get(vendor_name.lower())
                        if not vendor:
                            vendor = Vendor(name=vendor_name)
                            db.session.add(vendor)
                            db.session.flush()
                            vendor_cache[vendor_name.lower()] = vendor

                description = str(cell(idx_desc) or '').strip()
                if not description:
                    errors.append(f'Row {row_num}: Missing description')
                    continue

                # Approval column -> status flags. Non-admin uploaders always land
                # Pending, same as a normal non-admin Add Expense submission — a
                # spreadsheet can never self-approve on their behalf.
                approval_raw = str(cell(idx_approval) or '').strip().lower() if idx_approval is not None else ''
                status, is_approved, is_rejected, is_draft, rejection_reason = 'pending', False, False, False, None
                if is_admin and approval_raw == 'approved':
                    status, is_approved = 'confirmed', True
                elif is_admin and approval_raw == 'rejected':
                    status, is_rejected, rejection_reason = 'rejected', True, 'Bulk import: marked Rejected'
                elif is_admin and approval_raw == 'draft':
                    is_draft = True

                # Expense # -> upsert match. Exported sheets sometimes glue extra
                # badge text onto this cell (e.g. "EXP-10221 ... BOM Overhead"), so
                # only the first whitespace-delimited token counts as the number.
                raw_num = str(cell(idx_num) or '').strip() if idx_num is not None else ''
                clean_num = raw_num.split()[0] if raw_num and raw_num != '-' else ''
                existing = Expense.query.filter_by(expense_number=clean_num).first() if clean_num else None

                if existing:
                    if existing.is_bom_overhead:
                        errors.append(f'Row {row_num}: {clean_num} is a BOM Overhead expense — edit it individually, not via bulk upload.')
                        continue

                    linked_txn = ExpenseAccountTransaction.query.filter_by(expense_id=existing.id).first()
                    linked_account_id = linked_txn.account_id if linked_txn else None

                    existing.date = exp_date
                    existing.category_id = category.id
                    existing.vendor_id = vendor.id if vendor else None
                    existing.description = description
                    existing.amount = amount
                    existing.status = status
                    existing.is_approved = is_approved
                    existing.is_rejected = is_rejected
                    existing.is_draft = is_draft
                    existing.rejection_reason = rejection_reason
                    if is_approved:
                        existing.approved_by = current_user.id
                        existing.approved_at = datetime.utcnow()

                    if linked_account_id:
                        _sync_expense_account_transaction(existing, linked_account_id, is_approved)

                    updated += 1
                else:
                    expense_number, next_num = get_unique_expense_number(settings, settings.next_number)
                    settings.next_number = next_num
                    exp = Expense(
                        expense_number=expense_number,
                        date=exp_date,
                        category_id=category.id,
                        vendor_id=vendor.id if vendor else None,
                        description=description,
                        amount=amount,
                        status=status,
                        is_approved=is_approved,
                        is_rejected=is_rejected,
                        is_draft=is_draft,
                        rejection_reason=rejection_reason,
                        approved_by=current_user.id if is_approved else None,
                        approved_at=datetime.utcnow() if is_approved else None,
                        created_by=current_user.id,
                    )
                    db.session.add(exp)
                    created += 1

                db.session.commit()
            except Exception as e:
                db.session.rollback()
                errors.append(f'Row {row_num}: {str(e)}')

        if created or updated:
            log_activity('Accounting', 'Bulk Uploaded Expenses',
                        f'Created: {created}, Updated: {updated}, Errors: {len(errors)}')

        if created:
            flash(f'{created} expense(s) created.', 'success')
        if updated:
            flash(f'{updated} expense(s) updated.', 'success')
        if errors:
            shown = '; '.join(errors[:15])
            more = f' … and {len(errors) - 15} more' if len(errors) > 15 else ''
            flash(f'{len(errors)} row(s) skipped: {shown}{more}', 'warning')
        if not created and not updated and not errors:
            flash('No data rows found to import.', 'warning')

        return redirect(url_for('accounting.expenses'))

    return render_template('accounting/bulk_upload_expense.html')


@bp.route('/expense/download-sample')
@login_required
@permission_required('accounting', action='add')
def download_expense_sample():
    from openpyxl import Workbook
    from io import BytesIO
    from flask import send_file

    wb = Workbook()
    ws = wb.active
    ws.title = 'Expenses'

    headers = ['Expense #', 'Date', 'Category', 'Vendor', 'Description', 'Amount', 'Approval']
    ws.append(headers)
    sample_data = [
        ['', '01-04-2026', 'Fuel - Factory', '', 'Bike petrol', 250, 'Pending'],
        ['', '02-04-2026', 'Office Supplies', 'ABC Traders', 'Stationery purchase', 1200, 'Approved'],
        ['EXP-1050', '03-04-2026', 'Utilities', '', 'Electricity bill (updates existing EXP-1050 if it exists)', 5400, 'Approved'],
    ]
    for row in sample_data:
        ws.append(row)
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 40

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='sample_expenses.xlsx', as_attachment=True)


@bp.route('/expense/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('accounting', action='delete')
def delete_expense(id):
    from app.models import BOM
    from app.services.bom_versioning import BOMVersioningService
    
    expense = Expense.query.get_or_404(id)
    
    from app.models import ManufacturingOrder
    # Store info before deletion for BOM array / MO updates
    was_overhead = expense.is_bom_overhead if has_column('expenses', 'is_bom_overhead') else False
    bom_id = expense.bom_id if has_column('expenses', 'bom_id') else None
    product_id = expense.product_id if has_column('expenses', 'product_id') else None
    
    mo_id_to_reduce = None
    amount_to_reduce = expense.amount
    is_confirmed = expense.status == 'confirmed'
    if was_overhead and has_column('expenses', 'mo_id') and expense.mo_id:
        mo_id_to_reduce = expense.mo_id

    description = expense.description

    _delete_linked_expense_account_transaction(expense.id)
    _reverse_expense_payment_transfer(expense, current_user.id)
    db.session.delete(expense)

    # Also reduce actual_overhead_cost and total_cost in MO if it was linked and confirmed
    if mo_id_to_reduce and is_confirmed:
        linked_mo = ManufacturingOrder.query.get(mo_id_to_reduce)
        if linked_mo:
            linked_mo.actual_overhead_cost = max(0, (linked_mo.actual_overhead_cost or 0) - amount_to_reduce)
            linked_mo.total_cost = (linked_mo.actual_material_cost or 0) + (linked_mo.actual_labor_cost or 0) + linked_mo.actual_overhead_cost

    db.session.commit()
    
    # If deleted expense was overhead and confirmed, recalculate BOM
    if was_overhead and is_confirmed:
        bom_to_update = None
        if bom_id:
            bom_to_update = BOM.query.get(bom_id)
        elif product_id:
            bom_to_update = BOM.query.filter_by(product_id=product_id, is_active=True).first()
        
        if bom_to_update:
            try:
                # Use current_user.id if available, fallback to admin user
                user_id = None
                try:
                    if current_user and current_user.is_authenticated:
                        user_id = current_user.id
                except (AttributeError, TypeError):
                    pass
                
                if user_id is None:
                    from app.models import User as UserModel
                    admin_user = UserModel.query.filter_by(username='admin').first()
                    user_id = admin_user.id if admin_user else 1
                
                BOMVersioningService.create_bom_version(
                    bom=bom_to_update,
                    change_reason=f"Overhead expense deleted: {description}",
                    change_type='overhead_added',
                    created_by_id=user_id,
                    recalculate_overhead=True
                )
            except Exception as e:
                print(f"Error updating BOM after deleting expense: {e}")
    
    log_activity('Accounting', f'Deleted Expense: {expense.expense_number}', f'Description: {description}, Amount: {amount_to_reduce}')

    flash('Expense removed', 'success')
    return redirect(url_for('accounting.expenses'))


@bp.route('/expense-accounts/transaction/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('accounting', action='delete')
def delete_expense_account_debit_entry(id):
    """Deletes a standalone debit ('Add Money') ExpenseAccountTransaction.
    The account's balance is a live-computed property (see ExpenseAccount.balance),
    so removing the row IS the reversal — no separate bookkeeping entry needed.
    Refuses rows tied to a real Expense (delete_expense already cascades those)
    or either leg of a Transfer Funds pair (deleting only one side would break
    the double-entry balance — see transfer_funds_quick). If the entry was
    routed to a Sale invoice payment (see _sync_add_money_sale_transfer), that
    Payment is reversed and deleted too, so the invoice's paid_amount doesn't
    stay stuck counting money that no longer exists."""
    from app.models import ExpenseAccountTransaction

    txn = ExpenseAccountTransaction.query.get_or_404(id)
    if txn.expense_id is not None:
        flash('This entry belongs to an expense — delete the expense instead.', 'warning')
        return redirect(url_for('accounting.expenses', entry_view='debit'))
    if txn.transaction_type == 'transfer':
        flash('This is part of a fund transfer and cannot be deleted on its own.', 'warning')
        return redirect(url_for('accounting.expenses', entry_view='debit'))

    if txn.linked_payment_id:
        payment = Payment.query.get(txn.linked_payment_id)
        if payment:
            _reverse_and_delete_sale_payment(payment, current_user.id)

    account_name = txn.account.name if txn.account else 'account'
    amount = txn.amount
    db.session.delete(txn)
    db.session.commit()
    log_activity('Accounting', f'Deleted debit entry on {account_name}', f'Amount: {amount}')

    flash('Debit entry deleted — the account balance has been reversed.', 'success')
    return redirect(url_for('accounting.expenses', entry_view='debit'))


@bp.route('/expense-accounts/transaction/<int:id>/edit-quick', methods=['POST'])
@login_required
@permission_required('accounting', action='edit')
def edit_expense_account_debit_entry(id):
    """Edits a standalone debit ('Add Money') ExpenseAccountTransaction —
    account, amount, date, description, reference, and optionally a new bill
    image. Same safety guards as delete: refuses expense-linked or transfer
    rows. If this entry was routed to a Sale invoice payment, changing the
    amount adjusts that Payment (and the Sale's paid_amount) by the same
    delta via adjust_sale_payment — it never re-targets which Sale/Bill the
    transfer points at; that's create-only, in _sync_add_money_sale_transfer."""
    from app.models import ExpenseAccountTransaction, ExpenseAccount

    txn = ExpenseAccountTransaction.query.get_or_404(id)
    if txn.expense_id is not None:
        return jsonify({'success': False, 'message': 'This entry belongs to an expense — edit the expense instead.'})
    if txn.transaction_type == 'transfer':
        return jsonify({'success': False, 'message': 'This is part of a fund transfer and cannot be edited on its own.'})

    account_id = request.form.get('account_id', type=int)
    acct = ExpenseAccount.query.get(account_id) if account_id else None
    if not acct:
        return jsonify({'success': False, 'message': 'Select an account.'})

    try:
        amount = float(request.form.get('amount') or 0)
    except (TypeError, ValueError):
        amount = 0
    if amount <= 0:
        return jsonify({'success': False, 'message': 'Enter an amount greater than zero.'})

    if txn.linked_payment_id:
        payment = Payment.query.get(txn.linked_payment_id)
        if payment:
            delta = round(amount - payment.amount, 2)
            if abs(delta) > 0.009:
                from app.utils import adjust_sale_payment
                sale = Sale.query.get(payment.invoice_id)
                if sale and payment.is_approved:
                    adjust_sale_payment(sale, delta, current_user.id)
                payment.amount = amount

    bill_path = txn.bill_image_path
    new_bill = _save_fixed_expense_bill_image(request.files.get('bill_image'))
    if new_bill:
        bill_path = new_bill

    txn.account_id = acct.id
    txn.amount = amount
    txn.date = _parse_expense_date(request.form.get('date')) or txn.date
    txn.description = (request.form.get('description') or '').strip() or None
    txn.reference = (request.form.get('reference') or '').strip() or None
    txn.bill_image_path = bill_path
    txn.customer_id = request.form.get('customer_id', type=int) or None
    txn.warehouse_id = request.form.get('warehouse_id', type=int) or None
    db.session.commit()

    return jsonify({'success': True, 'message': 'Debit entry updated.'})


@bp.route('/expenses/bulk-delete', methods=['POST'])
@login_required
@permission_required('accounting', action='delete')
def bulk_delete_expenses():
    from app.models import BOM
    from app.services.bom_versioning import BOMVersioningService
    
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No expenses selected'}), 400
    
    deleted_count = 0
    errors = []
    boms_to_update = set()
    
    for expense_id in ids:
        expense = Expense.query.get(expense_id)
        if not expense:
            continue
            
        try:
            from app.models import ManufacturingOrder
            # Store info for BOM update
            was_overhead = expense.is_bom_overhead if has_column('expenses', 'is_bom_overhead') else False
            mo_id_to_reduce = None
            amount_to_reduce = expense.amount
            is_confirmed = expense.status == 'confirmed'
            if was_overhead:
                if has_column('expenses', 'bom_id') and expense.bom_id and is_confirmed:
                    boms_to_update.add(('bom', expense.bom_id))
                elif has_column('expenses', 'product_id') and expense.product_id and is_confirmed:
                    boms_to_update.add(('product', expense.product_id))
                if has_column('expenses', 'mo_id') and expense.mo_id:
                    mo_id_to_reduce = expense.mo_id

            _delete_linked_expense_account_transaction(expense.id)
            _reverse_expense_payment_transfer(expense, current_user.id)
            db.session.delete(expense)

            if mo_id_to_reduce and is_confirmed:
                linked_mo = ManufacturingOrder.query.get(mo_id_to_reduce)
                if linked_mo:
                    linked_mo.actual_overhead_cost = max(0, (linked_mo.actual_overhead_cost or 0) - amount_to_reduce)
                    linked_mo.total_cost = (linked_mo.actual_material_cost or 0) + (linked_mo.actual_labor_cost or 0) + linked_mo.actual_overhead_cost
                    
            deleted_count += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f'Error deleting Expense {expense.expense_number}: {str(e)}')
            
    if deleted_count > 0:
        db.session.commit()
        log_activity('Accounting', f'Bulk Deleted Expenses', f'Deleted {deleted_count} expenses.')
        
        # Recalculate BOMs if needed
        for type, id in boms_to_update:
            bom_to_update = None
            if type == 'bom':
                bom_to_update = BOM.query.get(id)
            else:
                bom_to_update = BOM.query.filter_by(product_id=id, is_active=True).first()
                
            if bom_to_update:
                try:
                    user_id = None
                    try:
                        if current_user and current_user.is_authenticated:
                            user_id = current_user.id
                    except: pass
                    
                    if user_id is None:
                        from app.models import User as UserModel
                        admin_user = UserModel.query.filter_by(username='admin').first()
                        user_id = admin_user.id if admin_user else 1
                        
                    BOMVersioningService.create_bom_version(
                        bom=bom_to_update,
                        change_reason=f"Bulk deletion of expenses including overheads",
                        change_type='overhead_added',
                        created_by_id=user_id,
                        recalculate_overhead=True
                    )
                except Exception as e:
                    print(f"Error updating BOM {bom_to_update.id} after bulk delete: {e}")
                    
    message = f'Successfully deleted {deleted_count} expenses.'
    if errors:
        return jsonify({'success': False, 'message': message, 'errors': errors}), 500
    return jsonify({'success': True, 'message': message})

@bp.route('/expenses/bulk-confirm', methods=['POST'])
@login_required
@permission_required('accounting', action='edit')
def bulk_confirm_expenses():
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Only admins can confirm expenses.'}), 403
        
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No expenses selected'}), 400
        
    from app.models import BOM, ManufacturingOrder
    from app.services.bom_versioning import BOMVersioningService
    
    confirmed_count = 0
    errors = []
    boms_to_update = set()
    
    for expense_id in ids:
        expense = Expense.query.get(expense_id)
        if not expense or expense.status == 'confirmed':
            continue
            
        try:
            expense.status = 'confirmed'
            
            # Apply side effects (MO updates)
            if expense.is_bom_overhead and expense.mo_id:
                mo = ManufacturingOrder.query.get(expense.mo_id)
                if mo:
                    mo.actual_overhead_cost = (mo.actual_overhead_cost or 0) + expense.amount
                    mo.total_cost = (mo.actual_material_cost or 0) + (mo.actual_labor_cost or 0) + mo.actual_overhead_cost
            
            # Track BOMs for update
            if expense.is_bom_overhead:
                if expense.bom_id:
                    boms_to_update.add(('bom', expense.bom_id))
                elif expense.product_id:
                    boms_to_update.add(('product', expense.product_id))
            
            confirmed_count += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f'Error confirming Expense {expense.expense_number}: {str(e)}')
            
    if confirmed_count > 0:
        db.session.commit()
        
        # Trigger BOM versioning
        for type_val, id_val in boms_to_update:
            bom_to_update = None
            if type_val == 'bom':
                bom_to_update = BOM.query.get(id_val)
            else:
                bom_to_update = BOM.query.filter_by(product_id=id_val, is_active=True).first()
                
            if bom_to_update:
                try:
                    BOMVersioningService.create_bom_version(
                        bom=bom_to_update,
                        change_reason=f"Bulk confirmation of overhead expenses",
                        change_type='overhead_added',
                        created_by_id=current_user.id,
                        recalculate_overhead=True
                    )
                except Exception as e:
                    print(f"Error updating BOM during bulk confirmation: {e}")
                    
    message = f'Successfully confirmed {confirmed_count} expenses.'
    if errors:
        return jsonify({'success': False, 'message': message, 'errors': errors}), 500
    return jsonify({'success': True, 'message': message})

@bp.route('/expense/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('accounting', action='edit')
def edit_expense(id):
    from app.models import ExpenseCategory, Vendor, BOM
    from werkzeug.utils import secure_filename
    import os
    
    expense = Expense.query.get_or_404(id)
    form = ExpenseForm(obj=expense)
    
    # Populate category choices
    categories = ExpenseCategory.query.filter_by(is_active=True).order_by(ExpenseCategory.name).all()
    form.category_id.choices = [(cat.id, cat.name) for cat in categories]
    
    # Populate vendor choices
    vendors = Vendor.query.filter_by(is_active=True).order_by(Vendor.name).all()
    form.vendor_id.choices = [(0, 'Select Vendor (Optional)')] + [(v.id, v.name) for v in vendors]

    # Populate customer / warehouse choices
    from app.models import Customer, Warehouse
    customers = Customer.query.filter_by(is_active=True).order_by(Customer.name).all()
    form.customer_id.choices = [(0, 'Select Customer (Optional)')] + [(c.id, c.name) for c in customers]
    warehouses = Warehouse.query.filter_by(is_active=True).order_by(Warehouse.name).all()
    form.warehouse_id.choices = [(0, 'Select Warehouse (Optional)')] + [(w.id, w.name) for w in warehouses]

    # Populate manufactured product choices (if column exists)
    from app.models import Product
    if has_column('products', 'is_manufactured'):
        manufactured_products = Product.query.filter_by(is_manufactured=True, is_active=True).order_by(Product.name).all()
    else:
        manufactured_products = []
    form.product_id.choices = [(0, 'Select Finished Product (Optional)')] + [(p.id, p.name) for p in manufactured_products]
    
    # Populate BOM choices
    boms = BOM.query.filter_by(is_active=True).order_by(BOM.name).all()
    form.bom_id.choices = [(0, 'Select BOM (Optional)')] + [(b.id, b.name) for b in boms]
    
    # Populate MO choices (only in-progress orders; no placeholder needed for multi-select)
    from app.models import ManufacturingOrder
    in_progress_mos = ManufacturingOrder.query.filter_by(status='In Progress').order_by(ManufacturingOrder.order_number).all()
    form.mo_id.choices = [(mo.id, f"{mo.order_number} — {mo.bom.product.name}") for mo in in_progress_mos]
    
    # Handle case where MO was deleted - store original mo_id before potential overwrite
    original_mo_id = expense.mo_id if has_column('expenses', 'mo_id') else None
    
    # Add any previously linked MO (deleted, completed, or in-progress) to choices
    if original_mo_id:
        existing_choice_ids = [c[0] for c in form.mo_id.choices]
        if original_mo_id not in existing_choice_ids:
            linked_mo = ManufacturingOrder.query.get(original_mo_id)
            if linked_mo:
                if linked_mo.status == 'In Progress':
                    form.mo_id.choices.append((linked_mo.id, f"{linked_mo.order_number} — {linked_mo.bom.product.name} (Current)"))
                else:
                    form.mo_id.choices.append((linked_mo.id, f"{linked_mo.order_number} — {linked_mo.bom.product.name} (Not In Progress)"))
            else:
                form.mo_id.choices.append((original_mo_id, f"MO ID {original_mo_id} (Deleted)"))
                
    if request.method == 'GET':
        if original_mo_id:
            # Set the data to include the original mo_id
            form.mo_id.data = [original_mo_id]
        
    # Populate Payment Method choices
    from app.models import PaymentMethod
    methods = PaymentMethod.query.filter_by(is_active=True).order_by(PaymentMethod.name).all()
    form.payment_method.choices = [(m.name, m.name) for m in methods]
    
    # Include current payment method in choices if it's inactive/missing
    if expense.payment_method and expense.payment_method not in [c[0] for c in form.payment_method.choices]:
        form.payment_method.choices.append((expense.payment_method, expense.payment_method + " (Inactive)"))
    
    if request.method == 'GET':
        # Set current vendor/customer/warehouse selection
        if expense.vendor_id:
            form.vendor_id.data = expense.vendor_id
        else:
            form.vendor_id.data = 0
        form.customer_id.data = expense.customer_id or 0
        form.warehouse_id.data = expense.warehouse_id or 0

        # product_id, bom_id, mo_id are SelectMultipleField — must be set as lists
        if has_column('expenses', 'product_id') and expense.product_id:
            form.product_id.data = [expense.product_id]
        else:
            form.product_id.data = []
        
        if has_column('expenses', 'bom_id') and expense.bom_id:
            form.bom_id.data = [expense.bom_id]
        else:
            form.bom_id.data = []
        
    if form.validate_on_submit():
        # Store old overhead state to detect changes
        old_is_overhead = expense.is_bom_overhead if has_column('expenses', 'is_bom_overhead') else False
        old_bom_id = expense.bom_id if has_column('expenses', 'bom_id') else None
        old_product_id = expense.product_id if has_column('expenses', 'product_id') else None
        old_mo_id = expense.mo_id if has_column('expenses', 'mo_id') else None
        old_amount = expense.amount
        is_confirmed = expense.status == 'confirmed'
        
        expense.date = form.date.data
        expense.category_id = form.category_id.data
        expense.vendor_id = form.vendor_id.data if form.vendor_id.data != 0 else None
        expense.customer_id = form.customer_id.data if form.customer_id.data != 0 else None
        expense.warehouse_id = form.warehouse_id.data if form.warehouse_id.data != 0 else None
        expense.description = form.description.data
        # Track if we are dealing with overhead and multiple targets
        if has_column('expenses', 'is_bom_overhead'):
            expense.is_bom_overhead = form.is_bom_overhead.data
            
        new_is_overhead = expense.is_bom_overhead if has_column('expenses', 'is_bom_overhead') else False
        pid_list = [p for p in (form.product_id.data or []) if p and p != 0]
        bid_list = [b for b in (form.bom_id.data or []) if b and b != 0]
        mo_list = [m for m in (form.mo_id.data or []) if m and m != 0]
        
        overhead_mode = request.form.get('overhead_mode', 'mo')
        targets = []
        if new_is_overhead:
            if overhead_mode == 'mo': targets = [('mo', m) for m in mo_list]
            elif overhead_mode == 'bulk': targets = [('product', p) for p in pid_list] + [('bom', b) for b in bid_list]

        num_targets = len(targets)
        created_expenses = []

        if new_is_overhead and num_targets > 1:
            divided_amount = form.amount.data / num_targets
            expense.amount = divided_amount
            new_amount = divided_amount
            
            # Apply first target to original expense
            first_type, first_id = targets[0]
            if has_column('expenses', 'product_id'): expense.product_id = first_id if first_type == 'product' else None
            if has_column('expenses', 'bom_id'): expense.bom_id = first_id if first_type == 'bom' else None
            if has_column('expenses', 'mo_id'): expense.mo_id = first_id if first_type == 'mo' else None
            
            # Handle bill image upload (Must be done before duplicating for New Expenses)
            if 'bill_image' in request.files:
                bill_file = request.files['bill_image']
                if bill_file and bill_file.filename:
                    import time, uuid
                    original_filename = secure_filename(bill_file.filename)
                    unique_prefix = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
                    filename = f"{unique_prefix}_{original_filename}"
                    
                    bill_path = os.path.join('app', 'static', 'uploads', 'bills', filename)
                    os.makedirs(os.path.dirname(bill_path), exist_ok=True)
                    bill_file.save(bill_path)
                    # Store path relative to project root
                    expense.bill_image_path = f"app/static/uploads/bills/{filename}"
                    
            # Handle monthly division
            if has_column('expenses', 'is_monthly_divided'):
                expense.is_monthly_divided = form.is_monthly_divided.data
                if form.is_monthly_divided.data:
                    expense.monthly_start_date = form.monthly_start_date.data
                    expense.monthly_end_date = form.monthly_end_date.data
                    expense.calculate_daily_amount()
                else:
                    expense.daily_amount = 0
            
            db.session.flush() # flush to save changes to current expense first
            
            # Create rest of expenses
            from app.routes.accounting import get_unique_expense_number
            from app.models import ExpenseSettings
            acc_settings = ExpenseSettings.query.first()
            if not acc_settings:
                acc_settings = ExpenseSettings()
                db.session.add(acc_settings)
                db.session.flush()
            next_expense_num = acc_settings.next_number
            
            for i in range(1, num_targets):
                t_type, t_id = targets[i]
                exp_num, next_expense_num = get_unique_expense_number(acc_settings, next_expense_num)
                
                exp_kwargs = {
                    'expense_number': exp_num,
                    'amount': divided_amount,
                    'is_bom_overhead': True,
                    'status': expense.status,
                    'created_by': current_user.id,
                    'date': expense.date,
                    'category_id': expense.category_id,
                    'vendor_id': expense.vendor_id,
                    'description': f"{expense.description} (Allocation {i+1}/{num_targets})",
                    'payment_method': expense.payment_method,
                    'reference': expense.reference,
                    'notes': expense.notes,
                }
                if has_column('expenses', 'product_id'): exp_kwargs['product_id'] = t_id if t_type == 'product' else None
                if has_column('expenses', 'bom_id'): exp_kwargs['bom_id'] = t_id if t_type == 'bom' else None
                if has_column('expenses', 'mo_id'): exp_kwargs['mo_id'] = t_id if t_type == 'mo' else None
                if has_column('expenses', 'bill_image_path'): exp_kwargs['bill_image_path'] = getattr(expense, 'bill_image_path', None)
                
                new_exp = Expense(**exp_kwargs)
                if has_column('expenses', 'is_monthly_divided') and expense.is_monthly_divided:
                    new_exp.is_monthly_divided = True
                    new_exp.monthly_start_date = expense.monthly_start_date
                    new_exp.monthly_end_date = expense.monthly_end_date
                    new_exp.calculate_daily_amount()
                db.session.add(new_exp)
                created_expenses.append((new_exp, t_type, t_id))
            acc_settings.next_number = next_expense_num
            
            # Update original expense description to show it's allocation 1
            expense.description = f"{form.description.data} (Allocation 1/{num_targets})"
            new_mo_id_val = expense.mo_id if has_column('expenses', 'mo_id') else None

        else:
            expense.amount = form.amount.data
            new_amount = expense.amount
            if has_column('expenses', 'product_id'):
                expense.product_id = pid_list[0] if (pid_list and (not new_is_overhead or overhead_mode == 'bulk')) else None
            if has_column('expenses', 'bom_id'):
                expense.bom_id = bid_list[0] if (bid_list and (not new_is_overhead or overhead_mode == 'bulk')) else None
            if has_column('expenses', 'mo_id'):
                expense.mo_id = mo_list[0] if (mo_list and (not new_is_overhead or overhead_mode == 'mo')) else None
            new_mo_id_val = expense.mo_id if has_column('expenses', 'mo_id') else None
            
            # Handle bill image upload
            if 'bill_image' in request.files:
                bill_file = request.files['bill_image']
                if bill_file and bill_file.filename:
                    filename = secure_filename(bill_file.filename)
                    bill_path = os.path.join('app', 'static', 'uploads', 'bills', filename)
                    os.makedirs(os.path.dirname(bill_path), exist_ok=True)
                    bill_file.save(bill_path)
                    expense.bill_image_path = bill_path.replace('\\', '/')
            
            # Handle monthly division
            if has_column('expenses', 'is_monthly_divided'):
                expense.is_monthly_divided = form.is_monthly_divided.data
                if form.is_monthly_divided.data:
                    expense.monthly_start_date = form.monthly_start_date.data
                    expense.monthly_end_date = form.monthly_end_date.data
                    expense.calculate_daily_amount()
                else:
                    expense.daily_amount = 0

        # Update Manufacturing Order costs if MO association or amount changed
        from app.models import ManufacturingOrder
        
        # 1. Revert from old MO if it was overhead, had an MO, and was confirmed
        if old_is_overhead and old_mo_id and is_confirmed:
            old_mo = ManufacturingOrder.query.get(old_mo_id)
            if old_mo:
                old_mo.actual_overhead_cost = max(0, (old_mo.actual_overhead_cost or 0) - old_amount)
                old_mo.total_cost = (old_mo.actual_material_cost or 0) + (old_mo.actual_labor_cost or 0) + old_mo.actual_overhead_cost

        # 2. Add to new MO if it is currently overhead, has an MO, and is confirmed (For Original Expense)
        if new_is_overhead and new_mo_id_val and is_confirmed:
            new_mo = ManufacturingOrder.query.get(new_mo_id_val)
            if new_mo:
                new_mo.actual_overhead_cost = (new_mo.actual_overhead_cost or 0) + new_amount
                new_mo.total_cost = (new_mo.actual_material_cost or 0) + (new_mo.actual_labor_cost or 0) + new_mo.actual_overhead_cost

        # 3. Add to new MOs for dynamically created expenses
        for new_exp, target_type, target_id in created_expenses:
            if target_type == 'mo' and new_exp.status == 'confirmed':
                new_mo = ManufacturingOrder.query.get(target_id)
                if new_mo:
                    new_mo.actual_overhead_cost = (new_mo.actual_overhead_cost or 0) + new_exp.amount
                    new_mo.total_cost = (new_mo.actual_material_cost or 0) + (new_mo.actual_labor_cost or 0) + new_mo.actual_overhead_cost

        # Keep the linked account transaction (if any) in sync with this
        # expense's current account/amount/description. Stays 'credit' —
        # switching an existing expense to Debit isn't offered here since that
        # would mean it's no longer really an expense.
        account_id = request.form.get('account_id', type=int)
        db.session.flush()
        _sync_expense_account_transaction(expense, account_id, is_confirmed)
        for new_exp, _target_type, _target_id in created_expenses:
            _sync_expense_account_transaction(new_exp, account_id, is_confirmed)

        # "Add this to Invoice/Purchase Payment" — same single-row restriction
        # as Add Expense: only applies to the original row, never to rows
        # spun off by a BOM-overhead split on this edit. An existing Expense
        # being edited is always a Credit row (Debit/Add Money entries never
        # create an Expense — see _add_money_from_expense_form), so "Add to
        # Invoice Payment" (sale) is rejected here too, same as in
        # add_expense()'s Credit path — EXCEPT re-saving a legacy expense
        # that was already sale-linked before this restriction existed, which
        # stays untouched rather than silently reassigned/dropped.
        is_admin = getattr(current_user, 'is_admin', False)
        transfer_type = request.form.get('payment_transfer_type')
        transfer_target_id = request.form.get('payment_transfer_target_id', type=int)
        if transfer_type == 'sale' and not (expense.linked_sale_id and expense.linked_sale_id == transfer_target_id):
            transfer_type = None
        if transfer_type in ('sale', 'bill') and transfer_target_id:
            _sync_expense_payment_transfer(expense, transfer_type, transfer_target_id,
                                           expense.amount, current_user.id, is_admin)
        else:
            _reverse_expense_payment_transfer(expense, current_user.id)

        db.session.commit()
        log_activity('Accounting', f'Updated Expense: {expense.expense_number}',
                    f'Amount: {expense.amount}, Description: {expense.description}')

        # Trigger BOM versioning if overhead status changed or if currently set as overhead
        # This handles both cases: adding overhead and removing overhead
        bom_to_update = None
        new_is_overhead = expense.is_bom_overhead if has_column('expenses', 'is_bom_overhead') else False
        overhead_status_changed = old_is_overhead != new_is_overhead

        if (overhead_status_changed or new_is_overhead) and is_confirmed:
            # Determine which BOM to update
            if new_is_overhead:
                # Expense is now overhead, find BOM to update
                if has_column('expenses', 'bom_id') and expense.bom_id:
                    bom_to_update = BOM.query.get(expense.bom_id)
                elif has_column('expenses', 'product_id') and expense.product_id:
                    bom_to_update = BOM.query.filter_by(product_id=expense.product_id, is_active=True).first()
            else:
                # Expense was overhead but now isn't, find old BOM to update
                if old_bom_id:
                    bom_to_update = BOM.query.get(old_bom_id)
                elif old_product_id:
                    bom_to_update = BOM.query.filter_by(product_id=old_product_id, is_active=True).first()
        
        if bom_to_update:
            from app.services.bom_versioning import BOMVersioningService
            from app.models import User as UserModel
            try:
                # Use current_user.id if available, fallback to admin user
                user_id = None
                try:
                    if current_user and current_user.is_authenticated:
                        user_id = current_user.id
                except (AttributeError, TypeError):
                    pass
                
                if user_id is None:
                    admin_user = UserModel.query.filter_by(username='admin').first()
                    user_id = admin_user.id if admin_user else 1
                
                if new_is_overhead:
                    change_reason = f"Overhead expense updated: {expense.description}"
                else:
                    change_reason = f"Overhead expense removed: {expense.description}"
                
                BOMVersioningService.create_bom_version(
                    bom=bom_to_update,
                    change_reason=change_reason,
                    change_type='overhead_added',
                    created_by_id=user_id,
                    recalculate_overhead=True
                )
            except Exception as e:
                print(f"Error creating BOM version: {e}")
        
        flash('Expense updated successfully', 'success')
        return redirect(url_for('accounting.expenses'))

    from app.models import ExpenseAccount, ExpenseAccountTransaction
    expense_accounts = ExpenseAccount.query.filter_by(is_active=True).order_by(ExpenseAccount.name).all()
    linked_txn = ExpenseAccountTransaction.query.filter_by(expense_id=expense.id).first()
    existing_account_id = linked_txn.account_id if linked_txn else None

    existing_payment_transfer = None
    if expense.linked_sale_id and expense.linked_sale:
        s = expense.linked_sale
        existing_payment_transfer = {'id': s.id, 'type': 'sale',
                                     'text': f"{s.invoice_number} — {s.customer.name if s.customer else 'No Customer'} (Due: PKR {s.balance_due:,.2f})"}
    elif expense.linked_bill_id and expense.linked_bill:
        b = expense.linked_bill
        balance = max(0.0, b.total - b.paid_amount - (b.cancelled_amount or 0))
        existing_payment_transfer = {'id': b.id, 'type': 'bill',
                                     'text': f"{b.bill_number} — {b.vendor.name if b.vendor else 'No Vendor'} (Due: PKR {balance:,.2f})"}

    return render_template('accounting/edit_expense.html', form=form, expense=expense,
                           expense_accounts=expense_accounts, existing_account_id=existing_account_id,
                           existing_payment_transfer=existing_payment_transfer)

@bp.route('/expense-categories')
@login_required
def expense_categories():
    from app.models import ExpenseCategory
    categories = ExpenseCategory.query.order_by(ExpenseCategory.name).all()
    return render_template('accounting/expense_categories.html', categories=categories)

@bp.route('/expense-category/add', methods=['GET', 'POST'])
@login_required
@permission_required('accounting', action='add')
def add_expense_category():
    from app.models import ExpenseCategory
    from app.forms import ExpenseCategoryForm
    
    form = ExpenseCategoryForm()
    if form.validate_on_submit():
        category = ExpenseCategory(
            name=form.name.data,
            description=form.description.data
        )
        db.session.add(category)
        db.session.commit()
        log_activity('Accounting', f'Created Expense Category: {category.name}', f'ID: {category.id}')
        flash('Expense category added successfully', 'success')
        return redirect(url_for('accounting.expense_categories'))
    
    return render_template('accounting/add_expense_category.html', form=form)

@bp.route('/expense/<int:id>/confirm', methods=['POST'])
@login_required
@permission_required('accounting', action='edit')
def confirm_expense(id):
    if not getattr(current_user, 'is_admin', False):
        flash('Only admins can confirm expenses.', 'danger')
        return redirect(url_for('accounting.expenses'))
        
    expense = Expense.query.get_or_404(id)
    if expense.is_approved:
        flash('Expense is already confirmed.', 'info')
        return redirect(url_for('accounting.expenses'))
        
    expense.status = 'confirmed'
    expense.is_approved = True
    expense.is_rejected = False
    expense.approved_by = current_user.id
    expense.approved_at = datetime.utcnow()
    
    # Apply side effects (MO updates)
    if expense.is_bom_overhead and expense.mo_id:
        from app.models import ManufacturingOrder
        mo = ManufacturingOrder.query.get(expense.mo_id)
        if mo:
            mo.actual_overhead_cost = (mo.actual_overhead_cost or 0) + expense.amount
            mo.total_cost = (mo.actual_material_cost or 0) + (mo.actual_labor_cost or 0) + mo.actual_overhead_cost
            
    # Apply BOM versioning if overhead
    if expense.is_bom_overhead:
        from app.models import BOM
        from app.services.bom_versioning import BOMVersioningService
        from app.models import User as UserModel
        
        bom_to_update = None
        if expense.bom_id:
            bom_to_update = BOM.query.get(expense.bom_id)
        elif expense.product_id:
            bom_to_update = BOM.query.filter_by(product_id=expense.product_id, is_active=True).first()
            
        if bom_to_update:
            try:
                BOMVersioningService.create_bom_version(
                    bom=bom_to_update,
                    change_reason=f"Overhead expense confirmed: {expense.description}",
                    change_type='overhead_added',
                    created_by_id=current_user.id,
                    recalculate_overhead=True
                )
            except Exception as e:
                print(f"Error updating BOM during confirmation: {e}")

    from app.models import ExpenseAccountTransaction
    linked_txn = ExpenseAccountTransaction.query.filter_by(expense_id=expense.id).first()
    if linked_txn:
        linked_txn.is_approved = True
        linked_txn.is_rejected = False
        linked_txn.approved_by = current_user.id
        linked_txn.approved_at = datetime.utcnow()

    db.session.commit()
    log_activity('Accounting', f'Confirmed Expense: {expense.expense_number}',
                f'Amount: {expense.amount}, Description: {expense.description}')
    flash(f'Expense {expense.expense_number} confirmed successfully.', 'success')
    return redirect(url_for('accounting.expenses'))

@bp.route('/expense/<int:id>/reject', methods=['POST'])
@login_required
@permission_required('accounting', action='edit')
def reject_expense(id):
    if not getattr(current_user, 'is_admin', False):
        flash('Only admins can reject expenses.', 'danger')
        return redirect(url_for('accounting.expenses'))
        
    expense = Expense.query.get_or_404(id)
    reason = request.form.get('reason', '')
    
    expense.status = 'rejected'
    expense.is_approved = False
    expense.is_rejected = True
    expense.rejection_reason = reason

    from app.models import ExpenseAccountTransaction
    linked_txn = ExpenseAccountTransaction.query.filter_by(expense_id=expense.id).first()
    if linked_txn:
        linked_txn.is_approved = False
        linked_txn.is_rejected = True
        linked_txn.rejection_reason = reason

    db.session.commit()
    log_activity('Accounting', f'Rejected Expense: {expense.expense_number}',
                f'Reason: {reason or "No reason provided"}')
    flash(f'Expense {expense.expense_number} rejected.', 'warning')
    return redirect(url_for('accounting.expenses'))

@bp.route('/expense-category/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('accounting', action='edit')
def edit_expense_category(id):
    from app.models import ExpenseCategory
    from app.forms import ExpenseCategoryForm
    
    category = ExpenseCategory.query.get_or_404(id)
    form = ExpenseCategoryForm(obj=category)
    
    if form.validate_on_submit():
        category.name = form.name.data
        category.description = form.description.data
        db.session.commit()
        log_activity('Accounting', f'Updated Expense Category: {category.name}', f'ID: {category.id}')
        flash('Expense category updated successfully', 'success')
        return redirect(url_for('accounting.expense_categories'))
    
    return render_template('accounting/edit_expense_category.html', form=form, category=category)

@bp.route('/expense-category/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('accounting', action='delete')
def delete_expense_category(id):
    from app.models import ExpenseCategory
    
    category = ExpenseCategory.query.get_or_404(id)
    # Check if category is being used
    if category.expenses:
        flash('Cannot delete category that has expenses associated with it', 'error')
        return redirect(url_for('accounting.expense_categories'))
    
    cat_name = category.name
    db.session.delete(category)
    db.session.commit()
    log_activity('Accounting', f'Deleted Expense Category: {cat_name}', f'ID: {id}')
    flash('Expense category deleted successfully', 'success')
    return redirect(url_for('accounting.expense_categories'))

# --- Expense Number Settings ---

@bp.route('/expense-settings', methods=['GET', 'POST'])
@login_required
def expense_settings():
    from app.models import ExpenseSettings
    from app.forms import ExpenseSettingsForm
    
    settings = ExpenseSettings.query.first()
    
    form = ExpenseSettingsForm(obj=settings)
    
    if form.validate_on_submit():
        if not settings:
            settings = ExpenseSettings()
            db.session.add(settings)
        
        settings.expense_prefix = form.expense_prefix.data or ''
        settings.expense_suffix = form.expense_suffix.data or ''
        settings.next_number = form.next_number.data or 1
        
        db.session.commit()
        log_activity('Settings', f'Updated Expense Number Settings',
                    f'Prefix: {settings.expense_prefix}, Next #: {settings.next_number}')
        flash('Expense number settings updated successfully.', 'success')
        return redirect(url_for('accounting.expense_settings'))
    
    return render_template('accounting/expense_settings.html', settings=settings, form=form)

# --- Payment Methods Management ---

@bp.route('/payment-methods')
@login_required
def payment_methods():
    from app.models import PaymentMethod
    methods = PaymentMethod.query.order_by(PaymentMethod.name).all()
    return render_template('accounting/payment_methods.html', methods=methods)

@bp.route('/payment-method/add', methods=['GET', 'POST'])
@login_required
@permission_required('accounting', action='add')
def add_payment_method():
    from app.models import PaymentMethod
    from app.forms import PaymentMethodForm
    
    form = PaymentMethodForm()
    if form.validate_on_submit():
        method = PaymentMethod(
            name=form.name.data,
            description=form.description.data
        )
        db.session.add(method)
        db.session.commit()
        log_activity('Accounting', f'Created Payment Method: {method.name}', f'ID: {method.id}')
        flash('Payment method added successfully', 'success')
        return redirect(url_for('accounting.payment_methods'))
    
    return render_template('accounting/add_payment_method.html', form=form)

@bp.route('/payment-method/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('accounting', action='edit')
def edit_payment_method(id):
    from app.models import PaymentMethod
    from app.forms import PaymentMethodForm
    
    method = PaymentMethod.query.get_or_404(id)
    form = PaymentMethodForm(obj=method)
    
    if form.validate_on_submit():
        method.name = form.name.data
        method.description = form.description.data
        db.session.commit()
        log_activity('Accounting', f'Updated Payment Method: {method.name}', f'ID: {method.id}')
        flash('Payment method updated successfully', 'success')
        return redirect(url_for('accounting.payment_methods'))
    
    return render_template('accounting/edit_payment_method.html', form=form, method=method)

@bp.route('/payment-method/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('accounting', action='delete')
def delete_payment_method(id):
    from app.models import PaymentMethod, Expense
    
    method = PaymentMethod.query.get_or_404(id)
    # Check if payment method is being used in expenses (string match)
    if Expense.query.filter_by(payment_method=method.name).first():
        flash('Cannot delete payment method that is associated with existing expenses', 'danger')
        return redirect(url_for('accounting.payment_methods'))
    
    method_name = method.name
    db.session.delete(method)
    db.session.commit()
    log_activity('Accounting', f'Deleted Payment Method: {method_name}', f'ID: {id}')
    flash('Payment method deleted successfully', 'success')
    return redirect(url_for('accounting.payment_methods'))


@bp.route('/bom/<int:bom_id>/reset-overhead', methods=['POST'])
@login_required
@permission_required('accounting', action='edit')
def reset_bom_overhead(bom_id):
    """Reset BOM overhead by marking all overhead expenses as non-overhead"""
    from app.models import BOM, Expense
    from app.services.bom_versioning import BOMVersioningService
    
    bom = BOM.query.get_or_404(bom_id)
    
    # Find all overhead expenses linked to this BOM
    overhead_expenses = Expense.query.filter(
        Expense.bom_id == bom_id,
        Expense.is_bom_overhead == True
    ).all()
    
    if not overhead_expenses:
        flash('No overhead expenses found for this BOM', 'info')
        return redirect(url_for('accounting.expenses'))
    
    # Delete all overhead expenses (so they don't appear in dashboard totals)
    expense_count = len(overhead_expenses)
    total_amount = sum(expense.amount for expense in overhead_expenses)
    
    for expense in overhead_expenses:
        db.session.delete(expense)
    
    db.session.commit()
    log_activity('Accounting', f'Reset BOM Overhead for BOM ID {bom_id}',
                f'Deleted {expense_count} expense(s), Total: PKR {total_amount}')

    # Recalculate BOM overhead (should be 0 now)
    try:
        # Safe user_id resolution
        user_id = None
        try:
            if current_user and current_user.is_authenticated:
                user_id = current_user.id
        except (AttributeError, TypeError):
            pass
        
        if user_id is None:
            from app.models import User
            admin_user = User.query.filter_by(username='admin').first()
            user_id = admin_user.id if admin_user else 1
        
        BOMVersioningService.create_bom_version(
            bom=bom,
            change_reason=f"BOM overhead reset: {expense_count} overhead expense(es) deleted (PKR {total_amount})",
            change_type='overhead_removed',
            created_by_id=user_id,
            recalculate_overhead=True
        )
        
        flash(f'BOM overhead reset successfully! Deleted {expense_count} overhead expense(es) totaling PKR {total_amount}. New overhead: PKR {bom.overhead_cost}', 'success')
    except Exception as e:
        flash(f'BOM overhead reset partially: {expense_count} expenses deleted but versioning failed: {str(e)}', 'warning')
        print(f"Error resetting BOM overhead: {e}")
    
    return redirect(url_for('accounting.expenses'))


# ══════════════════════════════════════════════════════════════════════════════
# Fixed (recurring, day-based) Expenses
#
# A self-contained module: it owns the fixed_expenses table and touches nothing
# else until someone presses Post, which is the only place a real Expense row is
# created — using the same numbering helper and status rules as Add Expense.
# ══════════════════════════════════════════════════════════════════════════════

def _fixed_expense_json(fx):
    """Serialise a template for the popup.

    For an auto-posting template the progress is read from the cycles actually
    written to the book — the same source the Expense list draws from — so the
    two screens always agree. The manual accrual counters only drive templates
    that are posted by hand.
    """
    today = datetime.utcnow().date()
    if fx.auto_post and fx.start_date and int(fx.cycles_posted or 0) > 0:
        n = int(fx.cycles_posted or 0)
        cyc_start, cyc_end = fx.cycle_window(n)
        cur_cycle_days = fx.cycle_length_for(n)
        gone = (today - cyc_start).days + 1
        day_in_cycle = 0 if gone < 0 else (cur_cycle_days if gone > cur_cycle_days else gone)
        cycles_completed = (n - 1) + (1 if day_in_cycle >= cur_cycle_days else 0)
        # Elapsed days/money from real dates rather than assuming every cycle
        # is the same length — correct for both fixed N-day cycles and
        # variable-length calendar-month cycles (mathematically equivalent to
        # the old cycles_completed*cycle_days formula when cycles ARE uniform).
        elapsed_days = (cyc_start - fx.start_date).days + day_in_cycle
        completed_total = float(fx.posted_amount or 0) - fx.cycle_total_for(n)
        incurred = round(completed_total + fx.per_day_amount_for(n) * day_in_cycle, 2)
    else:
        day_in_cycle = fx.day_in_cycle
        cycles_completed = fx.cycles_completed
        cur_cycle_days = fx.cycle_days
        # days_accrued/accrued_amount are already the exact lifetime running
        # totals (kept correct per-day by sync_accrual, including across
        # calendar-month rate changes) — no need to reconstruct them from a
        # single current-cycle rate.
        elapsed_days = int(fx.days_accrued or 0)
        incurred = round(float(fx.posted_amount or 0) + float(fx.accrued_amount or 0), 2)

    progress_pct = int(day_in_cycle * 100 / cur_cycle_days) if cur_cycle_days else 0

    return {
        'id': fx.id,
        'name': fx.name,
        'description': fx.description or '',
        'category_id': fx.category_id,
        'category_name': fx.category.name if fx.category else '',
        'vendor_id': fx.vendor_id,
        'vendor_name': fx.vendor.name if fx.vendor else '',
        'account_id': fx.expense_account_id,
        'account_name': fx.expense_account.name if fx.expense_account else '',
        'payment_method': fx.payment_method or '',
        'bill_image_path': fx.bill_image_path or '',
        'mode': fx.mode or 'divide',
        'amount': round(float(fx.amount or 0), 2),
        'days': fx.cycle_days,
        'days_raw': int(fx.days or 1),
        'cycle_type': fx.cycle_type or 'fixed_days',
        'start_date': fx.start_date.strftime('%Y-%m-%d') if fx.start_date else '',
        'is_active': bool(fx.is_active),
        'auto_post': bool(fx.auto_post),
        'cycles_posted': int(fx.cycles_posted or 0),
        'days_posted': int(fx.days_posted or 0),
        'per_day_amount': round(fx.per_day_amount, 2),
        'cycle_total': round(fx.cycle_total, 2),
        'day_in_cycle': day_in_cycle,
        'cycles_completed': cycles_completed,
        'cycle_progress_pct': progress_pct,
        'elapsed_days': elapsed_days,
        'incurred_amount': incurred,
        'days_accrued': int(fx.days_accrued or 0),
        'accrued_amount': round(float(fx.accrued_amount or 0), 2),
        'posted_amount': round(float(fx.posted_amount or 0), 2),
        'last_accrued_date': fx.last_accrued_date.strftime('%Y-%m-%d') if fx.last_accrued_date else '',
    }


def ensure_fixed_expense_rows():
    """Write each started cycle into the Expense book as ONE day-divided row.

    Shaped exactly like a divided expense added by hand from Add Expense
    (is_monthly_divided + daily_amount + the cycle window), so the list shows a
    single row with its per-day figure and progress bar, and the reports
    pro-rate it per day over the period being viewed.

    Idempotent: `cycles_posted` records how many cycles are already in the book,
    so refreshing a page can never write the same cycle twice. Never raises — a
    failure here must not break the page that called it.
    """
    from app.models import FixedExpense
    try:
        today = datetime.utcnow().date()
        templates = (FixedExpense.query
                     .filter(FixedExpense.is_active == True,
                             FixedExpense.auto_post == True)
                     .all())
        if not templates:
            return 0

        settings = ExpenseSettings.query.first()
        if not settings:
            settings = ExpenseSettings(expense_prefix='EXP-', expense_suffix='', next_number=1)
            db.session.add(settings)
            db.session.flush()
        next_num = settings.next_number

        created = 0
        for fx in templates:
            if not fx.start_date or not fx.category_id:
                continue                    # incomplete template — skip quietly
            started = fx.cycles_started(today)
            # As soon as the running cycle finishes, open the next one straight
            # away so the repetition is visible: it appears at Day 0 of N and
            # fills in a day at a time. Without this the row only turned up when
            # its first day arrived, and the list looked like it had stopped.
            if started > 0 and fx.cycle_window(started)[1] <= today:
                started += 1
            done = int(fx.cycles_posted or 0)
            if fx.cycle_total <= 0:
                continue

            # Cap the catch-up so a very old start date cannot flood the book
            # in one request; the rest is picked up on the next page load.
            while done < started and created < 60:
                done += 1
                cyc_start, cyc_end = fx.cycle_window(done)
                cyc_days = fx.cycle_length_for(done)
                cyc_total = fx.cycle_total_for(done)
                cyc_daily = fx.per_day_amount_for(done)
                expense_number, next_num = get_unique_expense_number(settings, next_num)
                exp = Expense(
                    expense_number=expense_number,
                    date=datetime.combine(cyc_start, datetime.min.time()),
                    amount=cyc_total,
                    description='%s - Fixed Expense (cycle %d of %d days)'
                                % (fx.name, done, cyc_days),
                    category_id=fx.category_id,
                    vendor_id=fx.vendor_id,
                    payment_method=fx.payment_method,
                    bill_image_path=fx.bill_image_path,
                    status='confirmed',
                    created_by=fx.created_by,
                    is_approved=True,
                    approved_by=fx.created_by,
                    approved_at=datetime.utcnow(),
                    is_rejected=False,
                    is_monthly_divided=True,
                    monthly_start_date=cyc_start,
                    monthly_end_date=cyc_end,
                    daily_amount=cyc_daily,
                    fixed_expense_id=fx.id,
                )
                db.session.add(exp)
                fx.cycles_posted = done
                # Cumulative days from start_date through the end of this cycle —
                # correct whether every cycle is the same length (fixed_days) or
                # not (calendar_month), unlike done * a-single-cycle-length.
                fx.days_posted = (cyc_end - fx.start_date).days + 1
                fx.posted_amount = float(fx.posted_amount or 0) + cyc_total
                # The cycle is in the book now; clear the manual accrual so the
                # same money can never be posted a second time by hand.
                fx.accrued_amount = 0
                created += 1

                # Charge this cycle against the linked Expense Account, if any
                # — same bridge Add/Edit Expense uses (credit = money out).
                # Needs exp.id, so flush before linking.
                if fx.expense_account_id:
                    db.session.flush()
                    _sync_expense_account_transaction(exp, fx.expense_account_id, True)

        if created:
            settings.next_number = next_num
            db.session.commit()
            try:
                log_activity('Accounting', 'Fixed Expenses posted automatically',
                             'Created %d cycle expense(s).' % created)
            except Exception:
                pass        # no request context (e.g. a script) — not worth failing over
        return created
    except Exception as e:
        db.session.rollback()
        try:
            from flask import current_app
            current_app.logger.error('Fixed expense auto-post failed: %s' % e)
        except Exception:
            pass
        return 0


def _sync_all_fixed_expenses():
    """Bring every active template up to today. Commits only if something moved."""
    from app.models import FixedExpense
    changed = 0
    for fx in FixedExpense.query.all():
        changed += fx.sync_accrual()
    if changed:
        db.session.commit()
    return changed


def _stop_fixed_expense(fx, stop_date=None):
    """Stopping must stop the money, not just future cycles.

    The cycle that is still open on `stop_date` is trimmed to the days actually
    used: its window ends on the stop date and its amount drops to
    per-day x days used, so the list, the report and the dashboard all stop
    growing. A cycle that had not started yet is removed outright.
    Caller commits.
    """
    from app.models import FixedExpense  # noqa: F401
    stop_date = stop_date or datetime.utcnow().date()
    open_rows = (Expense.query
                 .filter(Expense.fixed_expense_id == fx.id,
                         Expense.monthly_end_date != None,
                         Expense.monthly_end_date > stop_date)
                 .all())
    trimmed = 0
    for row in open_rows:
        if not row.monthly_start_date:
            continue
        if row.monthly_start_date > stop_date:
            db.session.delete(row)          # cycle never began
            fx.posted_amount = max(0.0, float(fx.posted_amount or 0) - float(row.amount or 0))
            fx.cycles_posted = max(0, int(fx.cycles_posted or 0) - 1)
            trimmed += 1
            continue
        used_days = (stop_date - row.monthly_start_date).days + 1
        per_day = float(row.daily_amount or 0)
        new_amount = round(per_day * used_days, 2)
        fx.posted_amount = max(0.0, float(fx.posted_amount or 0)
                               - float(row.amount or 0) + new_amount)
        row.amount = new_amount
        row.monthly_end_date = stop_date
        row.description = '%s [stopped on %s]' % (row.description, stop_date.strftime('%d-%m-%Y'))
        trimmed += 1
    fx.paused_on = stop_date
    return trimmed


def _restore_open_cycle(fx):
    """Undo a same-day stop: put the trimmed cycle back to its full window.

    Only used when a template is switched off and on again without a day
    passing, so the trim never really happened.
    """
    n = int(fx.cycles_posted or 0)
    if n <= 0:
        return False
    row = (Expense.query
           .filter_by(fixed_expense_id=fx.id)
           .order_by(Expense.id.desc())
           .first())
    if not row or not row.monthly_start_date:
        return False
    cyc_start, cyc_end = fx.cycle_window(n)
    if row.monthly_start_date != cyc_start or row.monthly_end_date == cyc_end:
        return False                      # not the trimmed row, or nothing to undo
    cyc_total = fx.cycle_total_for(n)
    fx.posted_amount = (max(0.0, float(fx.posted_amount or 0) - float(row.amount or 0))
                        + cyc_total)
    row.monthly_end_date = cyc_end
    row.amount = cyc_total
    row.description = row.description.split(' [stopped on ')[0]
    return True


def _resume_fixed_expense(fx, resume_date=None):
    """Switch a template back on without losing or inventing days.

    Same day as the stop — nothing actually elapsed, so the template is put back
    exactly as it was and the cycle that was trimmed or dropped is restored.

    Later — the days it spent switched off are not charged: the counters stay,
    and the start date is re-anchored so the NEXT cycle begins on the resume
    date instead of a fresh cycle 1 overlapping days already booked.
    """
    resume_date = resume_date or datetime.utcnow().date()

    if fx.paused_on and fx.paused_on == resume_date:
        _restore_open_cycle(fx)
        fx.paused_on = None
        return

    done = int(fx.cycles_posted or 0)
    if (fx.cycle_type or 'fixed_days') == 'calendar_month':
        # Calendar-month cycles aren't uniform-length, so there's no single
        # day-count that re-anchors start_date the way fixed_days does below.
        # Instead: the next cycle starts fresh right on the resume date (like
        # a new cycle 1, possibly a partial month), and cycle_base_n records
        # how many cycles were already posted before this anchor so future
        # cycle_window() calls keep numbering forward correctly.
        # days_posted is left as-is — it's informational bookkeeping only,
        # and gets a correct value again the moment the next cycle posts.
        fx.start_date = resume_date
        fx.cycle_base_n = done
    else:
        fx.start_date = resume_date - timedelta(days=done * fx.cycle_days)
        fx.days_posted = done * fx.cycle_days
    fx.days_accrued = 0
    fx.accrued_amount = 0
    fx.last_accrued_date = None
    fx.paused_on = None


def _save_fixed_expense_bill_image(file_storage):
    """Save an uploaded bill image for a Fixed Expense template under
    app/static/uploads/bills/ and return the project-root-relative path —
    same convention Add Expense and the Journal module already use. Returns
    None if no file was actually chosen."""
    if not file_storage or not file_storage.filename:
        return None
    import os
    import time
    import uuid
    from werkzeug.utils import secure_filename

    original_filename = secure_filename(file_storage.filename)
    unique_prefix = f"{int(time.time())}_{uuid.uuid4().hex[:8]}"
    filename = f"{unique_prefix}_{original_filename}"

    full_path = os.path.join('app', 'static', 'uploads', 'bills', filename)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    file_storage.save(full_path)
    return f"app/static/uploads/bills/{filename}"


def _read_fixed_expense_form(fx, form):
    """Apply submitted values onto a FixedExpense. Returns an error string or None."""
    name = (form.get('name') or '').strip()
    if not name:
        return 'Name is required.'

    mode = (form.get('mode') or 'divide').strip().lower()
    if mode not in ('divide', 'multiply'):
        mode = 'divide'

    cycle_type = (form.get('cycle_type') or 'fixed_days').strip().lower()
    if cycle_type not in ('fixed_days', 'calendar_month'):
        cycle_type = 'fixed_days'
    # Switching cycle_type on a template that has already run would reinterpret
    # its whole cycle history under a different set of rules (fixed N-day
    # windows vs. calendar months), which can desync cycles_posted/days_accrued
    # from what's actually in the Expense book. Once it's running, the choice
    # is locked — delete and recreate instead if a different schedule is
    # really needed.
    if fx.id and cycle_type != (fx.cycle_type or 'fixed_days'):
        if int(fx.cycles_posted or 0) > 0 or int(fx.days_accrued or 0) > 0:
            return ('Cycle type can\'t be changed after cycles have already run — '
                    'delete and recreate this Fixed Expense if you need a different reset schedule.')

    try:
        amount = float(form.get('amount') or 0)
    except (TypeError, ValueError):
        return 'Amount must be a number.'
    if amount <= 0:
        return 'Amount must be greater than zero.'

    try:
        days = int(form.get('days') or 0)
    except (TypeError, ValueError):
        return 'Days must be a whole number.'
    if days < 1:
        return 'Days must be at least 1.'

    start_raw = (form.get('start_date') or '').strip()
    if start_raw:
        try:
            start_date = datetime.strptime(start_raw, '%Y-%m-%d').date()
        except ValueError:
            return 'Start date is invalid.'
    else:
        start_date = datetime.utcnow().date()

    category_id = form.get('category_id') or None

    fx.name = name
    fx.description = (form.get('description') or '').strip() or None
    fx.category_id = int(category_id) if category_id else None
    # Vendor is not part of the popup form; only touch it when a value is
    # actually submitted, so editing never silently clears an existing one.
    if 'vendor_id' in form:
        vendor_id = form.get('vendor_id') or None
        fx.vendor_id = int(vendor_id) if vendor_id else None
    fx.mode = mode
    fx.cycle_type = cycle_type
    fx.amount = amount
    fx.days = days
    fx.start_date = start_date
    fx.is_active = str(form.get('is_active', '1')).lower() in ('1', 'true', 'on', 'yes')
    if 'auto_post' in form:
        fx.auto_post = str(form.get('auto_post')).lower() in ('1', 'true', 'on', 'yes')
    # Account/payment method are optional, applied to every cycle's Expense
    # row as it posts (see ensure_fixed_expense_rows/post_fixed_expense) —
    # never touching cycles already in the book. Uses the Expense module's own
    # ExpenseAccount, not fx.account_id (legacy, read-only Journal link).
    account_id = form.get('account_id') or None
    fx.expense_account_id = int(account_id) if account_id else None
    fx.payment_method = (form.get('payment_method') or '').strip() or None
    return None


@bp.route('/fixed-expenses', methods=['GET'])
@login_required
@permission_required('accounting', action='view')
def fixed_expenses_list():
    """Accrue up to today, then return every template as JSON for the popup."""
    from app.models import FixedExpense
    _sync_all_fixed_expenses()
    items = FixedExpense.query.order_by(FixedExpense.is_active.desc(),
                                        FixedExpense.name).all()
    return jsonify({
        'success': True,
        'items': [_fixed_expense_json(f) for f in items],
        'today': datetime.utcnow().date().strftime('%Y-%m-%d'),
    })


@bp.route('/fixed-expenses/create', methods=['POST'])
@login_required
@permission_required('accounting', action='add')
def create_fixed_expense():
    from app.models import FixedExpense
    fx = FixedExpense(created_by=current_user.id)
    err = _read_fixed_expense_form(fx, request.form)
    if err:
        return jsonify({'success': False, 'message': err}), 400

    bill_path = _save_fixed_expense_bill_image(request.files.get('bill_image'))
    if bill_path:
        fx.bill_image_path = bill_path

    db.session.add(fx)
    db.session.commit()
    fx.sync_accrual()
    db.session.commit()
    if fx.is_active:
        # Without this, a brand-new active+auto_post template sits with no
        # cycle posted until something else happens to call
        # ensure_fixed_expense_rows() (visiting Expenses/Dashboard/Reports,
        # or toggling it off and on) — it looked like the cycle "hadn't
        # started" even though the template itself was created correctly.
        ensure_fixed_expense_rows()

    log_activity('Accounting', 'Created Fixed Expense: ' + fx.name,
                 '%s - %s over %s day(s)' % (fx.mode, fx.amount, fx.days))
    return jsonify({'success': True,
                    'message': 'Fixed expense "%s" created.' % fx.name,
                    'item': _fixed_expense_json(fx)})


@bp.route('/fixed-expenses/<int:fx_id>/update', methods=['POST'])
@login_required
@permission_required('accounting', action='edit')
def update_fixed_expense(fx_id):
    from app.models import FixedExpense
    fx = FixedExpense.query.get_or_404(fx_id)
    was_active = fx.is_active

    err = _read_fixed_expense_form(fx, request.form)
    if err:
        return jsonify({'success': False, 'message': err}), 400

    # Only replace the bill image if a new one was actually chosen — leaves
    # the existing one alone otherwise, same as Add/Edit Expense.
    bill_path = _save_fixed_expense_bill_image(request.files.get('bill_image'))
    if bill_path:
        fx.bill_image_path = bill_path

    # A stop trims the open cycle; a resume starts a fresh cycle from today
    # rather than back-charging the time the template spent switched off.
    today = datetime.utcnow().date()
    if fx.is_active and not was_active:
        _resume_fixed_expense(fx, today)
    elif was_active and not fx.is_active:
        _stop_fixed_expense(fx, today)

    db.session.commit()
    fx.sync_accrual()
    db.session.commit()
    if fx.is_active:
        ensure_fixed_expense_rows()

    log_activity('Accounting', 'Updated Fixed Expense: ' + fx.name,
                 '%s - %s over %s day(s)' % (fx.mode, fx.amount, fx.days))
    return jsonify({'success': True,
                    'message': '"%s" updated.' % fx.name,
                    'item': _fixed_expense_json(fx)})


@bp.route('/fixed-expenses/<int:fx_id>/toggle', methods=['POST'])
@login_required
@permission_required('accounting', action='edit')
def toggle_fixed_expense(fx_id):
    from app.models import FixedExpense
    fx = FixedExpense.query.get_or_404(fx_id)

    today = datetime.utcnow().date()
    if fx.is_active:
        fx.sync_accrual()          # bank the days earned before pausing
        fx.is_active = False
        _stop_fixed_expense(fx, today)
        state = 'inactive'
    else:
        fx.is_active = True
        _resume_fixed_expense(fx, today)
        state = 'active'

    db.session.commit()
    if fx.is_active:
        ensure_fixed_expense_rows()
    log_activity('Accounting', 'Fixed Expense set %s: %s' % (state, fx.name))
    return jsonify({'success': True,
                    'message': '"%s" is now %s.' % (fx.name, state),
                    'item': _fixed_expense_json(fx)})


@bp.route('/fixed-expenses/<int:fx_id>/post', methods=['POST'])
@login_required
@permission_required('accounting', action='add')
def post_fixed_expense(fx_id):
    """Turn the accrued balance into one real Expense record."""
    from app.models import FixedExpense
    fx = FixedExpense.query.get_or_404(fx_id)
    if fx.auto_post:
        return jsonify({'success': False,
                        'message': 'This fixed expense is added to Expenses automatically — '
                                   'no manual posting needed.'}), 400
    fx.sync_accrual()

    pending = round(float(fx.accrued_amount or 0), 2)
    if pending <= 0:
        db.session.commit()
        return jsonify({'success': False,
                        'message': 'Nothing has accrued yet for this fixed expense.'}), 400

    if not fx.category_id:
        return jsonify({'success': False,
                        'message': 'Set a category on this fixed expense before posting.'}), 400

    settings = ExpenseSettings.query.first()
    if not settings:
        settings = ExpenseSettings(expense_prefix='EXP-', expense_suffix='', next_number=1)
        db.session.add(settings)
        db.session.flush()

    expense_number, next_num = get_unique_expense_number(settings, settings.next_number)
    settings.next_number = next_num

    is_admin = getattr(current_user, 'is_admin', False)
    days_note = '%s day(s) @ %s/day' % (fx.days_accrued, round(fx.per_day_amount, 2))
    expense = Expense(
        expense_number=expense_number,
        date=datetime.utcnow(),
        amount=pending,
        description='%s - fixed expense (%s)' % (fx.name, days_note),
        category_id=fx.category_id,
        vendor_id=fx.vendor_id,
        payment_method=fx.payment_method,
        bill_image_path=fx.bill_image_path,
        status='confirmed' if is_admin else 'pending',
        created_by=current_user.id,
        is_approved=is_admin,
        approved_by=current_user.id if is_admin else None,
        approved_at=datetime.utcnow() if is_admin else None,
        is_rejected=False,
        fixed_expense_id=fx.id,
    )
    db.session.add(expense)

    fx.posted_amount = float(fx.posted_amount or 0) + pending
    fx.accrued_amount = 0

    # Charge this posted amount against the linked Expense Account, if any.
    if fx.expense_account_id:
        db.session.flush()
        _sync_expense_account_transaction(expense, fx.expense_account_id, is_admin)

    db.session.commit()

    log_activity('Accounting', 'Posted Fixed Expense: ' + fx.name,
                 '%s - PKR %s (%s)' % (expense_number, pending, days_note))
    return jsonify({'success': True,
                    'message': 'Posted PKR {:,.2f} as {}.'.format(pending, expense_number),
                    'item': _fixed_expense_json(fx)})


@bp.route('/fixed-expenses/<int:fx_id>/delete', methods=['POST'])
@login_required
@permission_required('accounting', action='delete')
def delete_fixed_expense(fx_id):
    """Remove a template. Expenses already posted from it are left untouched."""
    from app.models import FixedExpense
    fx = FixedExpense.query.get_or_404(fx_id)
    name = fx.name
    # Detach the expenses it produced — they stay in the book on their own.
    Expense.query.filter_by(fixed_expense_id=fx.id).update({'fixed_expense_id': None})
    db.session.delete(fx)
    db.session.commit()
    log_activity('Accounting', 'Deleted Fixed Expense: ' + name)
    return jsonify({'success': True,
                    'message': '"%s" removed. Expenses already posted from it are unchanged.' % name})
